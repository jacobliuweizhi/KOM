# -*- coding: utf-8 -*-
from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


FINAL = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用\最终版本\KOM_投稿最终唯一总表_RECOVERED_FULL_20260610.xlsx")
EVIDENCE_DIR = FINAL.parent / "recovery_evidence_20260610"

REQUIRED_SHEETS = [
    "00_README_RECOVERED_FULL",
    "01_SOURCE_FILES",
    "02_FILE_DISCOVERY_MANIFEST",
    "03_SCRIPT_RECOVERY_METRICS",
    "04_CORE_NUMBERS_LOCKED",
    "05_MODULE_NAME_LOCK",
    "06_METHODS_PARAMETER_LOCK",
    "07_RESULTS_SENTENCE_LOCK",
    "08_MISSING_DATA_AUDIT",
    "09_FIGURE_READINESS",
    "10_DO_NOT_DELETE",
    "11_SOURCE_SHEET_INDEX",
    "12_QC_FINAL_STATUS",
    "RAW_Master_Long_Table",
    "RAW_KOMSim_TaskLevel",
    "RAW_Doctor_Prescription",
    "RAW_Expert_Prescription",
    "RAW_Expert_ICC",
    "RAW_Expert_Scores",
    "RAW_Treat_AblationScores",
    "RAW_Treat_RuleMetrics",
    "RAW_RAG_QueryLevel",
    "RAW_RAG_GoldDetail",
    "RAW_RAG_ManualReview",
    "RAW_RAG_Benchmark",
    "RAW_Evidence_PackCheck",
    "RAW_Evidence_Mapping",
    "RAW_OAKNet_Ablation",
    "RAW_OAKNet_Backbone",
    "RAW_Safety_Error_Log",
    "RAW_Case_Basic",
    "RAW_Patient_Fields",
    "RAW_Missingness",
    "RISK_metrics_latest",
]

REQUIRED_EVIDENCE_FILES = [
    "01_file_discovery/all_KOM_candidate_files.csv",
    "01_file_discovery/large_excel_candidates.csv",
    "01_file_discovery/zip_package_candidates.csv",
    "01_file_discovery/script_candidates.csv",
    "01_file_discovery/recycle_bin_check_report.md",
    "02_script_parsing/script_extracted_constants.csv",
    "02_script_parsing/script_extracted_paths.csv",
    "02_script_parsing/script_extracted_metrics.csv",
    "02_script_parsing/script_extracted_sheet_names.csv",
    "02_script_parsing/script_parsing_report.md",
    "QC_recovered_full_workbook.json",
    "QC_recovered_full_workbook.md",
]


def read_rows(wb, sheet_name: str, limit: int = 40):
    if sheet_name not in wb.sheetnames:
        return []
    out = []
    for i, row in enumerate(wb[sheet_name].iter_rows(values_only=True), start=1):
        out.append([v for v in row])
        if i >= limit:
            break
    return out


def main() -> None:
    sha256 = hashlib.sha256(FINAL.read_bytes()).hexdigest()
    wb = load_workbook(FINAL, read_only=True, data_only=True)
    sheets = wb.sheetnames
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in sheets]
    row_counts = {sheet: wb[sheet].max_row for sheet in REQUIRED_SHEETS if sheet in sheets}
    evidence_files = []
    for rel in REQUIRED_EVIDENCE_FILES:
        path = EVIDENCE_DIR / rel
        evidence_files.append(
            {
                "relative_path": rel,
                "exists": path.exists(),
                "size_bytes": path.stat().st_size if path.exists() else None,
            }
        )
    report = {
        "final_excel": str(FINAL),
        "exists": FINAL.exists(),
        "size_mb": round(FINAL.stat().st_size / 1024 / 1024, 3),
        "sha256": sha256,
        "sheet_count": len(sheets),
        "required_missing": missing,
        "required_row_counts": row_counts,
        "evidence_files_missing": [item for item in evidence_files if not item["exists"]],
        "evidence_files": evidence_files,
        "core_numbers_rows": read_rows(wb, "04_CORE_NUMBERS_LOCKED", 80),
        "risk_metrics_rows": read_rows(wb, "RISK_metrics_latest", 40),
        "rag_sample_rows": read_rows(wb, "RAG_query_level", 8) if "RAG_query_level" in sheets else read_rows(wb, "RAW_RAG_QueryLevel", 8),
        "qc_status_rows": read_rows(wb, "12_QC_FINAL_STATUS", 30),
    }
    wb.close()
    out = EVIDENCE_DIR / "independent_recovered_full_audit_20260610.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "size_mb": report["size_mb"],
                "sha256": sha256,
                "sheet_count": len(sheets),
                "required_missing_count": len(missing),
                "required_missing": missing,
                "evidence_missing_count": len(report["evidence_files_missing"]),
                "audit_json": str(out),
                "key_row_counts": row_counts,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
