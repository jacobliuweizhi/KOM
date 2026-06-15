from __future__ import annotations

import csv
import hashlib
import json
import os
import platform
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改")
SUBMISSION_DIR = PROJECT_ROOT / "投稿使用"
FINAL_DIR = SUBMISSION_DIR / "最终版本"
FINAL_XLSX = FINAL_DIR / "KOM_投稿最终唯一总表_LOCKED_20260610.xlsx"
ARCHIVE_ROOT = SUBMISSION_DIR / "原始数据与可复现归档_LOCKED_20260610"
SCRIPT_PATH = PROJECT_ROOT / "patch_raw_data_preservation_archive_20260610.py"

RISK_LOCK_DIR = PROJECT_ROOT / "KOMRisk_Final_Locked_PostDedup_20260610"
RISK_LOCK_ZIP = RISK_LOCK_DIR / "13_final_zip" / "KOMRisk_Final_Locked_PostDedup_20260610.zip"
RISK_FORMAL_RETRAIN_ZIP = PROJECT_ROOT / "KOMRisk_Formal_Retrain_FINAL_20260610.zip"
SOURCE_MASTER = SUBMISSION_DIR / "KOM_项目所有数据_最终总表_20260610.xlsx"

OAI_RAW_PATHS = [
    Path(r"C:\OAI研究项目\pythonProject1\OAI数据库\OAICompleteData_CSV"),
    Path(r"C:\OAI研究项目\pythonProject1\OAI数据库\OAICompleteData_SAS"),
    Path(r"C:\OAI数据库\OAICompleteData_CSV"),
    Path(r"C:\OAI数据库\OAICompleteData_SAS"),
]

SUBDIRS = [
    "00_README",
    "01_raw_data_manifest",
    "02_original_KOMSim_doctor_records",
    "03_original_expert_scoring",
    "04_original_KOMTreat_RAG_ablation",
    "05_original_KOMRisk_locked_packages",
    "06_original_RAG_retrieval_outputs",
    "07_figures_editable_and_png",
    "08_scripts_and_environment",
    "09_intermediate_final_outputs_kept",
    "10_data_availability_and_submission_notes",
    "99_QC_and_checksums",
]

NEW_SHEETS = [
    "29_RAW_DATA_ARCHIVE_INDEX",
    "30_REPRODUCIBILITY_ARCHIVE_INDEX",
    "31_DATA_AVAILABILITY_STATEMENT",
    "32_DO_NOT_DELETE_LIST",
    "33_SUBMISSION_SUPPORT_FILES",
]

FINAL_SHEET_ORDER = [
    "00_README",
    "01_FINAL_MASTER_INDEX",
    "02_PROJECT_CORE_NUMBERS",
    "03_MODULE_NAME_LOCK",
    "04_TOP_JOURNAL_METHOD_CHECKLIST",
    "05_STANDARDIZED_CASES",
    "06_KOM_PROFILE",
    "07_KOM_RAD_OAKNET",
    "08_KOM_RISK_LATEST",
    "09_KOM_KB",
    "10_KOM_RAG_COMPLETION",
    "11_KOM_MDT_RX_SAFE",
    "12_KOM_SCORE_EXPERT",
    "13_KOM_SCORE_RULE",
    "14_KOM_SCORE_ERROR",
    "15_KOM_SIM_CLINICIAN",
    "16_KOM_TREAT_ABLATION",
    "17_FINAL_MODEL_PERFORMANCE",
    "18_MAIN_RESULTS_LOCKED",
    "19_METHODS_PARAMETERS",
    "20_RESULTS_TEXT_BLOCKS",
    "21_RAG_MISSING_ITEMS",
    "22_MODULE_GAPS",
    "23_NUMERIC_TRACEABILITY",
    "24_SOURCE_FILE_INDEX",
    "25_VERSION_SELECTION_LOG",
    "26_OLD_VERSION_ARCHIVE_LOG",
    "27_FIGURE_TABLE_PLAN",
    "28_MANUSCRIPT_WRITING_LOCK",
    "29_RAW_DATA_ARCHIVE_INDEX",
    "30_REPRODUCIBILITY_ARCHIVE_INDEX",
    "31_DATA_AVAILABILITY_STATEMENT",
    "32_DO_NOT_DELETE_LIST",
    "33_SUBMISSION_SUPPORT_FILES",
    "99_QC_FINAL_STATUS",
]


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_dirs() -> None:
    ARCHIVE_ROOT.mkdir(parents=True, exist_ok=True)
    for name in SUBDIRS:
        (ARCHIVE_ROOT / name).mkdir(parents=True, exist_ok=True)


def sha256_file(path: Path, max_bytes: int | None = None) -> str:
    h = hashlib.sha256()
    read = 0
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
            read += len(chunk)
            if max_bytes is not None and read >= max_bytes:
                break
    if max_bytes is not None and path.stat().st_size > max_bytes:
        return f"partial_sha256_first_{max_bytes}_bytes:{h.hexdigest()}"
    return h.hexdigest()


def file_type_summary(path: Path) -> str:
    if not path.exists():
        return "path_not_found"
    if path.is_file():
        return path.suffix.lower() or "no_extension"
    counts: dict[str, int] = {}
    for root, _, files in os.walk(path):
        for name in files:
            suffix = Path(name).suffix.lower() or "no_extension"
            counts[suffix] = counts.get(suffix, 0) + 1
    if not counts:
        return "no_files"
    return "; ".join(f"{k}:{v}" for k, v in sorted(counts.items())[:30])


def dir_file_count_size(path: Path) -> tuple[int, int]:
    if not path.exists():
        return 0, 0
    if path.is_file():
        return 1, path.stat().st_size
    count = 0
    total = 0
    for root, _, files in os.walk(path):
        for name in files:
            p = Path(root) / name
            try:
                count += 1
                total += p.stat().st_size
            except OSError:
                pass
    return count, total


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({col: explicit(row.get(col)) for col in columns})


def explicit(value: Any) -> Any:
    if value is None:
        return "not_available"
    if isinstance(value, str) and not value.strip():
        return "not_available"
    return value


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def copy_or_index_file(src: Path, dst_dir: Path, purpose: str, max_copy_mb: int = 150) -> dict[str, Any]:
    src = src.resolve()
    dst_dir.mkdir(parents=True, exist_ok=True)
    if not src.exists():
        return {
            "module": purpose,
            "file_or_folder": src.name,
            "relative_or_absolute_path": str(src),
            "copied_or_indexed": "path_not_found",
            "sha256": "not_applicable",
            "purpose": purpose,
            "delete_allowed": False,
            "notes": "source path not found",
        }
    if src.is_dir():
        return {
            "module": purpose,
            "file_or_folder": src.name,
            "relative_or_absolute_path": str(src),
            "copied_or_indexed": "indexed_directory_not_copied",
            "sha256": "directory_manifest_only",
            "purpose": purpose,
            "delete_allowed": False,
            "notes": "directory is preserved in place; manifest records path only",
        }
    size_mb = src.stat().st_size / (1024 * 1024)
    digest = sha256_file(src)
    if size_mb <= max_copy_mb:
        target = dst_dir / src.name
        if target.exists() and sha256_file(target) == digest:
            status = "copied_existing_same_sha256"
        else:
            shutil.copy2(src, target)
            status = "copied"
        return {
            "module": purpose,
            "file_or_folder": src.name,
            "relative_or_absolute_path": str(target),
            "copied_or_indexed": status,
            "sha256": digest,
            "purpose": purpose,
            "delete_allowed": False,
            "notes": f"source={src}; size_mb={size_mb:.2f}",
        }
    return {
        "module": purpose,
        "file_or_folder": src.name,
        "relative_or_absolute_path": str(src),
        "copied_or_indexed": "indexed_large_file_not_copied",
        "sha256": digest,
        "purpose": purpose,
        "delete_allowed": False,
        "notes": f"large file indexed in place; size_mb={size_mb:.2f}",
    }


def copytree_or_index(src: Path, dst_dir: Path, purpose: str, max_total_mb: int = 800) -> dict[str, Any]:
    src = src.resolve()
    dst_dir.mkdir(parents=True, exist_ok=True)
    if not src.exists():
        return {
            "module": purpose,
            "file_or_folder": src.name,
            "relative_or_absolute_path": str(src),
            "copied_or_indexed": "path_not_found",
            "sha256": "not_applicable",
            "purpose": purpose,
            "delete_allowed": False,
            "notes": "source folder not found",
        }
    count, size = dir_file_count_size(src)
    size_mb = size / (1024 * 1024)
    target = dst_dir / src.name
    try:
        archive_root_resolved = ARCHIVE_ROOT.resolve()
        target_resolved = target.resolve()
        if target.exists() and str(target_resolved).startswith(str(archive_root_resolved)):
            shutil.rmtree(target)
    except OSError:
        pass
    return {
        "module": purpose,
        "file_or_folder": src.name,
        "relative_or_absolute_path": str(src),
        "copied_or_indexed": "indexed_directory_not_copied_windows_long_path_safe",
        "sha256": "directory_manifest_only",
        "purpose": purpose,
        "delete_allowed": False,
        "notes": f"directory preserved in original location and indexed to avoid Windows long-path copy failures; files={count}; size_mb={size_mb:.2f}",
    }


def safe_project_files() -> list[Path]:
    skip_parts = {
        "原始数据与可复现归档_LOCKED_20260610",
        "__pycache__",
        ".git",
        ".venv",
        "node_modules",
    }
    files: list[Path] = []
    for root, dirs, names in os.walk(PROJECT_ROOT):
        root_path = Path(root)
        dirs[:] = [d for d in dirs if d not in skip_parts and not d.startswith(".")]
        if any(part in skip_parts for part in root_path.parts):
            continue
        for name in names:
            files.append(root_path / name)
    return files


def find_by_keywords(files: list[Path], keywords: list[str], exts: set[str] | None = None, limit: int = 80) -> list[Path]:
    out: list[Path] = []
    lower_keywords = [k.lower() for k in keywords]
    for p in files:
        if exts and p.suffix.lower() not in exts:
            continue
        hay = str(p).lower()
        if any(k in hay for k in lower_keywords):
            out.append(p)
        if len(out) >= limit:
            break
    return out


def raw_data_manifest() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    for path in OAI_RAW_PATHS:
        count, total = dir_file_count_size(path)
        rows.append(
            {
                "data_category": "OAI_original_raw_data",
                "path": str(path),
                "exists": path.exists(),
                "file_count": count,
                "total_size_gb": round(total / (1024**3), 4),
                "file_type_summary": file_type_summary(path),
                "sha256_strategy": "directory_manifest_only_not_full_hash_due_to_raw_data_size",
                "is_original_data": True,
                "delete_allowed": False,
                "submission_allowed": False,
                "notes": "Do not move or delete. This workbook stores only the path-level manifest.",
            }
        )
    write_csv(
        ARCHIVE_ROOT / "01_raw_data_manifest" / "OAI_raw_data_manifest.csv",
        rows,
        [
            "data_category",
            "path",
            "exists",
            "file_count",
            "total_size_gb",
            "file_type_summary",
            "sha256_strategy",
            "is_original_data",
            "delete_allowed",
            "submission_allowed",
            "notes",
        ],
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "OAI_raw_data_summary"
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([explicit(r.get(c)) for c in cols])
    style_sheet(ws)
    wb.save(ARCHIVE_ROOT / "01_raw_data_manifest" / "OAI_raw_data_summary.xlsx")
    write_text(
        ARCHIVE_ROOT / "01_raw_data_manifest" / "OAI_raw_data_DO_NOT_DELETE.md",
        "# OAI raw data - DO NOT DELETE\n\n"
        "The OAI raw data paths listed in `OAI_raw_data_manifest.csv` are original or controlled-access data sources. "
        "They are indexed only and must not be deleted, moved, or redistributed with the manuscript unless the data use agreement explicitly permits it.\n",
    )
    reproducibility_rows = [
        {
            "module": "OAI_raw_data",
            "file_or_folder": Path(r["path"]).name,
            "relative_or_absolute_path": r["path"],
            "copied_or_indexed": "indexed_path_only" if r["exists"] else "path_not_found",
            "sha256": r["sha256_strategy"],
            "purpose": "original OAI source data manifest",
            "delete_allowed": False,
            "notes": r["notes"],
        }
        for r in rows
    ]
    return rows, reproducibility_rows


def build_manifests() -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    ensure_dirs()
    files = safe_project_files()
    raw_rows, repro_rows = raw_data_manifest()

    # KOMSim doctor records
    komsim_files = find_by_keywords(
        files,
        ["komsim", "医生", "doctor", "clinician", "human interaction", "prescription", "780", "clinician + kom", "clinician alone"],
        {".xlsx", ".xls", ".csv", ".json", ".jsonl", ".md", ".txt", ".docx"},
        limit=120,
    )
    if SOURCE_MASTER.exists() and SOURCE_MASTER not in komsim_files:
        komsim_files.append(SOURCE_MASTER)
    komsim_manifest: list[dict[str, Any]] = []
    for p in komsim_files:
        komsim_manifest.append(copy_or_index_file(p, ARCHIVE_ROOT / "02_original_KOMSim_doctor_records", "KOMSim_doctor_records"))
    write_csv(
        ARCHIVE_ROOT / "02_original_KOMSim_doctor_records" / "KOMSim_doctor_records_manifest.csv",
        komsim_manifest,
        ["module", "file_or_folder", "relative_or_absolute_path", "copied_or_indexed", "sha256", "purpose", "delete_allowed", "notes"],
    )
    write_text(
        ARCHIVE_ROOT / "02_original_KOMSim_doctor_records" / "KOMSim_doctor_records_summary.md",
        "# KOMSim doctor records archive summary\n\n"
        "- final_clinician_n = 26\n"
        "- final_prescription_records = 780\n"
        "- raw_records_delete_allowed = FALSE\n"
        f"- candidate_files_indexed_or_copied = {len(komsim_manifest)}\n"
        "- If raw per-clinician records are not among copied files, this remains WARNING and requires manual verification.\n",
    )
    repro_rows.extend(komsim_manifest)

    # Expert scoring
    expert_files = find_by_keywords(
        files,
        ["expert", "专家", "评分", "盲法", "icc", "kom-score", "orthopaedics", "sports medicine", "rehabilitation"],
        {".xlsx", ".xls", ".csv", ".json", ".jsonl", ".md", ".txt", ".docx"},
        limit=120,
    )
    expert_manifest = [copy_or_index_file(p, ARCHIVE_ROOT / "03_original_expert_scoring", "expert_scoring") for p in expert_files]
    write_csv(
        ARCHIVE_ROOT / "03_original_expert_scoring" / "expert_scoring_manifest.csv",
        expert_manifest,
        ["module", "file_or_folder", "relative_or_absolute_path", "copied_or_indexed", "sha256", "purpose", "delete_allowed", "notes"],
    )
    if not expert_manifest:
        write_text(
            ARCHIVE_ROOT / "03_original_expert_scoring" / "expert_raw_score_table_not_found_WARNING.md",
            "# WARNING: expert raw score table not found\n\n"
            "The locked summary includes six-expert blinded ICC results, but the raw expert scoring table was not found during the current archive scan. "
            "Set `requires_raw_expert_score_table_verification = TRUE` in the final workbook.\n",
        )
    repro_rows.extend(expert_manifest)

    # KOMTreat / RAG ablation
    treat_files = find_by_keywords(
        files,
        ["komtreat", "ablation", "rag消融", "no_rag", "no_mdt", "direct llm", "A_full", "B_no", "C_no", "D_bare"],
        {".xlsx", ".xls", ".csv", ".json", ".jsonl", ".md", ".txt", ".py"},
        limit=100,
    )
    treat_manifest = [copy_or_index_file(p, ARCHIVE_ROOT / "04_original_KOMTreat_RAG_ablation", "KOMTreat_RAG_ablation") for p in treat_files]
    write_csv(
        ARCHIVE_ROOT / "04_original_KOMTreat_RAG_ablation" / "KOMTreat_RAG_ablation_manifest.csv",
        treat_manifest,
        ["module", "file_or_folder", "relative_or_absolute_path", "copied_or_indexed", "sha256", "purpose", "delete_allowed", "notes"],
    )
    repro_rows.extend(treat_manifest)

    # KOMRisk locked packages
    risk_manifest: list[dict[str, Any]] = []
    risk_manifest.append(copytree_or_index(RISK_LOCK_DIR, ARCHIVE_ROOT / "05_original_KOMRisk_locked_packages", "KOMRisk_locked_package"))
    risk_manifest.append(copy_or_index_file(RISK_LOCK_ZIP, ARCHIVE_ROOT / "05_original_KOMRisk_locked_packages", "KOMRisk_locked_package"))
    risk_manifest.append(copy_or_index_file(RISK_FORMAL_RETRAIN_ZIP, ARCHIVE_ROOT / "05_original_KOMRisk_locked_packages", "KOMRisk_locked_package"))
    write_csv(
        ARCHIVE_ROOT / "05_original_KOMRisk_locked_packages" / "KOMRisk_locked_package_manifest.csv",
        risk_manifest,
        ["module", "file_or_folder", "relative_or_absolute_path", "copied_or_indexed", "sha256", "purpose", "delete_allowed", "notes"],
    )
    risk_report = "# KOMRisk locked package integrity report\n\n"
    risk_report += f"- KOMRisk_Final_Locked_PostDedup_20260610.zip exists: {RISK_LOCK_ZIP.exists()}\n"
    risk_report += f"- KOMRisk_Final_Locked_PostDedup_20260610 folder exists: {RISK_LOCK_DIR.exists()}\n"
    risk_report += f"- KOMRisk_Formal_Retrain_FINAL_20260610.zip exists: {RISK_FORMAL_RETRAIN_ZIP.exists()}\n"
    risk_report += "- FINAL_LOCKED_ACCEPTED: expected from locked package metadata / previous master.\n"
    risk_report += "- KXR READPRJ dedup PASS: expected from locked package metadata / previous master.\n"
    risk_report += "- SIDE mapping PASS: expected from locked package metadata / previous master.\n"
    risk_report += "- Endpoint B fixed-horizon binary: locked.\n"
    risk_report += "- Endpoint C ACCEPT_SUPPLEMENT: locked.\n"
    write_text(ARCHIVE_ROOT / "05_original_KOMRisk_locked_packages" / "KOMRisk_locked_package_integrity_report.md", risk_report)
    repro_rows.extend(risk_manifest)

    # RAG original outputs
    rag_files = find_by_keywords(
        files,
        ["kom-rag", "graphrag", "retrieval", "query", "gold", "topk", "top_k", "ndcg", "mrr", "precision", "recall", "hit@10", "citation", "faithfulness", "grounding", "stage4a", "stage3_retrieval"],
        {".xlsx", ".xls", ".csv", ".json", ".jsonl", ".md", ".txt", ".py"},
        limit=180,
    )
    rag_manifest = [copy_or_index_file(p, ARCHIVE_ROOT / "06_original_RAG_retrieval_outputs", "RAG_original_outputs") for p in rag_files]
    write_csv(
        ARCHIVE_ROOT / "06_original_RAG_retrieval_outputs" / "RAG_original_outputs_manifest.csv",
        rag_manifest,
        ["module", "file_or_folder", "relative_or_absolute_path", "copied_or_indexed", "sha256", "purpose", "delete_allowed", "notes"],
    )
    rag_names = " ".join(p.name.lower() for p in rag_files)
    missing_rag = []
    for label, keys in {
        "query benchmark": ["query"],
        "gold labels": ["gold"],
        "top-k run outputs": ["topk", "top_k"],
        "generation grounding metrics": ["grounding", "faithfulness"],
    }.items():
        if not any(k in rag_names for k in keys):
            missing_rag.append(label)
    write_text(
        ARCHIVE_ROOT / "06_original_RAG_retrieval_outputs" / "RAG_missing_original_outputs_report.md",
        "# RAG original outputs missing report\n\n"
        f"- candidate_files_indexed_or_copied = {len(rag_manifest)}\n"
        f"- missing_or_unverified_items = {', '.join(missing_rag) if missing_rag else 'none_detected_by_filename_scan'}\n"
        "- This report is filename/path based; absence here should be manually checked before manuscript submission.\n",
    )
    repro_rows.extend(rag_manifest)

    # Figures
    fig_files = find_by_keywords(
        files,
        ["figure", "fig", "图", "nature_figures", "pptx", "svg", "drawio"],
        {".png", ".tif", ".tiff", ".svg", ".pptx", ".ppt", ".ai", ".fig", ".drawio", ".pdf", ".xlsx", ".md"},
        limit=250,
    )
    fig_manifest = [copy_or_index_file(p, ARCHIVE_ROOT / "07_figures_editable_and_png", "figures_editable_and_png", max_copy_mb=60) for p in fig_files]
    write_csv(
        ARCHIVE_ROOT / "07_figures_editable_and_png" / "figure_source_manifest.csv",
        fig_manifest,
        ["module", "file_or_folder", "relative_or_absolute_path", "copied_or_indexed", "sha256", "purpose", "delete_allowed", "notes"],
    )
    missing_editable = []
    png_stems = {p.stem.lower() for p in fig_files if p.suffix.lower() == ".png"}
    editable_stems = {p.stem.lower() for p in fig_files if p.suffix.lower() in {".pptx", ".ppt", ".svg", ".ai", ".fig", ".drawio"}}
    for stem in sorted(png_stems - editable_stems):
        missing_editable.append({"figure_stem": stem, "missing_editable_source": True, "notes": "PNG found but no same-stem editable source found by filename scan"})
    write_csv(
        ARCHIVE_ROOT / "07_figures_editable_and_png" / "missing_editable_figure_sources.csv",
        missing_editable or [{"figure_stem": "none_detected", "missing_editable_source": False, "notes": "No missing same-stem editable sources detected by filename scan"}],
        ["figure_stem", "missing_editable_source", "notes"],
    )
    repro_rows.extend(fig_manifest)

    # Scripts and environment
    script_files = find_by_keywords(
        files,
        ["komrisk", "master", "rag", "komsim", "expert", "icc", "score", "make_", "patch_", "train", "lock", "evaluate"],
        {".py", ".ps1", ".bat", ".cmd", ".yaml", ".yml", ".toml", ".txt", ".md", ".json"},
        limit=220,
    )
    if SCRIPT_PATH not in script_files:
        script_files.append(SCRIPT_PATH)
    scripts_manifest = [copy_or_index_file(p, ARCHIVE_ROOT / "08_scripts_and_environment", "scripts_and_environment", max_copy_mb=20) for p in script_files]
    write_csv(
        ARCHIVE_ROOT / "08_scripts_and_environment" / "scripts_manifest.csv",
        scripts_manifest,
        ["module", "file_or_folder", "relative_or_absolute_path", "copied_or_indexed", "sha256", "purpose", "delete_allowed", "notes"],
    )
    pip_freeze = ""
    try:
        pip_freeze = subprocess.check_output([sys.executable, "-m", "pip", "freeze"], text=True, errors="replace", timeout=30)
    except Exception as exc:
        pip_freeze = f"pip_freeze_failed: {exc}"
    write_text(
        ARCHIVE_ROOT / "08_scripts_and_environment" / "environment_snapshot.txt",
        f"generated_at={now()}\n"
        f"python_executable={sys.executable}\n"
        f"python_version={sys.version}\n"
        f"platform={platform.platform()}\n\n"
        "[pip freeze]\n"
        f"{pip_freeze}\n",
    )
    write_text(
        ARCHIVE_ROOT / "08_scripts_and_environment" / "reproducibility_README.md",
        "# Reproducibility archive README\n\n"
        "This folder indexes or copies final scripts, environment information, locked model packages, RAG outputs, scoring materials, and manuscript support files. "
        "Original OAI data are not copied or deleted; only path-level manifests are stored.\n\n"
        "Run command used for this patch:\n\n"
        f"`{sys.executable} {SCRIPT_PATH}`\n",
    )
    repro_rows.extend(scripts_manifest)

    # Intermediate outputs kept
    final_support_candidates = find_by_keywords(
        files,
        ["final", "locked", "总表", "submission", "投稿", "manuscript", "手稿", "方法", "结果"],
        {".xlsx", ".docx", ".md", ".csv", ".json", ".zip"},
        limit=150,
    )
    interm_manifest = [copy_or_index_file(p, ARCHIVE_ROOT / "09_intermediate_final_outputs_kept", "intermediate_final_outputs_kept", max_copy_mb=100) for p in final_support_candidates]
    write_csv(
        ARCHIVE_ROOT / "09_intermediate_final_outputs_kept" / "intermediate_final_outputs_manifest.csv",
        interm_manifest,
        ["module", "file_or_folder", "relative_or_absolute_path", "copied_or_indexed", "sha256", "purpose", "delete_allowed", "notes"],
    )
    repro_rows.extend(interm_manifest)

    # Data availability notes
    cn_statement = (
        "本研究使用的 OAI 原始数据来自公开/受控访问队列，原始数据不随本文重复分发。"
        "研究中生成的分析表、模型输出、评分汇总、代码和处理流程已在本地归档，"
        "并可在符合数据使用协议和伦理要求的前提下按合理请求提供。"
        "医生实验和专家评分数据包含研究参与者信息，需在去标识化和审批后共享。"
    )
    en_statement = (
        "The original OAI data used in this study were obtained from a public or controlled-access cohort and are not redistributed with this manuscript. "
        "Derived analysis tables, model outputs, scoring summaries, code, and processing logs have been archived locally and may be made available upon reasonable request, subject to data use agreements and ethical restrictions. "
        "Clinician interaction and expert scoring data contain study participant information and can only be shared after de-identification and appropriate approval."
    )
    write_text(
        ARCHIVE_ROOT / "10_data_availability_and_submission_notes" / "DATA_AVAILABILITY_STATEMENT_DRAFT.md",
        "# Data availability statement draft\n\n"
        "## 中文\n\n"
        f"{cn_statement}\n\n"
        "## English\n\n"
        f"{en_statement}\n",
    )
    write_text(
        ARCHIVE_ROOT / "10_data_availability_and_submission_notes" / "DO_NOT_DELETE_LIST.md",
        "# DO NOT DELETE LIST\n\n"
        "- OAICompleteData_CSV\n"
        "- OAICompleteData_SAS\n"
        "- KOMSim raw doctor records\n"
        "- expert scoring raw tables\n"
        "- KOMRisk locked packages\n"
        "- RAG original retrieval outputs\n"
        "- final scripts\n"
        "- final Excel\n",
    )

    # Final QC
    status = {
        "generated_at": now(),
        "archive_root": str(ARCHIVE_ROOT),
        "OAI_raw_data_indexed": any(r["exists"] for r in raw_rows),
        "OAI_raw_data_deleted": False,
        "KOMSim_raw_records_indexed_or_copied": bool(komsim_manifest),
        "Expert_raw_scoring_indexed_or_copied": bool(expert_manifest),
        "KOMRisk_locked_package_preserved": RISK_LOCK_ZIP.exists() or RISK_LOCK_DIR.exists(),
        "RAG_raw_outputs_indexed_or_copied": bool(rag_manifest),
        "Scripts_preserved": bool(scripts_manifest),
        "DO_NOT_DELETE_LIST_exists": (ARCHIVE_ROOT / "10_data_availability_and_submission_notes" / "DO_NOT_DELETE_LIST.md").exists(),
        "requires_raw_expert_score_table_verification": not bool(expert_manifest),
        "rag_missing_items": missing_rag,
    }
    final_dir_extra_files = []
    if FINAL_DIR.exists():
        final_dir_extra_files = [
            p.name for p in FINAL_DIR.iterdir()
            if p.name != FINAL_XLSX.name
        ]
    status["final_version_extra_files"] = final_dir_extra_files
    status["final_version_folder_only_final_excel"] = len(final_dir_extra_files) == 0
    overall = "RAW_DATA_AND_REPRODUCIBILITY_ARCHIVE_READY"
    if not status["OAI_raw_data_indexed"] or not status["KOMRisk_locked_package_preserved"]:
        overall = "NEEDS_MANUAL_REVIEW"
    if not status["Expert_raw_scoring_indexed_or_copied"] or not status["RAG_raw_outputs_indexed_or_copied"]:
        overall = "NEEDS_MANUAL_REVIEW"
    status["overall"] = overall
    write_text(ARCHIVE_ROOT / "99_QC_and_checksums" / "archive_status.json", json.dumps(status, ensure_ascii=False, indent=2))
    write_text(
        ARCHIVE_ROOT / "99_QC_and_checksums" / "archive_QC_report.md",
        "# Archive QC report\n\n"
        f"- OAI raw data indexed = {'YES' if status['OAI_raw_data_indexed'] else 'NO'}\n"
        "- OAI raw data deleted = NO\n"
        f"- KOMSim raw records indexed/copied = {'YES' if status['KOMSim_raw_records_indexed_or_copied'] else 'WARNING'}\n"
        f"- Expert raw scoring indexed/copied = {'YES' if status['Expert_raw_scoring_indexed_or_copied'] else 'WARNING'}\n"
        f"- KOMRisk locked package preserved = {'YES' if status['KOMRisk_locked_package_preserved'] else 'NO'}\n"
        f"- RAG raw outputs indexed/copied = {'YES' if status['RAG_raw_outputs_indexed_or_copied'] else 'WARNING'}\n"
        f"- Scripts preserved = {'YES' if status['Scripts_preserved'] else 'WARNING'}\n"
        "- Final Excel updated with archive sheets = pending workbook update\n"
        f"- DO_NOT_DELETE_LIST exists = {'YES' if status['DO_NOT_DELETE_LIST_exists'] else 'NO'}\n"
        f"- Overall = {overall}\n",
    )
    return raw_rows, repro_rows, status


def style_sheet(ws) -> None:
    dark = PatternFill("solid", fgColor="1F4E78")
    white = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = dark
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
                cell.value = "not_available"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = max(len(str(c.value)) for c in ws[letter])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def add_table_sheet(wb, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    if title in wb.sheetnames:
        del wb[title]
    idx = wb.sheetnames.index("99_QC_FINAL_STATUS") if "99_QC_FINAL_STATUS" in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet(title, idx)
    ws.append(columns)
    for r in rows:
        ws.append([explicit(r.get(c)) for c in columns])
    if ws.max_row == 1:
        ws.append(["not_available" for _ in columns])
    style_sheet(ws)


def update_final_excel(raw_rows: list[dict[str, Any]], repro_rows: list[dict[str, Any]], status: dict[str, Any]) -> dict[str, Any]:
    if not FINAL_XLSX.exists():
        raise FileNotFoundError(FINAL_XLSX)
    backup_dir = ARCHIVE_ROOT / "09_intermediate_final_outputs_kept"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f"{FINAL_XLSX.stem}_before_raw_archive_patch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(FINAL_XLSX, backup)

    wb = load_workbook(FINAL_XLSX)
    for s in NEW_SHEETS + ["99_QC_FINAL_STATUS"]:
        if s in wb.sheetnames:
            del wb[s]

    add_table_sheet(
        wb,
        "29_RAW_DATA_ARCHIVE_INDEX",
        [
            {
                "data_category": r["data_category"],
                "path": r["path"],
                "exists": r["exists"],
                "file_count": r["file_count"],
                "size_gb": r["total_size_gb"],
                "is_original_data": r["is_original_data"],
                "delete_allowed": r["delete_allowed"],
                "submission_allowed": r["submission_allowed"],
                "manifest_file": str(ARCHIVE_ROOT / "01_raw_data_manifest" / "OAI_raw_data_manifest.csv"),
                "notes": r["notes"],
            }
            for r in raw_rows
        ],
        ["data_category", "path", "exists", "file_count", "size_gb", "is_original_data", "delete_allowed", "submission_allowed", "manifest_file", "notes"],
    )
    add_table_sheet(
        wb,
        "30_REPRODUCIBILITY_ARCHIVE_INDEX",
        repro_rows,
        ["module", "file_or_folder", "relative_or_absolute_path", "copied_or_indexed", "sha256", "purpose", "delete_allowed", "notes"],
    )
    cn_statement = (
        "本研究使用的 OAI 原始数据来自公开/受控访问队列，原始数据不随本文重复分发。"
        "研究中生成的分析表、模型输出、评分汇总、代码和处理流程已在本地归档，"
        "并可在符合数据使用协议和伦理要求的前提下按合理请求提供。"
        "医生实验和专家评分数据包含研究参与者信息，需在去标识化和审批后共享。"
    )
    en_statement = (
        "The original OAI data used in this study were obtained from a public or controlled-access cohort and are not redistributed with this manuscript. "
        "Derived analysis tables, model outputs, scoring summaries, code, and processing logs have been archived locally and may be made available upon reasonable request, subject to data use agreements and ethical restrictions. "
        "Clinician interaction and expert scoring data contain study participant information and can only be shared after de-identification and appropriate approval."
    )
    add_table_sheet(
        wb,
        "31_DATA_AVAILABILITY_STATEMENT",
        [
            {"language": "Chinese", "statement_type": "draft", "text": cn_statement, "shareability": "requires_data_use_agreement_and_ethics_review", "notes": "Manuscript-ready draft; author must confirm journal wording."},
            {"language": "English", "statement_type": "draft", "text": en_statement, "shareability": "requires_data_use_agreement_and_ethics_review", "notes": "Manuscript-ready draft; author must confirm journal wording."},
        ],
        ["language", "statement_type", "text", "shareability", "notes"],
    )
    do_not_delete = [
        ("OAICompleteData_CSV", "OAI original data", "FALSE", "Original data must never be deleted or moved for space saving."),
        ("OAICompleteData_SAS", "OAI original data", "FALSE", "Original data must never be deleted or moved for space saving."),
        ("KOMSim raw doctor records", "human participant/research records", "FALSE", "Preserve raw or indexed path/checksum."),
        ("expert scoring raw tables", "expert review source data", "FALSE", "Requires manual verification if raw table not found."),
        ("KOMRisk locked packages", "model reproducibility package", "FALSE", "PostDedup lock and formal retrain zip must be preserved."),
        ("RAG original retrieval outputs", "retrieval benchmark/source outputs", "FALSE", "Preserve query/gold/top-k/grounding outputs when available."),
        ("final scripts", "reproducibility scripts", "FALSE", "Do not delete final generation/training/evaluation scripts."),
        ("final Excel", "submission control workbook", "FALSE", "Locked workbook."),
    ]
    add_table_sheet(
        wb,
        "32_DO_NOT_DELETE_LIST",
        [{"path_or_item": a, "category": b, "delete_allowed": c, "notes": d} for a, b, c, d in do_not_delete],
        ["path_or_item", "category", "delete_allowed", "notes"],
    )
    support_rows = [
        {"material": "Final single master Excel", "current_location": str(FINAL_XLSX), "shareable_directly": True, "needs_deidentification": False, "needs_permission": False, "notes": "Submission control table."},
        {"material": "OAI raw data manifest", "current_location": str(ARCHIVE_ROOT / "01_raw_data_manifest"), "shareable_directly": False, "needs_deidentification": False, "needs_permission": True, "notes": "Manifest can be described; raw data not redistributed."},
        {"material": "KOMSim doctor records", "current_location": str(ARCHIVE_ROOT / "02_original_KOMSim_doctor_records"), "shareable_directly": False, "needs_deidentification": True, "needs_permission": True, "notes": "Human participant-related records."},
        {"material": "Expert scoring records", "current_location": str(ARCHIVE_ROOT / "03_original_expert_scoring"), "shareable_directly": False, "needs_deidentification": True, "needs_permission": True, "notes": "Raw table may require manual verification."},
        {"material": "KOMRisk locked package", "current_location": str(ARCHIVE_ROOT / "05_original_KOMRisk_locked_packages"), "shareable_directly": "conditional", "needs_deidentification": False, "needs_permission": "check_license", "notes": "Model package and lock reports."},
        {"material": "RAG retrieval outputs", "current_location": str(ARCHIVE_ROOT / "06_original_RAG_retrieval_outputs"), "shareable_directly": "conditional", "needs_deidentification": False, "needs_permission": "check_source_terms", "notes": "Some outputs may cite third-party evidence records."},
        {"material": "Figures editable and PNG", "current_location": str(ARCHIVE_ROOT / "07_figures_editable_and_png"), "shareable_directly": True, "needs_deidentification": False, "needs_permission": False, "notes": "Check embedded source images before journal upload."},
        {"material": "Scripts and environment", "current_location": str(ARCHIVE_ROOT / "08_scripts_and_environment"), "shareable_directly": True, "needs_deidentification": False, "needs_permission": False, "notes": "Remove local API keys if any before sharing."},
    ]
    add_table_sheet(wb, "33_SUBMISSION_SUPPORT_FILES", support_rows, ["material", "current_location", "shareable_directly", "needs_deidentification", "needs_permission", "notes"])

    qc_rows = [
        {"check": "OAI raw data indexed", "status": "YES" if status["OAI_raw_data_indexed"] else "NO", "details": str(ARCHIVE_ROOT / "01_raw_data_manifest" / "OAI_raw_data_manifest.csv")},
        {"check": "OAI raw data deleted", "status": "NO", "details": "No deletion was performed."},
        {"check": "KOMSim raw records indexed/copied", "status": "YES" if status["KOMSim_raw_records_indexed_or_copied"] else "WARNING", "details": str(ARCHIVE_ROOT / "02_original_KOMSim_doctor_records")},
        {"check": "Expert raw scoring indexed/copied", "status": "YES" if status["Expert_raw_scoring_indexed_or_copied"] else "WARNING", "details": str(ARCHIVE_ROOT / "03_original_expert_scoring")},
        {"check": "requires_raw_expert_score_table_verification", "status": status["requires_raw_expert_score_table_verification"], "details": "TRUE means raw expert table needs manual confirmation."},
        {"check": "KOMRisk locked package preserved", "status": "YES" if status["KOMRisk_locked_package_preserved"] else "NO", "details": str(ARCHIVE_ROOT / "05_original_KOMRisk_locked_packages")},
        {"check": "RAG raw outputs indexed/copied", "status": "YES" if status["RAG_raw_outputs_indexed_or_copied"] else "WARNING", "details": str(ARCHIVE_ROOT / "06_original_RAG_retrieval_outputs")},
        {"check": "Scripts preserved", "status": "YES" if status["Scripts_preserved"] else "WARNING", "details": str(ARCHIVE_ROOT / "08_scripts_and_environment")},
        {"check": "Final Excel updated with archive sheets", "status": "YES", "details": "; ".join(NEW_SHEETS)},
        {"check": "DO_NOT_DELETE_LIST exists", "status": "YES" if status["DO_NOT_DELETE_LIST_exists"] else "NO", "details": str(ARCHIVE_ROOT / "10_data_availability_and_submission_notes" / "DO_NOT_DELETE_LIST.md")},
        {"check": "Overall", "status": status["overall"], "details": str(ARCHIVE_ROOT / "99_QC_and_checksums" / "archive_status.json")},
        {"check": "Final version folder only final Excel", "status": "YES" if status.get("final_version_folder_only_final_excel") else "WARNING", "details": "; ".join(status.get("final_version_extra_files", [])) or "no extra files"},
    ]
    add_table_sheet(wb, "99_QC_FINAL_STATUS", qc_rows, ["check", "status", "details"])

    # Force final order for known sheets.
    ordered = [wb[s] for s in FINAL_SHEET_ORDER if s in wb.sheetnames]
    rest = [ws for ws in wb.worksheets if ws.title not in FINAL_SHEET_ORDER]
    wb._sheets = ordered + rest
    wb.save(FINAL_XLSX)
    wb.close()
    return {"backup": str(backup), "final_excel": str(FINAL_XLSX)}


def final_audit() -> dict[str, Any]:
    wb = load_workbook(FINAL_XLSX, read_only=True, data_only=True)
    sheetnames = wb.sheetnames
    row_counts = {ws.title: ws.max_row for ws in wb.worksheets}
    blanks = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None or (isinstance(c.value, str) and not c.value.strip()):
                    blanks += 1
    wb.close()
    status = json.loads((ARCHIVE_ROOT / "99_QC_and_checksums" / "archive_status.json").read_text(encoding="utf-8"))
    status["Final_Excel_archive_sheets"] = {s: s in sheetnames for s in NEW_SHEETS}
    status["Final_Excel_updated_with_archive_sheets"] = all(status["Final_Excel_archive_sheets"].values())
    status["workbook_sheet_count"] = len(sheetnames)
    status["workbook_blank_cell_count"] = blanks
    status["workbook_sheet_order_prefix_ok"] = sheetnames[: len(FINAL_SHEET_ORDER)] == FINAL_SHEET_ORDER
    status["final_excel_sha256"] = sha256_file(FINAL_XLSX)
    if status["Final_Excel_updated_with_archive_sheets"]:
        report = (ARCHIVE_ROOT / "99_QC_and_checksums" / "archive_QC_report.md").read_text(encoding="utf-8")
        report = report.replace("Final Excel updated with archive sheets = pending workbook update", "Final Excel updated with archive sheets = YES")
        write_text(ARCHIVE_ROOT / "99_QC_and_checksums" / "archive_QC_report.md", report)
    write_text(ARCHIVE_ROOT / "99_QC_and_checksums" / "archive_status.json", json.dumps(status, ensure_ascii=False, indent=2))
    return {"sheetnames": sheetnames, "row_counts": row_counts, "blank_cell_count": blanks, "status": status}


def main() -> None:
    raw_rows, repro_rows, status = build_manifests()
    update_info = update_final_excel(raw_rows, repro_rows, status)
    audit = final_audit()
    print(
        json.dumps(
            {
                "message": "KOM RAW DATA PRESERVATION PATCH COMPLETE",
                "final_excel": str(FINAL_XLSX),
                "archive_folder": str(ARCHIVE_ROOT),
                "update_info": update_info,
                "status": audit["status"],
                "sheet_count": len(audit["sheetnames"]),
                "blank_cell_count": audit["blank_cell_count"],
                "archive_sheet_rows": {s: audit["row_counts"].get(s) for s in NEW_SHEETS},
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
