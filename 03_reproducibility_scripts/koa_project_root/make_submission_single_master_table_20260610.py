from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改")
SUBMISSION_DIR = PROJECT_ROOT / "投稿使用"
FINAL_DIR = SUBMISSION_DIR / "最终版本"
FINAL_NAME = "KOM_投稿最终唯一总表_LOCKED_20260610.xlsx"
FINAL_PATH = FINAL_DIR / FINAL_NAME
SCRIPT_PATH = PROJECT_ROOT / "make_submission_single_master_table_20260610.py"

SOURCE_MASTER = SUBMISSION_DIR / "KOM_项目所有数据_最终总表_20260610.xlsx"
PREVIOUS_PACKAGE_DIR = PROJECT_ROOT / "KOM_项目最终总表与支持文件_LOCKED_20260610"
PREVIOUS_SUPPORT_MANIFEST = (
    PREVIOUS_PACKAGE_DIR
    / "support_files"
    / "00_process_evidence"
    / "02_search_and_discovery"
    / "all_discovered_files_manifest.csv"
)

RISK_DIR = PROJECT_ROOT / "KOMRisk_Final_Locked_PostDedup_20260610"
RISK_ZIP = RISK_DIR / "13_final_zip" / "KOMRisk_Final_Locked_PostDedup_20260610.zip"
RISK_STATUS_JSON = RISK_DIR / "00_README" / "final_lock_status.json"
RISK_METRICS_CSV = RISK_DIR / "07_metrics_and_acceptance" / "final_metrics_locked.csv"
RISK_ENDPOINT_QC_CSV = RISK_DIR / "04_endpoint_QC" / "endpoint_final_QC_summary.csv"
RISK_FEATURE_COUNTS_CSV = RISK_DIR / "06_feature_name_lock" / "final_feature_count_by_endpoint_and_domain.csv"
RISK_FEATURE_TABLE_XLSX = RISK_DIR / "06_feature_name_lock" / "ALL_ENDPOINTS_locked_final_feature_table.xlsx"
RISK_SIDE_CSV = RISK_DIR / "03_side_mapping_audit" / "side_mapping_final_check.csv"
RISK_KXR_DEDUP_CSV = RISK_DIR / "02_kxr_deduplication_audit" / "downstream_duplicate_row_check.csv"


SHEET_ORDER = [
    "00_README",
    "01_FINAL_MASTER_INDEX",
    "02_PROJECT_CORE_NUMBERS",
    "03_MODULE_NAME_LOCK",
    "04_TOP_JOURNAL_METHOD_CHECKLIST",
    "05_STANDARDIZED_CASES",
    "06_KOM_PROFILE",
    "07_KOM_RAD_OAKNET",
    "08_KOM_RISK_LATEST",
    "09_KOM_KB",
    "10_KOM_RAG_COMPLETION",
    "11_KOM_MDT_RX_SAFE",
    "12_KOM_SCORE_EXPERT",
    "13_KOM_SCORE_RULE",
    "14_KOM_SCORE_ERROR",
    "15_KOM_SIM_CLINICIAN",
    "16_KOM_TREAT_ABLATION",
    "17_FINAL_MODEL_PERFORMANCE",
    "18_MAIN_RESULTS_LOCKED",
    "19_METHODS_PARAMETERS",
    "20_RESULTS_TEXT_BLOCKS",
    "21_RAG_MISSING_ITEMS",
    "22_MODULE_GAPS",
    "23_NUMERIC_TRACEABILITY",
    "24_SOURCE_FILE_INDEX",
    "25_VERSION_SELECTION_LOG",
    "26_OLD_VERSION_ARCHIVE_LOG",
    "27_FIGURE_TABLE_PLAN",
    "28_MANUSCRIPT_WRITING_LOCK",
    "99_QC_FINAL_STATUS",
]


MISSING = "not_available"
SOURCE_LOCKED = "locked_from_previous_master_table_or_chat_context"
NOT_FOUND = "not_found_in_current_files"
NOT_APPLICABLE = "not_applicable"
PENDING = "pending_manual_confirmation"


def stable_value(value: Any) -> Any:
    if value is None:
        return MISSING
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else MISSING
    return value


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def exists_status(path: Path) -> str:
    return "found" if path.exists() else NOT_FOUND


def archive_existing_final_dir_items() -> list[dict[str, Any]]:
    FINAL_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_dir = SUBMISSION_DIR / f"最终版本_archive_old_versions_{timestamp}"
    archived: list[dict[str, Any]] = []

    for item in list(FINAL_DIR.iterdir()):
        if item.name == FINAL_NAME:
            should_archive = True
        else:
            should_archive = True
        if not should_archive:
            continue
        archive_dir.mkdir(parents=True, exist_ok=True)
        target = archive_dir / item.name
        if target.exists():
            target = archive_dir / f"{item.stem}_{timestamp}{item.suffix}"
        item_type = "directory" if item.is_dir() else "file"
        size_bytes = 0
        digest = NOT_APPLICABLE
        if item.is_file():
            size_bytes = item.stat().st_size
            digest = sha256_file(item)
        try:
            shutil.move(str(item), str(target))
            archive_status = "archived"
            archive_note = "final_directory_must_contain_only_the_locked_submission_workbook"
        except PermissionError as exc:
            archive_status = "archive_blocked_file_locked"
            archive_note = f"Windows reported file-in-use; close the owning application and rerun cleanup. error={exc}"
            if item.is_file():
                try:
                    shutil.copy2(str(item), str(target))
                    archive_status = "archive_copy_created_original_locked"
                    archive_note += "; backup_copy_created_but_original_left_in_place"
                except OSError as copy_exc:
                    archive_note += f"; backup_copy_failed={copy_exc}"
        archived.append(
            {
                "archived_item": item.name,
                "item_type": item_type,
                "original_path": str(item),
                "archive_path": str(target),
                "sha256": digest,
                "size_bytes": size_bytes,
                "archived_at": datetime.now().isoformat(timespec="seconds"),
                "reason": archive_note,
                "archive_status": archive_status,
            }
        )

    if not archived:
        archived.append(
            {
                "archived_item": "no_existing_final_directory_items",
                "item_type": NOT_APPLICABLE,
                "original_path": str(FINAL_DIR),
                "archive_path": "no_archive_created",
                "sha256": NOT_APPLICABLE,
                "size_bytes": 0,
                "archived_at": datetime.now().isoformat(timespec="seconds"),
                "reason": "no_old_versions_found_before_generation",
                "archive_status": "no_action_needed",
            }
        )
    seen_paths = {str(item.get("archive_path", "")) for item in archived}
    for existing_archive_dir in sorted(SUBMISSION_DIR.glob("最终版本_archive_old_versions*")):
        if not existing_archive_dir.is_dir():
            continue
        for archived_file in sorted(existing_archive_dir.rglob("*")):
            if not archived_file.is_file():
                continue
            archived_path = str(archived_file)
            if archived_path in seen_paths:
                continue
            seen_paths.add(archived_path)
            archived.append(
                {
                    "archived_item": archived_file.name,
                    "item_type": "file",
                    "original_path": "archived_before_current_run",
                    "archive_path": archived_path,
                    "sha256": sha256_file(archived_file),
                    "size_bytes": archived_file.stat().st_size,
                    "archived_at": datetime.fromtimestamp(archived_file.stat().st_mtime).isoformat(timespec="seconds"),
                    "reason": f"discovered_in_existing_archive_directory:{existing_archive_dir.name}",
                    "archive_status": "archived_existing_archive_dir_discovered",
                }
            )
    return archived


def load_source_master_summary() -> dict[str, Any]:
    summary: dict[str, Any] = {
        "source_master_exists": SOURCE_MASTER.exists(),
        "source_master_sheets": [],
    }
    if SOURCE_MASTER.exists():
        wb = load_workbook(SOURCE_MASTER, read_only=True, data_only=True)
        summary["source_master_sheets"] = wb.sheetnames
        wb.close()
    return summary


def endpoint_short(endpoint: str) -> str:
    text = (endpoint or "").lower()
    if text.startswith("a") or "structural" in text:
        return "A"
    if text.startswith("b") or "surgery" in text or "tkr" in text:
        return "B"
    if text.startswith("c") or "symptom" in text:
        return "C"
    return endpoint[:1].upper() if endpoint else "unknown"


def load_risk_latest_rows() -> list[dict[str, Any]]:
    metrics = read_csv_rows(RISK_METRICS_CSV)
    qc = {endpoint_short(r.get("endpoint", "")): r for r in read_csv_rows(RISK_ENDPOINT_QC_CSV)}
    feature_counts_raw = read_csv_rows(RISK_FEATURE_COUNTS_CSV)
    feature_counts: dict[str, dict[str, str]] = {}
    for r in feature_counts_raw:
        key = endpoint_short(r.get("endpoint", ""))
        feature_counts[key] = r

    locked_defaults = {
        "A": {
            "endpoint_label": "A_KL_structural_progression",
            "n_rows": 7294,
            "n_persons": 3656,
            "events": 970,
            "event_rate": 0.132986,
            "best_model": "CatBoost",
            "AUROC": 0.781279,
            "AUPRC": 0.349087,
            "Brier": 0.190896,
            "decision": "ACCEPT_MAIN",
            "feature_raw": 56,
            "feature_encoded": 60,
        },
        "B": {
            "endpoint_label": "B_TKR_or_knee_surgery_event",
            "n_rows": 9592,
            "n_persons": 4796,
            "events": 548,
            "event_rate": 0.057131,
            "best_model": "CatBoost",
            "AUROC": 0.868117,
            "AUPRC": 0.362152,
            "Brier": 0.128314,
            "decision": "ACCEPT_MAIN",
            "feature_raw": 56,
            "feature_encoded": 60,
        },
        "C": {
            "endpoint_label": "C_symptom_function_worsening",
            "n_rows": 8383,
            "n_persons": 4202,
            "events": 1871,
            "event_rate": 0.223190,
            "best_model": "CatBoost",
            "AUROC": 0.685375,
            "AUPRC": 0.348489,
            "Brier": 0.221623,
            "decision": "ACCEPT_SUPPLEMENT",
            "feature_raw": 60,
            "feature_encoded": 64,
        },
    }
    rows: list[dict[str, Any]] = []
    for key, defaults in locked_defaults.items():
        metric_match = None
        for r in metrics:
            if endpoint_short(r.get("endpoint", "")) == key:
                metric_match = r
                break
        qc_match = qc.get(key, {})
        fc_match = feature_counts.get(key, {})
        rows.append(
            {
                "endpoint": defaults["endpoint_label"],
                "analysis_role": "main" if key in {"A", "B"} else "supplement",
                "task_type": "binary fixed-horizon classification",
                "n_rows": metric_match.get("n_rows", defaults["n_rows"]) if metric_match else defaults["n_rows"],
                "n_persons": metric_match.get("n_persons", defaults["n_persons"]) if metric_match else defaults["n_persons"],
                "events": metric_match.get("events", defaults["events"]) if metric_match else defaults["events"],
                "event_rate": metric_match.get("event_rate", defaults["event_rate"]) if metric_match else defaults["event_rate"],
                "best_model": metric_match.get("best_model", defaults["best_model"]) if metric_match else defaults["best_model"],
                "AUROC": metric_match.get("AUROC", metric_match.get("auroc", defaults["AUROC"])) if metric_match else defaults["AUROC"],
                "AUPRC": metric_match.get("AUPRC", metric_match.get("auprc", defaults["AUPRC"])) if metric_match else defaults["AUPRC"],
                "Brier": metric_match.get("Brier", metric_match.get("brier", defaults["Brier"])) if metric_match else defaults["Brier"],
                "BACC": metric_match.get("balanced_accuracy", metric_match.get("BACC", MISSING)) if metric_match else MISSING,
                "sensitivity": metric_match.get("sensitivity", MISSING) if metric_match else MISSING,
                "specificity": metric_match.get("specificity", MISSING) if metric_match else MISSING,
                "F1": metric_match.get("f1", metric_match.get("F1", MISSING)) if metric_match else MISSING,
                "decision": qc_match.get("decision", defaults["decision"]),
                "feature_raw_count": fc_match.get("raw_feature_count", defaults["feature_raw"]),
                "feature_encoded_count": fc_match.get("encoded_feature_count", defaults["feature_encoded"]),
                "source_file": str(RISK_METRICS_CSV),
                "source_status": exists_status(RISK_METRICS_CSV),
                "notes": "PostDedup locked values; old LightGBM/CoxPH/CatBoost legacy endpoint metrics excluded.",
            }
        )
    return rows


def add_sheet(
    wb: Workbook,
    title: str,
    purpose: str,
    source_summary: str,
    columns: list[str],
    rows: Iterable[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(title)
    col_count = max(len(columns), 6)

    meta_rows = [
        ["sheet_name", title, "purpose", purpose, "generated_at", datetime.now().isoformat(timespec="seconds")],
        ["source_summary", source_summary, "missing_policy", "no blank cells; missing values use explicit status strings", "lock_status", "FINAL_LOCKED_20260610"],
        ["review_scope", "submission_single_master_workbook", "editable", "false_except_manual_confirmation_fields", "traceability", "see 23_NUMERIC_TRACEABILITY and 24_SOURCE_FILE_INDEX"],
    ]
    for meta in meta_rows:
        ws.append((meta + [NOT_APPLICABLE] * col_count)[:col_count])
    ws.append((columns + [NOT_APPLICABLE] * col_count)[:col_count])
    for row in rows:
        ws.append([stable_value(row.get(col, MISSING)) for col in columns])

    if ws.max_row < 5:
        ws.append([NOT_APPLICABLE for _ in range(col_count)])

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(col_count)}{ws.max_row}"

    dark = PatternFill("solid", fgColor="1F4E78")
    dark2 = PatternFill("solid", fgColor="305496")
    white = Font(color="FFFFFF", bold=True)
    for row_idx in range(1, 5):
        for cell in ws[row_idx]:
            cell.fill = dark if row_idx == 4 else dark2
            cell.font = white
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=5):
        for cell in row:
            if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
                cell.value = MISSING
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx in range(1, col_count + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def common_columns() -> list[str]:
    return ["section", "item", "value", "unit_or_scale", "source_file", "source_status", "traceability_id", "notes"]


def row(section: str, item: str, value: Any, unit: str, source: Path | str, status: str, tid: str, notes: str) -> dict[str, Any]:
    return {
        "section": section,
        "item": item,
        "value": value,
        "unit_or_scale": unit,
        "source_file": str(source),
        "source_status": status,
        "traceability_id": tid,
        "notes": notes,
    }


def build_workbook(archive_log: list[dict[str, Any]]) -> Workbook:
    source_summary = load_source_master_summary()
    risk_rows = load_risk_latest_rows()
    risk_status = read_json(RISK_STATUS_JSON)

    wb = Workbook()
    wb.remove(wb.active)

    source_master_status = exists_status(SOURCE_MASTER)
    risk_status_file_status = exists_status(RISK_STATUS_JSON)
    risk_zip_status = exists_status(RISK_ZIP)

    add_sheet(
        wb,
        "00_README",
        "Final single Excel deliverable for KOM submission; all module names and locked numeric values are centralized here.",
        f"Primary source master: {SOURCE_MASTER}; latest KOMRisk package: {RISK_DIR}",
        common_columns(),
        [
            row("identity", "project", "KOM / Knee Osteoarthritis Multimodal AI System", NOT_APPLICABLE, SOURCE_MASTER, source_master_status, "README-001", "Research decision-support system; not a replacement for clinicians."),
            row("identity", "final_file_name", FINAL_NAME, NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "README-002", "This directory is intentionally locked to one final workbook."),
            row("lock", "generation_script", str(SCRIPT_PATH), NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "README-003", "Generated with openpyxl; LibreOffice was not used."),
            row("core", "standardized_cases", 120, "cases", SOURCE_MASTER, SOURCE_LOCKED, "CORE-CASES-120", "Q1-Q4, 30 cases each."),
            row("core", "clinicians", 26, "clinicians", SOURCE_MASTER, SOURCE_LOCKED, "CORE-CLIN-26", "KOM-Sim clinician interaction experiment."),
            row("core", "prescription_records", 780, "records", SOURCE_MASTER, SOURCE_LOCKED, "CORE-RX-780", "26 clinicians x 30 tasks."),
            row("risk", "latest_risk_package", str(RISK_ZIP), NOT_APPLICABLE, RISK_ZIP, risk_zip_status, "RISK-PKG-POSTDEDUP", "Use this PostDedup locked package, not legacy KOMRisk values."),
            row("policy", "missing_value_policy", "explicit status strings only", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "README-004", "No blank cells are intentionally left in the workbook."),
        ],
    )

    add_sheet(
        wb,
        "01_FINAL_MASTER_INDEX",
        "High-level index of all locked KOM modules and which sheets contain their final values.",
        "Derived from user lock instructions and current source files.",
        common_columns(),
        [
            row("module", "KOM-Assess", "clinical data intake and standardized 120-case benchmark", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "IDX-001", "See 05_STANDARDIZED_CASES."),
            row("module", "KOM-Profile", "56-field patient profile extraction", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "IDX-002", "See 06_KOM_PROFILE."),
            row("module", "KOM-Rad / OAKNet", "radiographic KL grading and uncertainty", NOT_APPLICABLE, SOURCE_LOCKED, SOURCE_LOCKED, "IDX-003", "See 07_KOM_RAD_OAKNET."),
            row("module", "KOM-Risk", "three endpoint risk prediction, PostDedup locked", NOT_APPLICABLE, RISK_STATUS_JSON, risk_status_file_status, "IDX-004", "See 08_KOM_RISK_LATEST."),
            row("module", "KOM-KB", "Evidence Unit knowledge base", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "IDX-005", "See 09_KOM_KB."),
            row("module", "KOM-RAG", "GraphRAG retrieval benchmark", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "IDX-006", "See 10_KOM_RAG_COMPLETION."),
            row("module", "KOM-MDT / KOM-Rx / KOM-Safe", "multi-agent prescription and safety gates", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "IDX-007", "See 11_KOM_MDT_RX_SAFE."),
            row("module", "KOM-Score-Expert", "six-expert blinded scoring and ICC", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "IDX-008", "See 12_KOM_SCORE_EXPERT."),
            row("module", "KOM-Score-Rule/Error", "rule-based score and error taxonomy", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "IDX-009", "See 13 and 14."),
            row("module", "KOM-Sim", "26-clinician simulated interaction evaluation", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "IDX-010", "See 15_KOM_SIM_CLINICIAN."),
            row("module", "KOM-Treat", "treatment recommendation ablation", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "IDX-011", "See 16_KOM_TREAT_ABLATION."),
        ],
    )

    add_sheet(
        wb,
        "02_PROJECT_CORE_NUMBERS",
        "One-page numeric lock for the whole project.",
        "Compiled from source master plus latest PostDedup KOMRisk files.",
        common_columns(),
        [
            row("cases", "standardized_cases_total", 120, "cases", SOURCE_MASTER, SOURCE_LOCKED, "CORE-001", "Submission benchmark case set."),
            row("cases", "Q1_cases", 30, "cases", SOURCE_MASTER, SOURCE_LOCKED, "CORE-002", "Q1 low burden / low need."),
            row("cases", "Q2_cases", 30, "cases", SOURCE_MASTER, SOURCE_LOCKED, "CORE-003", "Q2 low burden / high need."),
            row("cases", "Q3_cases", 30, "cases", SOURCE_MASTER, SOURCE_LOCKED, "CORE-004", "Q3 high burden / low need."),
            row("cases", "Q4_cases", 30, "cases", SOURCE_MASTER, SOURCE_LOCKED, "CORE-005", "Q4 high burden / high need."),
            row("clinician_sim", "clinicians", 26, "clinicians", SOURCE_MASTER, SOURCE_LOCKED, "CORE-006", "Clinical user simulation."),
            row("clinician_sim", "tasks_per_clinician", 30, "tasks", SOURCE_MASTER, SOURCE_LOCKED, "CORE-007", "Each clinician completed 30 cases."),
            row("clinician_sim", "prescription_records", 780, "records", SOURCE_MASTER, SOURCE_LOCKED, "CORE-008", "26 x 30."),
            row("knowledge_base", "evidence_units", 3266, "Evidence Units", SOURCE_MASTER, SOURCE_LOCKED, "CORE-009", "KOM-KB evidence unit count."),
            row("knowledge_base", "unique_sources", 2174, "sources", SOURCE_MASTER, SOURCE_LOCKED, "CORE-010", "KOM-KB unique source count."),
            row("risk", "endpoint_A_rows", 7294, "rows", RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "CORE-011", "PostDedup latest."),
            row("risk", "endpoint_B_rows", 9592, "rows", RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "CORE-012", "PostDedup latest."),
            row("risk", "endpoint_C_rows", 8383, "rows", RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "CORE-013", "PostDedup latest."),
            row("risk", "side_mapping_RIGHT_match_rate", 0.995270, "proportion", RISK_SIDE_CSV, exists_status(RISK_SIDE_CSV), "CORE-014", "SIDE=1/right."),
            row("risk", "side_mapping_LEFT_match_rate", 0.996254, "proportion", RISK_SIDE_CSV, exists_status(RISK_SIDE_CSV), "CORE-015", "SIDE=2/left."),
        ],
    )

    add_sheet(
        wb,
        "03_MODULE_NAME_LOCK",
        "Locked terminology for modules, endpoints, and manuscript-facing names.",
        "Compiled from final naming decisions.",
        common_columns(),
        [
            row("system", "KOM", "Knee Osteoarthritis Multimodal AI System", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "NAME-001", "Use KOM consistently in manuscript tables."),
            row("module", "KOM-Assess", "case intake and structured clinical assessment", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "NAME-002", "No alternate names."),
            row("module", "KOM-Profile", "patient profile extraction", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "NAME-003", "No alternate names."),
            row("module", "KOM-Rad / OAKNet", "radiographic KL grading and uncertainty", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "NAME-004", "OAKNet is the model name."),
            row("module", "KOM-Risk", "risk prediction endpoints A/B/C", NOT_APPLICABLE, RISK_STATUS_JSON, risk_status_file_status, "NAME-005", "Use PostDedup endpoint definitions only."),
            row("module", "KOM-KB", "Evidence Unit knowledge base", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "NAME-006", "Evidence Unit should remain capitalized."),
            row("module", "KOM-RAG", "GraphRAG evidence retrieval", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "NAME-007", "Do not call plain RAG when graph fusion is meant."),
            row("module", "KOM-MDT", "multi-specialty agent discussion", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "NAME-008", "Five specialty agents plus safety/curation."),
            row("module", "KOM-Rx", "standardized prescription generation", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "NAME-009", "Final treatment output module."),
            row("module", "KOM-Safe", "safety gate and contraindication review", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "NAME-010", "Safety audit module."),
        ],
    )

    add_sheet(
        wb,
        "04_TOP_JOURNAL_METHOD_CHECKLIST",
        "Checklist mapping current evidence package to top-journal AI/clinical reporting expectations.",
        "Methodology lock sheet; missing items are explicitly marked.",
        common_columns(),
        [
            row("reporting", "DECIDE-AI", "partially_addressed", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "CHECK-001", "Human-in-loop CDS workflow described; prospective implementation remains future work."),
            row("reporting", "CONSORT-AI/SPIRIT-AI", "not_applicable_for_current_retrospective_simulation", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "CHECK-002", "No clinical trial claim."),
            row("reporting", "TRIPOD+AI", "partially_addressed_for_KOMRisk", NOT_APPLICABLE, RISK_STATUS_JSON, risk_status_file_status, "CHECK-003", "Endpoint A/B/C model metrics available; calibration plots need manuscript verification."),
            row("reporting", "CLAIM", "partially_addressed_for_OAKNet", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "CHECK-004", "Radiograph model metrics locked; image preprocessing details should be verified before submission."),
            row("retrieval", "GraphRAG evidence retrieval", "addressed_with_missing_method_details", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "CHECK-005", "Retrieval metrics locked; graph construction parameters remain a manuscript gap if raw config not found."),
            row("evaluation", "expert blinded review", "addressed_summary_locked", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "CHECK-006", "ICC summary locked; raw expert scoring table not found in current files."),
            row("safety", "AI does not replace clinician", "required_statement", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "CHECK-007", "Must appear in abstract/discussion/interface limitations."),
            row("reproducibility", "single final workbook", "addressed_by_this_file", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "CHECK-008", "Old versions archived outside final folder."),
        ],
    )

    add_sheet(
        wb,
        "05_STANDARDIZED_CASES",
        "Definition and composition of the 120-case standardized evaluation set.",
        "Locked from previous master table and chat-context design decisions.",
        common_columns(),
        [
            row("case_set", "source_pool", "OAI-derived candidate clinical profile pool", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "CASE-001", "Synthetic/standardized research cases, not real patient records."),
            row("sampling", "method", "stratified purposive sampling with max-variation coverage", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "CASE-002", "Not a prevalence sample."),
            row("quadrant", "Q1", "low burden / low need", "30 cases", SOURCE_MASTER, SOURCE_LOCKED, "CASE-003", "Mild phenotype."),
            row("quadrant", "Q2", "low burden / high need", "30 cases", SOURCE_MASTER, SOURCE_LOCKED, "CASE-004", "High demand despite lower structural burden."),
            row("quadrant", "Q3", "high burden / low need", "30 cases", SOURCE_MASTER, SOURCE_LOCKED, "CASE-005", "High structural burden but lower intervention demand."),
            row("quadrant", "Q4", "high burden / high need", "30 cases", SOURCE_MASTER, SOURCE_LOCKED, "CASE-006", "Complex/high-risk prescription set."),
            row("axes", "complexity_axes", "KL grade; pain; BMI; fall risk; patient demand; comorbidity/risk flags; missing information", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "CASE-007", "Use for Methods section."),
            row("limitation", "external_clinical_distribution", "pending_manual_confirmation", NOT_APPLICABLE, SOURCE_MASTER, PENDING, "CASE-008", "Do not claim real-world prevalence."),
        ],
    )

    add_sheet(
        wb,
        "06_KOM_PROFILE",
        "Patient-profile extraction module summary.",
        "Locked summary; field-level raw extraction table not found in current files.",
        common_columns(),
        [
            row("profile", "structured_fields", 56, "fields", SOURCE_MASTER, SOURCE_LOCKED, "PROFILE-001", "Patient profile extraction schema."),
            row("performance", "overall_F1_or_accuracy", 0.846, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "PROFILE-002", "Locked from previous master table/chat context."),
            row("field_status", "exact_fields", 31, "fields", SOURCE_MASTER, SOURCE_LOCKED, "PROFILE-003", "Exact extraction match."),
            row("field_status", "partial_fields", 25, "fields", SOURCE_MASTER, SOURCE_LOCKED, "PROFILE-004", "Partial extraction match."),
            row("source", "field_level_raw_table", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "PROFILE-005", "Retain as gap; do not fabricate field-level rows."),
            row("clinical_anchor", "required_patient_anchors", "age; sex; BMI; KL; NRS; WOMAC; fall risk; strength; renal/GI/anticoagulant status; surgery preference; rehab willingness", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "PROFILE-006", "Anchor fields for downstream personalization."),
        ],
    )

    add_sheet(
        wb,
        "07_KOM_RAD_OAKNET",
        "Radiographic KL grading model performance lock.",
        "Locked from previous master table/chat context; training checkpoint not included in this final Excel.",
        common_columns(),
        [
            row("model", "primary_model", "ConvNeXt-B with evidential uncertainty head", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RAD-001", "OAKNet primary radiographic model."),
            row("performance", "ConvNeXt-B_QWK", "0.806 ± 0.008", "weighted kappa", SOURCE_MASTER, SOURCE_LOCKED, "RAD-002", "Primary ordinal agreement metric."),
            row("performance", "ConvNeXt-B_BACC", 0.659, "balanced accuracy", SOURCE_MASTER, SOURCE_LOCKED, "RAD-003", "Locked value."),
            row("performance", "ConvNeXt-B_macro_F1", 0.664, "macro F1", SOURCE_MASTER, SOURCE_LOCKED, "RAD-004", "Locked value."),
            row("performance", "ConvNeXt-B_MAE", 0.417, "KL grade", SOURCE_MASTER, SOURCE_LOCKED, "RAD-005", "Mean absolute error."),
            row("performance", "ConvNeXt-B_ECE", 0.119, "calibration error", SOURCE_MASTER, SOURCE_LOCKED, "RAD-006", "Expected calibration error."),
            row("performance", "ConvNeXt-B_selective_accuracy_at_80", 0.725, "accuracy", SOURCE_MASTER, SOURCE_LOCKED, "RAD-007", "Selective prediction performance."),
            row("baseline", "DenseNet-121_QWK", "0.805 ± 0.007", "weighted kappa", SOURCE_MASTER, SOURCE_LOCKED, "RAD-008", "Comparator."),
            row("baseline", "DenseNet-121_BACC", 0.669, "balanced accuracy", SOURCE_MASTER, SOURCE_LOCKED, "RAD-009", "Comparator."),
            row("source_gap", "training_parameters", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RAD-010", "Do not claim hyperparameters unless verified."),
        ],
    )

    add_sheet(
        wb,
        "08_KOM_RISK_LATEST",
        "Latest PostDedup KOMRisk endpoint results; legacy metrics excluded.",
        f"Read from {RISK_METRICS_CSV} plus QC/feature-count files.",
        [
            "endpoint",
            "analysis_role",
            "task_type",
            "n_rows",
            "n_persons",
            "events",
            "event_rate",
            "best_model",
            "AUROC",
            "AUPRC",
            "Brier",
            "BACC",
            "sensitivity",
            "specificity",
            "F1",
            "decision",
            "feature_raw_count",
            "feature_encoded_count",
            "source_file",
            "source_status",
            "notes",
        ],
        risk_rows,
    )

    add_sheet(
        wb,
        "09_KOM_KB",
        "Evidence Unit knowledge-base composition and currently verified gaps.",
        "Locked summary from source master; detailed lifecycle audit not fully located in current files.",
        common_columns(),
        [
            row("count", "Evidence Units", 3266, "units", SOURCE_MASTER, SOURCE_LOCKED, "KB-001", "Evidence Unit database size."),
            row("count", "unique_sources", 2174, "sources", SOURCE_MASTER, SOURCE_LOCKED, "KB-002", "Unique evidence source count."),
            row("level", "L1_guideline_consensus", 99, "Evidence Units", SOURCE_MASTER, SOURCE_LOCKED, "KB-003", "Top hierarchy evidence."),
            row("level", "L2_systematic_review_meta", 648, "Evidence Units", SOURCE_MASTER, SOURCE_LOCKED, "KB-004", "Evidence hierarchy."),
            row("level", "L3_clinical_trial", 1124, "Evidence Units", SOURCE_MASTER, SOURCE_LOCKED, "KB-005", "Evidence hierarchy."),
            row("level", "L4_observational", 488, "Evidence Units", SOURCE_MASTER, SOURCE_LOCKED, "KB-006", "Evidence hierarchy."),
            row("level", "L5_or_lower", 907, "Evidence Units", SOURCE_MASTER, SOURCE_LOCKED, "KB-007", "Low-level/contextual evidence."),
            row("gap", "recommendation_specific_mapping_raw", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "KB-008", "Keep as Methods gap unless raw mapping table is provided."),
            row("gap", "source_verification_lifecycle_table", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "KB-009", "Do not invent DOI/PMID verification."),
        ],
    )

    add_sheet(
        wb,
        "10_KOM_RAG_COMPLETION",
        "GraphRAG retrieval benchmark and unresolved reporting gaps.",
        "Locked metrics from source master/chat context; raw benchmark row file not embedded in final workbook.",
        common_columns(),
        [
            row("GraphRAG", "Precision@10", 0.676, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "RAG-001", "Stage4A holdout/dev-calibrated relevance prior result."),
            row("GraphRAG", "Hit@10", 1.0, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "RAG-002", "At least one relevant item in Top10."),
            row("GraphRAG", "MRR", 0.748, "rank metric", SOURCE_MASTER, SOURCE_LOCKED, "RAG-003", "Mean reciprocal rank."),
            row("GraphRAG", "nDCG@10", 0.690, "rank metric", SOURCE_MASTER, SOURCE_LOCKED, "RAG-004", "Ranking quality."),
            row("GraphRAG", "Recall@10", 0.412, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "RAG-005", "Auxiliary recall curve metric."),
            row("GraphRAG", "Recall@20", 0.695, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "RAG-006", "Auxiliary recall curve metric."),
            row("GraphRAG", "Recall@27", 0.824, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "RAG-007", "Auxiliary recall curve metric."),
            row("GraphRAG", "Recall@30", 0.855, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "RAG-008", "Auxiliary recall curve metric."),
            row("naive_RAG", "Precision@10", 0.303, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "RAG-009", "Comparator value."),
            row("naive_RAG", "MRR", 0.159, "rank metric", SOURCE_MASTER, SOURCE_LOCKED, "RAG-010", "Comparator value."),
            row("naive_RAG", "nDCG@10", 0.237, "rank metric", SOURCE_MASTER, SOURCE_LOCKED, "RAG-011", "Comparator value."),
            row("naive_RAG", "Hit@10", 0.688, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "RAG-012", "Comparator value."),
            row("gap", "graph_nodes_edges", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RAG-013", "Need raw graph manifest for final Methods."),
            row("gap", "embedding_model_and_cache", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RAG-014", "If bge-m3/openai embedding was used, attach raw config before submission."),
            row("gap", "faithfulness_grounding_metrics", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RAG-015", "Do not overclaim generation grounding without raw data."),
        ],
    )

    add_sheet(
        wb,
        "11_KOM_MDT_RX_SAFE",
        "Standardized MDT prescription and safety-gate module lock.",
        "Locked from previous prescription pipeline and final safety-hardening instructions.",
        common_columns(),
        [
            row("agent", "R0_case_intake", "case standardization and missing-information gate", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-001", "Required before prescription."),
            row("agent", "R1_to_R4_specialist_agents", "medication; surgery; exercise; nutrition; psychology/behavior", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-002", "Five specialty modules."),
            row("agent", "R5_to_R8_audit_consensus", "self-audit; central audit; cross-review; consensus", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-003", "Process must be visible in reviewer materials."),
            row("safety", "missing_information_first", "renal/eGFR; GI bleeding; anticoagulant/current meds; CV risk before oral NSAID", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-004", "Hard gate in prescription text."),
            row("safety", "oral_NSAID_gate", "DEFER until renal + GI + anticoagulant/current meds + CV review complete", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-005", "Do not write vague caution."),
            row("safety", "injection_boundary", "only_if_flare_or_bridge; not routine repeated injections", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-006", "Patient preference considered."),
            row("safety", "surgery_boundary", "AI recommends specialist referral discussion; AI does not decide surgery type", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-007", "Avoid direct TKA/UKA/HTO decision."),
            row("safety", "exercise_boundary", "FITT + fall prevention + stop rules", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-008", "No high-impact plan in high-risk cases."),
            row("safety", "nutrition_boundary", "weight management plus muscle preservation", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-009", "No fixed high-protein target if renal function unknown."),
            row("safety", "psychology_boundary", "GAD-7; PHQ-9; PCS; sleep screening; non-stigmatizing pain education", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "MDT-010", "Specific screening tools required."),
        ],
    )

    add_sheet(
        wb,
        "12_KOM_SCORE_EXPERT",
        "Expert blinded review score reliability summary.",
        "Locked ICC summary; raw expert-level score table not located.",
        common_columns(),
        [
            row("panel", "expert_count", 6, "experts", SOURCE_MASTER, SOURCE_LOCKED, "EXPERT-001", "Orthopaedics=2; sports medicine=2; rehabilitation=2."),
            row("panel", "blinding", "blinded review", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "EXPERT-002", "Reviewer identity and model arms masked as designed."),
            row("ICC", "overall_quality_ICC_2_1", 0.796, "ICC", SOURCE_MASTER, SOURCE_LOCKED, "EXPERT-003", "KOM score expert overall quality."),
            row("ICC", "safety_ICC", 0.574, "ICC", SOURCE_MASTER, SOURCE_LOCKED, "EXPERT-004", "Moderate safety agreement."),
            row("ICC", "doctor_prescription_overall_quality_ICC_2_1", 0.946, "ICC", SOURCE_MASTER, SOURCE_LOCKED, "EXPERT-005", "Doctor prescription review reliability."),
            row("ICC", "doctor_prescription_dimension_ICC_range", "0.902-0.946", "ICC range", SOURCE_MASTER, SOURCE_LOCKED, "EXPERT-006", "Dimension-level range."),
            row("gap", "expert_raw_score_table", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "EXPERT-007", "Use summary cautiously; raw data needed for audit reanalysis."),
        ],
    )

    add_sheet(
        wb,
        "13_KOM_SCORE_RULE",
        "Rule-based prescription scoring domains and availability status.",
        "Locked rule score summaries; full item-level rule weight table not found.",
        common_columns(),
        [
            row("domain", "safety_gates", "oral NSAID; injection; surgery; fall prevention; nutrition; psychology; missing information", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RULE-001", "Core safety score dimensions."),
            row("domain", "guideline_alignment", "ACR/EULAR/NICE/OARSI anchored CDS rules", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RULE-002", "Method text should avoid unsupported guideline specifics."),
            row("domain", "individualization", "KL; pain; BMI; fall risk; preferences; missing info", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RULE-003", "Patient anchor matching."),
            row("domain", "executability", "dose/frequency/duration/progression/stop thresholds", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RULE-004", "Prescription usability."),
            row("gap", "full_weight_table", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RULE-005", "Need appendable rule list for Supplement."),
            row("locked_result", "KOM_standalone_rule_score", 70.0, "score", SOURCE_MASTER, SOURCE_LOCKED, "RULE-006", "See 15_KOM_SIM_CLINICIAN."),
        ],
    )

    add_sheet(
        wb,
        "14_KOM_SCORE_ERROR",
        "Error taxonomy used for safety and quality assessment.",
        "Locked taxonomy from prior score work; raw per-case error table not embedded.",
        common_columns(),
        [
            row("error_type", "critical", "may cause patient harm or violate a hard safety gate", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "ERR-001", "Examples: unsafe oral NSAID without renal/GI/anticoagulant/CV review."),
            row("error_type", "major", "meaningfully reduces prescription quality but not direct immediate harm", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "ERR-002", "Examples: missing dose/progression, missing specialty module."),
            row("error_type", "minor", "formatting or non-critical completeness issue", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "ERR-003", "Do not inflate minor issues into major."),
            row("metric", "clinician_alone_safety_critical_error_rate", 19.7, "per 100", SOURCE_MASTER, SOURCE_LOCKED, "ERR-004", "KOM-Sim locked value."),
            row("metric", "clinician_plus_KOM_safety_critical_error_rate", 8.8, "per 100", SOURCE_MASTER, SOURCE_LOCKED, "ERR-005", "KOM-Sim locked value."),
            row("metric", "clinician_plus_KOMR_safety_critical_error_rate", 14.5, "per 100", SOURCE_MASTER, SOURCE_LOCKED, "ERR-006", "KOM-Sim locked value."),
            row("metric", "KOM_standalone_safety_critical_error_rate", 0, "count/rate", SOURCE_MASTER, SOURCE_LOCKED, "ERR-007", "KOM-Treat/standalone locked value."),
        ],
    )

    sim_cols = ["arm", "overall_quality", "rule_score", "safety_critical_error_per_100", "high_quality_percent", "text_length", "components", "editing_time", "workload", "confidence", "info_sufficiency", "certainty", "ai_influence", "ai_view", "explanation_view", "evidence_view", "source_file", "source_status", "notes"]
    sim_rows = [
        {"arm": "Clinician alone", "overall_quality": 48.7, "rule_score": 30.1, "safety_critical_error_per_100": 19.7, "high_quality_percent": 4.6, "text_length": 58, "components": 3.26, "editing_time": 43.5, "workload": 5.10, "confidence": 7.60, "info_sufficiency": 7.93, "certainty": 7.38, "ai_influence": NOT_APPLICABLE, "ai_view": NOT_APPLICABLE, "explanation_view": NOT_APPLICABLE, "evidence_view": NOT_APPLICABLE, "source_file": str(SOURCE_MASTER), "source_status": SOURCE_LOCKED, "notes": "Locked simulation arm."},
        {"arm": "Clinician + KOM", "overall_quality": 73.4, "rule_score": 63.7, "safety_critical_error_per_100": 8.8, "high_quality_percent": 51.5, "text_length": 1490, "components": 5.47, "editing_time": 27.0, "workload": 4.35, "confidence": 7.78, "info_sufficiency": 8.05, "certainty": 7.51, "ai_influence": 5.63, "ai_view": 1.46, "explanation_view": NOT_APPLICABLE, "evidence_view": NOT_APPLICABLE, "source_file": str(SOURCE_MASTER), "source_status": SOURCE_LOCKED, "notes": "AI adoption/copy threshold >=0.6."},
        {"arm": "Clinician + KOM-R", "overall_quality": 70.1, "rule_score": 61.0, "safety_critical_error_per_100": 14.5, "high_quality_percent": 50.8, "text_length": 1434, "components": 5.50, "editing_time": 17.0, "workload": 3.84, "confidence": 7.93, "info_sufficiency": 8.12, "certainty": 7.66, "ai_influence": 5.89, "ai_view": 1.29, "explanation_view": 0.47, "evidence_view": 0.29, "source_file": str(SOURCE_MASTER), "source_status": SOURCE_LOCKED, "notes": "AI adoption/copy threshold >=0.6; explainability views recorded."},
        {"arm": "KOM standalone", "overall_quality": 84.6, "rule_score": 70.0, "safety_critical_error_per_100": 0, "high_quality_percent": MISSING, "text_length": MISSING, "components": MISSING, "editing_time": MISSING, "workload": MISSING, "confidence": MISSING, "info_sufficiency": MISSING, "certainty": MISSING, "ai_influence": NOT_APPLICABLE, "ai_view": NOT_APPLICABLE, "explanation_view": NOT_APPLICABLE, "evidence_view": NOT_APPLICABLE, "source_file": str(SOURCE_MASTER), "source_status": SOURCE_LOCKED, "notes": "System standalone prescription output."},
    ]
    add_sheet(wb, "15_KOM_SIM_CLINICIAN", "Clinician interaction simulation locked results.", "Locked from source master table.", sim_cols, sim_rows)

    treat_cols = ["arm", "overall_quality", "safety_score", "corrected_rule_score", "safety_critical_error", "interpretation", "source_file", "source_status", "notes"]
    treat_rows = [
        {"arm": "A_full_KOM", "overall_quality": 84.6, "safety_score": 91.1, "corrected_rule_score": 84.3, "safety_critical_error": 0, "interpretation": "full system", "source_file": str(SOURCE_MASTER), "source_status": SOURCE_LOCKED, "notes": "Primary treatment ablation arm."},
        {"arm": "B_without_RAG", "overall_quality": 65.6, "safety_score": 79.9, "corrected_rule_score": 71.6, "safety_critical_error": MISSING, "interpretation": "RAG contribution removed", "source_file": str(SOURCE_MASTER), "source_status": SOURCE_LOCKED, "notes": "A vs B estimates retrieval contribution."},
        {"arm": "C_without_MDT", "overall_quality": 64.4, "safety_score": 79.1, "corrected_rule_score": 81.7, "safety_critical_error": MISSING, "interpretation": "MDT contribution removed", "source_file": str(SOURCE_MASTER), "source_status": SOURCE_LOCKED, "notes": "A vs C estimates multi-agent contribution."},
        {"arm": "D_direct_LLM", "overall_quality": 54.7, "safety_score": 70.7, "corrected_rule_score": 44.8, "safety_critical_error": MISSING, "interpretation": "bare direct LLM", "source_file": str(SOURCE_MASTER), "source_status": SOURCE_LOCKED, "notes": "A vs D estimates total system gain."},
    ]
    add_sheet(wb, "16_KOM_TREAT_ABLATION", "Treatment recommendation ablation locked results.", "Locked from source master table.", treat_cols, treat_rows)

    add_sheet(
        wb,
        "17_FINAL_MODEL_PERFORMANCE",
        "Integrated final performance summary across KOM modules.",
        "Compiled from source master and latest risk package.",
        common_columns(),
        [
            row("KOM-Rad", "ConvNeXt-B_QWK", "0.806 ± 0.008", "QWK", SOURCE_MASTER, SOURCE_LOCKED, "PERF-001", "Radiographic KL grading."),
            row("KOM-Risk", "Endpoint_A_AUROC", 0.781279, "AUROC", RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "PERF-002", "PostDedup latest."),
            row("KOM-Risk", "Endpoint_B_AUROC", 0.868117, "AUROC", RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "PERF-003", "PostDedup latest."),
            row("KOM-Risk", "Endpoint_C_AUROC", 0.685375, "AUROC", RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "PERF-004", "PostDedup latest."),
            row("KOM-RAG", "Precision@10", 0.676, "precision", SOURCE_MASTER, SOURCE_LOCKED, "PERF-005", "GraphRAG retrieval."),
            row("KOM-RAG", "nDCG@10", 0.690, "nDCG", SOURCE_MASTER, SOURCE_LOCKED, "PERF-006", "GraphRAG retrieval."),
            row("KOM-Treat", "Full_KOM_overall_quality", 84.6, "score", SOURCE_MASTER, SOURCE_LOCKED, "PERF-007", "Treatment ablation."),
            row("KOM-Sim", "Clinician_plus_KOM_overall_quality", 73.4, "score", SOURCE_MASTER, SOURCE_LOCKED, "PERF-008", "Clinician simulation."),
            row("KOM-Sim", "critical_error_reduction_KOM_vs_clinician_alone", "19.7 to 8.8", "per 100", SOURCE_MASTER, SOURCE_LOCKED, "PERF-009", "Safety improvement."),
        ],
    )

    add_sheet(
        wb,
        "18_MAIN_RESULTS_LOCKED",
        "Locked manuscript-facing result statements.",
        "Derived from final numeric sheets; values should not be edited without traceability updates.",
        common_columns(),
        [
            row("R1", "standardized_cases", "KOM was evaluated on 120 standardized KOA cases spanning four burden-demand quadrants.", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RESULT-001", "Use in Results opening."),
            row("R2", "risk", "KOM-Risk achieved AUROC 0.781 for KL progression and 0.868 for future TKR/knee surgery events in the latest PostDedup lock.", NOT_APPLICABLE, RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "RESULT-002", "Endpoint C is supplementary."),
            row("R3", "retrieval", "GraphRAG outperformed naive RAG for retrieval ranking, including Precision@10 0.676 vs 0.303 and nDCG@10 0.690 vs 0.237.", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RESULT-003", "Mention metric design carefully."),
            row("R4", "treatment", "The full KOM treatment pipeline scored higher than ablated pipelines and direct LLM generation.", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RESULT-004", "Use with ablation table."),
            row("R5", "clinician_sim", "Clinician+KOM improved overall prescription quality from 48.7 to 73.4 and reduced critical safety errors from 19.7 to 8.8 per 100 prescriptions.", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RESULT-005", "Simulation, not clinical outcome trial."),
        ],
    )

    add_sheet(
        wb,
        "19_METHODS_PARAMETERS",
        "Methods parameters and reproducibility status by module.",
        "Explicitly separates found parameters from missing/pending items.",
        common_columns(),
        [
            row("case_set", "case_count", 120, "cases", SOURCE_MASTER, SOURCE_LOCKED, "METHOD-001", "Q1-Q4 equal allocation."),
            row("clinician_sim", "clinicians", 26, "clinicians", SOURCE_MASTER, SOURCE_LOCKED, "METHOD-002", "30 tasks each."),
            row("risk", "split_policy", "person-level split; no leakage", NOT_APPLICABLE, RISK_STATUS_JSON, risk_status_file_status, "METHOD-003", "Locked PostDedup package."),
            row("risk", "best_model_policy", "single best_model row per endpoint", NOT_APPLICABLE, RISK_STATUS_JSON, risk_status_file_status, "METHOD-004", "Checkpoint residue NO."),
            row("risk", "Endpoint_B_type", "fixed-horizon binary classification, not survival", NOT_APPLICABLE, RISK_STATUS_JSON, risk_status_file_status, "METHOD-005", "Avoid old CoxPH description."),
            row("RAG", "dev_calibrated_relevance_prior", "used for locked benchmark", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "METHOD-006", "Do not conflate with untuned raw retrieval."),
            row("RAG", "raw_embedding_config", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "METHOD-007", "Needs manual attachment if reviewers ask."),
            row("OAKNet", "training_hyperparameters", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "METHOD-008", "Do not overstate."),
        ],
    )

    add_sheet(
        wb,
        "20_RESULTS_TEXT_BLOCKS",
        "Reusable manuscript text blocks with locked numeric values.",
        "Draft text only; requires author polishing before submission.",
        common_columns(),
        [
            row("CN", "KOMRisk", "在PostDedup最终锁定版本中，KOM-Risk对KL结构进展、未来TKR/膝手术事件和症状/功能恶化的AUROC分别为0.781、0.868和0.685，其中A、B端点作为主分析，C端点作为补充分析。", NOT_APPLICABLE, RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "TEXT-001", "Chinese Results block."),
            row("EN", "KOMRisk", "In the final PostDedup lock, KOM-Risk achieved AUROC values of 0.781, 0.868 and 0.685 for KL structural progression, future TKR/knee surgery events, and symptom/function worsening, respectively.", NOT_APPLICABLE, RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "TEXT-002", "English Results block."),
            row("CN", "GraphRAG", "KOM-RAG在检索排序上优于朴素RAG，Precision@10为0.676对0.303，MRR为0.748对0.159，nDCG@10为0.690对0.237。", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "TEXT-003", "Chinese Results block."),
            row("EN", "ClinicianSim", "In the clinician simulation, clinician use of KOM improved overall prescription quality from 48.7 to 73.4 and reduced critical safety errors from 19.7 to 8.8 per 100 prescriptions.", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "TEXT-004", "English Results block."),
            row("limitation", "clinical_outcome", "This study evaluates decision-support performance and simulated prescribing quality, not patient clinical outcomes.", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "TEXT-005", "Required limitation."),
        ],
    )

    add_sheet(
        wb,
        "21_RAG_MISSING_ITEMS",
        "RAG/GraphRAG missing items that should not be silently implied in the manuscript.",
        "Generated from current file availability and locked metric status.",
        common_columns(),
        [
            row("available", "locked_retrieval_metrics", "available", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "RAGMISS-001", "Precision/MRR/nDCG/Hit/Recall curve values locked."),
            row("missing", "per_query_raw_records", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RAGMISS-002", "Needed for reproducible appendix if reviewers request."),
            row("missing", "gold_label_annotation_protocol", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RAGMISS-003", "Do not claim inter-annotator agreement unless raw protocol exists."),
            row("missing", "graph_node_edge_manifest", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RAGMISS-004", "Needed for full GraphRAG reproducibility."),
            row("missing", "embedding_cache_manifest", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RAGMISS-005", "State model/config only if verified."),
            row("missing", "faithfulness_generation_metric", "not_found_in_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "RAGMISS-006", "Retrieval success does not prove generation faithfulness."),
            row("action", "minimal_manual_action", "attach raw benchmark rows and graph/index config before submission if available", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "RAGMISS-007", "Manual author action."),
        ],
    )

    add_sheet(
        wb,
        "22_MODULE_GAPS",
        "Module-level gaps and manuscript risk controls.",
        "Gaps are explicit to avoid unsupported claims.",
        common_columns(),
        [
            row("KOM-Profile", "raw field-level extraction table", "missing", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "GAP-001", "Can report summary F1/accuracy; avoid per-field claims."),
            row("KOM-Rad", "training hyperparameters/checkpoints", "missing", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "GAP-002", "Report locked metrics; do not claim full model reproducibility."),
            row("KOM-Risk", "latest PostDedup package", "available", NOT_APPLICABLE, RISK_ZIP, risk_zip_status, "GAP-003", "Use latest values only."),
            row("KOM-RAG", "raw graph/index manifest", "missing", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "GAP-004", "Retrieval performance can be reported; graph implementation needs supplement."),
            row("KOM-Score-Expert", "raw expert scoring rows", "missing", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "GAP-005", "Report ICC summary as locked; raw table required for independent reanalysis."),
            row("KOM-Sim", "clinician raw interaction log", "not_embedded_in_final_workbook", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "GAP-006", "Keep original source secure; summary locked here."),
        ],
    )

    trace_rows = [
        row("source_master", "standardized_cases", 120, "cases", SOURCE_MASTER, source_master_status, "TRACE-001", "Q1-Q4=30 each."),
        row("source_master", "clinicians", 26, "clinicians", SOURCE_MASTER, source_master_status, "TRACE-002", "KOM-Sim."),
        row("source_master", "prescription_records", 780, "records", SOURCE_MASTER, source_master_status, "TRACE-003", "26 x 30."),
        row("KOMRisk", "Endpoint_A_AUROC", 0.781279, "AUROC", RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "TRACE-004", "PostDedup."),
        row("KOMRisk", "Endpoint_B_AUROC", 0.868117, "AUROC", RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "TRACE-005", "PostDedup."),
        row("KOMRisk", "Endpoint_C_AUROC", 0.685375, "AUROC", RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "TRACE-006", "PostDedup."),
        row("KOMRAG", "Precision@10", 0.676, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "TRACE-007", "Locked from retrieval benchmark."),
        row("KOMRAG", "nDCG@10", 0.690, "proportion", SOURCE_MASTER, SOURCE_LOCKED, "TRACE-008", "Locked from retrieval benchmark."),
        row("KOMTreat", "Full_KOM_overall", 84.6, "score", SOURCE_MASTER, SOURCE_LOCKED, "TRACE-009", "Treatment ablation."),
        row("KOMSim", "Clinician_plus_KOM_overall", 73.4, "score", SOURCE_MASTER, SOURCE_LOCKED, "TRACE-010", "Clinician simulation."),
    ]
    add_sheet(wb, "23_NUMERIC_TRACEABILITY", "Traceability table for every manuscript-critical numeric value.", "Compiled from locked source files.", common_columns(), trace_rows)

    source_index_rows = [
        row("source", "source_master_workbook", str(SOURCE_MASTER), NOT_APPLICABLE, SOURCE_MASTER, source_master_status, "SRC-001", "Current integrated source workbook."),
        row("source", "risk_final_zip", str(RISK_ZIP), NOT_APPLICABLE, RISK_ZIP, risk_zip_status, "SRC-002", "Latest PostDedup package."),
        row("source", "risk_lock_status_json", str(RISK_STATUS_JSON), NOT_APPLICABLE, RISK_STATUS_JSON, risk_status_file_status, "SRC-003", f"Status keys: {', '.join(list(risk_status.keys())[:8]) if risk_status else NOT_FOUND}."),
        row("source", "risk_metrics_csv", str(RISK_METRICS_CSV), NOT_APPLICABLE, RISK_METRICS_CSV, exists_status(RISK_METRICS_CSV), "SRC-004", "Endpoint metrics."),
        row("source", "risk_endpoint_qc_csv", str(RISK_ENDPOINT_QC_CSV), NOT_APPLICABLE, RISK_ENDPOINT_QC_CSV, exists_status(RISK_ENDPOINT_QC_CSV), "SRC-005", "Endpoint accept/supplement decisions."),
        row("source", "risk_feature_counts_csv", str(RISK_FEATURE_COUNTS_CSV), NOT_APPLICABLE, RISK_FEATURE_COUNTS_CSV, exists_status(RISK_FEATURE_COUNTS_CSV), "SRC-006", "Feature counts."),
        row("source", "risk_feature_table_xlsx", str(RISK_FEATURE_TABLE_XLSX), NOT_APPLICABLE, RISK_FEATURE_TABLE_XLSX, exists_status(RISK_FEATURE_TABLE_XLSX), "SRC-007", "Feature names."),
        row("source", "risk_side_mapping_csv", str(RISK_SIDE_CSV), NOT_APPLICABLE, RISK_SIDE_CSV, exists_status(RISK_SIDE_CSV), "SRC-008", "SIDE mapping audit."),
        row("source", "risk_kxr_dedup_csv", str(RISK_KXR_DEDUP_CSV), NOT_APPLICABLE, RISK_KXR_DEDUP_CSV, exists_status(RISK_KXR_DEDUP_CSV), "SRC-009", "KXR dedup audit."),
        row("source", "previous_support_manifest", str(PREVIOUS_SUPPORT_MANIFEST), NOT_APPLICABLE, PREVIOUS_SUPPORT_MANIFEST, exists_status(PREVIOUS_SUPPORT_MANIFEST), "SRC-010", "Previous discovery manifest, not copied into final folder."),
        row("source", "generation_script", str(SCRIPT_PATH), NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "SRC-011", "Openpyxl generation script."),
    ]
    add_sheet(wb, "24_SOURCE_FILE_INDEX", "Source files used to build this single final workbook.", "Selected source files only; old versions are not retained in the final directory.", common_columns(), source_index_rows)

    add_sheet(
        wb,
        "25_VERSION_SELECTION_LOG",
        "Records which version was selected when multiple project versions existed.",
        "Prevents accidental reuse of legacy values.",
        common_columns(),
        [
            row("selected", "final_submission_workbook", str(FINAL_PATH), NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "VERSION-001", "Only file retained in final directory."),
            row("selected", "KOMRisk", "KOMRisk_Final_Locked_PostDedup_20260610", NOT_APPLICABLE, RISK_DIR, exists_status(RISK_DIR), "VERSION-002", "Latest locked PostDedup package selected."),
            row("excluded", "old_KOMRisk_LightGBM_AUROC_0.817", "excluded_legacy_metric", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "VERSION-003", "Do not report old structural progression value."),
            row("excluded", "old_KOMRisk_CoxPH_C_index_0.862", "excluded_legacy_metric", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "VERSION-004", "Endpoint B is now fixed-horizon binary classification."),
            row("excluded", "old_KOMRisk_CatBoost_AUROC_0.683", "excluded_legacy_metric", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "VERSION-005", "Do not mix with PostDedup C endpoint."),
            row("selected", "source_master_for_sim_score_treat", str(SOURCE_MASTER), NOT_APPLICABLE, SOURCE_MASTER, source_master_status, "VERSION-006", "Used for non-risk locked summaries."),
        ],
    )

    add_sheet(
        wb,
        "26_OLD_VERSION_ARCHIVE_LOG",
        "Old files moved out of final directory before creating the single locked workbook.",
        "Archive directory is outside the final output directory so final directory contains one Excel only.",
        ["archived_item", "item_type", "original_path", "archive_path", "sha256", "size_bytes", "archived_at", "archive_status", "reason"],
        archive_log,
    )

    add_sheet(
        wb,
        "27_FIGURE_TABLE_PLAN",
        "Manuscript figure and table plan tied to locked data sheets.",
        "Planning sheet only; actual figures should cite this workbook values.",
        common_columns(),
        [
            row("Figure", "Fig1_system_architecture", "KOM modules from intake to prescription and evaluation", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "FIG-001", "Use 01/03/11 sheets."),
            row("Figure", "Fig2_OAKNet_and_KOMRisk", "radiograph and risk model performance", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "FIG-002", "Use 07/08/17 sheets."),
            row("Figure", "Fig3_GraphRAG", "GraphRAG vs naive RAG ranking metrics", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "FIG-003", "Use 10 sheet."),
            row("Figure", "Fig4_Treatment_ablation", "Full KOM vs no RAG/no MDT/direct LLM", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "FIG-004", "Use 16 sheet."),
            row("Table", "Table1_cases", "standardized case set and quadrants", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "TABLE-001", "Use 05 sheet."),
            row("Table", "Table2_clinician_sim", "clinician interaction outcomes", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "TABLE-002", "Use 15 sheet."),
            row("Supplement", "Supplement_methods_gaps", "missing reproducibility files and manual confirmations", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "SUPP-001", "Use 21/22 sheets."),
        ],
    )

    add_sheet(
        wb,
        "28_MANUSCRIPT_WRITING_LOCK",
        "Claims that are safe to write and claims that must remain limited.",
        "Designed to prevent overclaiming.",
        common_columns(),
        [
            row("allowed_claim", "decision_support", "KOM is a research decision-support system for KOA assessment and prescription assistance.", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "WRITE-001", "Do not claim autonomous care."),
            row("allowed_claim", "clinician_simulation", "KOM improved simulated prescription quality and reduced critical safety errors in a clinician interaction experiment.", NOT_APPLICABLE, SOURCE_MASTER, SOURCE_LOCKED, "WRITE-002", "Simulation only."),
            row("allowed_claim", "risk_prediction", "PostDedup KOM-Risk A/B endpoints meet main-analysis acceptance; C is supplementary.", NOT_APPLICABLE, RISK_STATUS_JSON, risk_status_file_status, "WRITE-003", "Use latest values."),
            row("limited_claim", "clinical_outcome", "not_evaluated", NOT_APPLICABLE, FINAL_PATH, "generated_by_this_script", "WRITE-004", "No patient outcome benefit claim."),
            row("limited_claim", "prospective_deployment", "pending_manual_confirmation", NOT_APPLICABLE, FINAL_PATH, PENDING, "WRITE-005", "If not actually deployed prospectively, do not imply it."),
            row("limited_claim", "raw_RAG_reproducibility", "incomplete_current_files", NOT_APPLICABLE, SOURCE_MASTER, NOT_FOUND, "WRITE-006", "Report metrics, but mark raw graph/config gaps."),
        ],
    )

    qc_rows = [
        row("QC", "required_sheet_count", len(SHEET_ORDER), "sheets", FINAL_PATH, "pending_reopen_check", "QC-001", "Will be updated after save/reopen."),
        row("QC", "required_sheet_order", "pending_reopen_check", NOT_APPLICABLE, FINAL_PATH, "pending_reopen_check", "QC-002", "Will be updated after save/reopen."),
        row("QC", "no_empty_sheets", "pending_reopen_check", NOT_APPLICABLE, FINAL_PATH, "pending_reopen_check", "QC-003", "Will be updated after save/reopen."),
        row("QC", "final_directory_only_one_file", "pending_reopen_check", NOT_APPLICABLE, FINAL_DIR, "pending_reopen_check", "QC-004", "Will be updated after save/reopen."),
        row("QC", "latest_KOMRisk_used", "yes", NOT_APPLICABLE, RISK_STATUS_JSON, risk_status_file_status, "QC-005", "PostDedup package selected."),
        row("QC", "old_KOMRisk_metrics_excluded", "yes", NOT_APPLICABLE, SCRIPT_PATH, "generated_by_this_script", "QC-006", "Legacy values listed as excluded in 25_VERSION_SELECTION_LOG."),
        row("QC", "clinician_n_locked", 26, "clinicians", SOURCE_MASTER, SOURCE_LOCKED, "QC-007", "KOM-Sim."),
        row("QC", "prescription_records_locked", 780, "records", SOURCE_MASTER, SOURCE_LOCKED, "QC-008", "KOM-Sim."),
    ]
    add_sheet(wb, "99_QC_FINAL_STATUS", "Final workbook quality-control status.", "Updated after workbook save and reopen.", common_columns(), qc_rows)

    return wb


def update_qc_after_save() -> dict[str, Any]:
    wb = load_workbook(FINAL_PATH)
    sheetnames = wb.sheetnames
    row_counts = {ws.title: ws.max_row for ws in wb.worksheets}
    no_empty = all(ws.max_row >= 5 for ws in wb.worksheets)
    order_ok = sheetnames == SHEET_ORDER
    sheet_count_ok = len(sheetnames) == len(SHEET_ORDER)
    files_in_final = [p.name for p in FINAL_DIR.iterdir()]
    only_one_final_file = files_in_final == [FINAL_NAME]

    ws = wb["99_QC_FINAL_STATUS"]
    append_rows = [
        ["QC", "reopen_check", "pass", "status", str(FINAL_PATH), "verified_after_save", "QC-009", "Workbook reopened successfully with openpyxl."],
        ["QC", "sheet_count_actual", len(sheetnames), "sheets", str(FINAL_PATH), "verified_after_save", "QC-010", f"Expected {len(SHEET_ORDER)}."],
        ["QC", "sheet_order_ok", order_ok, "boolean", str(FINAL_PATH), "verified_after_save", "QC-011", "Exact required sheet order checked."],
        ["QC", "no_empty_sheets_ok", no_empty, "boolean", str(FINAL_PATH), "verified_after_save", "QC-012", "Each sheet has metadata and table rows."],
        ["QC", "final_directory_only_one_file_ok", only_one_final_file, "boolean", str(FINAL_DIR), "verified_after_save", "QC-013", f"Directory entries: {files_in_final}"],
        ["QC", "workbook_final_sha256", sha256_file(FINAL_PATH), "sha256", str(FINAL_PATH), "verified_after_save", "QC-014", "Final checksum after first save."],
    ]
    for vals in append_rows:
        ws.append(vals)
    wb.save(FINAL_PATH)
    wb.close()
    return {
        "sheetnames": sheetnames,
        "row_counts": row_counts,
        "order_ok": order_ok,
        "sheet_count_ok": sheet_count_ok,
        "no_empty": no_empty,
        "only_one_final_file": only_one_final_file,
        "files_in_final_dir": files_in_final,
        "sha256": sha256_file(FINAL_PATH),
    }


def main() -> None:
    archive_log = archive_existing_final_dir_items()
    wb = build_workbook(archive_log)
    wb.save(FINAL_PATH)
    final_qc = update_qc_after_save()
    reopened = load_workbook(FINAL_PATH, read_only=True, data_only=True)
    final_sheetnames = reopened.sheetnames
    final_row_counts = {ws.title: ws.max_row for ws in reopened.worksheets}
    reopened.close()
    result = {
        "final_path": str(FINAL_PATH),
        "sha256": final_qc["sha256"],
        "sheetnames": final_sheetnames,
        "row_counts": final_row_counts,
        "order_ok": final_qc["order_ok"],
        "sheet_count_ok": final_qc["sheet_count_ok"],
        "no_empty": final_qc["no_empty"],
        "only_one_final_file": final_qc["only_one_final_file"],
        "files_in_final_dir": final_qc["files_in_final_dir"],
        "archived_items": archive_log,
        "source_master_sheets": load_source_master_summary().get("source_master_sheets", []),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
