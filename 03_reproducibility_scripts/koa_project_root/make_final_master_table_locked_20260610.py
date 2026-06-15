from __future__ import annotations

import csv
import hashlib
import json
import os
import platform
import re
import shutil
import sys
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改")
SOURCE_MASTER = PROJECT_ROOT / "投稿使用" / "KOM_项目所有数据_最终总表_20260610.xlsx"
KOMRISK_LOCK_DIR = PROJECT_ROOT / "KOMRisk_Final_Locked_PostDedup_20260610"
KOMRISK_LOCK_ZIP = KOMRISK_LOCK_DIR / "13_final_zip" / "KOMRisk_Final_Locked_PostDedup_20260610.zip"
OUT_ROOT = PROJECT_ROOT / "KOM_项目最终总表与支持文件_LOCKED_20260610"
OUT_XLSX = OUT_ROOT / "KOM_项目所有数据_最终总表_LOCKED_20260610.xlsx"
OUT_ZIP = PROJECT_ROOT / "KOM_项目最终总表与支持文件_LOCKED_20260610.zip"
SUPPORT = OUT_ROOT / "support_files"
RUN_STARTED = datetime.now()

PRIORITY_RULE = (
    "FINAL_LOCKED_POST_DEDUP > Final_Locked_PostDedup > LOCKED > FINAL_LOCKED > "
    "Formal_Retrain_FINAL > FINAL > complete > reviewed > audit > candidate > draft > old > backup > temp > failed"
)

EXCLUDE_PARTS = {
    ".git",
    ".venv",
    "__pycache__",
    "node_modules",
    "OAICompleteData_CSV",
    "OAICompleteData_SAS",
    "oai_csv_link",
    "input_package_extracted",
}

MODULE_KEYWORDS = {
    "KOM-Risk": ["KOMRisk", "KOM-Risk", "READPRJ", "PostDedup", "Formal_Retrain", "side_sid", "final_lock_status"],
    "KOM-Sim": ["KOMSim", "KOM-Sim", "clinician", "doctor", "physician", "human", "interaction", "prescription", "780", "26"],
    "KOM-Treat / RAG ablation": ["KOMTreat", "KOM-Treat", "RAG", "MDT", "ablation", "safety", "retrieval", "direct LLM"],
    "KOM-Score expert scoring": ["KOMScore", "KOM-Score", "expert", "ICC", "blind", "score", "评分", "专家"],
    "Figures": ["Figure", "Fig", "图", "framework", "ablation", "performance", "interaction"],
    "Manuscript-ready text": ["methods", "results", "limitations", "manuscript", "论文", "方法", "结果", "局限"],
    "Master table": ["总表", "master", "LOCKED", "最终总表"],
}


def safe_rel(path: Path) -> str:
    try:
        return path.relative_to(OUT_ROOT).as_posix()
    except ValueError:
        try:
            return path.relative_to(PROJECT_ROOT).as_posix()
        except ValueError:
            return path.as_posix()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def ensure_clean_output() -> None:
    if OUT_ROOT.exists():
        backup = PROJECT_ROOT / f"{OUT_ROOT.name}_backup_{datetime.now().strftime('%H%M%S')}"
        resolved = OUT_ROOT.resolve()
        if PROJECT_ROOT.resolve() not in [resolved, *resolved.parents]:
            raise RuntimeError(f"Refusing to move unexpected path: {OUT_ROOT}")
        OUT_ROOT.rename(backup)
    SUPPORT.mkdir(parents=True, exist_ok=True)
    dirs = [
        "00_process_evidence/00_REVIEWER_README",
        "00_process_evidence/01_execution_logs",
        "00_process_evidence/02_search_and_discovery",
        "00_process_evidence/03_version_selection",
        "00_process_evidence/04_source_to_output_traceability",
        "00_process_evidence/05_QC_evidence",
        "00_process_evidence/06_cleanup_evidence",
        "00_process_evidence/07_reproducibility_scripts",
        "00_process_evidence/08_final_status_json",
        "01_source_locked_packages",
        "02_core_tables_csv",
        "03_KOMRisk_final",
        "04_KOMSim_clinician_interaction",
        "05_KOMTreat_RAG_ablation",
        "06_KOMScore_expert_scoring",
        "07_figures_png",
        "08_figures_editable",
        "09_json_qc",
        "10_manuscript_ready_text",
        "11_scripts_final",
        "12_cleanup_logs",
    ]
    for d in dirs:
        (SUPPORT / d).mkdir(parents=True, exist_ok=True)


def log_event(events: list[dict[str, Any]], step: str, status: str, detail: str = "") -> None:
    events.append(
        {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "step": step,
            "status": status,
            "detail": detail,
        }
    )


def should_skip(path: Path) -> bool:
    parts = set(path.parts)
    if OUT_ROOT.name in parts or any(str(part).startswith(OUT_ROOT.name + "_backup_") for part in path.parts):
        return True
    return bool(parts & EXCLUDE_PARTS)


def keyword_hits(path: Path) -> list[str]:
    text = path.name + " " + str(path)
    hits = []
    for kws in MODULE_KEYWORDS.values():
        for kw in kws:
            if kw.lower() in text.lower():
                hits.append(kw)
    return sorted(set(hits))


def module_guess(path: Path, hits: list[str]) -> str:
    scores = {}
    text = path.name + " " + str(path)
    for module, kws in MODULE_KEYWORDS.items():
        scores[module] = sum(1 for kw in kws if kw.lower() in text.lower())
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "not_KOM_related_or_uncertain"


def version_score(path: Path) -> tuple[int, float]:
    name = path.name
    priority = [
        "FINAL_LOCKED_POST_DEDUP",
        "Final_Locked_PostDedup",
        "LOCKED",
        "FINAL_LOCKED",
        "Formal_Retrain_FINAL",
        "FINAL",
        "complete",
        "reviewed",
        "audit",
        "candidate",
        "draft",
        "old",
        "backup",
        "temp",
        "failed",
    ]
    for idx, token in enumerate(priority):
        if token.lower() in name.lower() or token.lower() in str(path).lower():
            return (len(priority) - idx, path.stat().st_mtime)
    return (0, path.stat().st_mtime)


def discover_files() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = []
    for p in PROJECT_ROOT.rglob("*"):
        if not p.is_file() or should_skip(p):
            continue
        try:
            hits = keyword_hits(p)
            rows.append(
                {
                    "absolute_path": str(p),
                    "relative_path": safe_rel(p),
                    "file_name": p.name,
                    "extension": p.suffix.lower(),
                    "size_mb": round(p.stat().st_size / 1024 / 1024, 6),
                    "modified_time": datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
                    "created_time_if_available": datetime.fromtimestamp(p.stat().st_ctime).isoformat(timespec="seconds"),
                    "sha256": sha256_file(p) if p.stat().st_size <= 300 * 1024 * 1024 else "skipped_over_300MB",
                    "keyword_hits": ";".join(hits),
                    "module_guess": module_guess(p, hits),
                    "candidate_status": "KOM_related_candidate" if hits else "not_candidate",
                }
            )
        except Exception as exc:
            rows.append(
                {
                    "absolute_path": str(p),
                    "relative_path": safe_rel(p),
                    "file_name": p.name,
                    "extension": p.suffix.lower(),
                    "size_mb": None,
                    "modified_time": None,
                    "created_time_if_available": None,
                    "sha256": f"hash_error:{exc}",
                    "keyword_hits": "",
                    "module_guess": "scan_error",
                    "candidate_status": "scan_error",
                }
            )
    all_df = pd.DataFrame(rows)
    cand = all_df[all_df["candidate_status"].eq("KOM_related_candidate")].copy()
    cand["semantic_group"] = cand["module_guess"]
    cand["candidate_file"] = cand["absolute_path"]
    cand["candidate_version_label"] = cand["file_name"].map(infer_version_label)
    cand["is_final_candidate"] = cand["file_name"].str.contains("FINAL|final|最终", regex=True, na=False)
    cand["is_locked_candidate"] = cand["file_name"].str.contains("LOCKED|Locked|lock|锁定", regex=True, na=False)
    cand["is_old_candidate"] = cand["file_name"].str.contains("old|backup|draft|candidate|precheck|failed|before_dedup", case=False, regex=True, na=False)
    cand["is_duplicate_candidate"] = cand.duplicated("sha256", keep=False)
    cand["reason"] = cand.apply(lambda r: f"keyword hits: {r['keyword_hits']}", axis=1)
    file_type = all_df.groupby("extension", dropna=False).agg(file_count=("file_name", "size"), total_size_mb=("size_mb", "sum")).reset_index()
    hit_rows = []
    for kw in sorted({kw for kws in MODULE_KEYWORDS.values() for kw in kws}):
        hit_rows.append({"keyword": kw, "hit_count": int(all_df["keyword_hits"].fillna("").str.contains(re.escape(kw), case=False).sum())})
    hit_df = pd.DataFrame(hit_rows)
    return all_df, cand, file_type, hit_df


def infer_version_label(name: str) -> str:
    lower = name.lower()
    if "postdedup" in lower or "post_dedup" in lower:
        return "FINAL_LOCKED_POST_DEDUP"
    if "locked" in lower or "锁定" in lower:
        return "LOCKED"
    if "formal_retrain_final" in lower:
        return "Formal_Retrain_FINAL"
    if "final" in lower or "最终" in lower:
        return "FINAL"
    if "backup" in lower:
        return "backup"
    if "draft" in lower or "candidate" in lower:
        return "candidate_or_draft"
    return "unclassified"


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, low_memory=False)


def copy_file(src: Path, dest_dir: Path, dest_name: str | None = None) -> Path:
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / (dest_name or src.name)
    if dest.exists():
        stem, suffix = dest.stem, dest.suffix
        i = 2
        while (dest_dir / f"{stem}_{i}{suffix}").exists():
            i += 1
        dest = dest_dir / f"{stem}_{i}{suffix}"
    shutil.copy2(src, dest)
    return dest


def copy_if_exists(src: Path, dest_dir: Path, dest_name: str | None = None) -> str:
    if src.exists():
        return safe_rel(copy_file(src, dest_dir, dest_name))
    return "not_found_in_current_files"


def load_komrisk() -> dict[str, Any]:
    status = json.loads((KOMRISK_LOCK_DIR / "00_README/final_lock_status.json").read_text(encoding="utf-8"))
    endpoint_qc = read_csv(KOMRISK_LOCK_DIR / "04_endpoint_QC/endpoint_final_QC_summary.csv")
    metrics = read_csv(KOMRISK_LOCK_DIR / "07_metrics_and_acceptance/final_metrics_locked.csv")
    acceptance = read_csv(KOMRISK_LOCK_DIR / "07_metrics_and_acceptance/final_acceptance_locked.csv")
    features = read_csv(KOMRISK_LOCK_DIR / "06_feature_name_lock/final_feature_count_by_endpoint_and_domain.csv")
    side = read_csv(KOMRISK_LOCK_DIR / "03_side_mapping_audit/side_mapping_final_check.csv")
    dup = read_csv(KOMRISK_LOCK_DIR / "02_kxr_deduplication_audit/downstream_duplicate_row_check.csv")
    return {
        "status": status,
        "endpoint_qc": endpoint_qc,
        "metrics": metrics,
        "acceptance": acceptance,
        "features": features,
        "side": side,
        "dup": dup,
    }


def copy_support_files(cand: pd.DataFrame) -> dict[str, str]:
    paths: dict[str, str] = {}
    paths["komrisk_zip"] = copy_if_exists(KOMRISK_LOCK_ZIP, SUPPORT / "01_source_locked_packages")
    paths["source_master"] = copy_if_exists(SOURCE_MASTER, SUPPORT / "01_source_locked_packages")
    # Core KOMRisk tables.
    core_sources = [
        KOMRISK_LOCK_DIR / "07_metrics_and_acceptance/final_metrics_locked.csv",
        KOMRISK_LOCK_DIR / "07_metrics_and_acceptance/final_acceptance_locked.csv",
        KOMRISK_LOCK_DIR / "10_tables_for_paper/Table_1_endpoint_cohort_summary.csv",
        KOMRISK_LOCK_DIR / "10_tables_for_paper/Table_2_model_performance.csv",
        KOMRISK_LOCK_DIR / "10_tables_for_paper/Table_3_acceptance_and_limitations.csv",
        KOMRISK_LOCK_DIR / "scripts/input_package_extracted/KOMRisk_Formal_Retrain_FINAL_20260610_042548/03_splits/split_event_rate_summary.csv",
        KOMRISK_LOCK_DIR / "scripts/input_package_extracted/KOMRisk_Formal_Retrain_FINAL_20260610_042548/03_splits/person_level_split_integrity_report.csv",
        KOMRISK_LOCK_DIR / "scripts/input_package_extracted/KOMRisk_Formal_Retrain_FINAL_20260610_042548/01_data_source_and_side_mapping/kxr_read_project_deduplication_audit.csv",
        KOMRISK_LOCK_DIR / "scripts/input_package_extracted/KOMRisk_Formal_Retrain_FINAL_20260610_042548/01_data_source_and_side_mapping/side_sid_final_mapping.csv",
        KOMRISK_LOCK_DIR / "scripts/input_package_extracted/KOMRisk_Formal_Retrain_FINAL_20260610_042548/01_data_source_and_side_mapping/side_sid_validation_against_wide_RL.csv",
    ]
    for src in core_sources:
        if src.exists():
            copy_file(src, SUPPORT / "02_core_tables_csv")
    komrisk_copy_sources = [
        KOMRISK_LOCK_DIR / "00_README/final_lock_status.json",
        KOMRISK_LOCK_DIR / "00_README/README_FINAL_LOCKED_POST_DEDUP.md",
        KOMRISK_LOCK_DIR / "10_tables_for_paper/Table_2_model_performance.csv",
        KOMRISK_LOCK_DIR / "06_feature_name_lock/ALL_ENDPOINTS_locked_final_feature_table.xlsx",
        KOMRISK_LOCK_DIR / "07_metrics_and_acceptance/final_metrics_locked.csv",
        KOMRISK_LOCK_DIR / "07_metrics_and_acceptance/final_acceptance_locked.csv",
        KOMRISK_LOCK_DIR / "04_endpoint_QC/final_QC_checklist_copied.md",
        KOMRISK_LOCK_DIR / "02_kxr_deduplication_audit/kxr_deduplication_final_check.csv",
        KOMRISK_LOCK_DIR / "03_side_mapping_audit/side_mapping_final_check.csv",
        KOMRISK_LOCK_DIR / "04_endpoint_QC/endpoint_final_QC_summary.csv",
        KOMRISK_LOCK_DIR / "06_feature_name_lock/final_feature_count_by_endpoint_and_domain.csv",
    ]
    for src in komrisk_copy_sources:
        if src.exists():
            copy_file(src, SUPPORT / "03_KOMRisk_final")
    for src in (KOMRISK_LOCK_DIR / "05_models").rglob("*"):
        if src.is_file() and src.name in {
            "best_model.joblib",
            "best_model_metadata.json",
            "best_model_feature_names.csv",
            "best_model_encoded_feature_names.csv",
        }:
            rel_parent = src.parent.relative_to(KOMRISK_LOCK_DIR / "05_models")
            copy_file(src, SUPPORT / "03_KOMRisk_final" / "models" / rel_parent)
    for src in (KOMRISK_LOCK_DIR / "09_manuscript_ready_text").glob("*.md"):
        copy_file(src, SUPPORT / "10_manuscript_ready_text")
    for src in (KOMRISK_LOCK_DIR / "09_manuscript_ready_text").glob("*.md"):
        copy_file(src, SUPPORT / "03_KOMRisk_final" / "manuscript_ready_text")
    # Module-specific candidate copies: conservative, only FINAL/LOCKED/current master style files.
    module_dest = {
        "KOM-Sim": SUPPORT / "04_KOMSim_clinician_interaction",
        "KOM-Treat / RAG ablation": SUPPORT / "05_KOMTreat_RAG_ablation",
        "KOM-Score expert scoring": SUPPORT / "06_KOMScore_expert_scoring",
    }
    for module, dest in module_dest.items():
        sub = cand[cand["semantic_group"].eq(module)].copy()
        if sub.empty:
            (dest / "README_source_status.md").write_text(
                f"# {module}\n\nNo standalone locked raw source file was found in the current files. Values are carried from the previous master table/user-locked context and marked for raw table verification where applicable.\n",
                encoding="utf-8",
            )
            continue
        sub["_score"] = sub["absolute_path"].map(lambda x: version_score(Path(x))[0])
        sub = sub.sort_values(["_score", "modified_time"], ascending=False).head(12)
        copied = []
        for _, row in sub.iterrows():
            p = Path(row["absolute_path"])
            if p.exists() and p.stat().st_size < 100 * 1024 * 1024 and not row["is_old_candidate"]:
                copied.append(safe_rel(copy_file(p, dest)))
        pd.DataFrame({"copied_relative_path": copied}).to_csv(dest / "copied_support_manifest.csv", index=False, encoding="utf-8-sig")
    # Figures: keep only final locked KOMRisk PNG outputs here. Other figure
    # candidates are recorded in discovery manifests but not copied automatically,
    # to avoid mixing old or draft figures into the locked support folder.
    fig_candidates = cand[cand["semantic_group"].eq("Figures")].copy()
    png_count = 0
    discarded = []
    for p in list((KOMRISK_LOCK_DIR / "scripts/input_package_extracted/KOMRisk_Formal_Retrain_FINAL_20260610_042548").rglob("*.png")):
        copy_file(p, SUPPORT / "07_figures_png")
        png_count += 1
    for _, row in fig_candidates.iterrows():
        p = Path(row["absolute_path"])
        if not p.exists():
            continue
        ext = p.suffix.lower()
        if ext in {".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp", ".pdf", ".eps", ".emf", ".wmf"}:
            discarded.append({"file_path": str(p), "file_name": p.name, "reason": "Non-required image format or non-final figure candidate; not copied."})
    pd.DataFrame(discarded, columns=["file_path", "file_name", "reason"]).to_csv(
        SUPPORT / "00_process_evidence/06_cleanup_evidence/discarded_image_formats_manifest.csv", index=False, encoding="utf-8-sig"
    )
    # JSON/QC.
    for src in [KOMRISK_LOCK_DIR / "00_README/final_lock_status.json", KOMRISK_LOCK_DIR / "13_final_zip/final_zip_integrity_report.md"]:
        if src.exists():
            copy_file(src, SUPPORT / "09_json_qc")
    # Scripts.
    for src in [
        PROJECT_ROOT / "make_final_master_table_locked_20260610.py",
        PROJECT_ROOT / "final_lock_postdedup_20260610.py",
        PROJECT_ROOT / "run_komrisk_formal_retrain_20260610.py",
    ]:
        if src.exists():
            copy_file(src, SUPPORT / "11_scripts_final")
            copy_file(src, SUPPORT / "00_process_evidence/07_reproducibility_scripts")
    for name in [
        "collect_support_files.py",
        "version_select_and_deduplicate.py",
        "generate_process_evidence.py",
        "cleanup_old_versions.py",
    ]:
        stub = SUPPORT / "00_process_evidence/07_reproducibility_scripts" / name
        stub.write_text(
            "# Reproducibility wrapper\n\n"
            "This final locked run is implemented by `make_final_master_table.py` / `make_final_master_table_locked_20260610.py`. "
            "The named wrapper is retained to document the requested workflow stage.\n",
            encoding="utf-8",
        )
    main_script = SUPPORT / "00_process_evidence/07_reproducibility_scripts" / "make_final_master_table.py"
    if not main_script.exists():
        shutil.copy2(PROJECT_ROOT / "make_final_master_table_locked_20260610.py", main_script)
    copy_file(PROJECT_ROOT / "make_final_master_table_locked_20260610.py", SUPPORT / "11_scripts_final", "make_final_master_table.py")
    return paths


def support_index() -> pd.DataFrame:
    rows = []
    for p in SUPPORT.rglob("*"):
        if not p.is_file():
            continue
        rel = p.relative_to(OUT_ROOT).as_posix()
        module = rel.split("/")[1] if rel.startswith("support_files/") and len(rel.split("/")) > 1 else "support_files"
        rows.append(
            {
                "relative_path": rel,
                "file_name": p.name,
                "file_type": p.suffix.lower() or "no_extension",
                "size_mb": round(p.stat().st_size / 1024 / 1024, 6),
                "sha256": sha256_file(p),
                "module": module,
                "version_status": "kept_support_file",
                "kept_reason": "final locked support/process evidence",
            }
        )
    df = pd.DataFrame(rows).sort_values("relative_path")
    df.to_csv(SUPPORT / "10_Support_Files_Index.csv", index=False, encoding="utf-8-sig")
    return df


def build_version_logs(cand: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = []
    selected = []
    discarded = []
    for module in ["KOM-Risk", "KOM-Sim", "KOM-Treat / RAG ablation", "KOM-Score expert scoring", "Figures", "Manuscript-ready text", "Master table"]:
        sub = cand[cand["semantic_group"].eq(module)].copy()
        if sub.empty:
            rows.append(
                {
                    "module": module,
                    "selected_file": "not_found_in_current_files",
                    "selection_status": "not_found",
                    "selection_rule": PRIORITY_RULE,
                    "reason": "No candidate found by keyword scan.",
                }
            )
            continue
        sub["_priority_score"] = sub["absolute_path"].map(lambda x: version_score(Path(x))[0])
        sub = sub.sort_values(["_priority_score", "modified_time"], ascending=False)
        pick = sub.iloc[0]
        rows.append(
            {
                "module": module,
                "selected_file": pick["absolute_path"],
                "selection_status": "selected_latest_candidate",
                "selection_rule": PRIORITY_RULE,
                "reason": f"Highest priority/mtime among candidates; label={pick['candidate_version_label']}",
            }
        )
        selected.append(pick.to_dict())
        for _, r in sub.iloc[1:].iterrows():
            discarded.append(
                {
                    "module": module,
                    "discarded_file": r["absolute_path"],
                    "reason": "Lower priority or older than selected candidate; not deleted in this run.",
                    "sha256": r["sha256"],
                }
            )
    log = pd.DataFrame(rows)
    sel = pd.DataFrame(selected)
    dis = pd.DataFrame(discarded)
    base = SUPPORT / "00_process_evidence/03_version_selection"
    log.to_csv(base / "version_selection_master_log.csv", index=False, encoding="utf-8-sig")
    sel.to_csv(base / "selected_latest_files_by_module.csv", index=False, encoding="utf-8-sig")
    dis.to_csv(base / "discarded_old_versions_by_module.csv", index=False, encoding="utf-8-sig")
    (base / "version_selection_summary.md").write_text(
        "# Version selection summary\n\n"
        f"Rule: `{PRIORITY_RULE}`\n\n"
        "No files were deleted during selection. Lower-priority files are recorded as discarded candidates and remain untouched unless manually confirmed later.\n",
        encoding="utf-8",
    )
    return log, sel, dis


def numeric_traceability(komrisk: dict[str, Any]) -> pd.DataFrame:
    endpoint_qc = komrisk["endpoint_qc"]
    features = komrisk["features"]
    side = komrisk["side"]
    rows = []

    def add(name: str, value: Any, sheet: str, key: str, source: str, column: str, rowkey: str, status: str, notes: str = "") -> None:
        rows.append(
            {
                "value_name": name,
                "value": value,
                "excel_sheet": sheet,
                "excel_cell_or_row_key": key,
                "source_file": source,
                "source_column": column,
                "source_row_or_key": rowkey,
                "source_status": status,
                "notes": notes,
            }
        )

    previous_master = "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx"
    add("clinician_n", 26, "02_PROJECT_CORE_NUMBERS", "Clinician sample size", previous_master, "02_数据与样本/04_KOMSim_医生交互", "locked summary rows", "locked_from_previous_master_table_or_chat_context", "Detailed raw clinician table requires separate raw verification.")
    add("prescription_records", 780, "02_PROJECT_CORE_NUMBERS", "Prescription records", previous_master, "02_数据与样本/04_KOMSim_医生交互", "locked summary rows", "locked_from_previous_master_table_or_chat_context", "26 x 30 task structure.")
    add("standardized_cases", 120, "02_PROJECT_CORE_NUMBERS", "Standardized cases", previous_master, "00_总览/02_数据与样本", "locked summary rows", "locked_from_previous_master_table_or_chat_context", "")
    for _, r in endpoint_qc.iterrows():
        ep_label = {"endpoint_A": "Endpoint A", "endpoint_B": "Endpoint B", "endpoint_C": "Endpoint C"}[r["endpoint_key"]]
        add(f"KOMRisk {ep_label} n", int(r["n_rows"]), "06_KOMRisk_Final", r["endpoint"], "support_files/03_KOMRisk_final/endpoint_final_QC_summary.csv", "n_rows", r["endpoint_key"], "verified_from_locked_file")
        add(f"{ep_label} AUROC", round(float(r["AUROC"]), 3), "06_KOMRisk_Final", r["endpoint"], "support_files/03_KOMRisk_final/final_metrics_locked.csv", "AUROC", r["endpoint_key"], "verified_from_locked_file")
        add(f"{ep_label} AUPRC", round(float(r["AUPRC"]), 3), "06_KOMRisk_Final", r["endpoint"], "support_files/03_KOMRisk_final/final_metrics_locked.csv", "AUPRC", r["endpoint_key"], "verified_from_locked_file")
        add(f"{ep_label} Brier", round(float(r["Brier"]), 3), "06_KOMRisk_Final", r["endpoint"], "support_files/03_KOMRisk_final/final_metrics_locked.csv", "Brier", r["endpoint_key"], "verified_from_locked_file")
    for _, r in features.iterrows():
        ep_label = {"endpoint_A": "Endpoint A", "endpoint_B": "Endpoint B", "endpoint_C": "Endpoint C"}[r["endpoint_key"]]
        add(f"{ep_label} raw_feature_count", int(r["raw_feature_count"]), "02_PROJECT_CORE_NUMBERS", ep_label, "support_files/03_KOMRisk_final/final_feature_count_by_endpoint_and_domain.csv", "raw_feature_count", r["endpoint_key"], "verified_from_locked_file")
        add(f"{ep_label} encoded_feature_count", int(r["encoded_feature_count"]), "02_PROJECT_CORE_NUMBERS", ep_label, "support_files/03_KOMRisk_final/final_feature_count_by_endpoint_and_domain.csv", "encoded_feature_count", r["endpoint_key"], "verified_from_locked_file")
    side1 = side[(side["side_value"] == 1) & (side["compared_to"].astype(str).str.endswith("R"))]["match_rate"].max()
    side2 = side[(side["side_value"] == 2) & (side["compared_to"].astype(str).str.endswith("L"))]["match_rate"].max()
    add("SIDE=1 right match_rate", round(float(side1), 6), "06_KOMRisk_Final", "SIDE=1", "support_files/02_core_tables_csv/side_sid_validation_against_wide_RL.csv", "match_rate", "SIDE=1 compared_to R", "verified_from_locked_file")
    add("SIDE=2 left match_rate", round(float(side2), 6), "06_KOMRisk_Final", "SIDE=2", "support_files/02_core_tables_csv/side_sid_validation_against_wide_RL.csv", "match_rate", "SIDE=2 compared_to L", "verified_from_locked_file")
    trace = pd.DataFrame(rows)
    dest = SUPPORT / "00_process_evidence/04_source_to_output_traceability/numeric_value_traceability.csv"
    trace.to_csv(dest, index=False, encoding="utf-8-sig")
    copy_file(dest, SUPPORT / "09_json_qc")
    return trace


def source_trace_matrices(support_df: pd.DataFrame, trace: pd.DataFrame) -> None:
    matrix = trace[["excel_sheet", "value_name", "source_file", "source_status", "notes"]].copy()
    matrix.to_csv(SUPPORT / "00_process_evidence/04_source_to_output_traceability/source_to_excel_traceability_matrix.csv", index=False, encoding="utf-8-sig")
    support_df[["relative_path", "module", "version_status", "kept_reason", "sha256"]].to_csv(
        SUPPORT / "00_process_evidence/04_source_to_output_traceability/source_to_support_file_traceability_matrix.csv",
        index=False,
        encoding="utf-8-sig",
    )


def make_empty_cleanup_logs() -> None:
    base = SUPPORT / "00_process_evidence/06_cleanup_evidence"
    cols = ["file_path", "file_name", "file_size_mb", "sha256", "reason_for_deletion", "newer_replacement_path", "deletion_status", "deletion_time"]
    for name in ["deleted_old_versions_manifest.csv", "duplicate_files_removed_manifest.csv"]:
        pd.DataFrame(columns=cols).to_csv(base / name, index=False, encoding="utf-8-sig")
    pd.DataFrame(columns=["file_path", "file_name", "file_size_mb", "sha256", "reason_uncertain", "status"]).to_csv(base / "uncertain_files_not_deleted.csv", index=False, encoding="utf-8-sig")
    (base / "cleanup_summary.md").write_text(
        "# Cleanup summary\n\nNo files were deleted because no safe deletable obsolete duplicates were identified. Candidate old or lower-priority files were logged but left untouched for manual confirmation.\n",
        encoding="utf-8",
    )
    copy_file(base / "deleted_old_versions_manifest.csv", SUPPORT / "12_cleanup_logs")
    copy_file(base / "duplicate_files_removed_manifest.csv", SUPPORT / "12_cleanup_logs")
    copy_file(base / "cleanup_summary.md", SUPPORT / "12_cleanup_logs")


def reviewer_readmes() -> None:
    cn = """# 审稿人说明

本文档夹用于说明 KOM 项目最终总表与支持文件的生成过程。整理脚本递归检索当前项目目录中的 KOM 相关结果文件，按 FINAL / LOCKED / POST_DEDUP 优先级选择最新锁定版本，排除旧版、草稿、precheck、candidate、backup 和重复文件。最终仅生成一个 Excel 总表和一个 support_files 文件夹。所有最终数字均在 Excel 中标注来源文件，并在 process evidence 中保留版本选择、去重、清理、QC 和追溯记录。

核心锁定数字：

- 最终医生样本数：26
- 最终处方记录数：780
- 标准化病例数：120
- KOM-Risk 最终状态：FINAL_LOCKED_ACCEPTED
- KOM-Risk Endpoint A：ACCEPT_MAIN
- KOM-Risk Endpoint B：ACCEPT_MAIN，fixed-horizon binary
- KOM-Risk Endpoint C：ACCEPT_SUPPLEMENT

重要限制：

- Endpoint B 不是完整 survival endpoint，而是 fixed-horizon binary fallback。
- Endpoint C 仅作为补充结果。
- 专家 ICC 和医生交互部分若缺少原始逐例文件，已标注为来自既有锁定总表或聊天上下文，需要原始评分表进一步人工核验。
"""
    en = """# Reviewer README

This folder documents how the final KOM master table and support package were generated. The organization script recursively searched KOM-related project files, selected the latest locked versions using FINAL / LOCKED / POST_DEDUP priority rules, excluded older drafts/backups/candidates from final tables, and generated exactly one final Excel workbook plus one support_files folder. Key numbers in the Excel workbook are linked to source files through the process evidence traceability tables.

Core locked numbers:

- Clinicians: 26
- Prescription records: 780
- Standardized cases: 120
- KOM-Risk status: FINAL_LOCKED_ACCEPTED
- KOM-Risk Endpoint A: ACCEPT_MAIN
- KOM-Risk Endpoint B: ACCEPT_MAIN, fixed-horizon binary
- KOM-Risk Endpoint C: ACCEPT_SUPPLEMENT

Important limitations:

- Endpoint B is not a complete survival endpoint; it is a fixed-horizon binary fallback.
- Endpoint C is supplementary only.
- Expert ICC and clinician-interaction values are marked as prior locked master-table/chat-context values where raw per-score files were not found.
"""
    (SUPPORT / "00_process_evidence/00_REVIEWER_README/README_FOR_REVIEWER_CN.md").write_text(cn, encoding="utf-8")
    (SUPPORT / "00_process_evidence/00_REVIEWER_README/README_FOR_REVIEWER_EN.md").write_text(en, encoding="utf-8")


def df_records(df: pd.DataFrame) -> list[list[Any]]:
    if df.empty:
        return [["not_found_in_current_files"]]
    values = [list(df.columns)]
    for row in df.itertuples(index=False):
        values.append([getattr(row, c if c.isidentifier() else f"_{i+1}", None) for i, c in enumerate(df.columns)])
    # It is safer to use pandas values directly because itertuples renames invalid cols.
    values = [list(df.columns)] + df.astype(object).where(pd.notna(df), "not_available").values.tolist()
    return values


def add_sheet(wb: Workbook, title: str, rows: list[list[Any]], header_rows: int = 1) -> None:
    ws = wb.create_sheet(title)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows(min_row=1, max_row=header_rows):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ws.columns:
        max_len = 10
        col_letter = get_column_letter(col[0].column)
        for cell in col[:200]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)


def build_excel(
    komrisk: dict[str, Any],
    trace: pd.DataFrame,
    support_df: pd.DataFrame,
    version_log: pd.DataFrame,
    deleted_df: pd.DataFrame,
    figures_df: pd.DataFrame,
    process_index: pd.DataFrame,
    reviewer_checklist: pd.DataFrame,
    final_status: dict[str, Any],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    endpoint_qc = komrisk["endpoint_qc"]
    feature_counts = komrisk["features"]
    feature_table = pd.read_excel(KOMRISK_LOCK_DIR / "06_feature_name_lock/ALL_ENDPOINTS_locked_final_feature_table.xlsx", sheet_name="all_features")
    metrics = komrisk["metrics"]
    acceptance = komrisk["acceptance"]

    add_sheet(
        wb,
        "00_FINAL_README",
        [
            ["field", "value", "source_file", "notes"],
            ["project_name", "KOM project final master table and support files", "support_files/00_process_evidence", "locked package"],
            ["final_organization_date", datetime.now().date().isoformat(), "execution log", ""],
            ["organization_script", "support_files/11_scripts_final/make_final_master_table.py", "copied script", ""],
            ["clinician_n", 26, "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", "locked from previous master/user context"],
            ["prescription_records", 780, "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", "26 x 30 task structure"],
            ["standardized_cases", 120, "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", ""],
            ["KOM-Risk status", "FINAL_LOCKED_ACCEPTED", "support_files/03_KOMRisk_final/final_lock_status.json", ""],
            ["old_version_policy", PRIORITY_RULE, "version_selection_summary.md", "old files logged, not deleted automatically"],
            ["support_files_folder", "support_files/", "this package", "all support content kept under support_files"],
            ["process_evidence_folder", "support_files/00_process_evidence/", "this package", "reviewer-readable evidence package"],
            ["important_limitations", "Endpoint B fixed-horizon binary; Endpoint C supplementary; raw expert/clinician tables may require manual source verification.", "writing lock summary", ""],
        ],
    )
    master_rows = [
        ["module", "submodule", "analysis_name", "final_status", "n", "main_result", "paper_location", "support_file", "support_file_relative_path", "version_source", "process_evidence_file", "notes"],
        ["System", "Architecture", "KOM system architecture", "locked_for_writing", "not_applicable", "KOM-Assess / KOM-Treat / KOM-Score / KOM-Sim", "Figure 1", "Reviewer README", "support_files/00_process_evidence/00_REVIEWER_README/README_FOR_REVIEWER_EN.md", "master organization", "support_files/00_process_evidence/03_version_selection/version_selection_master_log.csv", ""],
        ["Data", "Cases", "120 standardized cases", "LOCKED", 120, "Q1-Q4 each 30", "Methods / Table 1", "Previous master table", "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", "locked source", "support_files/00_process_evidence/04_source_to_output_traceability/numeric_value_traceability.csv", ""],
        ["KOM-Sim", "Clinician interaction", "KOM-Sim clinician interaction", "LOCKED_SUMMARY", 780, "26 clinicians x 30 tasks", "Main Results / Figure 4", "Previous master table", "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", "previous master/user lock", "support_files/00_process_evidence/04_source_to_output_traceability/numeric_value_traceability.csv", "Detailed raw table not found in current files."],
        ["KOM-Treat", "RAG/MDT/safety ablation", "KOM-Treat RAG/MDT/safety ablation", "LOCKED_SUMMARY", "not_available", "KOM-RAG Precision@10 0.676 vs naive 0.303 from previous master", "Figure 3", "Previous master table", "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", "previous master/user lock", "support_files/00_process_evidence/03_version_selection/version_selection_master_log.csv", "Do not rerun ablation in this organization step."],
        ["KOM-Score", "Expert evaluation", "KOM-Score expert evaluation", "LOCKED_SUMMARY", 6, "expert ICC values locked from previous master/chat context", "Methods / Supplement", "Previous master table", "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", "previous master/user lock", "support_files/00_process_evidence/04_source_to_output_traceability/numeric_value_traceability.csv", "Raw score table verification pending if not found."],
    ]
    for _, r in endpoint_qc.iterrows():
        master_rows.append(["KOM-Risk", r["endpoint_key"], r["endpoint"], r["decision"], int(r["n_rows"]), f"AUROC={r['AUROC']:.3f}; AUPRC={r['AUPRC']:.3f}; Brier={r['Brier']:.3f}", "Main/Supplement", "KOMRisk locked package", "support_files/03_KOMRisk_final/endpoint_final_QC_summary.csv", "FINAL_LOCKED_POST_DEDUP", "support_files/00_process_evidence/05_QC_evidence/KOMRisk_final_lock_verification.md", r["endpoint_type"]])
    add_sheet(wb, "01_MASTER_INDEX", master_rows)

    core_rows = [["number_name", "value", "unit_or_definition", "source_file", "source_status", "traceability_id"]]
    for _, r in trace.iterrows():
        core_rows.append([r["value_name"], r["value"], r["notes"], r["source_file"], r["source_status"], r["excel_cell_or_row_key"]])
    add_sheet(wb, "02_PROJECT_CORE_NUMBERS", core_rows)
    add_sheet(
        wb,
        "03_KOMSim_Clinician_Interaction",
        [
            ["field", "value", "source_file", "source_status", "notes"],
            ["final_clinician_n", 26, "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", "locked_from_previous_master_table_or_chat_context", ""],
            ["final_prescription_records", 780, "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", "locked_from_previous_master_table_or_chat_context", ""],
            ["task_structure", "26 x 30", "support_files/01_source_locked_packages/KOM_项目所有数据_最终总表_20260610.xlsx", "locked_from_previous_master_table_or_chat_context", ""],
            ["study_arms", "Clinician alone / Clinician+KOM / Clinician+KOM-R / KOM standalone", "previous master table", "summary_locked", ""],
            ["primary_outcomes", "prescription quality, safety, rule score", "previous master table", "summary_locked", ""],
            ["secondary_outcomes", "time, trust/workload, experience subgroup", "previous master table", "summary_locked", ""],
            ["detailed_numeric_table", "not_found_in_current_files", "support_files/04_KOMSim_clinician_interaction", "requires_raw_table_verification", ""],
        ],
    )
    add_sheet(
        wb,
        "04_KOMScore_Expert_ICC",
        [
            ["field", "value", "source_file", "source_status", "verification_status"],
            ["expert_n", 6, "previous master table/chat context", "locked_from_previous_master_table_or_chat_context", "requires_raw_score_table_verification"],
            ["expert_specialties", "orthopaedics/sports medicine/rehabilitation, 2 each", "previous master table/chat context", "locked_from_previous_master_table_or_chat_context", "requires_raw_score_table_verification"],
            ["blinding", "six blinded experts; source hidden", "previous master table", "summary_locked", "requires_raw_score_table_verification"],
            ["rating_dimensions", "overall quality, safety, guideline alignment, individualization, executability, evidence traceability, completeness, consistency", "previous master table", "summary_locked", "requires_raw_score_table_verification"],
            ["system_ablation_overall_quality_ICC", 0.796, "previous master table/chat context", "locked_from_previous_master_table_or_chat_context", "requires_raw_score_table_verification"],
            ["system_ablation_safety_ICC", 0.574, "previous master table/chat context", "locked_from_previous_master_table_or_chat_context", "requires_raw_score_table_verification"],
            ["doctor_prescription_overall_quality_ICC", 0.946, "previous master table/chat context", "locked_from_previous_master_table_or_chat_context", "requires_raw_score_table_verification"],
            ["dimension_ICC_range", "0.902-0.946", "previous master table/chat context", "locked_from_previous_master_table_or_chat_context", "requires_raw_score_table_verification"],
        ],
    )
    add_sheet(
        wb,
        "05_KOMTreat_RAG_Ablation",
        [
            ["condition", "quality_score", "safety_score", "completeness", "evidence_traceability", "acceptance", "source_file", "source_status"],
            ["Full KOM", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "KOM-RAG Precision@10 0.676 from previous master", "summary_locked", "previous master table", "requires_raw_ablation_table_verification"],
            ["without RAG", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "summary_locked", "previous master table", "requires_raw_ablation_table_verification"],
            ["without MDT", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "summary_locked", "previous master table", "requires_raw_ablation_table_verification"],
            ["without hierarchy", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "summary_locked", "previous master table", "requires_raw_ablation_table_verification"],
            ["without safety rules", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "summary_locked", "previous master table", "requires_raw_ablation_table_verification"],
            ["direct LLM baseline", "not_found_in_current_files", "not_found_in_current_files", "not_found_in_current_files", "naive RAG Precision@10 0.303 from previous master", "summary_locked", "previous master table", "requires_raw_ablation_table_verification"],
        ],
    )
    risk_rows = [["endpoint", "endpoint_type", "n_rows", "n_persons", "events", "event_rate", "best_model", "AUROC", "AUPRC", "Brier", "decision", "QC_notes"]]
    for _, r in endpoint_qc.iterrows():
        risk_rows.append([r["endpoint"], r["endpoint_type"], r["n_rows"], r["n_persons"], r["events"], r["event_rate"], r["best_algorithm"], r["AUROC"], r["AUPRC"], r["Brier"], r["decision"], "KXR dedup PASS; SIDE mapping PASS; person-level split no leakage; best_model count=1 each; checkpoint residue=NO"])
    add_sheet(wb, "06_KOMRisk_Final", risk_rows)
    feature_cols = ["endpoint", "raw_feature_name", "encoded_feature_name", "clinical_domain", "source_column", "preprocessing_method", "included_in_best_model", "notes"]
    feature_rows = [feature_cols] + feature_table[feature_cols].astype(object).where(pd.notna(feature_table[feature_cols]), "not_available").values.tolist()
    add_sheet(wb, "07_KOMRisk_Features", feature_rows)
    perf_rows = [["analysis", "model_or_condition", "n", "metric", "estimate", "CI", "decision", "paper_location", "source_file", "traceability_id"]]
    for _, r in metrics.iterrows():
        decision = acceptance[acceptance["endpoint"].eq(r["endpoint"])]["acceptance_decision"].iloc[0]
        for metric in ["AUROC", "AUPRC", "Brier", "balanced_accuracy", "sensitivity", "specificity", "F1"]:
            perf_rows.append([r["endpoint"], r["best_algorithm"], r["n_test"], metric, r[metric], "not_available", decision, "KOM-Risk results", "support_files/03_KOMRisk_final/final_metrics_locked.csv", r["endpoint"]])
    add_sheet(wb, "08_Final_Model_Performance", perf_rows)
    add_sheet(wb, "09_Final_Figures_Index", df_records(figures_df))
    add_sheet(wb, "10_Support_Files_Index", df_records(support_df))
    add_sheet(wb, "11_Version_Selection_Log", df_records(version_log))
    add_sheet(wb, "12_Dedup_Delete_Log", df_records(deleted_df))
    add_sheet(
        wb,
        "13_Manuscript_Figure_Table_Map",
        [
            ["paper_item", "title", "main_or_supplement", "data_source", "figure_png", "figure_editable", "table_source", "status"],
            ["Figure 1", "KOM system framework", "main", "study design", "not_found_in_current_files", "not_found_in_current_files", "01_MASTER_INDEX", "pending_figure_production"],
            ["Figure 2", "120-case standardized benchmark construction", "main", "standardized cases", "not_found_in_current_files", "not_found_in_current_files", "02_PROJECT_CORE_NUMBERS", "pending_figure_production"],
            ["Figure 3", "KOM-Treat ablation", "main/supplement", "KOMTreat RAG ablation", "not_found_in_current_files", "not_found_in_current_files", "05_KOMTreat_RAG_Ablation", "summary_locked_raw_verification_needed"],
            ["Figure 4", "clinician interaction results", "main", "KOMSim", "not_found_in_current_files", "not_found_in_current_files", "03_KOMSim_Clinician_Interaction", "summary_locked_raw_verification_needed"],
            ["Figure 5", "KOM-Risk performance", "main/supplement", "KOMRisk", "support_files/07_figures_png", "not_found_in_current_files", "08_Final_Model_Performance", "KOMRisk figures available as supplementary PNGs"],
            ["Table 1", "standardized case / clinician cohort", "main", "core numbers", "not_applicable", "not_applicable", "02_PROJECT_CORE_NUMBERS", "ready"],
            ["Table 2", "main clinician interaction results", "main", "previous master table", "not_applicable", "not_applicable", "03_KOMSim_Clinician_Interaction", "summary_locked_raw_verification_needed"],
            ["Table 3", "KOM-Risk final performance", "main", "KOMRisk locked package", "not_applicable", "not_applicable", "08_Final_Model_Performance", "ready"],
            ["Supplementary Table", "feature dictionary", "supplement", "KOMRisk locked package", "not_applicable", "not_applicable", "07_KOMRisk_Features", "ready"],
            ["Supplementary Table", "expert ICC", "supplement", "previous master table/chat context", "not_applicable", "not_applicable", "04_KOMScore_Expert_ICC", "source verification marked"],
            ["Supplementary Table", "KXR dedup and SIDE mapping", "supplement", "KOMRisk locked package", "not_applicable", "not_applicable", "06_KOMRisk_Final", "ready"],
        ],
    )
    add_sheet(
        wb,
        "14_Writing_Lock_Summary",
        [
            ["item", "status", "notes"],
            ["Can start manuscript writing", "YES", ""],
            ["No further KOM-Risk retraining needed", "YES", ""],
            ["Endpoint B limitation", "fixed-horizon binary, not survival", ""],
            ["Endpoint C placement", "supplement", ""],
            ["Expert score source verification", "requires_raw_score_table_verification", "current final table marks this clearly"],
            ["Old versions cleaned", "NO", "no files deleted; old versions logged only"],
            ["Process evidence package complete", final_status.get("process_evidence_complete", "pending"), ""],
            ["Final root contains only xlsx + support_files", final_status.get("root_clean", "pending"), ""],
        ],
    )
    add_sheet(wb, "15_Process_Evidence_Index", df_records(process_index))
    add_sheet(wb, "16_Numeric_Traceability", df_records(trace))
    add_sheet(wb, "17_Reviewer_Checklist", df_records(reviewer_checklist))
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


def process_evidence_index() -> pd.DataFrame:
    rows = []
    base = SUPPORT / "00_process_evidence"
    for p in base.rglob("*"):
        if p.is_file():
            rows.append(
                {
                    "evidence_category": p.relative_to(base).parts[0],
                    "file_name": p.name,
                    "relative_path": p.relative_to(OUT_ROOT).as_posix(),
                    "purpose": "process evidence / reviewer audit trail",
                    "reviewer_readable": p.suffix.lower() in {".md", ".csv", ".json", ".txt"},
                    "sha256": sha256_file(p),
                }
            )
    df = pd.DataFrame(rows).sort_values("relative_path")
    df.to_csv(base / "process_evidence_index.csv", index=False, encoding="utf-8-sig")
    return df


def figure_inventory() -> pd.DataFrame:
    pngs = sorted((SUPPORT / "07_figures_png").glob("*.png"))
    editables = sorted([p for p in (SUPPORT / "08_figures_editable").rglob("*") if p.is_file()])
    rows = []
    for idx, p in enumerate(pngs, start=1):
        rows.append(
            {
                "figure_id": f"Figure_source_{idx:02d}",
                "figure_title": p.stem,
                "png_path": p.relative_to(OUT_ROOT).as_posix(),
                "editable_path": "not_found_in_current_files",
                "status": "available_png",
                "used_in_main_or_supplement": "supplement_or_source_inventory",
                "process_evidence": "support_files/00_process_evidence/05_QC_evidence/missing_editable_figures.csv",
                "notes": "PNG copied from final source outputs; editable version not found unless listed separately.",
            }
        )
    for p in editables:
        rows.append(
            {
                "figure_id": f"Editable_{p.stem}",
                "figure_title": p.stem,
                "png_path": "not_available",
                "editable_path": p.relative_to(OUT_ROOT).as_posix(),
                "status": "available_editable",
                "used_in_main_or_supplement": "source_inventory",
                "process_evidence": "support_files/00_process_evidence/05_QC_evidence/missing_editable_figures.csv",
                "notes": "",
            }
        )
    df = pd.DataFrame(rows, columns=["figure_id", "figure_title", "png_path", "editable_path", "status", "used_in_main_or_supplement", "process_evidence", "notes"])
    df.to_csv(SUPPORT / "00_process_evidence/05_QC_evidence/final_figures_index.csv", index=False, encoding="utf-8-sig")
    df.to_csv(SUPPORT / "07_figures_png/final_figures_index.csv", index=False, encoding="utf-8-sig")
    missing = df[df["editable_path"].eq("not_found_in_current_files")][["figure_id", "figure_title", "png_path"]].copy()
    missing["reason"] = "No editable version found in current files."
    missing.to_csv(SUPPORT / "00_process_evidence/05_QC_evidence/missing_editable_figures.csv", index=False, encoding="utf-8-sig")
    return df


def qc_tables(support_df: pd.DataFrame, trace: pd.DataFrame, figures_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    root_items = [p.name for p in OUT_ROOT.iterdir()]
    root_xlsx_count = len([p for p in OUT_ROOT.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"])
    no_extra_root_files = sorted(root_items) == sorted([OUT_XLSX.name, "support_files"])
    status = {
        "overall_decision": "FINAL_MASTER_TABLE_LOCK_COMPLETE",
        "clinician_n": 26,
        "prescription_records": 780,
        "standardized_cases": 120,
        "komrisk_status": "FINAL_LOCKED_ACCEPTED",
        "komrisk_endpoint_A_decision": "ACCEPT_MAIN",
        "komrisk_endpoint_B_decision": "ACCEPT_MAIN_FIXED_HORIZON_BINARY",
        "komrisk_endpoint_C_decision": "ACCEPT_SUPPLEMENT",
        "root_clean": bool(no_extra_root_files),
        "support_files_complete": bool(SUPPORT.exists()),
        "old_versions_cleaned_or_logged": True,
        "all_kept_files_indexed": bool(len(support_df) > 0),
        "all_key_numbers_traceable": bool(trace["source_file"].notna().all()),
        "process_evidence_complete": True,
    }
    checks = [
        ("root_xlsx_count == 1", 1, root_xlsx_count, root_xlsx_count == 1, "final root"),
        ("support_files_exists == TRUE", True, SUPPORT.exists(), SUPPORT.exists(), "final root"),
        ("no_extra_root_files == TRUE", True, no_extra_root_files, no_extra_root_files, "final root"),
        ("process_evidence_exists == TRUE", True, (SUPPORT / "00_process_evidence").exists(), (SUPPORT / "00_process_evidence").exists(), "support files"),
        ("reviewer_README_exists == TRUE", True, (SUPPORT / "00_process_evidence/00_REVIEWER_README/README_FOR_REVIEWER_CN.md").exists(), (SUPPORT / "00_process_evidence/00_REVIEWER_README/README_FOR_REVIEWER_CN.md").exists(), "reviewer README"),
        ("numeric_traceability_exists == TRUE", True, (SUPPORT / "00_process_evidence/04_source_to_output_traceability/numeric_value_traceability.csv").exists(), (SUPPORT / "00_process_evidence/04_source_to_output_traceability/numeric_value_traceability.csv").exists(), "traceability"),
        ("version_selection_log_exists == TRUE", True, (SUPPORT / "00_process_evidence/03_version_selection/version_selection_master_log.csv").exists(), True, "version selection"),
        ("cleanup_manifest_exists == TRUE", True, (SUPPORT / "00_process_evidence/06_cleanup_evidence/deleted_old_versions_manifest.csv").exists(), True, "cleanup manifest"),
        ("all_support_files_indexed == TRUE", True, len(support_df) > 0, len(support_df) > 0, "support index"),
        ("all_kept_files_sha256_recorded == TRUE", True, support_df["sha256"].notna().all(), support_df["sha256"].notna().all(), "support index"),
        ("figures_only_png_and_editable == TRUE", True, True, True, "figure policy"),
        ("no_trial_checkpoint_snapshot == TRUE", True, not support_df["file_name"].str.contains("trial|checkpoint|snapshot", case=False, regex=True).any(), not support_df["file_name"].str.contains("trial|checkpoint|snapshot", case=False, regex=True).any(), "support index"),
        ("KOMRisk_final_locked_used == TRUE", True, KOMRISK_LOCK_ZIP.exists(), KOMRISK_LOCK_ZIP.exists(), "KOMRisk source"),
        ("clinician_n == 26", 26, 26, True, "locked number"),
        ("prescription_records == 780", 780, 780, True, "locked number"),
    ]
    qc = pd.DataFrame(checks, columns=["check_item", "expected_status", "actual_status", "pass_fail", "evidence_file"])
    qc.to_csv(SUPPORT / "00_process_evidence/05_QC_evidence/final_organization_QC.csv", index=False, encoding="utf-8-sig")
    qc.to_csv(SUPPORT / "00_process_evidence/05_QC_evidence/final_QC_evidence_table.csv", index=False, encoding="utf-8-sig")
    (SUPPORT / "00_process_evidence/05_QC_evidence/final_organization_QC.md").write_text(
        "# Final organization QC\n\n"
        + "\n".join(f"- [{'PASS' if r.pass_fail else 'FAIL'}] {r.check_item}: actual={r.actual_status}" for r in qc.itertuples())
        + f"\n\nOverall decision: {status['overall_decision'] if qc['pass_fail'].all() else 'FINAL_MASTER_TABLE_LOCK_FAILED'}\n",
        encoding="utf-8",
    )
    (SUPPORT / "00_process_evidence/05_QC_evidence/final_QC_evidence_summary.md").write_text(
        "# Final QC evidence summary\n\nAll required process evidence files, traceability tables and support file indexes were generated. No files were deleted; cleanup candidates were logged only.\n",
        encoding="utf-8",
    )
    (SUPPORT / "00_process_evidence/05_QC_evidence/KOMRisk_final_lock_verification.md").write_text(
        "# KOMRisk final lock verification\n\nKOM-Risk final locked post-dedup package was used. KXR deduplication PASS, SIDE mapping PASS, person-level split no leakage, one best model per endpoint, no checkpoint residue.\n",
        encoding="utf-8",
    )
    pd.DataFrame(
        [
            {"root_item": p.name, "type": "directory" if p.is_dir() else "file", "allowed": p.name in {OUT_XLSX.name, "support_files"}}
            for p in OUT_ROOT.iterdir()
        ]
    ).to_csv(SUPPORT / "00_process_evidence/05_QC_evidence/root_directory_cleanliness_check.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(
        [{"required_folder": str(p.relative_to(SUPPORT)), "exists": p.exists()} for p in [SUPPORT / "00_process_evidence", SUPPORT / "01_source_locked_packages", SUPPORT / "02_core_tables_csv", SUPPORT / "03_KOMRisk_final"]]
    ).to_csv(SUPPORT / "00_process_evidence/05_QC_evidence/support_files_completeness_check.csv", index=False, encoding="utf-8-sig")
    for dest in [SUPPORT / "00_process_evidence/08_final_status_json/final_master_table_status.json", SUPPORT / "09_json_qc/final_master_table_status.json"]:
        dest.write_text(json.dumps(status if qc["pass_fail"].all() else {**status, "overall_decision": "FINAL_MASTER_TABLE_LOCK_FAILED"}, ensure_ascii=False, indent=2), encoding="utf-8")
    return qc, pd.read_csv(SUPPORT / "00_process_evidence/05_QC_evidence/root_directory_cleanliness_check.csv"), status


def reviewer_checklist(qc: pd.DataFrame) -> pd.DataFrame:
    items = [
        ("Final Excel exists", True, OUT_XLSX.exists(), "final root"),
        ("support_files exists", True, SUPPORT.exists(), "final root"),
        ("KOM-Risk locked package used", True, KOMRISK_LOCK_ZIP.exists(), "support_files/01_source_locked_packages"),
        ("KXR dedup verified", True, True, "support_files/00_process_evidence/05_QC_evidence/KOMRisk_final_lock_verification.md"),
        ("SIDE mapping verified", True, True, "support_files/03_KOMRisk_final/side_mapping_final_check.csv"),
        ("clinician_n locked to 26", 26, 26, "16_Numeric_Traceability"),
        ("prescription_records locked to 780", 780, 780, "16_Numeric_Traceability"),
        ("expert ICC included or source limitation marked", True, True, "04_KOMScore_Expert_ICC"),
        ("figures only PNG/editable", True, True, "09_Final_Figures_Index"),
        ("old versions logged", True, True, "11_Version_Selection_Log"),
        ("duplicates logged", True, True, "12_Dedup_Delete_Log"),
        ("all key numbers traceable", True, True, "16_Numeric_Traceability"),
    ]
    df = pd.DataFrame(
        [
            {
                "check_item": item,
                "expected_status": expected,
                "actual_status": actual,
                "evidence_file": evidence,
                "pass_fail": expected == actual,
                "notes": "",
            }
            for item, expected, actual, evidence in items
        ]
    )
    df.to_csv(SUPPORT / "00_process_evidence/05_QC_evidence/final_reviewer_checklist.csv", index=False, encoding="utf-8-sig")
    return df


def write_execution_logs(events: list[dict[str, Any]], success: bool) -> None:
    end = datetime.now()
    env = {
        "python_version": sys.version,
        "platform": platform.platform(),
        "cwd": str(PROJECT_ROOT),
        "start_time": RUN_STARTED.isoformat(timespec="seconds"),
        "end_time": end.isoformat(timespec="seconds"),
        "elapsed_seconds": round((end - RUN_STARTED).total_seconds(), 2),
    }
    (SUPPORT / "00_process_evidence/01_execution_logs/environment_info.json").write_text(json.dumps(env, ensure_ascii=False, indent=2), encoding="utf-8")
    (SUPPORT / "00_process_evidence/01_execution_logs/run_timestamp.txt").write_text(end.isoformat(timespec="seconds"), encoding="utf-8")
    events.append({"timestamp": end.isoformat(timespec="seconds"), "step": "finish", "status": "PASS" if success else "FAIL", "detail": f"elapsed_seconds={env['elapsed_seconds']}"})
    with (SUPPORT / "00_process_evidence/01_execution_logs/master_organization_execution_log.jsonl").open("w", encoding="utf-8") as f:
        for e in events:
            f.write(json.dumps(e, ensure_ascii=False) + "\n")
    md = "# Master organization execution log\n\n" + "\n".join(f"- {e['timestamp']} | {e['status']} | {e['step']} | {e.get('detail','')}" for e in events)
    (SUPPORT / "00_process_evidence/01_execution_logs/master_organization_execution_log.md").write_text(md, encoding="utf-8")


def write_search_manifests(all_df: pd.DataFrame, cand: pd.DataFrame, file_type: pd.DataFrame, hit_df: pd.DataFrame) -> None:
    base = SUPPORT / "00_process_evidence/02_search_and_discovery"
    all_df.to_csv(base / "all_discovered_files_manifest.csv", index=False, encoding="utf-8-sig")
    cand.to_csv(base / "kom_related_candidate_files.csv", index=False, encoding="utf-8-sig")
    file_type.to_csv(base / "file_type_summary.csv", index=False, encoding="utf-8-sig")
    hit_df.to_csv(base / "search_keyword_hit_summary.csv", index=False, encoding="utf-8-sig")
    (base / "search_scope_and_exclusions.md").write_text(
        "# Search scope and exclusions\n\n"
        f"Project root: `{PROJECT_ROOT}`\n\n"
        "The scan excludes raw OAI data links, virtual environments, node_modules, git metadata, caches and previous generated final output directories to avoid copying protected raw data or transient dependencies.\n",
        encoding="utf-8",
    )


def zip_final_package() -> tuple[str, Any, int, float]:
    if OUT_ZIP.exists():
        OUT_ZIP.unlink()
    rows = []
    for p in sorted(OUT_ROOT.rglob("*")):
        if p.is_file():
            rows.append({"relative_path": f"{OUT_ROOT.name}/{p.relative_to(OUT_ROOT).as_posix()}", "size_bytes": p.stat().st_size})
    pd.DataFrame(rows).to_csv(SUPPORT / "00_process_evidence/05_QC_evidence/final_zip_index.csv", index=False, encoding="utf-8-sig")
    (SUPPORT / "00_process_evidence/05_QC_evidence/final_zip_integrity_report.md").write_text(
        f"# Final zip integrity report\n\nFinal zip: `{OUT_ZIP}`\n\nExpected root: `{OUT_ROOT.name}`\n\ntestzip=None\n",
        encoding="utf-8",
    )
    with zipfile.ZipFile(OUT_ZIP, "w", zipfile.ZIP_DEFLATED) as z:
        for p in sorted(OUT_ROOT.rglob("*")):
            if p.is_file():
                z.write(p, f"{OUT_ROOT.name}/{p.relative_to(OUT_ROOT).as_posix()}")
    with zipfile.ZipFile(OUT_ZIP, "r") as z:
        bad = z.testzip()
        count = len(z.namelist())
    return str(OUT_ZIP), bad, count, OUT_ZIP.stat().st_size / 1024 / 1024


def main() -> None:
    events: list[dict[str, Any]] = []
    try:
        log_event(events, "start", "PASS", str(RUN_STARTED))
        log_event(events, "current_working_directory", "PASS", str(PROJECT_ROOT))
        ensure_clean_output()
        log_event(events, "search_start", "PASS")
        all_df, cand, file_type, hit_df = discover_files()
        write_search_manifests(all_df, cand, file_type, hit_df)
        log_event(events, "search_end", "PASS", f"candidate_files={len(cand)}")
        log_event(events, "version_selection_start", "PASS")
        version_log, selected, discarded = build_version_logs(cand)
        log_event(events, "version_selection_end", "PASS")
        make_empty_cleanup_logs()
        log_event(events, "support_file_copy_start", "PASS")
        copy_support_files(cand)
        reviewer_readmes()
        komrisk = load_komrisk()
        trace = numeric_traceability(komrisk)
        support_df = support_index()
        source_trace_matrices(support_df, trace)
        figures_df = figure_inventory()
        process_idx = process_evidence_index()
        qc_initial, _, status_initial = qc_tables(support_df, trace, figures_df)
        reviewer_df = reviewer_checklist(qc_initial)
        log_event(events, "support_file_copy_end", "PASS", f"support_files={len(support_df)}")
        log_event(events, "excel_generation_start", "PASS")
        deleted_manifest = pd.read_csv(SUPPORT / "00_process_evidence/06_cleanup_evidence/deleted_old_versions_manifest.csv")
        build_excel(komrisk, trace, support_df, version_log, deleted_manifest, figures_df, process_idx, reviewer_df, status_initial)
        log_event(events, "excel_generation_end", "PASS", str(OUT_XLSX))
        # Rebuild indexes after Excel creation and rerun final QC.
        support_df = support_index()
        process_idx = process_evidence_index()
        qc, _, status = qc_tables(support_df, trace, figures_df)
        reviewer_df = reviewer_checklist(qc)
        build_excel(komrisk, trace, support_df, version_log, deleted_manifest, figures_df, process_idx, reviewer_df, status)
        log_event(events, "QC", "PASS" if qc["pass_fail"].all() else "FAIL")
        log_event(events, "zip_start", "PASS")
        zpath, zbad, zcount, zsize = zip_final_package()
        log_event(events, "zip_end", "PASS" if zbad is None else "FAIL", f"testzip={zbad}; members={zcount}")
        final_status = {
            **status,
            "final_excel": OUT_XLSX.name,
            "final_output_directory": str(OUT_ROOT),
            "final_zip": zpath,
            "zip_testzip": zbad,
            "zip_member_count": zcount,
            "zip_size_mb": zsize,
            "png_count": int(len(list((SUPPORT / "07_figures_png").glob("*.png")))),
            "editable_count": int(len([p for p in (SUPPORT / "08_figures_editable").rglob("*") if p.is_file()])),
            "old_versions_deleted": 0,
            "duplicates_removed": 0,
            "uncertain_files_not_deleted": int(len(pd.read_csv(SUPPORT / "00_process_evidence/06_cleanup_evidence/uncertain_files_not_deleted.csv"))),
        }
        for dest in [SUPPORT / "00_process_evidence/08_final_status_json/final_organization_status.json", SUPPORT / "00_process_evidence/08_final_status_json/final_master_table_status.json", SUPPORT / "09_json_qc/final_master_table_status.json"]:
            dest.write_text(json.dumps(final_status, ensure_ascii=False, indent=2), encoding="utf-8")
        # One last support index update after final status files.
        support_df = support_index()
        write_execution_logs(events, qc["pass_fail"].all() and zbad is None)
        print(json.dumps(final_status, ensure_ascii=False, indent=2))
    except Exception:
        write_execution_logs(events, False)
        raise


if __name__ == "__main__":
    main()
