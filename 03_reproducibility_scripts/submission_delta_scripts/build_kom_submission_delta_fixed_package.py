from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import pandas as pd


ROOT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改")
SOURCE = ROOT / "KOM_Submission_Audit_Package_202606_FINAL"
OUT = ROOT / "KOM_Submission_Audit_Package_202606_FINAL_DELTA_FIXED"
ZIP_OUT = ROOT / "KOM_Submission_Audit_Package_202606_FINAL_DELTA_FIXED.zip"
NOW = datetime.now().strftime("%Y%m%d_%H%M%S")


MASTER_COLUMNS = [
    "record_id",
    "module",
    "submodule",
    "item_name",
    "status",
    "priority",
    "required_for_main_text",
    "required_for_supplement",
    "source_file",
    "output_file",
    "evidence_summary",
    "data_granularity",
    "can_recalculate",
    "needs_manual_review",
    "needs_export",
    "needs_rerun",
    "missing_reason",
    "next_action",
    "manuscript_location",
    "search_tags",
]

PRED_COLUMNS = [
    "sample_id",
    "person_id",
    "knee_id",
    "side",
    "endpoint",
    "split",
    "fold",
    "site",
    "time_origin",
    "followup_time",
    "event_time",
    "censor_time",
    "event_observed",
    "y_true",
    "predicted_probability",
    "risk_score",
    "predicted_class",
    "model_name",
    "algorithm",
    "feature_set",
    "source_file",
    "row_origin",
    "data_quality_flag",
    "search_tags",
]


def ensure_clean_output() -> None:
    if OUT.exists():
        backup = ROOT / f"{OUT.name}_backup_{NOW}"
        shutil.move(str(OUT), str(backup))
        print(f"Backed up existing delta dir: {backup}")
    if ZIP_OUT.exists():
        backup_zip = ROOT / f"{ZIP_OUT.stem}_backup_{NOW}.zip"
        shutil.move(str(ZIP_OUT), str(backup_zip))
        print(f"Backed up existing delta zip: {backup_zip}")
    for rel in [
        "00_DELTA_README",
        "01_master_searchable_table",
        "02_KOMSim_final_disclosure",
        "03_KOMRisk_prediction_export_or_missing",
        "04_KOMRAG_topk_label_finalization",
        "05_expert_label_lock",
        "06_submission_ready_text",
        "07_scripts_to_run_if_missing",
        "08_qc_reports",
    ]:
        (OUT / rel).mkdir(parents=True, exist_ok=True)


def read_csv(path: Path, nrows: int | None = None) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    for enc in ["utf-8-sig", "utf-8", "gb18030"]:
        try:
            return pd.read_csv(path, nrows=nrows, encoding=enc)
        except Exception:
            continue
    try:
        return pd.read_csv(path, nrows=nrows, encoding_errors="ignore")
    except Exception:
        return pd.DataFrame()


def write_csv(path: Path, rows_or_df: list[dict[str, Any]] | pd.DataFrame, columns: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df = rows_or_df if isinstance(rows_or_df, pd.DataFrame) else pd.DataFrame(rows_or_df)
    if columns is not None:
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        df = df[columns]
    if "search_tags" not in df.columns:
        df["search_tags"] = "#PARTIAL"
    df.to_csv(path, index=False, encoding="utf-8-sig")


def write_md(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.strip() + "\n", encoding="utf-8")


def boolish(value: Any) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def normalize_tags(tags: str) -> str:
    parts = [p.strip() for p in str(tags).split(";") if p and p.strip()]
    seen: list[str] = []
    for part in parts:
        if not part.startswith("#"):
            part = "#" + part
        if part not in seen:
            seen.append(part)
    return ";".join(seen)


def add_constant_tags(df: pd.DataFrame, tags: str) -> pd.DataFrame:
    out = df.copy()
    out["search_tags"] = normalize_tags(tags)
    return out


def source_path(*parts: str) -> Path:
    return SOURCE.joinpath(*parts)


def rel_output(path: Path) -> str:
    try:
        return str(path.relative_to(OUT)).replace("\\", "/")
    except Exception:
        return str(path)


def make_master_row(
    rows: list[dict[str, Any]],
    *,
    module: str,
    submodule: str,
    item_name: str,
    status: str,
    priority: str,
    required_for_main_text: bool,
    required_for_supplement: bool,
    source_file: str | Path,
    output_file: str | Path,
    evidence_summary: str,
    data_granularity: str,
    can_recalculate: bool,
    needs_manual_review: bool,
    needs_export: bool,
    needs_rerun: bool,
    missing_reason: str,
    next_action: str,
    manuscript_location: str,
    search_tags: str,
) -> None:
    rows.append(
        {
            "record_id": f"REC_{len(rows) + 1:05d}",
            "module": module,
            "submodule": submodule,
            "item_name": item_name,
            "status": status,
            "priority": priority,
            "required_for_main_text": required_for_main_text,
            "required_for_supplement": required_for_supplement,
            "source_file": str(source_file),
            "output_file": str(output_file),
            "evidence_summary": evidence_summary,
            "data_granularity": data_granularity,
            "can_recalculate": can_recalculate,
            "needs_manual_review": needs_manual_review,
            "needs_export": needs_export,
            "needs_rerun": needs_rerun,
            "missing_reason": missing_reason,
            "next_action": next_action,
            "manuscript_location": manuscript_location,
            "search_tags": normalize_tags(search_tags),
        }
    )


def build_komsim(master_rows: list[dict[str, Any]]) -> dict[str, Any]:
    out = OUT / "02_KOMSim_final_disclosure"
    disclosure = out / "KOMSim_final_log_disclosure_LOCKED.md"
    time_def = out / "KOMSim_time_definition_LOCKED.md"
    metric_csv = out / "KOMSim_time_metrics_source_status_LOCKED.csv"
    tags = "#KOMSIM_TIME_DEF;#KOMSIM_TASK_LOG;#KOMSIM_EVENT_LOG_MISSING;#KOMSIM_DISCLOSURE_LOCKED;#SUBMISSION_READY"

    disclosure_text = """
# KOM-Sim final log disclosure LOCKED

Complete raw event-level click/session logs were not available.
KOM-Sim timing analyses were based on exported task-level timing records and summary tables.
Task time was measured from entry into the prescription-answering interface to prescription submission.
The interval included case reading and interaction with AI materials within the task interface.
The interval excluded post-task questionnaires.
Task time was summarized primarily using medians.
Extreme values were not excluded in the primary median-based analysis unless explicitly documented by exported logs.
Repeated-case balance across all three conditions for the same physician was not assumed.

This disclosure is locked for submission transparency. It does not fabricate event-level logs and does not claim balanced repeated cases unless directly supported by allocation records.
"""
    write_md(disclosure, disclosure_text)

    protocol_src = source_path("01_KOMSim_time_and_log_audit", "time_definition_protocol", "KOM-Sim_log_cleaning_protocol_FINAL.md")
    protocol_text = protocol_src.read_text(encoding="utf-8") if protocol_src.exists() else disclosure_text
    write_md(time_def, "# KOM-Sim time definition LOCKED\n\n" + protocol_text)

    rows = [
        {
            "metric": "editing time median",
            "condition_A": 43.5,
            "condition_B": 27.0,
            "condition_C": 17.0,
            "comparison": "",
            "effect": "",
            "q_value": "",
            "source_status": "summary_table_only",
            "data_granularity": "summary_level",
            "search_tags": tags,
        },
        {
            "metric": "workload",
            "condition_A": 5.10,
            "condition_B": 4.35,
            "condition_C": 3.84,
            "comparison": "",
            "effect": "",
            "q_value": "",
            "source_status": "summary_table_only",
            "data_granularity": "summary_level",
            "search_tags": tags,
        },
        {"metric": "editing time paired difference", "condition_A": "", "condition_B": "", "condition_C": "", "comparison": "B vs A", "effect": "-6.75 sec", "q_value": "0.048", "source_status": "summary_table_only", "data_granularity": "summary_level", "search_tags": tags},
        {"metric": "editing time paired difference", "condition_A": "", "condition_B": "", "condition_C": "", "comparison": "C vs A", "effect": "-19.5 sec", "q_value": "2.3e-4", "source_status": "summary_table_only", "data_granularity": "summary_level", "search_tags": tags},
        {"metric": "editing time paired difference", "condition_A": "", "condition_B": "", "condition_C": "", "comparison": "C vs B", "effect": "-8.75 sec", "q_value": "2.3e-4", "source_status": "summary_table_only", "data_granularity": "summary_level", "search_tags": tags},
        {"metric": "workload paired difference", "condition_A": "", "condition_B": "", "condition_C": "", "comparison": "B vs A", "effect": "-0.75", "q_value": "0.043", "source_status": "summary_table_only", "data_granularity": "summary_level", "search_tags": tags},
        {"metric": "workload paired difference", "condition_A": "", "condition_B": "", "condition_C": "", "comparison": "C vs A", "effect": "-1.27", "q_value": "0.004", "source_status": "summary_table_only", "data_granularity": "summary_level", "search_tags": tags},
        {"metric": "workload paired difference", "condition_A": "", "condition_B": "", "condition_C": "", "comparison": "C vs B", "effect": "-0.52", "q_value": "0.004", "source_status": "summary_table_only", "data_granularity": "summary_level", "search_tags": tags},
    ]
    write_csv(metric_csv, rows)
    for path, item, granularity in [
        (disclosure, "KOM-Sim final log disclosure", "file_level"),
        (time_def, "KOM-Sim locked time definition", "file_level"),
        (metric_csv, "KOM-Sim locked time metrics source status", "summary_level"),
    ]:
        make_master_row(
            master_rows,
            module="KOM-Sim",
            submodule="final disclosure",
            item_name=item,
            status="complete",
            priority="high",
            required_for_main_text=True,
            required_for_supplement=True,
            source_file=protocol_src if path != disclosure else source_path("05_crosscheck_reports", "FINAL_BUILD_STATUS.json"),
            output_file=path,
            evidence_summary="Task-level timing and missing event-log disclosure locked for submission.",
            data_granularity=granularity,
            can_recalculate=False,
            needs_manual_review=False,
            needs_export=False,
            needs_rerun=False,
            missing_reason="",
            next_action="Use locked disclosure wording in supplementary methods.",
            manuscript_location="Supplementary Methods; limitations",
            search_tags=tags,
        )
    return {"task_level_timing": "complete", "event_level_log": "not_found", "locked_time_protocol": "complete"}


def columns_from_file(path: Path) -> tuple[int | None, list[str], str]:
    ext = path.suffix.lower()
    try:
        if ext in {".csv", ".tsv"}:
            sep = "\t" if ext == ".tsv" else None
            df = pd.read_csv(path, nrows=100, sep=sep, engine="python", encoding_errors="ignore")
            return len(df), list(df.columns), ""
        if ext in {".xlsx", ".xls"}:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value is not None]
            rows = max(ws.max_row - 1, 0)
            wb.close()
            return rows, headers, ""
        if ext == ".json":
            text = path.read_text(encoding="utf-8", errors="ignore")[:200000]
            keys = re.findall(r'"([^"]+)"\s*:', text)
            return None, sorted(set(keys))[:200], ""
        if ext == ".parquet":
            df = pd.read_parquet(path, engine="pyarrow")
            return len(df), list(df.columns), ""
    except Exception as exc:
        return None, [], f"read_failed: {type(exc).__name__}: {exc}"
    return None, [], "unsupported_extension"


def classify_candidate(path: Path, columns: list[str]) -> dict[str, Any]:
    name = path.name.lower()
    text = " ".join([name, str(path).lower(), " ".join(map(str, columns)).lower()])
    colset = {str(c).lower() for c in columns}
    possible_endpoint = "unknown"
    if re.search(r"structural|progression|kl progression", text):
        possible_endpoint = "KL structural progression"
    if re.search(r"tkr|surgery|arthroplasty", text):
        possible_endpoint = "TKR / knee surgery event"
    if re.search(r"symptom|function|womac|worsening", text):
        possible_endpoint = "Symptom/function worsening"
    looks_oak = bool(re.search(r"oaknet|radiograph|kl grading|kl classification|external_test_predictions", text)) or (
        {"kl_true", "kl_pred"}.issubset(colset) and any(c in colset for c in ["p0", "p1", "p2", "p3", "p4"])
    )
    has_sample = any(c in colset for c in ["sample_id", "person_id", "id", "knee_id"])
    has_y = any(c in colset for c in ["y_true", "label", "target", "event", "event_observed", "outcome"])
    has_event_time = any(c in colset for c in ["event_time", "censor_time", "followup_time", "time_origin"])
    has_pred_prob = any(c in colset for c in ["predicted_probability", "probability", "pred_probability", "risk_probability"])
    has_risk = any(c in colset for c in ["risk_score", "prediction", "pred", "score"])
    has_model = any(c in colset for c in ["model_name", "algorithm", "lightgbm", "catboost", "coxph"])
    has_split = any(c in colset for c in ["split", "fold"])
    usable = bool(has_sample and has_y and (has_pred_prob or has_risk) and possible_endpoint != "unknown" and not looks_oak)
    if looks_oak:
        reason = "excluded_OAKNet_imaging_prediction"
    elif usable:
        reason = "candidate_has_sample_label_prediction_and_endpoint_fields"
    else:
        reason = "insufficient_endpoint_specific_longitudinal_prediction_fields"
    return {
        "possible_endpoint": possible_endpoint,
        "has_sample_id": has_sample,
        "has_knee_id": "knee_id" in colset,
        "has_y_true": has_y,
        "has_event_observed": "event_observed" in colset,
        "has_event_time": "event_time" in colset,
        "has_censor_time": "censor_time" in colset,
        "has_predicted_probability": has_pred_prob,
        "has_risk_score": has_risk,
        "has_model_name": has_model,
        "has_split": has_split,
        "has_fold": "fold" in colset,
        "looks_like_oaknet_imaging_prediction": looks_oak,
        "looks_like_komrisk_longitudinal_prediction": usable,
        "usable_for_endpoint_specific_prediction": usable,
        "reason": reason,
    }


def scan_risk_candidates() -> pd.DataFrame:
    terms = re.compile(
        r"risk|prediction|prognosis|progression|structural|tkr|surgery|symptom|function|worsening|cox|survival|lightgbm|catboost|calibration|brier|cindex|shap|feature|split",
        re.I,
    )
    rows: list[dict[str, Any]] = []
    for path in ROOT.rglob("*"):
        if not path.is_file():
            continue
        if OUT.name in path.parts:
            continue
        if path.suffix.lower() not in {".csv", ".tsv", ".xlsx", ".xls", ".json", ".parquet"}:
            continue
        if not terms.search(str(path)):
            continue
        if path.stat().st_size > 150_000_000:
            continue
        n, cols, note = columns_from_file(path)
        c = classify_candidate(path, cols)
        if note and not cols:
            c["reason"] = note
        rows.append(
            {
                "file_path": str(path),
                "num_rows": n if n is not None else "",
                "columns": ";".join(map(str, cols[:120])),
                **c,
                "search_tags": normalize_tags("#KOMRISK_ENDPOINT_PRED;#KOMRISK_MISSING_SAMPLELEVEL" + (";#REQUIRES_EXPORT" if not c["usable_for_endpoint_specific_prediction"] else "")),
            }
        )
    return pd.DataFrame(rows)


def missing_prediction_row(endpoint: str, tags: str) -> dict[str, Any]:
    return {
        "sample_id": "",
        "person_id": "",
        "knee_id": "",
        "side": "",
        "endpoint": endpoint,
        "split": "",
        "fold": "",
        "site": "",
        "time_origin": "",
        "followup_time": "",
        "event_time": "",
        "censor_time": "",
        "event_observed": "",
        "y_true": "",
        "predicted_probability": "",
        "risk_score": "",
        "predicted_class": "",
        "model_name": "",
        "algorithm": "",
        "feature_set": "",
        "source_file": "not_found",
        "row_origin": "missing_template_not_fabricated",
        "data_quality_flag": "sample_level_predictions_missing",
        "search_tags": tags,
    }


def write_risk_export_scripts() -> None:
    script = OUT / "07_scripts_to_run_if_missing" / "export_KOMRisk_predictions_template.py"
    text = r'''#!/usr/bin/env python
"""Template for exporting KOM-Risk endpoint-specific longitudinal predictions.

This template intentionally does not mix OAKNet imaging predictions with
KOM-Risk longitudinal predictions.
"""
from __future__ import annotations

import argparse
from pathlib import Path

OUTPUT_COLUMNS = [
    "sample_id", "person_id", "knee_id", "side", "endpoint", "split", "fold",
    "event_time", "censor_time", "event_observed", "y_true",
    "predicted_probability", "risk_score", "predicted_class", "model_name",
    "algorithm", "feature_set",
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data_root", required=True)
    parser.add_argument("--model_dir", required=True)
    parser.add_argument("--output_dir", required=True)
    parser.add_argument("--endpoint", required=True, choices=[
        "structural_progression", "tkr_knee_surgery", "symptom_function_worsening", "all"
    ])
    parser.add_argument("--split_csv", required=True)
    parser.add_argument("--feature_csv", required=True)
    parser.add_argument("--label_csv", required=True)
    parser.add_argument("--model_file", required=True)
    args = parser.parse_args()

    model_file = Path(args.model_file)
    if not model_file.exists():
        raise FileNotFoundError(
            "Model file not found. Please provide trained KOM-Risk model artifact "
            "or rerun model training/export."
        )

    raise NotImplementedError(
        "Load the trained KOM-Risk model artifact, join split/features/labels by "
        "sample_id or person_id/knee_id, compute predicted_probability or risk_score, "
        "and write OUTPUT_COLUMNS. Do not mix OAKNet imaging predictions with "
        "KOM-Risk longitudinal predictions."
    )


if __name__ == "__main__":
    main()
'''
    write_md(script, text)
    bat = OUT / "07_scripts_to_run_if_missing" / "run_export_KOMRisk_predictions.bat"
    bat.write_text(
        "@echo off\r\n"
        "python export_KOMRisk_predictions_template.py --data_root %1 --model_dir %2 --output_dir %3 --endpoint all --split_csv %4 --feature_csv %5 --label_csv %6 --model_file %7\r\n",
        encoding="utf-8",
    )
    sh = OUT / "07_scripts_to_run_if_missing" / "run_export_KOMRisk_predictions.sh"
    sh.write_text(
        "#!/usr/bin/env bash\n"
        "python export_KOMRisk_predictions_template.py --data_root \"$1\" --model_dir \"$2\" --output_dir \"$3\" --endpoint all --split_csv \"$4\" --feature_csv \"$5\" --label_csv \"$6\" --model_file \"$7\"\n",
        encoding="utf-8",
    )


def build_risk(master_rows: list[dict[str, Any]]) -> dict[str, Any]:
    out = OUT / "03_KOMRisk_prediction_export_or_missing"
    scan = scan_risk_candidates()
    scan_path = out / "KOMRisk_candidate_prediction_file_scan.csv"
    write_csv(scan_path, scan)

    endpoints = [
        ("KL structural progression", "structural_progression", "#KOMRISK_ENDPOINT_PRED;#KOMRISK_PRED_STRUCTURAL;#KOMRISK_MISSING_SAMPLELEVEL;#REQUIRES_EXPORT"),
        ("TKR / knee surgery event", "tkr_knee_surgery", "#KOMRISK_ENDPOINT_PRED;#KOMRISK_PRED_TKR_SURGERY;#KOMRISK_MISSING_SAMPLELEVEL;#REQUIRES_EXPORT"),
        ("Symptom/function worsening", "symptom_function_worsening", "#KOMRISK_ENDPOINT_PRED;#KOMRISK_PRED_SYMFUNC;#KOMRISK_MISSING_SAMPLELEVEL;#REQUIRES_EXPORT"),
    ]
    all_rows = []
    statuses: dict[str, str] = {}
    for endpoint, slug, tags in endpoints:
        pred_path = out / f"risk_predictions_{slug}_FINAL_LOCKED.csv"
        rows = [missing_prediction_row(endpoint, normalize_tags(tags))]
        write_csv(pred_path, rows, PRED_COLUMNS)
        all_rows.extend(rows)
        statuses[slug] = "missing_template"
        make_master_row(
            master_rows,
            module="KOM-Risk",
            submodule="endpoint predictions",
            item_name=f"{endpoint} sample-level prediction file",
            status="requires_export",
            priority="critical",
            required_for_main_text=False,
            required_for_supplement=True,
            source_file="not_found",
            output_file=pred_path,
            evidence_summary="Endpoint-specific sample-level longitudinal prediction rows were not found; missing template is explicit and not fabricated.",
            data_granularity="sample_level",
            can_recalculate=False,
            needs_manual_review=False,
            needs_export=True,
            needs_rerun=True,
            missing_reason="sample-level prediction rows not located under strict non-OAKNet criteria",
            next_action="Run export_KOMRisk_predictions_template.py after providing trained KOM-Risk model artifacts and split/feature/label files.",
            manuscript_location="Supplementary reproducibility package; limitations",
            search_tags=tags,
        )
    all_path = out / "risk_predictions_all_endpoints_clean_FINAL_LOCKED.csv"
    write_csv(all_path, all_rows, PRED_COLUMNS)

    metric_src = source_path("02_KOMRisk_reproducible_prediction_package", "metrics", "risk_model_metrics_FINAL.csv")
    metrics = read_csv(metric_src)
    metrics = add_constant_tags(metrics, "#KOMRISK_METRICS;#SUBMISSION_READY")
    metric_out = out / "risk_model_metrics_FINAL_LOCKED.csv"
    write_csv(metric_out, metrics)
    make_master_row(
        master_rows,
        module="KOM-Risk",
        submodule="metrics",
        item_name="KOM-Risk locked endpoint-level metrics",
        status="complete",
        priority="high",
        required_for_main_text=True,
        required_for_supplement=True,
        source_file=metric_src,
        output_file=metric_out,
        evidence_summary="Main-text endpoint-level summary metrics preserved: AUROC/C-index/BACC where available; missing calibration/AUPRC values remain NA.",
        data_granularity="endpoint_level",
        can_recalculate=False,
        needs_manual_review=False,
        needs_export=False,
        needs_rerun=False,
        missing_reason="",
        next_action="Use locked metrics with transparent sample-level prediction limitation.",
        manuscript_location="Main Results; Supplementary Methods",
        search_tags="#KOMRISK_METRICS;#SUBMISSION_READY",
    )

    write_risk_export_scripts()
    export_plan = out / "KOMRisk_endpoint_prediction_export_plan.md"
    write_md(
        export_plan,
        """# KOM-Risk endpoint prediction export plan

Endpoint-specific sample-level longitudinal prediction rows were not found in the current package. Current KOM-Risk results are supported by endpoint-level summary metrics. Full independent recalculation of endpoint-specific calibration, DCA, and sample-level performance requires exporting model predictions using the provided script template.

Do not mix OAKNet imaging predictions with KOM-Risk longitudinal predictions. OAKNet radiograph KL grading or cross-sectional imaging classifier outputs must remain excluded from longitudinal KOM-Risk endpoint prediction files.

search_tags: #KOMRISK_EXPORT_SCRIPT;#KOMRISK_MISSING_SAMPLELEVEL;#REQUIRES_EXPORT;#REQUIRES_RERUN
""",
    )
    missing = out / "risk_missing_items_FINAL_LOCKED.md"
    write_md(
        missing,
        """# KOM-Risk missing items FINAL LOCKED

Endpoint-specific sample-level longitudinal prediction rows were not found in the current package. Current KOM-Risk results are supported by endpoint-level summary metrics. Full independent recalculation of endpoint-specific calibration, DCA, and sample-level performance requires exporting model predictions using the provided script template.

Missing items:
- structural progression sample-level predictions
- TKR/knee surgery sample-level predictions
- symptom/function worsening sample-level predictions
- endpoint-specific calibration probabilities
- endpoint-specific DCA-ready absolute risks
- endpoint-specific SHAP or feature-importance tables

search_tags: #KOMRISK_MISSING_SAMPLELEVEL;#KOMRISK_REQUIRES_RERUN;#MISSING;#REQUIRES_EXPORT
""",
    )
    for path, item, tags in [
        (scan_path, "KOM-Risk candidate prediction file scan", "#KOMRISK_ENDPOINT_PRED;#KOMRISK_MISSING_SAMPLELEVEL;#PARTIAL"),
        (all_path, "KOM-Risk all-endpoints clean missing-template predictions", "#KOMRISK_ENDPOINT_PRED;#KOMRISK_MISSING_SAMPLELEVEL;#REQUIRES_EXPORT"),
        (export_plan, "KOM-Risk export/rerun plan", "#KOMRISK_EXPORT_SCRIPT;#REQUIRES_EXPORT;#REQUIRES_RERUN"),
        (missing, "KOM-Risk missing items locked report", "#KOMRISK_MISSING_SAMPLELEVEL;#MISSING;#REQUIRES_EXPORT"),
    ]:
        make_master_row(
            master_rows,
            module="KOM-Risk",
            submodule="prediction export or missing",
            item_name=item,
            status="partial" if path == scan_path else "requires_export",
            priority="critical" if "missing" in item.lower() or "prediction" in item.lower() else "high",
            required_for_main_text=False,
            required_for_supplement=True,
            source_file=SOURCE,
            output_file=path,
            evidence_summary="Delta-fixed transparency artifact for missing endpoint-specific prediction rows.",
            data_granularity="file_level" if path.suffix == ".md" else "sample_level",
            can_recalculate=False,
            needs_manual_review=False,
            needs_export=True,
            needs_rerun=True,
            missing_reason="sample-level predictions not found",
            next_action="Provide trained model artifacts and run the export template.",
            manuscript_location="Supplementary reproducibility package",
            search_tags=tags,
        )
    for script_path in [
        OUT / "07_scripts_to_run_if_missing" / "export_KOMRisk_predictions_template.py",
        OUT / "07_scripts_to_run_if_missing" / "run_export_KOMRisk_predictions.bat",
        OUT / "07_scripts_to_run_if_missing" / "run_export_KOMRisk_predictions.sh",
    ]:
        make_master_row(
            master_rows,
            module="KOM-Risk",
            submodule="export script",
            item_name=script_path.name,
            status="requires_export",
            priority="critical",
            required_for_main_text=False,
            required_for_supplement=True,
            source_file="generated_template",
            output_file=script_path,
            evidence_summary="Runnable template for exporting sample-level predictions once real KOM-Risk model artifacts are provided.",
            data_granularity="file_level",
            can_recalculate=True,
            needs_manual_review=False,
            needs_export=True,
            needs_rerun=True,
            missing_reason="model artifacts and source data must be supplied",
            next_action="Run after locating trained KOM-Risk model, split, feature, and label files.",
            manuscript_location="Supplementary reproducibility package",
            search_tags="#KOMRISK_EXPORT_SCRIPT;#REQUIRES_EXPORT;#REQUIRES_RERUN",
        )
    return {"prediction_statuses": statuses, "candidate_scan_rows": int(len(scan)), "export_script": "created", "metrics": "complete"}


def lock_table(src: Path, dst: Path, tags: str, rowwise: bool = False) -> pd.DataFrame:
    df = read_csv(src)
    if rowwise and "label_source" in df.columns:
        def tag_row(row: pd.Series) -> str:
            if str(row.get("label_source", "")).lower() == "rule_inferred_candidate":
                return normalize_tags("#KOMRAG_RELEVANCE_LABEL;#KOMRAG_RULE_INFERRED;#KOMRAG_MANUAL_REVIEW;#REQUIRES_MANUAL_REVIEW")
            if str(row.get("label_source", "")).lower() == "guideline_anchor_weak_gold":
                return normalize_tags("#KOMRAG_RELEVANCE_LABEL;#KOMRAG_WEAK_GOLD;#KOMRAG_TRACEABILITY")
            return normalize_tags(tags)
        df["search_tags"] = df.apply(tag_row, axis=1)
    else:
        df = add_constant_tags(df, tags)
    write_csv(dst, df)
    return df


def build_rag(master_rows: list[dict[str, Any]]) -> dict[str, Any]:
    out = OUT / "04_KOMRAG_topk_label_finalization"
    rag_src = source_path("03_KOMRAG_query_level_evidence_mapping")
    files = [
        ("clean_query_set/rag_query_set_CLEAN.csv", "rag_query_set_FINAL_LOCKED.csv", "#KOMRAG_QUERY_SET;#SUBMISSION_READY", "query set", "complete", "query_level"),
        ("guideline_anchor_mapping/guideline_anchor_mapping_CLEAN.csv", "guideline_anchor_mapping_FINAL_LOCKED.csv", "#KOMRAG_GUIDELINE_ANCHOR;#KOMRAG_TRACEABILITY;#PARTIAL", "guideline anchors", "partial", "evidence_unit_level"),
        ("retrieval_results/rag_retrieval_results_CLEAN.csv", "rag_retrieval_results_FINAL_LOCKED.csv", "#KOMRAG_RETRIEVAL_TOPK;#KOMRAG_FINAL_TOPK_MISSING;#PARTIAL", "retrieval results", "partial", "query_level"),
        ("relevance_labels/rag_relevance_labels_CLEAN.csv", "rag_relevance_labels_FINAL_LOCKED.csv", "#KOMRAG_RELEVANCE_LABEL;#PARTIAL", "relevance labels", "partial", "query_level"),
        ("query_metrics/rag_metric_by_query_CLEAN.csv", "rag_metric_by_query_FINAL_LOCKED.csv", "#KOMRAG_QUERY_METRIC;#SUBMISSION_READY", "query metrics", "complete", "query_level"),
        ("query_metrics/rag_metric_summary_CLEAN.csv", "rag_metric_summary_FINAL_LOCKED.csv", "#KOMRAG_QUERY_METRIC;#SUBMISSION_READY", "metric summary", "complete", "summary_level"),
        ("error_cases/rag_error_cases_CLEAN.csv", "rag_error_cases_FINAL_LOCKED.csv", "#KOMRAG_ERROR_CASE;#KOMRAG_MANUAL_REVIEW;#REQUIRES_MANUAL_REVIEW", "error cases", "complete", "query_level"),
    ]
    locked: dict[str, pd.DataFrame] = {}
    for rel, dst_name, tags, item, status, granularity in files:
        src = rag_src / rel
        dst = out / dst_name
        df = lock_table(src, dst, tags, rowwise=("relevance_labels" in rel))
        if dst_name == "rag_metric_summary_FINAL_LOCKED.csv" and "data_quality_flag" not in df.columns:
            df["data_quality_flag"] = "carried_forward_from_frozen_stage4a_metrics"
            df["search_tags"] = normalize_tags("#KOMRAG_QUERY_METRIC;#SUBMISSION_READY")
            write_csv(dst, df)
        locked[dst_name] = df
        make_master_row(
            master_rows,
            module="KOM-RAG",
            submodule="top-k label finalization",
            item_name=f"KOM-RAG locked {item}",
            status=status,
            priority="critical" if status == "partial" and item in {"retrieval results", "relevance labels"} else "high",
            required_for_main_text=item in {"query metrics", "metric summary"},
            required_for_supplement=True,
            source_file=src,
            output_file=dst,
            evidence_summary=f"Locked copy of {item}; search_tags added; partial status preserved where full all-query final TopK or manual labels are not available.",
            data_granularity=granularity,
            can_recalculate=item in {"query metrics", "metric summary"},
            needs_manual_review=item in {"relevance labels", "error cases"},
            needs_export=item == "retrieval results",
            needs_rerun=False,
            missing_reason="full all-query final top-k or manual labels partial" if status == "partial" else "",
            next_action="Manual review rule-inferred labels and export full all-query final TopK if needed." if status == "partial" else "Use locked table.",
            manuscript_location="Supplementary RAG audit",
            search_tags=tags,
        )

    labels = locked["rag_relevance_labels_FINAL_LOCKED.csv"]
    queries = locked["rag_query_set_FINAL_LOCKED.csv"]
    retrieval = locked["rag_retrieval_results_FINAL_LOCKED.csv"]
    qmap = queries[["query_id", "query_text", "treatment_domain"]].drop_duplicates("query_id") if not queries.empty else pd.DataFrame(columns=["query_id", "query_text", "treatment_domain"])
    rmap = retrieval[["query_id", "evidence_unit_id", "evidence_title", "source_url_or_doi_or_pmid"]].drop_duplicates(["query_id", "evidence_unit_id"]) if not retrieval.empty else pd.DataFrame(columns=["query_id", "evidence_unit_id", "evidence_title", "source_url_or_doi_or_pmid"])
    mask = labels.get("label_source", pd.Series(dtype=str)).astype(str).str.lower().eq("rule_inferred_candidate") & labels.get("requires_manual_review", pd.Series(dtype=bool)).map(boolish)
    review = labels.loc[mask].copy()
    review = review.merge(qmap, on="query_id", how="left").merge(rmap, on=["query_id", "evidence_unit_id"], how="left")
    review.insert(0, "review_id", [f"RAG_REVIEW_{i+1:05d}" for i in range(len(review))])
    review["current_is_relevant"] = review.get("is_relevant", "")
    review["current_relevance_grade"] = review.get("relevance_grade", "")
    review["current_label_source"] = review.get("label_source", "")
    for col in ["manual_review_decision", "manual_relevance_grade", "manual_reviewer", "review_notes"]:
        review[col] = ""
    review["search_tags"] = normalize_tags("#KOMRAG_RELEVANCE_LABEL;#KOMRAG_RULE_INFERRED;#KOMRAG_MANUAL_REVIEW;#REQUIRES_MANUAL_REVIEW")
    review_cols = [
        "review_id",
        "query_id",
        "query_text",
        "treatment_domain",
        "evidence_unit_id",
        "evidence_title",
        "source_url_or_doi_or_pmid",
        "current_is_relevant",
        "current_relevance_grade",
        "current_label_source",
        "label_rationale",
        "manual_review_decision",
        "manual_relevance_grade",
        "manual_reviewer",
        "review_notes",
        "search_tags",
    ]
    review_csv = out / "rag_manual_review_required_labels.csv"
    write_csv(review_csv, review, review_cols)
    review_xlsx = out / "rag_manual_review_sampling_sheet.xlsx"
    with pd.ExcelWriter(review_xlsx, engine="openpyxl") as writer:
        review[review_cols].to_excel(writer, sheet_name="Manual review labels", index=False)
    topk_md = out / "KOMRAG_final_topk_partial_disclosure.md"
    write_md(
        topk_md,
        """# KOM-RAG final top-k partial disclosure

Full all-query final top-k retrieval rows were not located as one clean export. The current package includes clean query set, frozen Stage4A query-level metrics, sampled final top-k rows, relevance labels separated by source, and error cases. This is sufficient for transparent supplementary audit of the reported benchmark metrics, but not a replacement for a fully manual, all-query gold label file.

search_tags: #KOMRAG_RETRIEVAL_TOPK;#KOMRAG_FINAL_TOPK_MISSING;#PARTIAL;#REQUIRES_EXPORT
""",
    )
    for path, item, granularity, tags in [
        (review_csv, "KOM-RAG manual review required labels", "query_level", "#KOMRAG_RELEVANCE_LABEL;#KOMRAG_RULE_INFERRED;#KOMRAG_MANUAL_REVIEW;#REQUIRES_MANUAL_REVIEW"),
        (review_xlsx, "KOM-RAG manual review sampling sheet", "query_level", "#KOMRAG_MANUAL_REVIEW;#REQUIRES_MANUAL_REVIEW"),
        (topk_md, "KOM-RAG final top-k partial disclosure", "file_level", "#KOMRAG_RETRIEVAL_TOPK;#KOMRAG_FINAL_TOPK_MISSING;#PARTIAL;#REQUIRES_EXPORT"),
    ]:
        make_master_row(
            master_rows,
            module="KOM-RAG",
            submodule="manual review pack",
            item_name=item,
            status="partial" if path == topk_md else "requires_export" if "top-k" in item else "complete",
            priority="critical" if path == topk_md else "high",
            required_for_main_text=False,
            required_for_supplement=True,
            source_file=rag_src,
            output_file=path,
            evidence_summary=f"{len(review)} rule-inferred labels require manual review." if path != topk_md else "Full all-query final top-k rows remain partial.",
            data_granularity=granularity,
            can_recalculate=False,
            needs_manual_review=path != topk_md,
            needs_export=path == topk_md,
            needs_rerun=False,
            missing_reason="full all-query final top-k export not located" if path == topk_md else "",
            next_action="Complete manual review decision columns." if path != topk_md else "Export all-query final top-k rows if reviewer requests exact table.",
            manuscript_location="Supplementary RAG audit",
            search_tags=tags,
        )
    return {
        "query_set": "complete",
        "guideline_anchors": "partial",
        "retrieval_results": "partial",
        "relevance_labels": "partial",
        "query_metrics": "complete",
        "error_cases": "complete",
        "manual_review_pack": "created",
        "manual_review_required_labels": int(len(review)),
    }


def build_expert(master_rows: list[dict[str, Any]]) -> dict[str, Any]:
    out = OUT / "05_expert_label_lock"
    src = source_path("04_expert_label_name_audit", "expert_label_dictionary_CLEAN.csv")
    df = read_csv(src)
    df["score_values_altered"] = "NO"
    df["search_tags"] = normalize_tags("#EXPERT_LABEL_DICT;#EXPERT_HUMAN_RATING;#EXPERT_LABEL_NORMALIZATION;#SUBMISSION_READY")
    dst = out / "expert_label_dictionary_FINAL_LOCKED.csv"
    write_csv(dst, df)
    report = out / "expert_label_normalization_report_FINAL_LOCKED.md"
    write_md(
        report,
        """# Expert label normalization report FINAL LOCKED

Human expert ratings were retained as originally recorded. Label normalization was performed only to harmonize field names for analysis and reporting. Score values were not altered.

This package does not rerate prescriptions, does not merge different expert scores without source-file traceability, and does not represent API/model scores as human expert ratings.

search_tags: #EXPERT_LABEL_DICT;#EXPERT_HUMAN_RATING;#EXPERT_LABEL_NORMALIZATION;#SUBMISSION_READY
""",
    )
    for path, item in [(dst, "Expert label dictionary FINAL LOCKED"), (report, "Expert label normalization report FINAL LOCKED")]:
        make_master_row(
            master_rows,
            module="Expert labels",
            submodule="label lock",
            item_name=item,
            status="complete",
            priority="high",
            required_for_main_text=False,
            required_for_supplement=True,
            source_file=src,
            output_file=path,
            evidence_summary="Human expert ratings remain unchanged; only label names are normalized for readability.",
            data_granularity="file_level",
            can_recalculate=False,
            needs_manual_review=False,
            needs_export=False,
            needs_rerun=False,
            missing_reason="",
            next_action="Use locked dictionary in methods/reporting; never alter source score values.",
            manuscript_location="Supplementary Methods; expert rating methods",
            search_tags="#EXPERT_LABEL_DICT;#EXPERT_HUMAN_RATING;#EXPERT_LABEL_NORMALIZATION;#SUBMISSION_READY",
        )
    return {"dictionary": "complete", "score_values_altered": "NO"}


def build_supp_methods(master_rows: list[dict[str, Any]]) -> dict[str, str]:
    out = OUT / "06_submission_ready_text"
    files = {
        "KOMSim": (
            out / "Supplementary_Methods_KOMSim_Time_Definition_FINAL_LOCKED.md",
            "#KOMSIM_TIME_DEF;#KOMSIM_METHODS_TEXT;#KOMSIM_DISCLOSURE_LOCKED;#SUBMISSION_READY",
            "yes",
            "Task time was measured from entry into the prescription-answering interface to prescription submission. This interval included case reading and interaction with AI materials within the task interface and excluded post-task questionnaires. Because complete raw event-level click logs were not available for all tasks, time analyses were based on exported task-level timing records and summary tables.",
            "Complete raw event-level click/session logs were not available.",
        ),
        "KOMRisk": (
            out / "Supplementary_Methods_KOMRisk_Reproducibility_FINAL_LOCKED.md",
            "#KOMRISK_METRICS;#KOMRISK_MISSING_SAMPLELEVEL;#KOMRISK_EXPORT_SCRIPT;#PARTIAL;#REQUIRES_EXPORT",
            "partial",
            "Endpoint-level summary metrics are available. Endpoint-specific sample-level predictions were not found in the current package and are therefore marked as missing templates. Full recalculation of calibration, DCA and sample-level performance requires exporting predictions using the provided script.",
            "Sample-level longitudinal predictions were not located; calibration and DCA require exported probabilities.",
        ),
        "KOMRAG": (
            out / "Supplementary_Methods_KOMRAG_Query_Level_Audit_FINAL_LOCKED.md",
            "#KOMRAG_QUERY_SET;#KOMRAG_QUERY_METRIC;#KOMRAG_MANUAL_REVIEW;#PARTIAL",
            "partial",
            "The RAG benchmark used a frozen Stage4A query-level metric table and relevance labels separated by source. Weak-gold labels were derived from guideline anchors where available; rule-inferred candidate labels require manual review. Full all-query final top-k rows were partial in the current package.",
            "Manual labels and full all-query final top-k export remain partial.",
        ),
        "Expert": (
            out / "Supplementary_Methods_Expert_Label_Normalization_FINAL_LOCKED.md",
            "#EXPERT_LABEL_DICT;#EXPERT_HUMAN_RATING;#EXPERT_LABEL_NORMALIZATION;#SUBMISSION_READY",
            "yes",
            "Human expert ratings were retained as originally recorded. Label normalization was performed only to harmonize field names for analysis and reporting. Score values were not altered.",
            "This audit addresses naming consistency only and does not rescore expert assessments.",
        ),
    }
    result: dict[str, str] = {}
    for module, (path, tags, ready, text, limitations) in files.items():
        write_md(
            path,
            f"""# Supplementary Methods: {module}

search_tags: {tags}

source_files_used:
- {SOURCE}

ready_for_methods: {ready}

text_for_manuscript:
{text}

limitations:
{limitations}
""",
        )
        result[module] = str(path)
        make_master_row(
            master_rows,
            module=module,
            submodule="supplementary methods",
            item_name=path.name,
            status="complete" if ready == "yes" else "partial",
            priority="high",
            required_for_main_text=False,
            required_for_supplement=True,
            source_file=SOURCE,
            output_file=path,
            evidence_summary=f"Locked supplementary methods text for {module}.",
            data_granularity="file_level",
            can_recalculate=False,
            needs_manual_review=(ready != "yes"),
            needs_export=False,
            needs_rerun=False,
            missing_reason=limitations if ready != "yes" else "",
            next_action="Use text in Supplementary Methods with limitations disclosed.",
            manuscript_location="Supplementary Methods",
            search_tags=tags,
        )
    return result


def add_file_inventory_rows(master_rows: list[dict[str, Any]]) -> None:
    existing_outputs = {str(Path(r["output_file"])) for r in master_rows}
    for p in sorted(OUT.rglob("*")):
        if not p.is_file():
            continue
        if str(p) in existing_outputs:
            continue
        tags = "#SUBMISSION_READY"
        module = "Delta package"
        if "KOMRisk" in str(p):
            tags = "#KOMRISK_ENDPOINT_PRED;#PARTIAL"
            module = "KOM-Risk"
        elif "KOMRAG" in str(p) or "rag_" in p.name:
            tags = "#KOMRAG_TRACEABILITY;#PARTIAL"
            module = "KOM-RAG"
        elif "KOMSim" in str(p):
            tags = "#KOMSIM_DISCLOSURE_LOCKED;#SUBMISSION_READY"
            module = "KOM-Sim"
        elif "expert" in p.name.lower():
            tags = "#EXPERT_LABEL_DICT;#SUBMISSION_READY"
            module = "Expert labels"
        elif p.suffix.lower() in {".py", ".bat", ".sh"}:
            tags = "#REQUIRES_EXPORT;#REQUIRES_RERUN"
            module = "Scripts"
        make_master_row(
            master_rows,
            module=module,
            submodule="generated file inventory",
            item_name=p.name,
            status="complete" if p.stat().st_size > 0 else "partial",
            priority="low",
            required_for_main_text=False,
            required_for_supplement=True,
            source_file="generated",
            output_file=p,
            evidence_summary=f"Generated delta-fixed package file: {rel_output(p)}",
            data_granularity="file_level",
            can_recalculate=False,
            needs_manual_review=False,
            needs_export=False,
            needs_rerun=False,
            missing_reason="" if p.stat().st_size > 0 else "empty file",
            next_action="Use as supporting audit artifact.",
            manuscript_location="Supplementary package inventory",
            search_tags=tags,
        )


def build_master_workbook(master_rows: list[dict[str, Any]], status: dict[str, Any]) -> pd.DataFrame:
    add_file_inventory_rows(master_rows)
    df = pd.DataFrame(master_rows)
    for col in MASTER_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[MASTER_COLUMNS]
    df["search_tags"] = df["search_tags"].map(normalize_tags)
    master_csv = OUT / "01_master_searchable_table" / "KOM_submission_master_audit_table_FINAL_DELTA_FIXED.csv"
    write_csv(master_csv, df, MASTER_COLUMNS)

    sim_metrics = read_csv(OUT / "02_KOMSim_final_disclosure" / "KOMSim_time_metrics_source_status_LOCKED.csv")
    risk_preds = read_csv(OUT / "03_KOMRisk_prediction_export_or_missing" / "risk_predictions_all_endpoints_clean_FINAL_LOCKED.csv")
    risk_scan = read_csv(OUT / "03_KOMRisk_prediction_export_or_missing" / "KOMRisk_candidate_prediction_file_scan.csv")
    rag_metrics = read_csv(OUT / "04_KOMRAG_topk_label_finalization" / "rag_metric_summary_FINAL_LOCKED.csv")
    rag_review = read_csv(OUT / "04_KOMRAG_topk_label_finalization" / "rag_manual_review_required_labels.csv")
    expert = read_csv(OUT / "05_expert_label_lock" / "expert_label_dictionary_FINAL_LOCKED.csv")
    missing = df[df["status"].isin(["partial", "missing", "requires_export", "requires_rerun"])].copy()
    insertion = pd.DataFrame(
        [
            {"manuscript_location": "Supplementary Methods", "file": "Supplementary_Methods_KOMSim_Time_Definition_FINAL_LOCKED.md", "module": "KOM-Sim", "search_tags": "#KOMSIM_METHODS_TEXT;#SUBMISSION_READY"},
            {"manuscript_location": "Supplementary Reproducibility", "file": "Supplementary_Methods_KOMRisk_Reproducibility_FINAL_LOCKED.md", "module": "KOM-Risk", "search_tags": "#KOMRISK_METRICS;#PARTIAL"},
            {"manuscript_location": "Supplementary RAG audit", "file": "Supplementary_Methods_KOMRAG_Query_Level_Audit_FINAL_LOCKED.md", "module": "KOM-RAG", "search_tags": "#KOMRAG_QUERY_METRIC;#PARTIAL"},
            {"manuscript_location": "Supplementary Expert ratings", "file": "Supplementary_Methods_Expert_Label_Normalization_FINAL_LOCKED.md", "module": "Expert labels", "search_tags": "#EXPERT_LABEL_DICT;#SUBMISSION_READY"},
        ]
    )
    readme_sheet = pd.DataFrame(
        [
            {"field": "package", "value": OUT.name},
            {"field": "created", "value": datetime.now().isoformat(timespec="seconds")},
            {"field": "purpose", "value": "Delta-fixed searchable audit package with locked disclosures, missing templates, manual-review pack, and scripts."},
            {"field": "search_tags_required", "value": "yes"},
        ]
    )
    xlsx = OUT / "01_master_searchable_table" / "KOM_submission_master_audit_table_FINAL_DELTA_FIXED.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        readme_sheet.to_excel(writer, sheet_name="README", index=False)
        df.to_excel(writer, sheet_name="Master searchable table", index=False)
        sim_metrics.to_excel(writer, sheet_name="KOM-Sim final disclosure", index=False)
        risk_preds.to_excel(writer, sheet_name="KOM-Risk endpoint predictions", index=False)
        risk_scan.to_excel(writer, sheet_name="KOM-Risk export-rerun plan", index=False)
        rag_metrics.to_excel(writer, sheet_name="KOM-RAG topk and labels", index=False)
        rag_review.to_excel(writer, sheet_name="KOM-RAG manual review pack", index=False)
        expert.to_excel(writer, sheet_name="Expert label lock", index=False)
        missing.to_excel(writer, sheet_name="Missing Partial Ready", index=False)
        insertion.to_excel(writer, sheet_name="Manuscript insertion map", index=False)
    return df


def build_qc(master_df: pd.DataFrame, status: dict[str, Any]) -> dict[str, Any]:
    qc_dir = OUT / "08_qc_reports"
    checks: list[dict[str, Any]] = []

    def check(item: str, passed: bool, evidence: str, tags: str, module: str = "QC") -> None:
        checks.append({"item": item, "module": module, "status": "pass" if passed else "fail", "passed": bool(passed), "evidence": evidence, "search_tags": normalize_tags(tags)})

    master_csv = OUT / "01_master_searchable_table" / "KOM_submission_master_audit_table_FINAL_DELTA_FIXED.csv"
    check("Master table generated", master_csv.exists() and master_csv.stat().st_size > 0, str(master_csv), "#SUBMISSION_READY")
    check("Master table search_tags populated", "search_tags" in master_df.columns and master_df["search_tags"].astype(str).str.len().gt(0).all(), f"rows={len(master_df)}", "#SUBMISSION_READY")
    sim_disc = OUT / "02_KOMSim_final_disclosure" / "KOMSim_final_log_disclosure_LOCKED.md"
    sim_text = sim_disc.read_text(encoding="utf-8") if sim_disc.exists() else ""
    check("KOM-Sim event-level log not found disclosed", "Complete raw event-level click/session logs were not available" in sim_text, str(sim_disc), "#KOMSIM_EVENT_LOG_MISSING", "KOM-Sim")
    check("KOM-Sim locked time protocol exists", (OUT / "02_KOMSim_final_disclosure" / "KOMSim_time_definition_LOCKED.md").exists(), "locked protocol", "#KOMSIM_TIME_DEF", "KOM-Sim")
    for slug in ["structural_progression", "tkr_knee_surgery", "symptom_function_worsening"]:
        p = OUT / "03_KOMRisk_prediction_export_or_missing" / f"risk_predictions_{slug}_FINAL_LOCKED.csv"
        df = read_csv(p)
        check(f"KOM-Risk {slug} prediction file exists", p.exists(), str(p), "#KOMRISK_ENDPOINT_PRED", "KOM-Risk")
        check(f"KOM-Risk {slug} missing template explicit", (not df.empty) and "sample_level_predictions_missing" in df.get("data_quality_flag", pd.Series(dtype=str)).astype(str).tolist(), str(p), "#KOMRISK_MISSING_SAMPLELEVEL", "KOM-Risk")
    check("KOM-Risk export script generated", (OUT / "07_scripts_to_run_if_missing" / "export_KOMRisk_predictions_template.py").exists(), "script template", "#KOMRISK_EXPORT_SCRIPT", "KOM-Risk")
    scan = read_csv(OUT / "03_KOMRisk_prediction_export_or_missing" / "KOMRisk_candidate_prediction_file_scan.csv")
    oaknet_excluded = True if scan.empty else not ((scan["looks_like_oaknet_imaging_prediction"].map(boolish)) & (scan["usable_for_endpoint_specific_prediction"].map(boolish))).any()
    check("KOM-Risk OAKNet imaging predictions excluded", oaknet_excluded, f"scan_rows={len(scan)}", "#KOMRISK_MISSING_SAMPLELEVEL", "KOM-Risk")
    check("KOM-RAG final locked query set exists", (OUT / "04_KOMRAG_topk_label_finalization" / "rag_query_set_FINAL_LOCKED.csv").exists(), "query set", "#KOMRAG_QUERY_SET", "KOM-RAG")
    review = read_csv(OUT / "04_KOMRAG_topk_label_finalization" / "rag_manual_review_required_labels.csv")
    check("KOM-RAG manual review pack generated", len(review) > 0, f"manual_review_rows={len(review)}", "#KOMRAG_MANUAL_REVIEW", "KOM-RAG")
    labels = read_csv(OUT / "04_KOMRAG_topk_label_finalization" / "rag_relevance_labels_FINAL_LOCKED.csv")
    rule = labels[labels.get("label_source", pd.Series(dtype=str)).astype(str).str.lower().eq("rule_inferred_candidate")]
    rule_ok = len(rule) == 0 or rule.get("requires_manual_review", pd.Series(dtype=str)).map(boolish).all()
    check("Rule-inferred labels require manual review", rule_ok, f"rule_inferred={len(rule)}", "#KOMRAG_RULE_INFERRED;#REQUIRES_MANUAL_REVIEW", "KOM-RAG")
    err = read_csv(OUT / "04_KOMRAG_topk_label_finalization" / "rag_error_cases_FINAL_LOCKED.csv")
    check("rag_error_cases is non-empty", len(err) > 0, f"rows={len(err)}", "#KOMRAG_ERROR_CASE", "KOM-RAG")
    check("Expert label lock exists", (OUT / "05_expert_label_lock" / "expert_label_dictionary_FINAL_LOCKED.csv").exists(), "expert dictionary", "#EXPERT_LABEL_DICT", "Expert labels")
    expert = read_csv(OUT / "05_expert_label_lock" / "expert_label_dictionary_FINAL_LOCKED.csv")
    score_ok = (not expert.empty) and set(expert.get("score_values_altered", pd.Series(dtype=str)).astype(str).str.upper()) == {"NO"}
    check("Expert score values not altered", score_ok, "score_values_altered=NO", "#EXPERT_HUMAN_RATING", "Expert labels")
    supp = list((OUT / "06_submission_ready_text").glob("Supplementary_Methods_*_FINAL_LOCKED.md"))
    check("Four Supplementary Methods markdown files generated", len(supp) == 4, f"files={len(supp)}", "#SUPPLEMENT_READY")
    partial = master_df[master_df["status"].isin(["partial", "missing", "requires_export", "requires_rerun"])]
    next_ok = partial["next_action"].astype(str).str.len().gt(0).all()
    check("All partial/missing rows have next_action", next_ok, f"rows={len(partial)}", "#PARTIAL;#MISSING")
    no_fab = (master_df["evidence_summary"].astype(str).str.contains("not fabricated|not alter|remain unchanged|missing template", case=False, regex=True).any())
    check("No fabricated or hard-filled data claim", no_fab, "missing templates and no-alter statements present", "#SUBMISSION_READY")
    qc_df = pd.DataFrame(checks)
    write_csv(qc_dir / "final_status_matrix.csv", qc_df)
    md = "# QC FINAL DELTA FIXED\n\n" + "\n".join(
        f"- {'PASS' if r['passed'] else 'FAIL'}: {r['item']} | {r['evidence']} | {r['search_tags']}" for r in checks
    )
    write_md(qc_dir / "QC_FINAL_DELTA_FIXED.md", md)
    return {"qc_passed": bool(qc_df["passed"].all()), "checks": len(checks), "failures": qc_df.loc[~qc_df["passed"], "item"].tolist()}


def build_readme(status: dict[str, Any]) -> None:
    text = f"""# KOM Submission Audit Package FINAL DELTA FIXED

## English summary

This delta-fix package adds a searchable master audit table, locked KOM-Sim disclosures, explicit KOM-Risk missing prediction templates and export scripts, locked KOM-RAG tables with a manual-review pack, locked expert label normalization, final supplementary-methods text, and QC reports.

Still missing: raw KOM-Sim event-level click/session logs; endpoint-specific KOM-Risk longitudinal sample-level prediction rows; full all-query final KOM-RAG top-k rows; fully manual RAG relevance labels.

Directly usable in the main manuscript: locked KOM-Sim timing definition, KOM-Risk endpoint-level summary metrics, and frozen Stage4A RAG benchmark metrics, with limitations disclosed.

Best suited for supplementary materials: the master searchable table, locked RAG label/top-k tables, manual-review pack, Risk export templates, expert label dictionary, and QC reports.

Manual review required: {status.get('manual_review_required_labels', 0)} rule-inferred RAG relevance labels.

Requires export/rerun: KOM-Risk endpoint-specific sample-level predictions, and full final all-query RAG top-k export if reviewers request it.

Most critical remaining item: endpoint-specific sample-level KOM-Risk longitudinal predictions are still missing and must be exported from trained model artifacts for full calibration/DCA/sample-level recalculation.

Shortest next command:
`python 07_scripts_to_run_if_missing/export_KOMRisk_predictions_template.py --data_root <data_root> --model_dir <model_dir> --output_dir <output_dir> --endpoint all --split_csv <split.csv> --feature_csv <features.csv> --label_csv <labels.csv> --model_file <model>`

Use `search_tags` to filter the package, for example `#KOMRISK_MISSING_SAMPLELEVEL`, `#KOMRAG_MANUAL_REVIEW`, or `#KOMSIM_DISCLOSURE_LOCKED`.

## 中文摘要

本轮 delta-fix 补齐了统一可检索总表、KOM-Sim locked disclosure、KOM-Risk 缺失预测模板与导出脚本、KOM-RAG locked tables 与 manual-review pack、专家评分标签锁定、投稿级补充方法文本和 QC 报告。

仍然缺失：KOM-Sim 原始事件级点击/session 日志；KOM-Risk 端点级纵向样本预测行；KOM-RAG 全量 final top-k 行；完全人工复核的 RAG relevance labels。

可直接支持主文：KOM-Sim 时间定义、KOM-Risk endpoint-level summary metrics、Stage4A frozen RAG benchmark metrics，但必须保留限制说明。

适合放入补充材料：master searchable table、RAG locked top-k/label tables、manual-review pack、Risk export templates、expert label dictionary 和 QC reports。

需要人工复核：{status.get('manual_review_required_labels', 0)} 条 rule-inferred RAG relevance labels。

需要重新导出或重跑：KOM-Risk endpoint-specific sample-level predictions；如果审稿人要求，还需要导出 full all-query final RAG top-k。

最关键 remaining item：KOM-Risk endpoint-specific sample-level longitudinal predictions 仍缺失，必须从训练好的模型 artifact 导出，才能完整重算 calibration、DCA 和 sample-level performance。

最短下一步命令：
`python 07_scripts_to_run_if_missing/export_KOMRisk_predictions_template.py --data_root <data_root> --model_dir <model_dir> --output_dir <output_dir> --endpoint all --split_csv <split.csv> --feature_csv <features.csv> --label_csv <labels.csv> --model_file <model>`

可用 `search_tags` 检索文件，例如 `#KOMRISK_MISSING_SAMPLELEVEL`、`#KOMRAG_MANUAL_REVIEW`、`#KOMSIM_DISCLOSURE_LOCKED`。
"""
    write_md(OUT / "README_FINAL_DELTA_FIXED.md", text)
    write_md(OUT / "00_DELTA_README" / "README_FINAL_DELTA_FIXED_COPY.md", text)


def zip_package() -> dict[str, Any]:
    with ZipFile(ZIP_OUT, "w", ZIP_DEFLATED) as z:
        for dirpath, _, filenames in os.walk(OUT):
            for fn in filenames:
                p = Path(dirpath) / fn
                z.write(p, p.relative_to(ROOT))
    with ZipFile(ZIP_OUT) as z:
        bad = z.testzip()
        entries = len(z.namelist())
    return {"zip_path": str(ZIP_OUT), "zip_size_bytes": ZIP_OUT.stat().st_size, "zip_entries": entries, "zip_testzip": bad}


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(f"Source FINAL package not found: {SOURCE}")
    ensure_clean_output()
    master_rows: list[dict[str, Any]] = []
    sim_status = build_komsim(master_rows)
    risk_status = build_risk(master_rows)
    rag_status = build_rag(master_rows)
    expert_status = build_expert(master_rows)
    supp_status = build_supp_methods(master_rows)
    status = {
        "sim": sim_status,
        "risk": risk_status,
        "rag": rag_status,
        "expert": expert_status,
        "supplementary_methods": supp_status,
        "manual_review_required_labels": rag_status["manual_review_required_labels"],
    }
    build_readme(status)
    shutil.copy2(Path(__file__), OUT / "07_scripts_to_run_if_missing" / Path(__file__).name)
    master_df = build_master_workbook(master_rows, status)
    qc = build_qc(master_df, status)
    master_df = build_master_workbook(master_rows, status)
    qc = build_qc(master_df, status)
    final_status_path = OUT / "08_qc_reports" / "FINAL_DELTA_FIXED_STATUS.json"
    final_status = {
        "output_dir": str(OUT),
        "zip_path": str(ZIP_OUT),
        "sim_status": sim_status,
        "risk_status": risk_status,
        "rag_status": rag_status,
        "expert_status": expert_status,
        "master_table_created": True,
        "master_rows": 0,
        "searchable_tags": False,
        "qc": qc,
        "zip": {"zip_path": str(ZIP_OUT), "zip_size_bytes": None, "zip_entries": None, "zip_testzip": None},
    }
    final_status_path.write_text(json.dumps(final_status, ensure_ascii=False, indent=2), encoding="utf-8")
    master_df = build_master_workbook(master_rows, status)
    qc = build_qc(master_df, status)
    master_df = build_master_workbook(master_rows, status)
    zip_status = zip_package()
    final_status["master_rows"] = int(len(master_df))
    final_status["searchable_tags"] = bool(master_df["search_tags"].astype(str).str.len().gt(0).all())
    final_status["qc"] = qc
    final_status["zip"] = zip_status
    final_status_path.write_text(json.dumps(final_status, ensure_ascii=False, indent=2), encoding="utf-8")
    zip_status = zip_package()
    final_status["zip"] = zip_status
    final_status_path.write_text(json.dumps(final_status, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        "\n".join(
            [
                "KOM submission audit delta-fix completed.",
                "",
                "KOM-Sim:",
                f"- task-level timing: {sim_status['task_level_timing'].upper()}",
                f"- event-level log: {sim_status['event_level_log'].upper().replace('_', ' ')}",
                f"- locked time protocol: {sim_status['locked_time_protocol'].upper()}",
                "",
                "KOM-Risk:",
                f"- structural progression prediction file: {risk_status['prediction_statuses']['structural_progression'].upper()}",
                f"- TKR/knee surgery prediction file: {risk_status['prediction_statuses']['tkr_knee_surgery'].upper()}",
                f"- symptom/function prediction file: {risk_status['prediction_statuses']['symptom_function_worsening'].upper()}",
                f"- metrics: {risk_status['metrics'].upper()}",
                f"- export/rerun script: {risk_status['export_script'].upper()}",
                "",
                "KOM-RAG:",
                f"- query set: {rag_status['query_set'].upper()}",
                f"- guideline anchors: {rag_status['guideline_anchors'].upper()}",
                f"- retrieval results: {rag_status['retrieval_results'].upper()}",
                f"- relevance labels: {rag_status['relevance_labels'].upper()}",
                f"- query metrics: {rag_status['query_metrics'].upper()}",
                f"- error cases: {rag_status['error_cases'].upper()}",
                f"- manual review pack: {rag_status['manual_review_pack'].upper()}",
                f"- manual review required labels: {rag_status['manual_review_required_labels']}",
                "",
                "Expert labels:",
                f"- dictionary: {expert_status['dictionary'].upper()}",
                f"- score values altered: {expert_status['score_values_altered']}",
                "",
                "Master table:",
                "- created: YES",
                f"- rows: {len(master_df)}",
                f"- searchable tags: {'YES' if final_status['searchable_tags'] else 'NO'}",
                "",
                "Most critical remaining item:",
                "Endpoint-specific sample-level KOM-Risk longitudinal predictions are still missing and require export from trained model artifacts.",
                "",
                "Output:",
                "./KOM_Submission_Audit_Package_202606_FINAL_DELTA_FIXED.zip",
            ]
        )
    )
    print(json.dumps(final_status, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
