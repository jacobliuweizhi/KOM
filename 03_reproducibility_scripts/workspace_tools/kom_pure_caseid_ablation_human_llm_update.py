import csv
import json
import math
import os
import re
import shutil
import statistics
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path("C:/OAI\u7814\u7a76\u9879\u76ee/pythonProject1/KOM\u8fd4\u4fee\u4fee\u6539")
WORKBOOK = ROOT / "\u6295\u7a3f\u4f7f\u7528" / "KOM_\u7eaf\u51c0\u7248.xlsx"
LOCAL = ROOT / "\u672c\u5730\u5316" / "koa_mdt_agents"
ENV_PATH = LOCAL / ".env"
ABLATION_CSV = LOCAL / "data/processed/ablation/ablation_full_120cases.csv"
OUT_DIR = ROOT / "\u6295\u7a3f\u4f7f\u7528" / f"KOM_pure_update_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
RAW_DIR = OUT_DIR / "human_interaction_llm_gpt4o_raw"


def load_env(path: Path) -> dict:
    env = {}
    if path.exists():
        for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def safe_float(value, default=0.0):
    try:
        if value in (None, "", "NA", "not_applicable_no_gold_set"):
            return default
        return float(value)
    except Exception:
        return default


def safe_bool(value):
    return str(value).strip().lower() in {"true", "yes", "1", "pass"}


def text_has(text: str, terms) -> int:
    t = str(text or "").lower()
    return sum(1 for term in terms if term.lower() in t)


def rule_score_ablation(rows, arm_code, arm_id):
    samples = len(rows)
    text = "\n".join(str(r.get("output_prefix", "")) for r in rows)
    domain_counts = [safe_float(r.get("domain_coverage_count")) for r in rows]
    critical_counts = [safe_float(r.get("critical_error_count")) for r in rows]
    action_hits = [safe_float(r.get("actionability_hit_count")) for r in rows]
    topk = [safe_float(r.get("retrieval_topk_count")) for r in rows]
    guideline = [safe_float(r.get("guideline_anchor_id_count")) for r in rows]
    direct_mentions = [safe_float(r.get("direct_evidence_id_mentions")) for r in rows]
    faith = [safe_float(r.get("faithfulness_proxy")) for r in rows]
    safety_visible = any(safe_bool(r.get("safety_gate_visible")) for r in rows)

    mean_domain = statistics.mean(domain_counts) if domain_counts else 0
    mean_critical = statistics.mean(critical_counts) if critical_counts else 0
    mean_action = statistics.mean(action_hits) if action_hits else 0
    mean_topk = statistics.mean(topk) if topk else 0
    mean_guideline = statistics.mean(guideline) if guideline else 0
    mean_direct = statistics.mean(direct_mentions) if direct_mentions else 0
    mean_faith = statistics.mean(faith) if faith else 0

    safety = 8
    if safety_visible or text_has(text, ["NSAID", "\u7981\u5fcc", "\u80be", "\u80c3\u80a0", "\u6297\u51dd", "CV", "\u5b89\u5168"]):
        safety += 7
    if mean_critical == 0:
        safety += 5
    elif mean_critical <= 1:
        safety += 1
    safety += min(5, text_has(text, ["\u533b\u751f", "\u590d\u6838", "\u8f6c\u8bca", "\u505c\u6b62", "\u968f\u8bbf", "\u4e0d\u76f4\u63a5\u51b3\u5b9a"]))
    safety = max(0, min(25, safety - 5 * max(0, mean_critical - 1)))

    completeness = max(0, min(25, mean_domain / 6 * 25))

    anchors = text_has(text, ["KL", "BMI", "WOMAC", "NRS", "\u75bc\u75db", "\u529f\u80fd", "\u8dcc\u5012", "\u5e73\u8861", "\u808c\u529b", "\u76ee\u6807", "\u504f\u597d"])
    individual = min(8, anchors) + min(8, mean_action) + min(4, text_has(text, ["FITT", "\u8fdb\u9636", "\u8c03\u6574", "\u76d1\u6d4b", "\u9608\u503c"])) + min(3, mean_domain / 2)
    if mean_critical == 0:
        individual += 2
    individual = max(0, min(25, individual))

    evidence_terms = text_has(text, ["guideline", "\u6307\u5357", "evidence", "\u8bc1\u636e", "meta", "RCT", "\u968f\u673a", "\u7cfb\u7edf\u7efc\u8ff0", "KOA-EU"])
    evidence = min(6, mean_topk / 3) + min(6, mean_guideline * 1.5) + min(5, evidence_terms) + min(4, mean_direct) + min(4, mean_faith * 6)
    if arm_code == "B_no_rag":
        evidence = min(evidence, 8)
    if arm_code == "D_bare":
        evidence = min(evidence, 4)
    evidence = max(0, min(25, evidence))

    total = safety + completeness + individual + evidence
    design_warning = "No"
    if arm_code == "A_full" and mean_topk == 0:
        design_warning = "Source ablation file marks full-system retrieval indicators as zero; rule score uses available text and stored indicators only."
    if arm_code == "B_no_rag" and mean_topk > 0:
        design_warning = "Source ablation file marks no-RAG arm retrieval indicators as present; RAG-specific rule score is capped."
    return {
        "sample_count": samples,
        "safety_gate_score": round(safety, 2),
        "five_domain_score": round(completeness, 2),
        "individualization_score": round(individual, 2),
        "evidence_traceability_score": round(evidence, 2),
        "total_rule_score": round(total, 2),
        "mean_existing_computer_score": round(statistics.mean([safe_float(r.get("computer_total_score")) for r in rows]), 2),
        "rule_basis": "Deterministic rule score from ablation source metrics plus available output text prefix; not LLM.",
        "design_warning": design_warning,
    }


def add_or_reset_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def write_ablation_rule_sheets(wb):
    rows = list(csv.DictReader(ABLATION_CSV.open("r", encoding="utf-8-sig", newline="")))
    target_arms = {
        "\u5b8c\u6574\u7cfb\u7edf_\u4f18\u5316\u7248": "A_full",
        "\u65e0RAG": "B_no_rag",
        "\u65e0MDT": "C_no_mdt",
        "B0_\u88f8\u6a21\u578b\u81ea\u7531\u76f4\u51fa": "D_bare",
    }
    # Align with the Stage4B LLM test, which used three repeats.
    grouped = {}
    for r in rows:
        if r.get("arm_id") not in target_arms:
            continue
        if str(r.get("sample")) not in {"1", "2", "3"}:
            continue
        case_id = r.get("case_id")
        arm_id = r.get("arm_id")
        grouped.setdefault((case_id, arm_id), []).append(r)

    detail_headers = [
        "case_id",
        "arm_code",
        "arm_id",
        "sample_count",
        "safety_gate_score",
        "five_domain_score",
        "individualization_score",
        "evidence_traceability_score",
        "total_rule_score",
        "mean_existing_computer_score",
        "rule_basis",
        "design_warning",
    ]
    ws = add_or_reset_sheet(wb, "\u6d88\u878d_\u56db\u81c2_\u89c4\u5219\u6307\u6807")
    ws.append(detail_headers)
    detail_rows = []
    for (case_id, arm_id), rs in sorted(grouped.items()):
        arm_code = target_arms[arm_id]
        score = rule_score_ablation(rs, arm_code, arm_id)
        row = {"case_id": case_id, "arm_code": arm_code, "arm_id": arm_id, **score}
        detail_rows.append(row)
        ws.append([row.get(h) for h in detail_headers])

    summary_headers = [
        "arm_code",
        "arm_id",
        "n_cases",
        "mean_total_rule_score",
        "sd_total_rule_score",
        "mean_safety",
        "mean_completeness",
        "mean_individualization",
        "mean_evidence_traceability",
        "direction_note",
    ]
    ws2 = add_or_reset_sheet(wb, "\u6d88\u878d_\u56db\u81c2_\u89c4\u5219_\u81c2\u7ea7\u6c47\u603b")
    ws2.append(summary_headers)
    summary = []
    for arm_code in ["A_full", "B_no_rag", "C_no_mdt", "D_bare"]:
        rs = [r for r in detail_rows if r["arm_code"] == arm_code]
        vals = [r["total_rule_score"] for r in rs]
        note = "Deterministic reproducible rule score; compare with LLM semantic score."
        if arm_code == "B_no_rag":
            note = "RAG-related evidence traceability capped because this is the no-RAG arm."
        if arm_code == "D_bare":
            note = "Bare output expected to score lowest for completeness, actionability and traceability."
        item = {
            "arm_code": arm_code,
            "arm_id": rs[0]["arm_id"] if rs else "",
            "n_cases": len(rs),
            "mean_total_rule_score": round(statistics.mean(vals), 2) if vals else "",
            "sd_total_rule_score": round(statistics.stdev(vals), 2) if len(vals) > 1 else 0,
            "mean_safety": round(statistics.mean([r["safety_gate_score"] for r in rs]), 2) if rs else "",
            "mean_completeness": round(statistics.mean([r["five_domain_score"] for r in rs]), 2) if rs else "",
            "mean_individualization": round(statistics.mean([r["individualization_score"] for r in rs]), 2) if rs else "",
            "mean_evidence_traceability": round(statistics.mean([r["evidence_traceability_score"] for r in rs]), 2) if rs else "",
            "direction_note": note,
        }
        summary.append(item)
        ws2.append([item.get(h) for h in summary_headers])
    return {"detail_rows": len(detail_rows), "summary": summary}


def parse_json_object(text):
    try:
        return json.loads(text)
    except Exception:
        pass
    match = re.search(r"\{.*\}", text or "", re.S)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            return {}
    return {}


def api_call(base_url, api_key, model, prompt, timeout=90):
    url = base_url.rstrip("/") + "/chat/completions"
    body = {
        "model": model,
        "temperature": 0.5,
        "messages": [
            {
                "role": "system",
                "content": "You are a strict clinical-prescription quality evaluator for knee osteoarthritis research. Return JSON only.",
            },
            {"role": "user", "content": prompt},
        ],
    }
    data = json.dumps(body, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8", errors="replace"))


def build_prompt(item):
    prescription = (item.get("prescription") or "")[:4500]
    return f"""
Score this doctor-written knee osteoarthritis prescription as a research quality-control task.
Do not provide medical advice. Do not rewrite the prescription. Score only what is written.

Case_ID: {item.get('case_id')}
Quadrant: {item.get('quadrant')}
Mode/Arm: {item.get('mode')}
Doctor seniority: {item.get('seniority')}
Specialty: {item.get('specialty')}

Dimensions:
1. safety_0_10: medication/injection/surgery safety boundaries, contraindication handling, and clinician review.
2. guideline_alignment_0_10: alignment with KOA conservative care, exercise, weight management, topical-first medication, injection caution, and referral boundaries.
3. individualization_0_10: use of case-specific symptoms, function, risk, goals, preferences, comorbidity and missing information.
4. actionability_0_10: concrete dose/frequency/intensity/timing/follow-up/escalation rules.
5. faithfulness_0_10: claims remain supported by the provided case and standard KOA care, without unsupported certainty.
6. hallucination_rate_0_1: 0 means no evident unsupported clinical invention; 1 means severe unsupported invention.
7. overall_0_100: integrated quality score.
Also count critical_error_count, major_error_count, minor_error_count.

Use strict but fair scoring. A short or vague prescription should not receive a high actionability or completeness-related score.
Return one JSON object exactly with keys:
safety_0_10, guideline_alignment_0_10, individualization_0_10, actionability_0_10, faithfulness_0_10, hallucination_rate_0_1, overall_0_100, critical_error_count, major_error_count, minor_error_count, pass_status, comments.

Prescription:
{prescription}
""".strip()


def score_one(item, base_url, api_key, model):
    raw_path = RAW_DIR / f"row{item['excel_row']:04d}_{item['participant']}_{item['case_id']}_{item['mode']}.json"
    if raw_path.exists():
        raw = json.loads(raw_path.read_text(encoding="utf-8"))
        return raw
    prompt = build_prompt(item)
    last_error = None
    for attempt in range(1, 4):
        try:
            response = api_call(base_url, api_key, model, prompt)
            content = response["choices"][0]["message"].get("content", "")
            parsed = parse_json_object(content)
            raw = {
                "status": "ok",
                "attempt": attempt,
                "model": model,
                "temperature": 0.5,
                "excel_row": item["excel_row"],
                "participant": item["participant"],
                "task_number": item["task_number"],
                "case_id": item["case_id"],
                "quadrant": item["quadrant"],
                "mode": item["mode"],
                "prescription_chars": len(item.get("prescription") or ""),
                "parsed": parsed,
                "raw_content": content,
                "response_model": response.get("model"),
                "created": response.get("created"),
            }
            raw_path.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")
            return raw
        except Exception as exc:
            last_error = repr(exc)
            time.sleep(1.5 * attempt)
    raw = {
        "status": "error",
        "model": model,
        "excel_row": item["excel_row"],
        "participant": item["participant"],
        "task_number": item["task_number"],
        "case_id": item["case_id"],
        "quadrant": item["quadrant"],
        "mode": item["mode"],
        "parsed": {},
        "error": last_error,
    }
    raw_path.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")
    return raw


def get_or_add_columns(ws, header_row, names):
    existing = {ws.cell(header_row, c).value: c for c in range(1, ws.max_column + 1)}
    for name in names:
        if name not in existing:
            col = ws.max_column + 1
            ws.cell(header_row, col).value = name
            existing[name] = col
    return existing


def to_num(parsed, key, lo, hi):
    try:
        value = float(parsed.get(key))
        return max(lo, min(hi, value))
    except Exception:
        return None


def run_human_llm_scoring(wb):
    env = load_env(ENV_PATH)
    base_url = env.get("XIAOAI_BASE_URL") or env.get("OPENAI_BASE_URL") or "https://xiaoai.plus/v1"
    api_key = env.get("XIAOAI_API_KEY") or env.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("API key not found in .env")
    model = "gpt-4o"
    ws = wb["\u4eba\u673a\u4ea4\u4e92\u539f\u59cb"]
    header_row = 4
    headers = {ws.cell(header_row, c).value: c for c in range(1, ws.max_column + 1)}
    colmap = get_or_add_columns(
        ws,
        header_row,
        [
            "LLM_\u8bc4\u4ef7\u5206",
            "LLM_\u5b89\u5168\u6027",
            "LLM_\u6307\u5357\u4e00\u81f4",
            "LLM_\u4e2a\u4f53\u5316",
            "LLM_\u53ef\u6267\u884c\u6027",
            "LLM_\u5fe0\u5b9e\u5ea6",
            "LLM_\u5e7b\u89c9\u7387",
            "LLM_critical",
            "LLM_major",
            "LLM_minor",
            "LLM\u6a21\u578b",
            "LLM_raw\u8def\u5f84",
            "LLM_\u8bc4\u4ef7\u5907\u6ce8",
        ],
    )
    headers = {ws.cell(header_row, c).value: c for c in range(1, ws.max_column + 1)}
    items = []
    for r in range(header_row + 1, ws.max_row + 1):
        prescription = ws.cell(r, headers["\u533b\u751f\u5904\u65b9\u539f\u6587"]).value
        if prescription is None or str(prescription).strip() == "":
            continue
        items.append(
            {
                "excel_row": r,
                "participant": str(ws.cell(r, headers["\u53c2\u4e0e\u8005ID"]).value or ""),
                "task_number": str(ws.cell(r, headers["\u4efb\u52a1\u5e8f\u53f7"]).value or ""),
                "case_id": str(ws.cell(r, headers["\u75c5\u4f8bID"]).value or ""),
                "quadrant": str(ws.cell(r, headers["\u8c61\u9650"]).value or ""),
                "mode": str(ws.cell(r, headers["\u6a21\u5f0f"]).value or ""),
                "seniority": str(ws.cell(r, headers["\u533b\u751f\u5e74\u8d44"]).value or ""),
                "specialty": str(ws.cell(r, headers["\u4e13\u79d1"]).value or ""),
                "prescription": str(prescription),
            }
        )
    results = []
    progress_path = OUT_DIR / "human_interaction_llm_gpt4o_progress.json"
    with ThreadPoolExecutor(max_workers=8) as ex:
        futs = [ex.submit(score_one, item, base_url, api_key, model) for item in items]
        done = 0
        for fut in as_completed(futs):
            raw = fut.result()
            results.append(raw)
            done += 1
            if done % 20 == 0 or done == len(items):
                progress_path.write_text(
                    json.dumps(
                        {
                            "completed": done,
                            "total": len(items),
                            "ok": sum(1 for x in results if x.get("status") == "ok"),
                            "error": sum(1 for x in results if x.get("status") != "ok"),
                            "updated_at": datetime.now().isoformat(timespec="seconds"),
                        },
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )

    by_row = {int(r["excel_row"]): r for r in results}
    for r, raw in by_row.items():
        parsed = raw.get("parsed") or {}
        rel_raw = str((RAW_DIR / f"row{r:04d}_{raw.get('participant')}_{raw.get('case_id')}_{raw.get('mode')}.json").relative_to(ROOT / "\u6295\u7a3f\u4f7f\u7528"))
        ws.cell(r, colmap["LLM_\u8bc4\u4ef7\u5206"]).value = to_num(parsed, "overall_0_100", 0, 100)
        ws.cell(r, colmap["LLM_\u5b89\u5168\u6027"]).value = to_num(parsed, "safety_0_10", 0, 10)
        ws.cell(r, colmap["LLM_\u6307\u5357\u4e00\u81f4"]).value = to_num(parsed, "guideline_alignment_0_10", 0, 10)
        ws.cell(r, colmap["LLM_\u4e2a\u4f53\u5316"]).value = to_num(parsed, "individualization_0_10", 0, 10)
        ws.cell(r, colmap["LLM_\u53ef\u6267\u884c\u6027"]).value = to_num(parsed, "actionability_0_10", 0, 10)
        ws.cell(r, colmap["LLM_\u5fe0\u5b9e\u5ea6"]).value = to_num(parsed, "faithfulness_0_10", 0, 10)
        ws.cell(r, colmap["LLM_\u5e7b\u89c9\u7387"]).value = to_num(parsed, "hallucination_rate_0_1", 0, 1)
        ws.cell(r, colmap["LLM_critical"]).value = to_num(parsed, "critical_error_count", 0, 99)
        ws.cell(r, colmap["LLM_major"]).value = to_num(parsed, "major_error_count", 0, 99)
        ws.cell(r, colmap["LLM_minor"]).value = to_num(parsed, "minor_error_count", 0, 99)
        ws.cell(r, colmap["LLM\u6a21\u578b"]).value = model
        ws.cell(r, colmap["LLM_raw\u8def\u5f84"]).value = rel_raw
        ws.cell(r, colmap["LLM_\u8bc4\u4ef7\u5907\u6ce8"]).value = (parsed.get("comments") if isinstance(parsed, dict) else "") or ("API error" if raw.get("status") != "ok" else "")

    detail = add_or_reset_sheet(wb, "\u4eba\u673a\u4ea4\u4e92_LLM\u8bc4\u5206\u660e\u7ec6")
    detail_headers = [
        "Participant_ID",
        "Task_Number",
        "Case_ID",
        "Quadrant",
        "Mode",
        "Seniority",
        "Specialty",
        "Evaluator_Model",
        "Safety_0_10",
        "Guideline_Alignment_0_10",
        "Individualization_0_10",
        "Actionability_0_10",
        "Faithfulness_0_10",
        "Hallucination_Rate_0_1",
        "Overall_0_100",
        "Critical_Error_Count",
        "Major_Error_Count",
        "Minor_Error_Count",
        "Pass_Status",
        "Comments",
        "Raw_Response_Path",
        "Use_Note",
    ]
    detail.append(detail_headers)
    for item in items:
        raw = by_row[item["excel_row"]]
        parsed = raw.get("parsed") or {}
        raw_rel = str((RAW_DIR / f"row{item['excel_row']:04d}_{item['participant']}_{item['case_id']}_{item['mode']}.json").relative_to(ROOT / "\u6295\u7a3f\u4f7f\u7528"))
        detail.append(
            [
                item["participant"],
                item["task_number"],
                item["case_id"],
                item["quadrant"],
                item["mode"],
                item["seniority"],
                item["specialty"],
                model,
                to_num(parsed, "safety_0_10", 0, 10),
                to_num(parsed, "guideline_alignment_0_10", 0, 10),
                to_num(parsed, "individualization_0_10", 0, 10),
                to_num(parsed, "actionability_0_10", 0, 10),
                to_num(parsed, "faithfulness_0_10", 0, 10),
                to_num(parsed, "hallucination_rate_0_1", 0, 1),
                to_num(parsed, "overall_0_100", 0, 100),
                to_num(parsed, "critical_error_count", 0, 99),
                to_num(parsed, "major_error_count", 0, 99),
                to_num(parsed, "minor_error_count", 0, 99),
                parsed.get("pass_status"),
                parsed.get("comments"),
                raw_rel,
                "Auxiliary LLM reference only; primary human-interaction quality evidence remains deterministic Computer_Total rule score.",
            ]
        )

    trend = add_or_reset_sheet(wb, "\u4eba\u673a\u4ea4\u4e92_LLM\u8c61\u9650\u6a21\u5f0f\u8d8b\u52bf")
    trend.append(["Quadrant", "Mode", "n", "Mean_Overall_0_100", "SD_Overall_0_100", "Mean_Safety", "Mean_Actionability", "Mean_Hallucination_Rate"])
    bins = {}
    for raw in results:
        parsed = raw.get("parsed") or {}
        q = raw.get("quadrant") or ""
        m = raw.get("mode") or ""
        bins.setdefault((q, m), []).append(parsed)
    for (q, m), ps in sorted(bins.items()):
        overall = [to_num(p, "overall_0_100", 0, 100) for p in ps if to_num(p, "overall_0_100", 0, 100) is not None]
        safety = [to_num(p, "safety_0_10", 0, 10) for p in ps if to_num(p, "safety_0_10", 0, 10) is not None]
        action = [to_num(p, "actionability_0_10", 0, 10) for p in ps if to_num(p, "actionability_0_10", 0, 10) is not None]
        halluc = [to_num(p, "hallucination_rate_0_1", 0, 1) for p in ps if to_num(p, "hallucination_rate_0_1", 0, 1) is not None]
        trend.append(
            [
                q,
                m,
                len(ps),
                round(statistics.mean(overall), 2) if overall else "",
                round(statistics.stdev(overall), 2) if len(overall) > 1 else 0,
                round(statistics.mean(safety), 2) if safety else "",
                round(statistics.mean(action), 2) if action else "",
                round(statistics.mean(halluc), 4) if halluc else "",
            ]
        )
    return {
        "items": len(items),
        "ok": sum(1 for x in results if x.get("status") == "ok"),
        "error": sum(1 for x in results if x.get("status") != "ok"),
        "raw_dir": str(RAW_DIR),
        "model": model,
    }


def validate_workbook(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    errors = []
    expected_absent = [
        "\u5904\u65b9_AI\u8bc4\u4ef7",
        "\u5904\u65b9_AI\u8bc4\u4ef7_ICC",
        "\u5904\u65b9_AI\u8bc4\u4ef7_ICC_\u6a21\u578b\u5bf9\u6bd4",
        "\u5904\u65b9_AI\u8bc4\u4ef7_mini\u65e7\u7ed3\u679c",
        "\u5904\u65b9_AI\u8bc4\u4ef7_ICC_mini\u65e7\u7ed3\u679c",
        "\u5904\u65b9_AI\u8bc4\u4ef7_gpt4o\u65e0rubric\u65e7",
        "\u5904\u65b9_AI\u8bc4\u4ef7_ICC_gpt4o\u65e0rubric\u65e7",
    ]
    for s in expected_absent:
        if s in wb.sheetnames:
            errors.append(f"old standalone sheet still present: {s}")
    required = [
        "\u6d88\u878d_\u56db\u81c2_\u8bc4\u5206\u660e\u7ec6",
        "\u6d88\u878d_\u56db\u81c2_\u89c4\u5219\u6307\u6807",
        "\u6d88\u878d_\u56db\u81c2_\u89c4\u5219_\u81c2\u7ea7\u6c47\u603b",
        "\u4eba\u673a\u4ea4\u4e92_LLM\u8bc4\u5206\u660e\u7ec6",
        "\u4eba\u673a\u4ea4\u4e92_LLM\u8c61\u9650\u6a21\u5f0f\u8d8b\u52bf",
        "\u68c0\u7d22_StageA_GraphRAG_benchmark",
    ]
    for s in required:
        if s not in wb.sheetnames:
            errors.append(f"required sheet missing: {s}")
    counts = {}
    if "\u6d88\u878d_\u56db\u81c2_\u89c4\u5219\u6307\u6807" in wb.sheetnames:
        counts["ablation_rule_rows"] = wb["\u6d88\u878d_\u56db\u81c2_\u89c4\u5219\u6307\u6807"].max_row - 1
        if counts["ablation_rule_rows"] != 480:
            errors.append(f"ablation rule rows should be 480, got {counts['ablation_rule_rows']}")
    if "\u4eba\u673a\u4ea4\u4e92_LLM\u8bc4\u5206\u660e\u7ec6" in wb.sheetnames:
        counts["human_llm_rows"] = wb["\u4eba\u673a\u4ea4\u4e92_LLM\u8bc4\u5206\u660e\u7ec6"].max_row - 1
        if counts["human_llm_rows"] != 780:
            errors.append(f"human LLM rows should be 780, got {counts['human_llm_rows']}")
    return {"status": "PASS" if not errors else "FAIL", "errors": errors, "counts": counts, "sheet_count": len(wb.sheetnames)}


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    backup = OUT_DIR / "KOM_\u7eaf\u51c0\u7248.before_human_llm_gpt4o.xlsx"
    if not backup.exists():
        shutil.copy2(WORKBOOK, backup)
    wb = openpyxl.load_workbook(WORKBOOK)
    ablation_audit = write_ablation_rule_sheets(wb)
    human_audit = run_human_llm_scoring(wb)
    wb.save(WORKBOOK)
    validation = validate_workbook(WORKBOOK)
    report = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "workbook": str(WORKBOOK),
        "backup": str(backup),
        "ablation_rule": ablation_audit,
        "human_llm": human_audit,
        "validation": validation,
    }
    (OUT_DIR / "tasks3_4_completion_audit.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if validation["status"] != "PASS":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
