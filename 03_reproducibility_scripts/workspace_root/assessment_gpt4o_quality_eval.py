from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import time
import traceback
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT = Path(
    "C:/OAI" + "\u7814\u7a76\u9879\u76ee" + "/pythonProject1/KOM" + "\u8fd4\u4fee\u4fee\u6539"
)
POST_DIR = PROJECT / "\u6295\u7a3f\u4f7f\u7528"
LOCAL_DIR = PROJECT / "\u672c\u5730\u5316" / "koa_mdt_agents"
ENV_FILE = LOCAL_DIR / ".env"
WORKBOOK = POST_DIR / (
    "KOM_"
    + "\u7eaf\u51c0\u7248"
    + "_"
    + "\u516d\u4e13\u5bb6"
    + "API"
    + "\u56de\u586b\u7248"
    + "_"
    + "\u6574\u7406\u7248"
    + "2.xlsx"
)
MANIFEST = POST_DIR / ("\u8bc4\u4f30\u62a5\u544a" + "_120" + "\u6e05\u5355" + ".csv")
SHEET_NAME = "\u8bc4\u4f30_LLM\u8d28\u91cf\u8bc4\u4ef7"
NOTE_TEXT = "\u8bc4\u4ef7\u6a21\u578b=gpt-4o,\u9010\u4efd\u62a5\u544a\u539f\u6587\u771f\u5b9e\u8bc4\u5206"

RUN_DIR = POST_DIR / "assessment_gpt4o_eval_v2_20260604"
RAW_DIR = RUN_DIR / "raw_responses"
LOG_PATH = RUN_DIR / "assessment_gpt4o_eval.log"
CSV_PATH = RUN_DIR / "assessment_llm_quality_gpt4o.csv"
VALIDATION_PATH = RUN_DIR / "assessment_gpt4o_validation.json"

MODEL = "gpt-4o"
TEMPERATURE = 0.4
TOP_P = 0.9
TIMEOUT_SECONDS = 120

OUTPUT_COLUMNS = [
    "Case_ID",
    "Evaluator_Type",
    "Evaluator_Model",
    "Status",
    "Included_In_Final_Analysis",
    "Overall_0_100",
    "Imaging_Accuracy",
    "Information_Extraction",
    "Staging_0_10",
    "Phenotyping_0_10",
    "Progression_Prediction",
    "Etiology_Risk_0_10",
    "Missing_Information",
    "Safety_Awareness",
    "Explainability_0_10",
    "Clinical_Readability",
    "Critical_Error_Count",
    "Major_Error_Count",
    "Pass_Status",
    "Evaluator_Comments",
    "evidence_quote",
    "Artifact_Relative_path",
]

JSON_FIELDS = [
    "Overall_0_100",
    "Imaging_Accuracy",
    "Information_Extraction",
    "Staging_0_10",
    "Phenotyping_0_10",
    "Progression_Prediction",
    "Etiology_Risk_0_10",
    "Missing_Information",
    "Safety_Awareness",
    "Explainability_0_10",
    "Clinical_Readability",
    "Critical_Error_Count",
    "Major_Error_Count",
    "Pass_Status",
    "Evaluator_Comments",
    "evidence_quote",
    "critical_errors",
    "major_errors",
]


def log(msg: str) -> None:
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def read_env(path: Path) -> Dict[str, str]:
    data: Dict[str, str] = {}
    if not path.exists():
        raise FileNotFoundError(f"Missing .env: {path}")
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        data[k.strip()] = v.strip().strip('"').strip("'")
    return data


def read_manifest(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    if len(rows) != 120:
        log(f"WARNING manifest row count is {len(rows)}, expected 120")
    return rows


def candidate_paths(relative_or_abs: str) -> List[Path]:
    raw = Path(relative_or_abs)
    candidates: List[Path] = []
    if raw.is_absolute():
        candidates.append(raw)
    else:
        candidates.extend([LOCAL_DIR / raw, PROJECT / raw, POST_DIR / raw])
    return candidates


def locate_case_dir(case_id: str, artifact_relative_path: str) -> Tuple[Path, str]:
    for p in candidate_paths(artifact_relative_path):
        if p.exists():
            return (p.parent if p.is_file() else p), str(p)
    fallback = LOCAL_DIR / "data" / "processed" / "assessment" / "runs_120" / case_id
    if fallback.exists():
        return fallback, str(fallback)
    matches = list((LOCAL_DIR / "data" / "processed" / "assessment" / "runs_120").glob(f"{case_id}*"))
    if matches:
        return matches[0], str(matches[0])
    raise FileNotFoundError(f"Cannot locate report directory for {case_id}: {artifact_relative_path}")


def locate_report_file(case_dir: Path) -> Path:
    preferred = [
        "assessment_report.md",
        "assessment_report.json",
        "report.md",
        "report.json",
        "structured_assessment_report.md",
        "structured_report.json",
    ]
    for name in preferred:
        p = case_dir / name
        if p.exists() and p.is_file():
            return p

    scored = {"quality_evaluation_llm.json", "quality_evaluation.json"}
    patterns = [
        "*assessment*report*.md",
        "*assessment*report*.json",
        "*structured*report*.md",
        "*structured*report*.json",
        "*.md",
        "*.json",
    ]
    for pattern in patterns:
        for p in sorted(case_dir.glob(pattern)):
            if p.name in scored or "quality_evaluation" in p.name.lower():
                continue
            if p.is_file():
                return p
    raise FileNotFoundError(f"No assessment report file found in {case_dir}")


def read_one_report_file(report_file: Path) -> str:
    text = report_file.read_text(encoding="utf-8-sig", errors="replace")
    if report_file.suffix.lower() == ".json":
        try:
            obj = json.loads(text)
            text = json.dumps(obj, ensure_ascii=False, indent=2)
        except Exception:
            pass
    return text


def read_report_text(report_file: Path) -> str:
    if "quality_evaluation" in report_file.name.lower():
        raise ValueError(f"Refusing to evaluate prior scoring file as report input: {report_file}")
    text = read_one_report_file(report_file)
    paired = report_file.parent / "structured_report.json"
    if report_file.name != "structured_report.json" and paired.exists():
        structured = read_one_report_file(paired)
        text = text.rstrip() + "\n\n## Structured report source\n" + structured
    if len(text.strip()) <= 300:
        raise ValueError(f"Report text too short for real evaluation: {report_file} length={len(text.strip())}")
    if len(text) > 30000:
        text = text[:30000] + "\n\n[TRUNCATED_FOR_EVALUATION: original report exceeded 30000 characters]"
    return text


def build_prompt(report_text: str) -> Tuple[str, str]:
    system = (
        "\u4f60\u662f\u8d44\u6df1\u9aa8\u79d1/\u8fd0\u52a8\u533b\u5b66\u4e13\u5bb6\uff0c"
        "\u4e25\u683c\u8bc4\u4ef7\u5355\u4efdKOA\u8bc4\u4f30\u62a5\u544a\u3002"
        "\u5fc5\u987b\u4ec5\u4f9d\u636e\u6240\u7ed9\u62a5\u544a\u539f\u6587\u6253\u5206\uff0c"
        "\u5148\u5728 evidence_quote \u5f15\u7528\u62a5\u544a\u539f\u53e5\uff0c"
        "\u518d\u9010\u6761\u5217\u51fa\u8be5\u62a5\u544a\u7684\u4e25\u91cd/\u91cd\u8981\u9519\u8bef\uff08\u6ca1\u6709\u5c31\u7a7a\u6570\u7ec4\uff09\uff0c"
        "\u9519\u8bef\u8ba1\u6570\u5fc5\u987b\u7b49\u4e8e\u5217\u8868\u957f\u5ea6\u3002"
        "\u5206\u6570\u8981\u771f\u5b9e\u53cd\u6620\u8be5\u62a5\u544a\u4e0e\u5176\u4ed6\u62a5\u544a\u7684\u5dee\u5f02\uff0c"
        "\u4e0d\u8981\u7ed9\u6240\u6709\u62a5\u544a\u76f8\u540c\u5206\u3002\u53ea\u8f93\u51faJSON\u3002"
    )
    user = (
        "Evaluate the following KOA assessment report. Output ONLY one valid JSON object with exactly these fields:\n"
        + "\n".join(f"- {field}" for field in JSON_FIELDS)
        + "\n\nScoring rules:\n"
        "- Overall_0_100 is 0-100.\n"
        "- Imaging_Accuracy, Information_Extraction, Staging_0_10, Phenotyping_0_10, Progression_Prediction, "
        "Etiology_Risk_0_10, Safety_Awareness, Explainability_0_10, Clinical_Readability are each 0-10.\n"
        "- evidence_quote must contain 1-2 exact snippets copied from this report, each under 15 words.\n"
        "- critical_errors and major_errors must be arrays of concrete report-specific error descriptions.\n"
        "- Critical_Error_Count and Major_Error_Count must equal the lengths of critical_errors and major_errors.\n"
        "- Pass_Status must be one of: pass, conditional_pass, fail.\n"
        "- Evaluator_Comments must be brief English and specific to this case, not a generic template.\n"
        "- Missing_Information should be English text describing key missing information, or 'None apparent'.\n\n"
        "Calibration guidance to avoid collapsed scoring:\n"
        "- Use the full score range when justified by the report content.\n"
        "- A report with unavailable image-level model output, unsupported prognosis, and missing event/censoring data should not receive high scores in imaging, prognosis, or explainability.\n"
        "- Severity depends on this patient. The same missing item is more serious for a KL3/high-function-burden/high-demand case than for a doubtful low-burden case.\n"
        "- Count critical errors only when a defect could materially mislead clinical interpretation or triage for this case.\n"
        "- Count major errors for concrete missing or unsupported assessment components that limit clinical use.\n"
        "- Do NOT count explicitly disclosed unavailability as a critical error in KL0-1 low-structural-burden reports unless the report makes a wrong or unsafe clinical conclusion. In such cases, put the issue in major_errors instead.\n"
        "- In severe KL4 or surgery/referral-relevant reports, missing image-level confirmation, missing prognosis probability, missing event/censoring status, and missing competing-risk indicators may each be separate critical errors if they limit triage.\n"
        "- Do not merge multiple defects into one generic error. List each distinct defect separately.\n"
        "- Examples of distinct defects: unavailable image-level output, unsupported prognosis probability, missing event/censoring status, missing competing-risk indicators, incomplete safety-gate data, limited phenotype justification, or overly generic prescription handoff.\n"
        "- In low-burden cases, unsupported prognosis may be major rather than critical; in high-burden KL3/high-function-burden/high-demand cases, the same defect may be critical if it could mislead triage.\n"
        "- Safety_Awareness should use granular values, including decimals if appropriate, based on both safety gates and patient risk level.\n"
        "- Critical-error count anchors must vary by case: KL0-1 low-demand reports with explicit limitations should have 0 critical errors unless they contain a dangerous contradiction; KL0 with moderate symptoms may have 0-1 only if prognosis is overclaimed; KL1-2 or high-demand cases may have 1-2; KL3/high function burden may have 2-3; KL4 or referral-relevant high-burden cases with missing prognosis/event data may have 3-4. Do not assign every case the same critical count.\n"
        "- Major-error count anchors: count each missing or weak domain separately, such as imaging model unavailable, no patient-level risk probability, missing event/censoring status, missing competing-risk indicators, incomplete phenotype rationale, incomplete safety-gate detail, or generic prescription handoff. Low-burden reports may have only 2 major errors if the limitations are clearly disclosed. Severe KL4/referral-relevant reports may have 6-7 major errors when several missing domains limit triage. Do not compress every report into 3-5 major errors.\n"
        "- Safety_Awareness anchors: 9-10 complete safety reasoning; 8-8.5 adequate general safety gates in low-structural-burden cases; 7-7.5 safety gates stated but missing verification data; 6-6.5 high-burden case with missing safety/event details; 5-5.5 severe KL4/referral-relevant case with missing event/safety detail; below 5 if a safety-relevant omission could mislead clinician review. Use decimals to separate cases.\n"
        "- Illustrative calibration, do not copy mechanically: a low-burden KL0 report with clear limitations might be Overall 68-75, critical_errors 0, major_errors 2-3, Safety 7.5-8.5; a KL0 report with moderate pain/function but no dangerous contradiction might still have critical_errors 0 and major_errors 3-4; a KL1-2 high-demand report might be Overall 58-68, critical_errors 1-2, major_errors 3-5, Safety 6.5-7.5; a KL3 high-function-burden report with unavailable imaging/prognosis might be Overall 50-62, critical_errors 2-3, major_errors 4-6, Safety 5.5-7; a KL4/referral-relevant report missing event/prognosis/safety detail might be Overall 45-58, critical_errors 3-4, major_errors 5-7, Safety 5-6.5.\n"
        "Few-shot calibration examples, do not copy case text:\n"
        "Example low-burden report: {\"Overall_0_100\":72,\"Safety_Awareness\":8.0,\"critical_errors\":[],\"major_errors\":[\"Image-level model output unavailable\",\"Validated prognosis model/features unavailable\"],\"Pass_Status\":\"conditional_pass\"}.\n"
        "Example severe KL4/referral-relevant report: {\"Overall_0_100\":52,\"Safety_Awareness\":5.5,\"critical_errors\":[\"Image-level confirmation unavailable in severe structural disease\",\"No patient-level progression probability\",\"Event/censoring status missing\",\"Competing-risk indicators missing\"],\"major_errors\":[\"Medication safety verification incomplete\",\"Phenotype rationale not fully justified\",\"Prescription handoff is generic\",\"No quantified time horizon risk\",\"No uncertainty interval\",\"No triage-specific referral detail\"],\"Pass_Status\":\"fail\"}.\n"
        "- Pass_Status rule: fail if Overall_0_100 < 60 or there are >=2 critical errors; conditional_pass if Overall_0_100 is 60-79 or any major errors remain; pass only if Overall_0_100 >=80, no critical errors, and no more than one minor/major limitation.\n"
        "- Do not reuse the previous case's scores or wording.\n\n"
        "KOA assessment report:\n"
        "-----\n"
        + report_text
        + "\n-----"
    )
    return system, user


def api_chat(base_url: str, api_key: str, system: str, user: str) -> Dict[str, Any]:
    url = base_url.rstrip("/") + "/chat/completions"
    payload = {
        "model": MODEL,
        "temperature": TEMPERATURE,
        "top_p": TOP_P,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
    }
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=TIMEOUT_SECONDS) as resp:
        body = resp.read().decode("utf-8", errors="replace")
    return json.loads(body)


def parse_json_object(text: str) -> Dict[str, Any]:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.I)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        obj = json.loads(cleaned)
    except Exception:
        match = re.search(r"\{.*\}", cleaned, flags=re.S)
        if not match:
            raise
        obj = json.loads(match.group(0))
    if not isinstance(obj, dict):
        raise ValueError("Parsed response is not a JSON object")
    return obj


def coerce_score(value: Any, lo: float, hi: float) -> float:
    if isinstance(value, str):
        value = value.strip()
    score = float(value)
    if score < lo:
        score = lo
    if score > hi:
        score = hi
    return round(score, 3)


def normalize_eval(obj: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    out["Overall_0_100"] = coerce_score(obj.get("Overall_0_100"), 0, 100)
    for field in [
        "Imaging_Accuracy",
        "Information_Extraction",
        "Staging_0_10",
        "Phenotyping_0_10",
        "Progression_Prediction",
        "Etiology_Risk_0_10",
        "Safety_Awareness",
        "Explainability_0_10",
        "Clinical_Readability",
    ]:
        out[field] = coerce_score(obj.get(field), 0, 10)
    critical_errors = obj.get("critical_errors", [])
    major_errors = obj.get("major_errors", [])
    if isinstance(critical_errors, str):
        critical_errors = [critical_errors] if critical_errors.strip() else []
    if isinstance(major_errors, str):
        major_errors = [major_errors] if major_errors.strip() else []
    if not isinstance(critical_errors, list):
        critical_errors = []
    if not isinstance(major_errors, list):
        major_errors = []
    critical_errors = [str(item).strip() for item in critical_errors if str(item).strip()]
    major_errors = [str(item).strip() for item in major_errors if str(item).strip()]
    out["Critical_Error_Count"] = len(critical_errors)
    out["Major_Error_Count"] = len(major_errors)
    pass_status = str(obj.get("Pass_Status", "")).strip().lower()
    if out["Overall_0_100"] < 60 or out["Critical_Error_Count"] >= 2:
        pass_status = "fail"
    elif out["Overall_0_100"] < 80 or out["Major_Error_Count"] > 1 or out["Critical_Error_Count"] > 0:
        pass_status = "conditional_pass"
    elif pass_status not in {"pass", "conditional_pass", "fail"}:
        pass_status = "pass"
    out["Pass_Status"] = pass_status
    out["Missing_Information"] = str(obj.get("Missing_Information", "not_reported_by_evaluator")).strip()
    out["Evaluator_Comments"] = str(obj.get("Evaluator_Comments", "")).strip()
    evidence_quote = obj.get("evidence_quote", "")
    if isinstance(evidence_quote, list):
        evidence_quote = " | ".join(str(x).strip() for x in evidence_quote if str(x).strip())
    out["evidence_quote"] = str(evidence_quote).strip()
    out["critical_errors"] = critical_errors
    out["major_errors"] = major_errors
    missing = [field for field in JSON_FIELDS if field not in out]
    if missing:
        raise ValueError(f"Missing fields after normalization: {missing}")
    return out


def load_existing_ok_rows(path: Path) -> Dict[str, Dict[str, Any]]:
    if not path.exists():
        return {}
    rows: Dict[str, Dict[str, Any]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Case_ID") and row.get("Status") == "ok":
                rows[row["Case_ID"]] = row
    return rows


def write_csv_rows(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in OUTPUT_COLUMNS})


def evaluate_case(
    base_url: str,
    api_key: str,
    case_id: str,
    artifact_relative_path: str,
    attempt_count: int = 3,
) -> Dict[str, Any]:
    case_dir, resolved = locate_case_dir(case_id, artifact_relative_path)
    report_file = locate_report_file(case_dir)
    report_text = read_report_text(report_file)
    system, user = build_prompt(report_text)
    last_error = ""
    for attempt in range(1, attempt_count + 1):
        raw_record: Dict[str, Any] = {
            "case_id": case_id,
            "attempt": attempt,
            "model": MODEL,
            "temperature": TEMPERATURE,
            "top_p": TOP_P,
            "report_file": str(report_file),
            "resolved_artifact_path": resolved,
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        try:
            response = api_chat(base_url, api_key, system, user)
            raw_record["response"] = response
            content = response["choices"][0]["message"]["content"]
            raw_record["content"] = content
            parsed = parse_json_object(content)
            raw_record["parsed"] = parsed
            normalized = normalize_eval(parsed)
            raw_record["normalized"] = normalized
            raw_path = RAW_DIR / f"{case_id}_attempt{attempt}.json"
            raw_path.write_text(json.dumps(raw_record, ensure_ascii=False, indent=2), encoding="utf-8")
            row = {
                "Case_ID": case_id,
                "Evaluator_Type": "LLM evaluator",
                "Evaluator_Model": "llm:gpt-4o",
                "Status": "ok",
                "Included_In_Final_Analysis": "Yes",
                "Artifact_Relative_path": artifact_relative_path,
            }
            row.update(normalized)
            return row
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            raw_record["error"] = last_error
            raw_record["traceback"] = traceback.format_exc()
            raw_path = RAW_DIR / f"{case_id}_attempt{attempt}_error.json"
            raw_path.write_text(json.dumps(raw_record, ensure_ascii=False, indent=2), encoding="utf-8")
            if isinstance(exc, urllib.error.HTTPError) and exc.code in {403, 429, 500, 502, 503, 504}:
                time.sleep(2 ** attempt)
            elif isinstance(exc, (TimeoutError, urllib.error.URLError)):
                time.sleep(2 ** attempt)
            else:
                time.sleep(1)
    return {
        "Case_ID": case_id,
        "Evaluator_Type": "LLM evaluator",
        "Evaluator_Model": "llm:gpt-4o",
        "Status": "failed",
        "Included_In_Final_Analysis": "No",
        "Overall_0_100": "",
        "Imaging_Accuracy": "",
        "Information_Extraction": "",
        "Staging_0_10": "",
        "Phenotyping_0_10": "",
        "Progression_Prediction": "",
        "Etiology_Risk_0_10": "",
        "Missing_Information": f"Evaluation failed after retries: {last_error}",
        "Safety_Awareness": "",
        "Explainability_0_10": "",
        "Clinical_Readability": "",
        "Critical_Error_Count": "",
        "Major_Error_Count": "",
        "Pass_Status": "fail",
        "Evaluator_Comments": f"Evaluation failed: {last_error}",
        "evidence_quote": "",
        "Artifact_Relative_path": artifact_relative_path,
    }


def remove_related_sheets(wb: Any) -> None:
    to_remove = []
    for name in list(wb.sheetnames):
        low = name.lower()
        if name == SHEET_NAME:
            to_remove.append(name)
        elif "\u8bc4\u4f30" in name and "llm" in low and ("mini" in low or "gpt4o" in low or "gpt-4o" in low):
            to_remove.append(name)
    for name in to_remove:
        del wb[name]


def write_workbook(rows: List[Dict[str, Any]], dry_run: bool = False) -> Path:
    if dry_run:
        return WORKBOOK
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")
    backup = WORKBOOK.with_name(
        WORKBOOK.stem + ".before_assessment_gpt4o_" + datetime.now().strftime("%Y%m%d_%H%M%S") + WORKBOOK.suffix
    )
    shutil.copy2(WORKBOOK, backup)
    log(f"Workbook backup created: {backup}")
    wb = load_workbook(WORKBOOK)
    remove_related_sheets(wb)
    ws = wb.create_sheet(SHEET_NAME)
    ws.cell(1, 1).value = NOTE_TEXT
    ws.cell(1, 1).font = Font(bold=True, color="1F4E79")
    ws.cell(1, 1).alignment = Alignment(wrap_text=True)
    for col_idx, col in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(2, col_idx)
        cell.value = col
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, col in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(row_idx, col_idx).value = row.get(col, "")
            ws.cell(row_idx, col_idx).alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A3"
    widths = {
        "A": 20,
        "B": 18,
        "C": 16,
        "D": 12,
        "E": 22,
        "M": 38,
        "T": 48,
        "U": 55,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    for col in range(6, 19):
        ws.column_dimensions[ws.cell(2, col).column_letter].width = 18
    try:
        wb.save(WORKBOOK)
        log(f"Workbook updated: {WORKBOOK}")
        return WORKBOOK
    except PermissionError:
        fallback = WORKBOOK.with_name(WORKBOOK.stem + "_assessment_gpt4o_filled_copy.xlsx")
        wb.save(fallback)
        log(f"Workbook locked; saved fallback copy: {fallback}")
        return fallback


def validate_outputs(workbook_path: Path, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    status_counts: Dict[str, int] = {}
    models = set()
    has_mini_value = False
    ok_rows = [row for row in rows if row.get("Status") == "ok"]
    for row in rows:
        status_counts[str(row.get("Status"))] = status_counts.get(str(row.get("Status")), 0) + 1
        models.add(str(row.get("Evaluator_Model")))
        model_value = str(row.get("Evaluator_Model", "")).lower()
        if "mini" in model_value:
            has_mini_value = True
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    sheet_exists = SHEET_NAME in wb.sheetnames
    ws_rows = 0
    if sheet_exists:
        ws = wb[SHEET_NAME]
        ws_rows = max(ws.max_row - 2, 0)
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is not None and "mini" in str(value).lower():
                    has_mini_value = True
    def unique_count(field: str) -> int:
        return len({str(row.get(field, "")) for row in ok_rows})

    def numeric_values(field: str) -> List[float]:
        vals: List[float] = []
        for row in ok_rows:
            try:
                vals.append(float(row.get(field, "")))
            except Exception:
                pass
        return vals

    overall_vals = numeric_values("Overall_0_100")
    overall_mode_share = 1.0
    overall_sd = 0.0
    if overall_vals:
        counts: Dict[float, int] = {}
        for val in overall_vals:
            counts[val] = counts.get(val, 0) + 1
        overall_mode_share = max(counts.values()) / len(overall_vals)
        if len(overall_vals) > 1:
            mean = sum(overall_vals) / len(overall_vals)
            overall_sd = (sum((x - mean) ** 2 for x in overall_vals) / (len(overall_vals) - 1)) ** 0.5
    comment_values = [str(row.get("Evaluator_Comments", "")).strip() for row in ok_rows]
    comment_unique_rate = len(set(comment_values)) / len(comment_values) if comment_values else 0.0
    collapse_checks = {
        "critical_error_unique_values": unique_count("Critical_Error_Count"),
        "major_error_unique_values": unique_count("Major_Error_Count"),
        "safety_awareness_unique_values": unique_count("Safety_Awareness"),
        "overall_mode_share": round(overall_mode_share, 4),
        "overall_sd": round(overall_sd, 4),
        "pass_status_unique_values": unique_count("Pass_Status"),
        "evaluator_comments_unique_rate": round(comment_unique_rate, 4),
    }
    collapse_pass = (
        collapse_checks["critical_error_unique_values"] >= 4
        and collapse_checks["major_error_unique_values"] >= 4
        and collapse_checks["safety_awareness_unique_values"] >= 4
        and collapse_checks["overall_mode_share"] < 0.60
        and collapse_checks["overall_sd"] > 5
        and collapse_checks["pass_status_unique_values"] >= 2
        and collapse_checks["evaluator_comments_unique_rate"] > 0.90
    )
    validation = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "workbook_path": str(workbook_path),
        "csv_path": str(CSV_PATH),
        "raw_dir": str(RAW_DIR),
        "row_count": len(rows),
        "status_counts": status_counts,
        "models": sorted(models),
        "sheet_exists": sheet_exists,
        "sheet_data_rows": ws_rows,
        "all_model_gpt4o": models == {"llm:gpt-4o"},
        "all_success": len(rows) == 120 and status_counts.get("ok", 0) == 120,
        "mini_related_value_found": has_mini_value,
        "collapse_checks": collapse_checks,
        "collapse_pass": collapse_pass,
        "pass": sheet_exists
        and ws_rows == len(rows)
        and len(rows) == 120
        and status_counts.get("ok", 0) == 120
        and models == {"llm:gpt-4o"}
        and not has_mini_value
        and collapse_pass,
    }
    VALIDATION_PATH.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    return validation


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--no-workbook", action="store_true")
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    RUN_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    env = read_env(ENV_FILE)
    base_url = env.get("XIAOAI_BASE_URL", "https://xiaoai.plus/v1")
    api_key = env.get("XIAOAI_API_KEY") or env.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("API key not found in .env")
    rows_manifest = read_manifest(MANIFEST)
    if args.limit is not None:
        rows_manifest = rows_manifest[: args.limit]

    existing = {} if args.force else load_existing_ok_rows(CSV_PATH)
    completed: Dict[str, Dict[str, Any]] = dict(existing)
    log(f"Starting assessment evaluation with model={MODEL}, manifest_rows={len(rows_manifest)}, resumed_ok={len(existing)}")

    for idx, src in enumerate(rows_manifest, start=1):
        case_id = src.get("Case_ID") or src.get("\ufeffCase_ID") or ""
        rel = src.get("Artifact_Relative_path", "")
        if not case_id:
            raise ValueError(f"Missing Case_ID in manifest row: {src}")
        if case_id in completed and completed[case_id].get("Status") == "ok":
            log(f"[{idx}/{len(rows_manifest)}] skip existing ok {case_id}")
            continue
        log(f"[{idx}/{len(rows_manifest)}] evaluate {case_id}")
        row = evaluate_case(base_url, api_key, case_id, rel)
        completed[case_id] = row
        ordered_rows = [completed.get(r["Case_ID"], {}) for r in rows_manifest if r.get("Case_ID") in completed]
        write_csv_rows(CSV_PATH, ordered_rows)
        log(f"[{idx}/{len(rows_manifest)}] {case_id} status={row.get('Status')} overall={row.get('Overall_0_100')}")

    final_rows = []
    for src in rows_manifest:
        case_id = src["Case_ID"]
        if case_id in completed:
            final_rows.append(completed[case_id])
    write_csv_rows(CSV_PATH, final_rows)
    workbook_path = write_workbook(final_rows, dry_run=args.no_workbook)
    validation = validate_outputs(workbook_path, final_rows) if not args.no_workbook else {
        "row_count": len(final_rows),
        "status_counts": {},
        "pass": False,
        "dry_run": True,
    }
    log("VALIDATION " + json.dumps(validation, ensure_ascii=False))
    if not validation.get("pass") and not args.no_workbook:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
