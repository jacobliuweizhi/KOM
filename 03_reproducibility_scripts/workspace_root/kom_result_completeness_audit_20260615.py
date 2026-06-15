from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
import shutil
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


STATUS_ORDER = {
    "PASS": 0,
    "NOT_APPLICABLE": 1,
    "FOUND_NEEDS_FILL": 2,
    "FOUND_NEEDS_SOURCE_CONFIRMATION": 3,
    "NEEDS_SOURCE_CONFIRMATION": 3,
    "MISSING_CAN_RECOMPUTE": 4,
    "MISSING_REQUIRES_RETRAIN": 5,
    "MISSING_REQUIRES_EXPERIMENT": 6,
}

RESULT_SHEET_NAMES = {
    "README_UPDATE",
    "RESULT_CHECKPOINTS",
    "FOUND_RESULTS",
    "MISSING_RESULTS",
    "RECOMPUTE_TASKS",
    "SOURCE_FILE_MANIFEST",
    "MODULE_STATUS",
    "FIGURE_DATA_STATUS",
    "UPDATE_LOG",
}

TEXT_EXTS = {".csv", ".tsv", ".txt", ".md", ".json", ".jsonl", ".py", ".r", ".yml", ".yaml", ".log"}
TARGET_EXTS = {
    ".xlsx", ".xlsm", ".xls", ".csv", ".tsv", ".txt", ".md", ".json", ".jsonl",
    ".py", ".r", ".yml", ".yaml", ".log", ".zip", ".pkl", ".joblib", ".pt", ".pth",
    ".ckpt", ".onnx", ".parquet", ".svg", ".png", ".pdf", ".pptx", ".docx",
}
SKIP_DIRS = {
    ".git", ".venv", "venv", "node_modules", "__pycache__", ".pytest_cache",
    "site-packages", "$RECYCLE.BIN", "System Volume Information",
    "input_package_extracted", "00_unzipped", "z",
}

CORE_MODULES = [
    ("A", "Standardized cases"),
    ("B", "KOM-Profile"),
    ("C", "KOM-Rad / OAK-Net"),
    ("D", "KOM-Risk"),
    ("E", "SHAP / Feature lineage"),
    ("F", "KOM-KB evidence database"),
    ("G", "KOM-RAG / GraphRAG"),
    ("H", "KOM-MDT / KOM-Treat"),
    ("I", "KOM-Safe"),
    ("J", "KOM-Score"),
    ("K", "KOM-Sim human interaction"),
    ("L", "Main figures and source data"),
]


def getenv_path(name: str, required: bool = True) -> Path | None:
    value = os.environ.get(name)
    if not value:
        if required:
            raise RuntimeError(f"Missing required environment variable: {name}")
        return None
    return Path(value)


def sha256_file(path: Path, max_size: int = 750 * 1024 * 1024) -> str:
    try:
        if path.stat().st_size > max_size:
            return "sha256_skipped_large_file"
        h = hashlib.sha256()
        with path.open("rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception as exc:
        return f"sha256_failed:{exc}"


def safe_str(value: Any, limit: int = 32000) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value)
    text = text.replace("\x00", "")
    if len(text) > limit:
        return text[: limit - 15] + "...[truncated]"
    return text


def norm_text(value: Any) -> str:
    return safe_str(value, 500000).lower()


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        fieldnames = keys
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: safe_str(row.get(k, "")) for k in fieldnames})


def read_text_limited(path: Path, max_chars: int = 500_000) -> tuple[str, str]:
    try:
        raw = path.read_bytes()
        status = "full"
        if len(raw) > max_chars * 4:
            raw = raw[: max_chars * 4]
            status = "truncated"
        return raw.decode("utf-8", errors="ignore")[:max_chars], status
    except Exception as exc:
        return "", f"read_failed:{exc}"


def method_sections(method_text: str) -> tuple[list[dict[str, Any]], dict[str, tuple[int, int]]]:
    lines = method_text.splitlines()
    heads: list[tuple[str, int, str]] = []
    for idx, line in enumerate(lines, start=1):
        m = re.match(r"^(S\d+)\.\s*(.*)$", line.strip())
        if m:
            heads.append((m.group(1), idx, m.group(2)))
    ranges: dict[str, tuple[int, int]] = {}
    rows: list[dict[str, Any]] = []
    for i, (section, start, title) in enumerate(heads):
        end = heads[i + 1][1] - 1 if i + 1 < len(heads) else len(lines)
        ranges[section] = (start, end)
        rows.append({"section": section, "title": title, "line_start": start, "line_end": end})
    return rows, ranges


def cp(
    checkpoint_id: str,
    module: str,
    section: str,
    title: str,
    requirement: str,
    result_type: str,
    terms: list[str],
    sheets: list[str] | None = None,
    required: bool = True,
    default_missing: str = "MISSING_CAN_RECOMPUTE",
    recompute_plan: str = "",
    no_retrain_reason: str = "",
) -> dict[str, Any]:
    return {
        "checkpoint_id": checkpoint_id,
        "module": module,
        "method_section": section,
        "checkpoint_title": title,
        "requirement": requirement,
        "result_type": result_type,
        "terms": terms,
        "expected_sheets": sheets or [],
        "required": required,
        "default_missing_status": default_missing,
        "recompute_plan": recompute_plan,
        "no_retrain_reason": no_retrain_reason,
    }


def checkpoint_definitions() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    rows.extend([
        cp("A01", "A. Standardized cases", "S2", "120-case lock",
           "120 de-identified standardized KOA cases must be locked and traceable.",
           "count/result", ["120", "standardized_cases", "standardized cases", "case lock", "case_id"],
           ["04_CORE_NUMBERS_LOCKED", "SUM_163_05_STANDARDIZED_CASE", "RAW_Case_Basic"]),
        cp("A02", "A. Standardized cases", "S2", "Candidate case manifest",
           "Candidate manifest should include source ID, knee side, visit, imaging label, symptoms/function, BMI, comorbidities and treatment variables.",
           "audit table", ["candidate", "manifest", "source id", "knee side", "visit", "case"],
           ["RAW_Case_Basic", "SUM_163_05_STANDARDIZED_CASE"]),
        cp("A03", "A. Standardized cases", "S2", "Inclusion/exclusion audit",
           "Inclusion and exclusion records must support target knee, core-variable completeness, side consistency and task value.",
           "audit table", ["inclusion", "exclusion", "target knee", "side consistency", "complete"],
           ["SUM_163_05_STANDARDIZED_CASE"], default_missing="MISSING_CAN_RECOMPUTE",
           recompute_plan="Recreate from locked case manifest and OAI-derived case fields if raw manifest exists."),
        cp("A04", "A. Standardized cases", "S2", "Coverage matrix",
           "Coverage must describe structure severity, pain/function, metabolic load, safety risk, goals and treatment scenarios.",
           "audit table", ["coverage matrix", "burden", "quadrant", "structure", "pain", "function", "BMI", "safety"],
           ["RAW_Case_Basic", "SUM_163_05_STANDARDIZED_CASE"]),
        cp("A05", "A. Standardized cases", "S2", "Duplicate and similarity checks",
           "Duplicate subject, duplicate knee and high-similarity case checks must be archived.",
           "audit table", ["duplicate", "similarity", "same subject", "same knee", "dedup"],
           ["SUM_163_05_STANDARDIZED_CASE"], default_missing="MISSING_CAN_RECOMPUTE",
           recompute_plan="Run duplicate/similarity audit from locked case table; do not alter original OAI files."),
        cp("A06", "A. Standardized cases", "S2", "Source-variable mapping",
           "Each key case field should map back to OAI/source variables.",
           "source map", ["source variable", "mapping", "oai", "source_file", "field"],
           ["RAW_Patient_Fields", "01_SOURCE_FILES"]),
        cp("A07", "A. Standardized cases", "S2", "Case template lock table",
           "Unified doctor-task template must be version frozen.",
           "locked table", ["template", "case template", "lock", "doctor task", "version"],
           ["RAW_Case_Basic"], default_missing="MISSING_CAN_RECOMPUTE"),
    ])
    rows.extend([
        cp("B01", "B. KOM-Profile", "S3", "Field schema and dictionary",
           "KOM-Profile schema-first dictionary must define names, types, units, source variables, missing codes and downstream use.",
           "schema/source map", ["KOM-Profile", "schema", "field", "dictionary", "source variable"],
           ["RAW_Patient_Fields", "SUM_164_06_KOM_PROFILE"]),
        cp("B02", "B. KOM-Profile", "S3", "Source-truth field evaluation",
           "Source-truth fields require exact/partial/wrong/missing/not-available evaluation.",
           "metric/raw table", ["exact", "partial", "wrong", "missing", "field extraction", "0.846"],
           ["SUM_164_06_KOM_PROFILE", "RAW_Figure_Value_Crosscheck"],
           default_missing="MISSING_REQUIRES_EXPERIMENT",
           no_retrain_reason="This is an extraction/annotation audit, not a model retraining task."),
        cp("B03", "B. KOM-Profile", "S3", "Rule-derived field audit",
           "Rule-derived fields require input, rule logic and output audit.",
           "audit table", ["rule-derived", "rule", "derived", "risk tag", "management need"],
           ["RAW_Patient_Fields"], default_missing="MISSING_CAN_RECOMPUTE"),
        cp("B04", "B. KOM-Profile", "S3", "Conflict audit",
           "Source-variable and case-text conflicts must be recorded and not silently overwritten.",
           "audit table", ["conflict", "source variable", "case text", "manual review"],
           ["RAW_Missingness"], default_missing="MISSING_CAN_RECOMPUTE"),
        cp("B05", "B. KOM-Profile", "S3", "Profile output records",
           "Structured profile outputs should exist for locked cases.",
           "raw/output records", ["profile.json", "generated_profiles", "patient profile", "Q1-", "Q2-"],
           ["RAW_Patient_Fields"]),
    ])
    rows.extend([
        cp("C01", "C. KOM-Rad / OAK-Net", "S4", "Image integrity and side-label audit",
           "X-ray image file integrity, de-identification, knee-side mapping, label completeness and quality audit must be retained.",
           "audit table", ["OAKNet", "image", "side", "label", "integrity", "quality", "KL"],
           ["SUM_165_07_KOM_RAD_OAKNET", "RISK_side_mapping"], default_missing="MISSING_CAN_RECOMPUTE"),
        cp("C02", "C. KOM-Rad / OAK-Net", "S4", "Preprocessing configuration and image index",
           "Orientation, ROI/crop, normalization, size standardization and side-alignment config/index must be saved.",
           "config/index", ["preprocess", "ROI", "crop", "normalization", "image index", "OAKNet"],
           ["SUM_165_07_KOM_RAD_OAKNET"], default_missing="MISSING_CAN_RECOMPUTE"),
        cp("C03", "C. KOM-Rad / OAK-Net", "S4", "Training split and hyperparameters",
           "Person-level split, seed, model architecture, optimizer, learning rate, scheduler, batch size, epoch and checkpoint selection must be saved.",
           "training records", ["training", "split", "seed", "optimizer", "learning", "checkpoint", "epoch"],
           ["oaknet_training_history_long", "SUM_165_07_KOM_RAD_OAKNET"]),
        cp("C04", "C. KOM-Rad / OAK-Net", "S4", "OAK-Net performance metrics",
           "QWK, balanced accuracy, macro-F1, MAE, ECE and selective accuracy must be available.",
           "metrics", ["QWK", "BACC", "macro_F1", "MAE", "ECE", "selective", "OAKNET_metrics"],
           ["OAKNET_metrics", "oaknet_metrics_summary", "source_A_series_Ablation_Trajec"]),
        cp("C05", "C. KOM-Rad / OAK-Net", "S4", "Image-level test predictions",
           "Test predictions are required to support confusion matrix and image-level diagnostics.",
           "raw predictions", ["image-level", "prediction", "test predictions", "pred_label", "true_label"],
           ["OAKNET_missing_image_level_pred"], default_missing="MISSING_REQUIRES_RETRAIN",
           no_retrain_reason="No --allow-retrain flag was provided; if checkpoint exists, inference can be attempted under --allow-recompute."),
        cp("C06", "C. KOM-Rad / OAK-Net", "S4", "Confusion matrix source",
           "Confusion matrix source data must be available, not only a rendered figure.",
           "source data", ["confusion", "matrix", "OAKNet", "pred", "true"],
           [], default_missing="MISSING_REQUIRES_RETRAIN"),
        cp("C07", "C. KOM-Rad / OAK-Net", "S4", "Calibration and selective accuracy source",
           "Calibration bins/curve and selective accuracy source data must be available.",
           "source data", ["calibration", "reliability", "selective", "coverage", "accuracy_after_abstention"],
           ["source_Figure2a_reliability_A1_", "source_Figure2a_reliability_A7_", "source_Figure2b_selective_predi"]),
        cp("C08", "C. KOM-Rad / OAK-Net", "S4", "Grad-CAM/saliency explanation audit",
           "Explanation images must identify model version, input image, target label, prediction and method.",
           "figure/audit", ["Grad-CAM", "CAM", "saliency", "explanation", "model version"],
           ["09_FIGURE_READINESS"], default_missing="MISSING_REQUIRES_RETRAIN"),
    ])
    rows.extend([
        cp("D01", "D. KOM-Risk", "S5", "Endpoint definitions and labels",
           "Endpoint A/B/C definitions, inclusion/exclusion and label scripts must be retained.",
           "endpoint audit", ["endpoint", "label", "KL structural progression", "TKR", "symptom"],
           ["SUM_166_08_KOM_RISK_LATEST", "RISK_status_json"]),
        cp("D02", "D. KOM-Risk", "S5", "Raw/encoded feature dictionary and lineage",
           "Raw feature list, encoded feature list, dictionary and feature-lineage table must exist.",
           "feature lineage", ["feature", "encoded", "dictionary", "lineage", "raw feature"],
           ["RISK_feature_lock", "SUM_147_07_KOMRisk_Features"]),
        cp("D03", "D. KOM-Risk", "S5", "READPRJ deduplication audit",
           "KXR READPRJ priority deduplication and side mapping audit must be available.",
           "audit table", ["READPRJ", "dedup", "side", "kxr", "SIDE"],
           ["RISK_kxr_dedup", "RISK_side_mapping", "RISK_status_json"]),
        cp("D04", "D. KOM-Risk", "S5", "Person-level split and leakage audit",
           "Subject/person-level split and leakage checks must prevent future/end-point proxy variables.",
           "audit table", ["person-level", "split", "leakage", "future", "proxy"],
           ["RISK_status_json"], default_missing="MISSING_CAN_RECOMPUTE"),
        cp("D05", "D. KOM-Risk", "S5", "Risk model metrics",
           "AUROC, AUPRC, Brier, calibration and DCA must be present for locked models.",
           "metrics", ["AUROC", "AUPRC", "Brier", "calibration", "DCA", "CatBoost"],
           ["RISK_metrics_latest", "RISK_calib_A", "RISK_DCA_A", "komrisk_mode_performance"]),
        cp("D06", "D. KOM-Risk", "S5", "Sample-level risk predictions",
           "Continuous risk probabilities and sample-level predictions must be retained.",
           "raw predictions", ["prediction", "probability", "RISK_pred_A", "RISK_pred_B", "RISK_pred_C"],
           ["RISK_pred_A", "RISK_pred_B", "RISK_pred_C"]),
        cp("D07", "D. KOM-Risk", "S5", "Threshold/risk-stratification file",
           "Preset threshold file is needed before reporting low/mid/high risk strata.",
           "threshold table", ["threshold", "risk strata", "risk stratification", "komrisk_threshold_metrics"],
           ["komrisk_threshold_metrics"], default_missing="MISSING_CAN_RECOMPUTE"),
    ])
    rows.extend([
        cp("E01", "E. SHAP / Feature lineage", "S5", "Locked-model SHAP values",
           "SHAP must be based on locked model and locked encoded matrix.",
           "explainability", ["SHAP", "mean_abs_shap", "locked model", "encoded matrix"],
           ["komrisk_shap_top", "OLD_053_Expert_干预_SHAP解释", "OLD_092_Expert_干预_SHAP特征"],
           default_missing="FOUND_NEEDS_SOURCE_CONFIRMATION"),
        cp("E02", "E. SHAP / Feature lineage", "S5", "SHAP feature-to-source mapping",
           "Each SHAP feature must map to model input, raw feature, original column, source file, visit, side and endpoint-overlap flag.",
           "feature lineage", ["SHAP", "raw feature", "source file", "visit", "side", "endpoint overlap", "lineage"],
           ["RISK_feature_lock", "SUM_147_07_KOMRisk_Features"]),
        cp("E03", "E. SHAP / Feature lineage", "S5", "Final SHAP figure validity",
           "Old SHAP figures with features outside the locked universe cannot be used as final figures.",
           "figure QC", ["final SHAP", "feature universe", "locked", "invalid", "old"],
           ["RISK_imp_A", "RISK_imp_B", "RISK_imp_C"], default_missing="MISSING_CAN_RECOMPUTE"),
    ])
    rows.extend([
        cp("F01", "F. KOM-KB evidence database", "S6", "PubMed/search lifecycle audit",
           "Search date, database, query, hits, dedup, screening, full-text assessment, included source count and evidence-unit count must be archived.",
           "search audit", ["PubMed", "search", "dedup", "screening", "evidence unit", "included"],
           ["SUM_167_09_KOM_KB"], default_missing="MISSING_REQUIRES_EXPERIMENT"),
        cp("F02", "F. KOM-KB evidence database", "S6", "Evidence unit database and schema",
           "Evidence unit schema must include ID, title, source, year, type, level, domain, population, intervention, outcome, safety, use, specialty and audit status.",
           "database/schema", ["evidence unit", "KOA-EU", "evidence_id", "evidence level", "specialty", "safety"],
           ["SUM_167_09_KOM_KB", "RAW_Evidence_Mapping", "RAW_RAG_GoldDetail"]),
        cp("F03", "F. KOM-KB evidence database", "S6", "L1-L7 evidence counts",
           "L1-L7 hierarchy counts may be reported only if the detailed export supports them.",
           "count table", ["L1", "L2", "L3", "L4", "L5", "L6", "L7", "guideline", "systematic"],
           ["SUM_167_09_KOM_KB"], default_missing="FOUND_NEEDS_SOURCE_CONFIRMATION"),
    ])
    rows.extend([
        cp("G01", "G. KOM-RAG / GraphRAG", "S7", "RAG query set",
           "Query set must include patient profile, treatment domain and prescribing intent.",
           "raw query table", ["query", "patient", "domain", "intent", "RAG_query_level"],
           ["RAW_RAG_QueryLevel", "RAG_query_level"]),
        cp("G02", "G. KOM-RAG / GraphRAG", "S7", "Graph nodes and edges",
           "GraphRAG must expose patient-feature, treatment-question, domain, guideline-anchor, evidence-unit, safety, agent and recommendation nodes/edges.",
           "graph source", ["graph", "node", "edge", "source", "target", "GraphRAG"],
           ["graph_rag_nodes_edges", "RAG_missing_items"], default_missing="MISSING_CAN_RECOMPUTE"),
        cp("G03", "G. KOM-RAG / GraphRAG", "S7", "Naive RAG baseline",
           "Naive baseline must use the same query set and evidence pool without graph/routing constraints.",
           "baseline metrics", ["naive", "baseline", "same query", "same evidence", "GraphRAG"],
           ["RAW_RAG_Benchmark", "SUM_168_10_KOM_RAG_COMPLETIO"]),
        cp("G04", "G. KOM-RAG / GraphRAG", "S7", "Gold relevance labels",
           "Gold relevance sets must be locked before retrieval comparison and not derived from GraphRAG results.",
           "gold labels", ["gold", "relevance", "label", "0", "1", "2", "RAG_gold_detail"],
           ["RAW_RAG_GoldDetail", "RAG_gold_detail", "RAW_RAG_QueryLevel"]),
        cp("G05", "G. KOM-RAG / GraphRAG", "S7", "Retrieval metrics",
           "Precision@10, Recall@K, Hit@10, MRR and nDCG@10 must be reported from locked retrieval evaluation.",
           "metrics", ["Precision@10", "Recall", "Hit@10", "MRR", "nDCG@10"],
           ["RAW_RAG_Benchmark", "RAG_metrics_summary", "SUM_168_10_KOM_RAG_COMPLETIO"]),
        cp("G06", "G. KOM-RAG / GraphRAG", "S7", "Generation faithfulness/citation audit",
           "Citation support, faithfulness, unsupported-claim rate and conflict handling require separate generation-level audit.",
           "generation audit", ["faithfulness", "citation_support", "unsupported claim", "conflict handling"],
           ["RAG_missing_items", "blocked_real_gaps"], default_missing="MISSING_REQUIRES_EXPERIMENT",
           no_retrain_reason="This requires a generation audit/annotation run, not model retraining."),
    ])
    rows.extend([
        cp("H01", "H. KOM-MDT / KOM-Treat", "S8", "R0-R8 agent prompt records",
           "Each agent run must save prompt version, model version, input hash, output hash, evidence unit IDs and safety-rule triggers.",
           "prompt/run log", ["R0", "R1", "R2", "R3", "R4", "R5", "R6", "R7", "R8", "prompt", "input hash"],
           ["SUM_169_11_KOM_MDT_RX_SAFE"], default_missing="FOUND_NEEDS_FILL"),
        cp("H02", "H. KOM-MDT / KOM-Treat", "S8", "Structured KOM-Rx outputs",
           "KOM-Rx must contain clinical summary, exercise, weight/nutrition, drugs, injection boundary, psychology, assistive/brace, follow-up, referral, safety and evidence links.",
           "output records", ["KOM-Rx", "prescription", "evidence link", "follow-up", "referral", "safety"],
           ["SUM_169_11_KOM_MDT_RX_SAFE", "RAW_Treat_AblationScores"]),
        cp("H03", "H. KOM-MDT / KOM-Treat", "S8", "No unsupported R8 additions",
           "R8 must not add recommendations unsupported by earlier evidence/specialist consensus.",
           "audit table", ["R8", "unsupported", "evidence", "arbitration", "delete"],
           ["RAW_Safety_Error_Log"], default_missing="MISSING_REQUIRES_EXPERIMENT"),
        cp("H04", "H. KOM-MDT / KOM-Treat", "S11", "Treatment ablation design",
           "Full KOM, without RAG, without MDT and Direct LLM arms must use same cases, output schema and scoring pipeline.",
           "experiment table", ["Full KOM", "without RAG", "without MDT", "Direct LLM", "ablation"],
           ["RAW_Treat_AblationScores", "SUM_174_16_KOM_TREAT_ABLATIO"]),
        cp("H05", "H. KOM-MDT / KOM-Treat", "S11", "Paired/FDR ablation statistics",
           "Paired comparisons and FDR correction must be retained.",
           "statistics", ["paired", "FDR", "q", "Wilcoxon", "ablation"],
           ["RAW_Treat_AblationScores", "SUM_145_05_KOMTreat_RAG_Abla"], default_missing="MISSING_CAN_RECOMPUTE"),
    ])
    rows.extend([
        cp("I01", "I. KOM-Safe", "S9", "Safety-rule library",
           "Rules must cover NSAIDs, GI, renal, CV, anticoagulation, injection, exercise load, falls, brace, surgery/referral, red flags and monitoring.",
           "rule table", ["NSAID", "renal", "GI", "CV", "anticoag", "injection", "fall", "red flag"],
           ["SUM_169_11_KOM_MDT_RX_SAFE", "SCORE_rule_raw"], default_missing="FOUND_NEEDS_FILL"),
        cp("I02", "I. KOM-Safe", "S9", "PASS/WARN/FAIL safety records",
           "Each rule should output PASS/WARN/FAIL plus trigger, text, patient factor, action, modified segment and severity.",
           "audit table", ["PASS", "WARN", "FAIL", "trigger", "severity", "modified", "safety"],
           ["RAW_Safety_Error_Log", "SCORE_error_log"]),
        cp("I03", "I. KOM-Safe", "S9", "Safety-critical repair/human review",
           "Safety-critical FAIL outputs must enter revision or human-review workflow.",
           "audit workflow", ["safety-critical", "FAIL", "human review", "revision", "resolved"],
           ["RAW_Safety_Error_Log"], default_missing="MISSING_REQUIRES_EXPERIMENT"),
    ])
    rows.extend([
        cp("J01", "J. KOM-Score", "S10", "Expert raw scores",
           "Blinded expert scores from orthopaedics, sports medicine and rehabilitation must be archived.",
           "raw scoring", ["expert", "blind", "ORTHO", "SPORT", "REHAB", "score"],
           ["RAW_Expert_Scores", "SCORE_expert_raw", "SUM_170_12_KOM_SCORE_EXPERT"]),
        cp("J02", "J. KOM-Score", "S10", "Expert reliability",
           "ICC(2,1) must be reported for continuous scores.",
           "statistics", ["ICC", "ICC_2_1", "overall_quality", "safety_ICC"],
           ["RAW_Expert_ICC", "SCORE_expert_ICC", "SUM_170_12_KOM_SCORE_EXPERT"]),
        cp("J03", "J. KOM-Score", "S10", "Rule score outputs",
           "Rule scoring should cover structural completeness, safety gates, individualization and executability.",
           "rule scoring", ["rule", "safety gate", "individualization", "executability"],
           ["SCORE_rule_raw", "SUM_171_13_KOM_SCORE_RULE"]),
        cp("J04", "J. KOM-Score", "S10", "Error taxonomy and kappa/agreement",
           "Safety-critical, clinically relevant and minor errors require classification and agreement/kappa where applicable.",
           "error scoring", ["safety-critical", "clinically relevant", "minor", "kappa", "agreement"],
           ["SCORE_error_log", "SUM_172_14_KOM_SCORE_ERROR"], default_missing="MISSING_CAN_RECOMPUTE"),
        cp("J05", "J. KOM-Score", "S10", "Training/arbitration materials",
           "Scorer training materials, scoring sheet, error examples, arbitration records and scoring scripts must be archived.",
           "training/audit", ["training", "arbitration", "scoring script", "error example"],
           ["RAW_Expert_Scores"], default_missing="MISSING_REQUIRES_EXPERIMENT"),
    ])
    rows.extend([
        cp("K01", "K. KOM-Sim human interaction", "S12", "Participant metadata",
           "Doctor specialty, clinical years, KOA years, AI tool experience and guideline familiarity must be recorded.",
           "participant table", ["doctor", "clinician", "specialty", "KOA", "years", "AI"],
           ["RAW_KOMSim_TaskLevel", "hci_raw_case_level", "SIM_doctor_level_summary"]),
        cp("K02", "K. KOM-Sim human interaction", "S12", "Task assignment and conditions",
           "Case order, task allocation and conditions Clinician alone, +KOM, +KOM-R must be retained.",
           "raw task table", ["condition", "Clinician alone", "KOM-R", "case_order", "stage"],
           ["RAW_KOMSim_TaskLevel", "hci_case_level"]),
        cp("K03", "K. KOM-Sim human interaction", "S12", "Doctor-task raw records",
           "Each doctor-task record must save prescription text, condition, case ID, doctor ID, time, components, view counts, copy/adoption behavior, workload, confidence, scores and errors.",
           "raw records", ["prescription", "edit", "view", "copy", "workload", "confidence", "case_id"],
           ["RAW_KOMSim_TaskLevel", "hci_raw_all_metrics", "hci_raw_events", "hci_raw_workload"]),
        cp("K04", "K. KOM-Sim human interaction", "S12", "Final record filter log",
           "final_record_filter_log must document duplicates, empty prescriptions, system-output rows, test rows, missing scores and final included records.",
           "filter log", ["final_record_filter_log", "duplicate", "empty", "included", "filter"],
           ["SIM_final_record_filter_log"], default_missing="MISSING_CAN_RECOMPUTE"),
        cp("K05", "K. KOM-Sim human interaction", "S12", "Primary and secondary comparisons",
           "Clinician+KOM vs alone and KOM-R vs KOM comparisons should be retained with paired/mixed model logic.",
           "statistics", ["Clinician + KOM", "Clinician alone", "KOM-R", "mixed", "paired"],
           ["SUM_173_15_KOM_SIM_CLINICIAN", "SIM_task_level_raw"]),
        cp("K06", "K. KOM-Sim human interaction", "S12", "Experience/LOWESS analysis",
           "KOA experience vs assisted benefit should have doctor-level points and LOWESS source.",
           "source data", ["LOWESS", "experience", "benefit", "KOA_experience_years"],
           ["SIM_experience_points", "SIM_LOWESS_benefit", "SIM_LOWESS_rationale"]),
    ])
    rows.extend([
        cp("L01", "L. Main figures and source data", "S13", "Main/figure source-data manifest",
           "Each main or supplementary figure must retain source CSV/table, script, QC and editable/bitmap exports.",
           "manifest", ["figure", "source data", "script", "QC", "svg", "png", "pdf"],
           ["09_FIGURE_READINESS", "OLD_008_Figure_Metadata_Inde", "SUM_149_09_Final_Figures_Ind"]),
        cp("L02", "L. Main figures and source data", "S13", "Figure denominators",
           "Each figure must declare denominators such as cases, doctors, doctor-task records, queries, evidence units or features.",
           "QC metadata", ["denominator", "n_cases", "n_doctors", "queries", "evidence units", "features"],
           ["RAW_Figure_Value_Crosscheck", "SUM_185_27_FIGURE_TABLE_PLAN"], default_missing="MISSING_CAN_RECOMPUTE"),
        cp("L03", "L. Main figures and source data", "S14", "Supplementary method figures S1-S13",
           "Method workflow figures S1-S13 are placeholders/recommended figures; data-bearing panels must have source data if used.",
           "figure plan", ["Figure S1", "Figure S2", "Figure S3", "Figure S4", "Figure S5", "Figure S13"],
           ["SUM_185_27_FIGURE_TABLE_PLAN"], required=False, default_missing="NOT_APPLICABLE"),
        cp("L04", "L. Main figures and source data", "S13", "No bitmap-only figure rule",
           "Figures may not exist only as non-editable bitmaps.",
           "QC rule", ["bitmap", "editable", "svg", "pdf", "source data"],
           ["09_FIGURE_READINESS", "ALL_FIGURES_MANIFEST"], default_missing="MISSING_CAN_RECOMPUTE"),
    ])
    return rows


def workbook_sources(primary: Path, secondary: Path | None) -> list[tuple[str, Path]]:
    sources = [("primary_master", primary)]
    if secondary and secondary.exists():
        sources.append(("figure_master", secondary))
    return sources


def scan_workbooks(sources: list[tuple[str, Path]]) -> tuple[list[dict[str, Any]], dict[str, str], list[dict[str, Any]], set[str]]:
    inventory: list[dict[str, Any]] = []
    blobs: dict[str, str] = {}
    result_rows: list[dict[str, Any]] = []
    referenced_paths: set[str] = set()
    path_re = re.compile(r"[A-Za-z]:\\[^\\/:*?\"<>|\r\n]+(?:\\[^\\/:*?\"<>|\r\n]+)+")

    for source_label, path in sources:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            preview: list[str] = []
            blob_parts = [source_label, ws.title]
            nonempty = 0
            header = ""
            max_blob_rows = 300
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = [safe_str(v, 700) for v in row]
                if any(values):
                    nonempty += 1
                    if not header:
                        header = " | ".join(values[:12])
                    if len(preview) < 3:
                        preview.append(" | ".join(values[:12]))
                    if len(" ".join(blob_parts)) < 850_000:
                        blob_parts.append(" | ".join(values[:60]))
                    for value in values:
                        for match in path_re.findall(value):
                            referenced_paths.add(match)
                if r_idx >= max_blob_rows:
                    if ws.max_row > max_blob_rows:
                        blob_parts.append(f"[sheet_scan_truncated_after_{max_blob_rows}_rows]")
                    break
            inventory.append({
                "workbook_label": source_label,
                "workbook_path": str(path),
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "nonempty_rows_scanned": nonempty,
                "header_preview": header,
                "preview": " || ".join(preview),
            })
            blobs[f"{source_label}::{ws.title}"] = "\n".join(blob_parts).lower()
        wb.close()

    summary_sheets = [
        "04_CORE_NUMBERS_LOCKED", "07_RESULTS_SENTENCE_LOCK", "RAW_Figure_Value_Crosscheck",
        "RAW_RAG_Benchmark", "RAG_metrics_summary", "RISK_metrics_latest", "OAKNET_metrics",
        "SUM_164_06_KOM_PROFILE", "SUM_165_07_KOM_RAD_OAKNET", "SUM_166_08_KOM_RISK_LATEST",
        "SUM_167_09_KOM_KB", "SUM_168_10_KOM_RAG_COMPLETIO", "SUM_169_11_KOM_MDT_RX_SAFE",
        "SUM_170_12_KOM_SCORE_EXPERT", "SUM_171_13_KOM_SCORE_RULE", "SUM_172_14_KOM_SCORE_ERROR",
        "SUM_173_15_KOM_SIM_CLINICIAN", "SUM_174_16_KOM_TREAT_ABLATIO", "SUM_176_18_MAIN_RESULTS_LOCK",
        "oaknet_metrics_summary", "oaknet_training_history_long", "komrisk_mode_performance",
        "komrisk_shap_top", "komrisk_threshold_metrics", "hci_case_rating_summary",
        "hci_condition_radar", "hci_workload_summary", "hci_final_survey_long",
        "graph_rag_nodes_edges", "methods_result_map", "blocked_real_gaps",
    ]
    for source_label, path in sources:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in summary_sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            max_rows = min(ws.max_row, 500)
            rows = list(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True))
            if not rows:
                continue
            header = [safe_str(v, 200) or f"col_{i+1}" for i, v in enumerate(rows[0])]
            for idx, row in enumerate(rows[1:], start=2):
                values = [safe_str(v, 1000) for v in row]
                if not any(values):
                    continue
                compact = {}
                for i, value in enumerate(values):
                    if value:
                        compact[header[i] if i < len(header) else f"col_{i+1}"] = value
                result_rows.append({
                    "workbook_label": source_label,
                    "source_sheet": sheet_name,
                    "source_row": idx,
                    "module_guess": guess_module(sheet_name + " " + json.dumps(compact, ensure_ascii=False)),
                    "result_name": infer_result_name(sheet_name, compact, idx),
                    "result_value": infer_result_value(compact),
                    "result_context": json.dumps(compact, ensure_ascii=False)[:2500],
                    "source_status": "existing_workbook_row",
                    "action": "fill_or_crosscheck_in_updated_master",
                })
        wb.close()
    return inventory, blobs, result_rows, referenced_paths


def guess_module(text: str) -> str:
    t = text.lower()
    if "profile" in t or "patient_fields" in t:
        return "B. KOM-Profile"
    if "oak" in t or "rad" in t or "convnext" in t or "densenet" in t:
        return "C. KOM-Rad / OAK-Net"
    if "risk" in t or "shap" in t or "dca" in t or "auroc" in t:
        return "D/E. KOM-Risk / SHAP"
    if "evidence" in t or "kb" in t:
        return "F. KOM-KB evidence database"
    if "rag" in t or "retrieval" in t or "graphrag" in t:
        return "G. KOM-RAG / GraphRAG"
    if "mdt" in t or "rx" in t or "treat" in t or "ablation" in t:
        return "H. KOM-MDT / KOM-Treat"
    if "safe" in t or "safety" in t:
        return "I. KOM-Safe"
    if "score" in t or "expert" in t or "icc" in t:
        return "J. KOM-Score"
    if "sim" in t or "hci" in t or "doctor" in t or "clinician" in t:
        return "K. KOM-Sim human interaction"
    if "figure" in t or "source_" in t:
        return "L. Main figures and source data"
    if "case" in t or "standardized" in t:
        return "A. Standardized cases"
    return "Unclassified"


def infer_result_name(sheet_name: str, compact: dict[str, str], idx: int) -> str:
    for key in ["item", "key", "metric", "metrics", "Outcome", "outcome", "section", "module", "checkpoint", "node", "condition", "arm"]:
        if compact.get(key):
            return f"{sheet_name}:{compact[key]}"
    return f"{sheet_name}:row_{idx}"


def infer_result_value(compact: dict[str, str]) -> str:
    for key in ["value", "AUROC", "AUPRC", "Brier", "qwk_mean", "mean_abs_shap", "threshold", "mean", "raw_auroc", "raw_qwk", "result"]:
        if compact.get(key):
            return compact[key]
    values = [v for v in compact.values() if v]
    return " | ".join(values[:5])


def should_skip_dir_name(name: str) -> bool:
    low = name.lower()
    if name in SKIP_DIRS or name.startswith("."):
        return True
    noisy_tokens = [
        "backup_", "_backup", "archive_old", "old_versions", "node_modules",
        "site-packages", "__pycache__", "runtime", "lib", "dist-info",
    ]
    return any(token in low for token in noisy_tokens)


def scan_files(roots: list[Path], referenced_paths: set[str]) -> tuple[list[dict[str, Any]], dict[str, str]]:
    rows: list[dict[str, Any]] = []
    blobs: dict[str, str] = {}
    seen: set[str] = set()

    def add_file(path: Path, origin: str) -> None:
        try:
            resolved_key = str(path.resolve()).lower()
        except Exception:
            resolved_key = str(path).lower()
        if resolved_key in seen or not path.exists() or not path.is_file():
            return
        seen.add(resolved_key)
        ext = path.suffix.lower()
        if ext not in TARGET_EXTS:
            return
        try:
            st = path.stat()
        except Exception:
            return
        text = ""
        read_status = "metadata_only"
        zip_hits: list[str] = []
        if ext in TEXT_EXTS and st.st_size <= 4 * 1024 * 1024:
            text, read_status = read_text_limited(path)
        elif ext == ".zip":
            try:
                with zipfile.ZipFile(path) as zf:
                    names = zf.namelist()
                    interesting = [n for n in names if any(k in n.lower() for k in [
                        "rag", "risk", "oak", "profile", "score", "sim", "hci", "safe",
                        "treat", "mdt", "prompt", "figure", "source", "metric", "shap",
                        "threshold", "case", "gold", "confusion", "calibration",
                    ])]
                    zip_hits = interesting[:350]
                    text = "\n".join(zip_hits)
                    read_status = f"zip_index_entries={len(names)};matched={len(interesting)}"
            except Exception as exc:
                read_status = f"zip_index_failed:{exc}"
        rel = ""
        for root in roots:
            try:
                rel = str(path.relative_to(root))
                break
            except Exception:
                continue
        blob = f"{path.name}\n{path}\n{rel}\n{text}".lower()
        score_terms = [
            "kom", "koa", "oak", "risk", "rag", "graphrag", "profile", "score", "safe",
            "sim", "hci", "mdt", "treat", "prompt", "agent", "figure", "source", "audit",
            "shap", "threshold", "case", "doctor", "expert", "clinician", "evidence",
            "calibration", "confusion", "grad", "cam", "prediction", "metric", "locked",
            "final",
        ]
        relevance = sum(1 for term in score_terms if term in blob)
        if relevance == 0 and origin != "workbook_referenced_path":
            return
        rows.append({
            "absolute_path": str(path),
            "relative_path": rel,
            "origin": origin,
            "file_name": path.name,
            "extension": ext,
            "size_bytes": st.st_size,
            "size_mb": round(st.st_size / 1024 / 1024, 4),
            "modified_time": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
            "sha256": sha256_file(path, max_size=250 * 1024 * 1024) if relevance > 0 or ext in {".xlsx", ".zip"} else "not_hashed_low_relevance",
            "module_guess": guess_module(blob),
            "relevance_score": relevance,
            "read_status": read_status,
            "zip_entry_hits": " | ".join(zip_hits[:25]),
            "notes": "",
        })
        blobs[str(path)] = blob

    recursive_scan = os.environ.get("FULL_RECURSIVE_SCAN", "").strip() == "1"
    shallow_subdirs = [
        "qc", "audit", "tables", "source_data", "ALL_FIGURES_ONE_FOLDER",
        "figures_no_title", "figures_composite_editable", "KOMRisk_Final_Locked_PostDedup_20260610",
        "KOMRisk_feature_lineage_audit_20260612_1546", "KOMRisk_feature_lineage_audit_20260612_1352",
        "KOMRisk_feature_universe_audit_20260612_1546", "KOM_Submission_Audit_Package_202606_FINAL",
        "KOM_Submission_Audit_Package_202606_FINAL_DELTA_FIXED", "KOM_项目最终总表与支持文件_LOCKED_20260610",
        "KOM_AllModules_Methods_Results_Report_20260614", "KOM_Manuscript_Figure_Record_Package_202606_PNG_REVIEW",
        "人机交互试验", "治疗智能体组", "评估智能体", "KOA患者数据",
    ]

    def shallow_candidates(root: Path) -> list[Path]:
        out: list[Path] = []
        try:
            out.extend([p for p in root.iterdir() if p.is_file()])
        except Exception:
            return out
        for rel in shallow_subdirs:
            sub = root / rel
            if not sub.exists() or not sub.is_dir():
                continue
            try:
                out.extend([p for p in sub.iterdir() if p.is_file()])
            except Exception:
                continue
            # Figure/source folders are useful and bounded enough for one extra level.
            if rel in {"source_data", "qc", "audit", "tables", "ALL_FIGURES_ONE_FOLDER", "figures_no_title"}:
                try:
                    for child in sub.iterdir():
                        if child.is_dir() and not should_skip_dir_name(child.name):
                            out.extend([p for p in child.iterdir() if p.is_file()])
                except Exception:
                    pass
        return out

    for root in roots:
        if not root or not root.exists():
            continue
        if not recursive_scan:
            for path in shallow_candidates(root):
                add_file(path, "shallow_scan")
            continue
        count = 0
        for current, dirs, files in os.walk(root, topdown=True):
            dirs[:] = [d for d in dirs if not should_skip_dir_name(d)]
            count += len(files)
            if count > 25000:
                rows.append({
                    "absolute_path": str(root),
                    "relative_path": "",
                    "origin": "scan_limit",
                    "file_name": "",
                    "extension": "",
                    "size_bytes": "",
                    "size_mb": "",
                    "modified_time": "",
                    "sha256": "",
                    "module_guess": "scan_limit",
                    "relevance_score": "",
                    "read_status": "scan_limit_after_25000_files",
                    "zip_entry_hits": "",
                    "notes": "Increase limit only if needed; current audit uses workbook-referenced paths and high-relevance files.",
                })
                break
            for name in files:
                add_file(Path(current) / name, "recursive_scan")

    for raw in referenced_paths:
        cleaned = raw.strip().strip('"').strip("'")
        try:
            add_file(Path(cleaned), "workbook_referenced_path")
        except Exception:
            continue

    return rows, blobs


def score_hits(terms: list[str], blobs: dict[str, str], limit: int = 8) -> list[dict[str, Any]]:
    scored = []
    low_terms = [t.lower() for t in terms if t]
    for key, blob in blobs.items():
        score = 0
        matched = []
        for term in low_terms:
            if term in blob:
                matched.append(term)
                score += max(1, min(5, len(term) // 4))
        if score:
            scored.append({"key": key, "score": score, "matched_terms": ";".join(matched[:12])})
    scored.sort(key=lambda x: (-x["score"], x["key"]))
    return scored[:limit]


def direct_sheet_hits(expected: list[str], workbook_blobs: dict[str, str]) -> list[dict[str, Any]]:
    out = []
    existing_lower = {k.lower(): k for k in workbook_blobs.keys()}
    for sheet in expected:
        for lower_key, original in existing_lower.items():
            if lower_key.endswith("::" + sheet.lower()) or sheet.lower() in lower_key:
                out.append({"key": original, "score": 50, "matched_terms": "expected_sheet"})
                break
    return out


def classify_checkpoint(
    checkpoint: dict[str, Any],
    wb_hits: list[dict[str, Any]],
    file_hits: list[dict[str, Any]],
    missing_markers: str,
) -> tuple[str, str, str]:
    has_direct_wb = any(h["score"] >= 50 for h in wb_hits)
    has_wb = bool(wb_hits)
    has_source = bool(file_hits)
    marker = missing_markers.lower()

    if not checkpoint["required"] and checkpoint["default_missing_status"] == "NOT_APPLICABLE":
        if has_wb or has_source:
            return "PASS", "Optional/planned item has supporting material.", "No action unless selected for final manuscript."
        return "NOT_APPLICABLE", "Method states placeholder/recommended rather than mandatory data-bearing result.", "Document if not used."

    if checkpoint["checkpoint_id"] in {"G06"}:
        if "faithfulness" in marker or "citation_support" in marker or "source table not recovered" in marker:
            return "MISSING_REQUIRES_EXPERIMENT", "Existing audit sheets explicitly mark generation-level grounding/citation audit as not recovered.", "Run separate generation audit; do not report as passed."

    if checkpoint["checkpoint_id"] in {"C05", "C06"}:
        if "oaknet_image_level_predictions" in marker or "image-level predictions" in marker:
            return checkpoint["default_missing_status"], "Existing missing audit marks OAKNet image-level predictions as not found.", "Recover predictions/checkpoint or retrain only with --allow-retrain."

    if checkpoint["checkpoint_id"] == "B02" and "source table not recovered" in marker:
        return "FOUND_NEEDS_SOURCE_CONFIRMATION", "Workbook contains summary metric, but figure workbook flags direct field-extraction source table as not recovered.", "Recover raw field-level evaluation table before manuscript result claim."

    if checkpoint["checkpoint_id"] == "H04" and "source table not recovered" in marker:
        if has_wb and has_source:
            return "FOUND_NEEDS_SOURCE_CONFIRMATION", "Ablation rows exist in one workbook, but figure workbook flags missing judge-summary source table.", "Confirm canonical ablation source before result claim."

    if checkpoint["checkpoint_id"] == "I01" and "dedicated safety challenge result table" in marker:
        if has_wb:
            return "FOUND_NEEDS_SOURCE_CONFIRMATION", "Safety module summary exists, but dedicated safety challenge table is marked unrecovered.", "Recover or run safety challenge audit."

    if has_direct_wb and has_source:
        return "PASS", "Expected workbook sheet exists and source/project evidence was found.", "No immediate action."
    if has_direct_wb and not has_source:
        return "FOUND_NEEDS_SOURCE_CONFIRMATION", "Expected workbook sheet exists but no independent source/locked package match was found.", "Find source file or downgrade manuscript claim."
    if has_wb and has_source:
        return "PASS", "Workbook and source evidence both matched checkpoint terms.", "No immediate action."
    if has_source and not has_wb:
        return "FOUND_NEEDS_FILL", "Source evidence exists but the current master workbook lacks an explicit checkpoint row.", "Fill into updated master workbook."
    if has_wb and not has_source:
        return "FOUND_NEEDS_SOURCE_CONFIRMATION", "Workbook contains result-like content but source/locked support was not found.", "Recover source evidence before final claim."

    default = checkpoint["default_missing_status"]
    if default == "FOUND_NEEDS_SOURCE_CONFIRMATION":
        return "FOUND_NEEDS_SOURCE_CONFIRMATION", "Only old or noncanonical summary appears likely; source is not confirmed.", "Recover locked source package."
    return default, "No matching workbook or source evidence found in this audit run.", checkpoint.get("recompute_plan") or "Create task according to missing status."


def build_match_table(
    checkpoints: list[dict[str, Any]],
    workbook_blobs: dict[str, str],
    file_blobs: dict[str, str],
    method_ranges: dict[str, tuple[int, int]],
) -> list[dict[str, Any]]:
    missing_markers = "\n".join(
        blob for key, blob in workbook_blobs.items()
        if any(name in key.lower() for name in ["missing", "blocked_real_gaps"])
    )
    rows = []
    for checkpoint in checkpoints:
        wb_hits = direct_sheet_hits(checkpoint["expected_sheets"], workbook_blobs)
        more_wb_hits = score_hits(checkpoint["terms"], workbook_blobs, limit=10)
        existing_keys = {h["key"] for h in wb_hits}
        wb_hits.extend([h for h in more_wb_hits if h["key"] not in existing_keys])
        file_hits = score_hits(checkpoint["terms"], file_blobs, limit=10)
        status, rationale, next_action = classify_checkpoint(checkpoint, wb_hits, file_hits, missing_markers)
        method_line_start, method_line_end = method_ranges.get(checkpoint["method_section"], ("", ""))
        rows.append({
            "checkpoint_id": checkpoint["checkpoint_id"],
            "module": checkpoint["module"],
            "method_section": checkpoint["method_section"],
            "method_line_start": method_line_start,
            "method_line_end": method_line_end,
            "checkpoint_title": checkpoint["checkpoint_title"],
            "requirement": checkpoint["requirement"],
            "result_type": checkpoint["result_type"],
            "required": checkpoint["required"],
            "status": status,
            "rationale": rationale,
            "workbook_hits": " || ".join([f"{h['key']} (score={h['score']}; {h['matched_terms']})" for h in wb_hits[:5]]),
            "source_file_hits": " || ".join([f"{h['key']} (score={h['score']}; {h['matched_terms']})" for h in file_hits[:5]]),
            "evidence_strength": evidence_strength(status, wb_hits, file_hits),
            "next_action": next_action,
            "recompute_plan": checkpoint.get("recompute_plan", ""),
            "no_retrain_reason": checkpoint.get("no_retrain_reason", ""),
        })
    return rows


def evidence_strength(status: str, wb_hits: list[dict[str, Any]], file_hits: list[dict[str, Any]]) -> str:
    if status == "PASS":
        return "workbook_plus_source"
    if status == "FOUND_NEEDS_FILL":
        return "source_only"
    if status == "FOUND_NEEDS_SOURCE_CONFIRMATION":
        return "workbook_or_old_summary_only"
    if status.startswith("MISSING"):
        return "not_found"
    return "not_applicable_or_optional"


def task_rows(match_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for row in match_rows:
        status = row["status"]
        if status in {"PASS", "NOT_APPLICABLE"}:
            continue
        if status == "FOUND_NEEDS_FILL":
            task_type = "fill_master_workbook"
            allowed = "allowed_now"
        elif status == "FOUND_NEEDS_SOURCE_CONFIRMATION":
            task_type = "source_confirmation"
            allowed = "manual_source_recovery_or_locked_package_confirmation"
        elif status == "MISSING_CAN_RECOMPUTE":
            task_type = "recompute"
            allowed = "requires_explicit_--allow-recompute"
        elif status == "MISSING_REQUIRES_RETRAIN":
            task_type = "retrain_or_inference_recovery"
            allowed = "blocked_without_--allow-retrain; try checkpoint/prediction recovery first"
        else:
            task_type = "experiment_or_human_audit"
            allowed = "requires_new_experiment_or_annotation_plan"
        rows.append({
            "checkpoint_id": row["checkpoint_id"],
            "module": row["module"],
            "status": status,
            "task_type": task_type,
            "allowed_action": allowed,
            "task": row["next_action"],
            "source_hint": row["source_file_hits"][:1500],
            "workbook_hint": row["workbook_hits"][:1500],
            "recompute_plan": row.get("recompute_plan", ""),
            "no_retrain_reason": row.get("no_retrain_reason", ""),
        })
    return rows


def module_status_rows(match_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in match_rows:
        groups[row["module"]].append(row)
    out = []
    for code, module_label in CORE_MODULES:
        module_rows = []
        for key, value in groups.items():
            if key.startswith(code + ".") or key.startswith(code):
                module_rows.extend(value)
        if not module_rows:
            continue
        counts = Counter(r["status"] for r in module_rows)
        worst = max(module_rows, key=lambda r: STATUS_ORDER.get(r["status"], 99))["status"]
        if any(r["status"] in {"MISSING_REQUIRES_RETRAIN", "MISSING_REQUIRES_EXPERIMENT"} and r["required"] for r in module_rows):
            verdict = "NOT_READY"
        elif any(r["status"] in {"FOUND_NEEDS_SOURCE_CONFIRMATION", "MISSING_CAN_RECOMPUTE", "FOUND_NEEDS_FILL"} and r["required"] for r in module_rows):
            verdict = "READY_WITH_WARNINGS"
        else:
            verdict = "READY_FOR_MANUSCRIPT_RESULTS"
        out.append({
            "module_code": code,
            "module": module_label,
            "checkpoint_count": len(module_rows),
            "pass": counts.get("PASS", 0),
            "not_applicable": counts.get("NOT_APPLICABLE", 0),
            "found_needs_fill": counts.get("FOUND_NEEDS_FILL", 0),
            "source_confirmation": counts.get("FOUND_NEEDS_SOURCE_CONFIRMATION", 0),
            "missing_can_recompute": counts.get("MISSING_CAN_RECOMPUTE", 0),
            "missing_requires_retrain": counts.get("MISSING_REQUIRES_RETRAIN", 0),
            "missing_requires_experiment": counts.get("MISSING_REQUIRES_EXPERIMENT", 0),
            "worst_status": worst,
            "module_verdict": verdict,
        })
    return out


def figure_status_rows(file_rows: list[dict[str, Any]], workbook_blobs: dict[str, str]) -> list[dict[str, Any]]:
    rows = []
    all_file_text = "\n".join([f"{r.get('absolute_path', '')} {r.get('file_name', '')}" for r in file_rows]).lower()
    all_wb_text = "\n".join(workbook_blobs.values())
    figure_items = [
        ("S1", "overall KOM workflow", ["workflow", "pipeline", "kom_graph_rag_pipeline", "overall"]),
        ("S2", "standardized case construction", ["case", "standardized", "coverage", "manifest"]),
        ("S3", "KOM-Profile extraction", ["profile", "fieldextraction", "patient_fields"]),
        ("S4", "OAK-Net model/training workflow", ["oaknet", "training", "workflow", "backbone"]),
        ("S5", "OAK-Net curves/confusion/calibration/CAM", ["oaknet", "training_curve", "reliability", "selective", "cam", "confusion"]),
        ("S6", "KOM-Risk endpoint/split/SHAP audit", ["risk", "endpoint", "shap", "dca", "calib"]),
        ("S7", "KOM-KB evidence unit construction", ["evidence", "kb", "koa-eu", "evidence unit"]),
        ("S8", "KOM-RAG gold/GraphRAG/baseline", ["graphrag", "rag", "gold", "graph_rag"]),
        ("S9", "KOM-MDT R0-R8 negotiation", ["mdt", "rx", "agent", "r0", "r8"]),
        ("S10", "KOM-Safe safety audit", ["safe", "safety", "warn", "fail"]),
        ("S11", "KOM-Score scoring workflow", ["score", "expert", "icc", "rule"]),
        ("S12", "KOM-Treat ablation matrix", ["treat", "ablation", "without rag", "without mdt"]),
        ("S13", "KOM-Sim clinician interaction", ["sim", "hci", "clinician", "doctor"]),
        ("MAIN", "main figure source-data bundle", ["all_figures_manifest", "source_data", "figure", "svg", "png", "pdf"]),
    ]
    for fig_id, title, terms in figure_items:
        file_hit = any(t.lower() in all_file_text for t in terms)
        wb_hit = any(t.lower() in all_wb_text for t in terms)
        if fig_id in {"S1", "S2", "S3", "S4", "S6", "S7", "S8", "S9", "S10", "S11", "S12", "S13"} and not (file_hit or wb_hit):
            status = "MISSING_CAN_RECOMPUTE"
            notes = "Method schematic can be generated from method/module records, but source file not found."
        elif fig_id == "S5" and not (file_hit or wb_hit):
            status = "MISSING_REQUIRES_RETRAIN"
            notes = "Data-bearing OAKNet panels require recovered image-level predictions/checkpoint or retraining."
        else:
            status = "PASS" if (file_hit or wb_hit) else "FOUND_NEEDS_SOURCE_CONFIRMATION"
            notes = "Found in workbook/file evidence." if status == "PASS" else "Figure claim exists but source bundle was not independently located."
        rows.append({
            "figure_id": fig_id,
            "figure_title": title,
            "status": status,
            "workbook_hit": wb_hit,
            "file_hit": file_hit,
            "required_if_used": True,
            "notes": notes,
        })
    return rows


def final_verdict(match_rows: list[dict[str, Any]]) -> str:
    required = [r for r in match_rows if str(r["required"]).lower() == "true" or r["required"] is True]
    if all(r["status"] in {"PASS", "NOT_APPLICABLE"} for r in required):
        return "READY_FOR_MANUSCRIPT_RESULTS"
    if any(r["status"] in {"MISSING_REQUIRES_RETRAIN", "MISSING_REQUIRES_EXPERIMENT"} for r in required):
        return "NOT_READY"
    return "READY_WITH_WARNINGS"


def copy_evidence_files(file_rows: list[dict[str, Any]], match_rows: list[dict[str, Any]], out_dir: Path) -> list[dict[str, Any]]:
    evidence_dir = out_dir / "organized_evidence"
    evidence_dir.mkdir(parents=True, exist_ok=True)
    wanted: dict[str, str] = {}
    for row in match_rows:
        if row["status"] not in {"PASS", "FOUND_NEEDS_FILL", "FOUND_NEEDS_SOURCE_CONFIRMATION"}:
            continue
        for part in row["source_file_hits"].split(" || "):
            m = re.match(r"(.+?) \(score=", part)
            if m:
                wanted[m.group(1)] = row["module"]
    copied = []
    for file_row in sorted(file_rows, key=lambda r: -int(r.get("relevance_score") or 0)):
        p = Path(file_row["absolute_path"])
        if str(p) not in wanted and len(copied) > 50:
            continue
        module = wanted.get(str(p), file_row.get("module_guess", "misc"))
        if len(copied) >= 80:
            break
        if not p.exists() or not p.is_file():
            continue
        size = p.stat().st_size
        if size > 20 * 1024 * 1024:
            copied.append({
                "source_path": str(p),
                "copied_path": "",
                "module": module,
                "copy_status": "not_copied_large_file_indexed_only",
                "size_bytes": size,
            })
            continue
        safe_module = re.sub(r"[^A-Za-z0-9._-]+", "_", module)[:60]
        dest_dir = evidence_dir / safe_module
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / (hashlib.md5(str(p).encode("utf-8")).hexdigest()[:8] + "_" + p.name)
        try:
            if not dest.exists():
                shutil.copy2(p, dest)
            copied.append({
                "source_path": str(p),
                "copied_path": str(dest),
                "module": module,
                "copy_status": "copied",
                "size_bytes": size,
            })
        except Exception as exc:
            copied.append({
                "source_path": str(p),
                "copied_path": str(dest),
                "module": module,
                "copy_status": f"copy_failed:{exc}",
                "size_bytes": size,
            })
    write_csv(out_dir / "organized_evidence_manifest.csv", copied)
    return copied


def workbook_table(ws, rows: list[dict[str, Any]], header: list[str]) -> None:
    ws.append(header)
    for row in rows:
        ws.append([row.get(h, "") for h in header])
    style_sheet(ws, len(rows) + 1, len(header))


def style_sheet(ws, max_row: int, max_col: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    if max_row >= 1 and max_col >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 12
        for cell in ws[col_letter][: min(max_row, 80)]:
            max_len = max(max_len, min(70, len(safe_str(cell.value, 500))))
        ws.column_dimensions[col_letter].width = max(12, min(55, max_len + 2))


def write_updated_workbook(
    primary_master: Path,
    output_xlsx: Path,
    readme_rows: list[dict[str, Any]],
    match_rows: list[dict[str, Any]],
    found_rows: list[dict[str, Any]],
    missing_rows: list[dict[str, Any]],
    task_rows_: list[dict[str, Any]],
    file_manifest_rows: list[dict[str, Any]],
    module_rows: list[dict[str, Any]],
    figure_rows: list[dict[str, Any]],
    update_log: list[dict[str, Any]],
) -> None:
    shutil.copy2(primary_master, output_xlsx)
    wb = load_workbook(output_xlsx)
    for sheet in RESULT_SHEET_NAMES:
        if sheet in wb.sheetnames:
            del wb[sheet]
    sheet_order = [
        ("README_UPDATE", readme_rows),
        ("RESULT_CHECKPOINTS", match_rows),
        ("FOUND_RESULTS", found_rows),
        ("MISSING_RESULTS", missing_rows),
        ("RECOMPUTE_TASKS", task_rows_),
        ("SOURCE_FILE_MANIFEST", file_manifest_rows[:15000]),
        ("MODULE_STATUS", module_rows),
        ("FIGURE_DATA_STATUS", figure_rows),
        ("UPDATE_LOG", update_log),
    ]
    for idx, (name, rows) in enumerate(sheet_order):
        ws = wb.create_sheet(name, idx)
        header = []
        for row in rows:
            for key in row:
                if key not in header:
                    header.append(key)
        if not header:
            header = ["message"]
            rows = [{"message": "no rows"}]
        workbook_table(ws, rows, header)
    wb.save(output_xlsx)
    wb.close()


def make_reports(
    out_dir: Path,
    verdict: str,
    input_rows: list[dict[str, Any]],
    module_rows: list[dict[str, Any]],
    match_rows: list[dict[str, Any]],
    task_rows_: list[dict[str, Any]],
    copied_rows: list[dict[str, Any]],
) -> tuple[dict[str, Any], str, str]:
    counts = Counter(r["status"] for r in match_rows)
    qc = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "overall_verdict": verdict,
        "status_counts": dict(counts),
        "module_status": module_rows,
        "required_checkpoint_count": sum(1 for r in match_rows if r["required"]),
        "missing_required_retrain_or_experiment": [
            r for r in match_rows
            if r["required"] and r["status"] in {"MISSING_REQUIRES_RETRAIN", "MISSING_REQUIRES_EXPERIMENT"}
        ],
        "source_confirmation_required": [r for r in match_rows if r["status"] == "FOUND_NEEDS_SOURCE_CONFIRMATION"],
        "input_manifest": input_rows,
        "organized_evidence_copied": len([r for r in copied_rows if r["copy_status"] == "copied"]),
        "organized_evidence_indexed_large": len([r for r in copied_rows if "large" in r["copy_status"]]),
        "policy": {
            "oai_raw_data_modified": False,
            "raw_files_deleted": False,
            "fabricated_results": False,
            "automatic_retrain": False,
            "automatic_experiment": False,
        },
    }
    qc_json = out_dir / "09_QC_result_completeness_report.json"
    qc_json.write_text(json.dumps(qc, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# KOM Result Completeness QC Report",
        "",
        f"- Generated at: {qc['generated_at']}",
        f"- Overall verdict: **{verdict}**",
        f"- Required checkpoints: {qc['required_checkpoint_count']}",
        f"- Status counts: {dict(counts)}",
        "",
        "## Module Status",
        "",
        "| Module | Verdict | PASS | Needs fill | Source confirmation | Recompute | Retrain | Experiment |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for row in module_rows:
        lines.append(
            f"| {row['module']} | {row['module_verdict']} | {row['pass']} | "
            f"{row['found_needs_fill']} | {row['source_confirmation']} | "
            f"{row['missing_can_recompute']} | {row['missing_requires_retrain']} | "
            f"{row['missing_requires_experiment']} |"
        )
    lines.extend([
        "",
        "## Hard Stops",
        "",
        "- No OAI raw data were modified.",
        "- No original files were deleted.",
        "- No missing result was promoted to PASS without workbook plus source evidence.",
        "- No retraining or new human experiment was run.",
        "",
        "## Highest Priority Remaining Items",
        "",
    ])
    for row in task_rows_[:30]:
        lines.append(f"- {row['checkpoint_id']} [{row['status']}]: {row['task']}")
    report_md = out_dir / "08_QC_result_completeness_report.md"
    report_md.write_text("\n".join(lines), encoding="utf-8")

    next_lines = [
        "# Project Next Actions",
        "",
        "This file lists only actions that remain after the current no-retrain/no-new-experiment audit.",
        "",
    ]
    for row in task_rows_:
        next_lines.append(f"## {row['checkpoint_id']} - {row['module']}")
        next_lines.append(f"- Status: {row['status']}")
        next_lines.append(f"- Action gate: {row['allowed_action']}")
        next_lines.append(f"- Task: {row['task']}")
        if row.get("recompute_plan"):
            next_lines.append(f"- Recompute plan: {row['recompute_plan']}")
        if row.get("no_retrain_reason"):
            next_lines.append(f"- No-retrain note: {row['no_retrain_reason']}")
        next_lines.append("")
    next_md = out_dir / "10_codex_next_actions.md"
    next_md.write_text("\n".join(next_lines), encoding="utf-8")
    return qc, str(report_md), str(qc_json)


def verify_outputs(out_dir: Path, output_xlsx: Path) -> dict[str, Any]:
    required_files = [
        "00_input_manifest.csv",
        "01_method_required_checkpoints.csv",
        "02_master_workbook_inventory.csv",
        "03_checkpoint_result_match_table.csv",
        "04_found_results_to_fill.csv",
        "05_missing_results_to_complete.csv",
        "06_recompute_or_retrain_task_list.csv",
        "07_updated_master_workbook.xlsx",
        "08_QC_result_completeness_report.md",
        "09_QC_result_completeness_report.json",
        "10_codex_next_actions.md",
    ]
    files = []
    for name in required_files:
        p = out_dir / name
        files.append({"file": name, "exists": p.exists(), "size_bytes": p.stat().st_size if p.exists() else 0})
    wb = load_workbook(output_xlsx, read_only=True, data_only=True)
    sheet_status = {name: (name in wb.sheetnames) for name in RESULT_SHEET_NAMES}
    row_counts = {name: wb[name].max_row for name in RESULT_SHEET_NAMES if name in wb.sheetnames}
    wb.close()
    return {
        "all_required_files_exist": all(f["exists"] and f["size_bytes"] > 0 for f in files),
        "files": files,
        "all_required_sheets_exist": all(sheet_status.values()),
        "sheet_status": sheet_status,
        "new_sheet_row_counts": row_counts,
    }


def main() -> None:
    method_txt = getenv_path("METHOD_TXT")
    primary_master = getenv_path("MASTER_XLSX")
    figure_master = getenv_path("FIGURE_XLSX", required=False)
    project_root = getenv_path("PROJECT_ROOT")
    final_dir = getenv_path("FINAL_DIR")
    workspace_root = getenv_path("WORKSPACE_ROOT", required=False)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_dir = final_dir / f"KOM_Result_Completeness_Audit_{timestamp}"
    suffix = 1
    while out_dir.exists():
        out_dir = final_dir / f"KOM_Result_Completeness_Audit_{timestamp}_{suffix:02d}"
        suffix += 1
    out_dir.mkdir(parents=True, exist_ok=True)

    method_text = method_txt.read_text(encoding="utf-8")
    method_rows, ranges = method_sections(method_text)
    checkpoints = checkpoint_definitions()

    sources = workbook_sources(primary_master, figure_master)
    inventory_rows, workbook_blobs, found_rows, referenced_paths = scan_workbooks(sources)

    roots = [project_root]
    if workspace_root and workspace_root.exists():
        roots.append(workspace_root)
    if figure_master:
        roots.append(figure_master.parent.parent)
    roots = list(dict.fromkeys([r for r in roots if r and r.exists()]))
    file_rows, file_blobs = scan_files(roots, referenced_paths)

    input_rows = []
    for label, path in sources:
        input_rows.append({
            "input_type": label,
            "path": str(path),
            "exists": path.exists(),
            "size_bytes": path.stat().st_size if path.exists() else "",
            "modified_time": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds") if path.exists() else "",
            "sha256": sha256_file(path),
        })
    input_rows.append({
        "input_type": "method_text",
        "path": str(method_txt),
        "exists": method_txt.exists(),
        "size_bytes": method_txt.stat().st_size,
        "modified_time": datetime.fromtimestamp(method_txt.stat().st_mtime).isoformat(timespec="seconds"),
        "sha256": sha256_file(method_txt),
    })
    for idx, root in enumerate(roots, start=1):
        input_rows.append({
            "input_type": f"scan_root_{idx}",
            "path": str(root),
            "exists": root.exists(),
            "size_bytes": "",
            "modified_time": "",
            "sha256": "",
        })

    # Promote method section rows into checkpoint records for audit traceability.
    checkpoint_rows = []
    for c in checkpoints:
        start, end = ranges.get(c["method_section"], ("", ""))
        checkpoint_rows.append({
            "checkpoint_id": c["checkpoint_id"],
            "module": c["module"],
            "method_section": c["method_section"],
            "method_line_start": start,
            "method_line_end": end,
            "checkpoint_title": c["checkpoint_title"],
            "requirement": c["requirement"],
            "result_type": c["result_type"],
            "required": c["required"],
            "expected_sheets": "; ".join(c["expected_sheets"]),
            "terms": "; ".join(c["terms"]),
        })

    match_rows = build_match_table(checkpoints, workbook_blobs, file_blobs, ranges)
    missing_rows = [
        row for row in match_rows
        if row["status"] in {
            "FOUND_NEEDS_SOURCE_CONFIRMATION",
            "MISSING_CAN_RECOMPUTE",
            "MISSING_REQUIRES_RETRAIN",
            "MISSING_REQUIRES_EXPERIMENT",
        }
    ]
    tasks = task_rows(match_rows)
    modules = module_status_rows(match_rows)
    figures = figure_status_rows(file_rows, workbook_blobs)
    verdict = final_verdict(match_rows)

    # Add checkpoint link to found rows by fuzzy module/status when possible.
    for row in found_rows:
        row["checkpoint_hint"] = ""
        module = row.get("module_guess", "")
        candidates = [m for m in match_rows if module and module.split(".")[0] in m["module"]]
        if candidates:
            row["checkpoint_hint"] = "; ".join(c["checkpoint_id"] for c in candidates[:5])

    write_csv(out_dir / "00_input_manifest.csv", input_rows)
    write_csv(out_dir / "01_method_required_checkpoints.csv", checkpoint_rows)
    write_csv(out_dir / "02_master_workbook_inventory.csv", inventory_rows)
    write_csv(out_dir / "03_checkpoint_result_match_table.csv", match_rows)
    write_csv(out_dir / "04_found_results_to_fill.csv", found_rows)
    write_csv(out_dir / "05_missing_results_to_complete.csv", missing_rows)
    write_csv(out_dir / "06_recompute_or_retrain_task_list.csv", tasks)
    write_csv(out_dir / "SOURCE_FILE_MANIFEST_full.csv", file_rows)
    write_csv(out_dir / "METHOD_SECTIONS.csv", method_rows)

    copied_rows = copy_evidence_files(file_rows, match_rows, out_dir)

    readme_rows = [
        {"field": "generated_at", "value": datetime.now().isoformat(timespec="seconds"), "notes": ""},
        {"field": "overall_verdict", "value": verdict, "notes": "Computed from required checkpoint statuses."},
        {"field": "primary_master", "value": str(primary_master), "notes": "Original workbook was copied; original file not modified."},
        {"field": "figure_master", "value": str(figure_master) if figure_master else "", "notes": "Second workbook included as evidence source."},
        {"field": "method_text", "value": str(method_txt), "notes": ""},
        {"field": "policy_no_oai_raw_modification", "value": "TRUE", "notes": "This audit only reads/copies evidence and writes a new output directory."},
        {"field": "policy_no_retrain", "value": "TRUE", "notes": "No --allow-retrain flag was provided."},
        {"field": "policy_no_new_experiment", "value": "TRUE", "notes": "Missing experiment items are task-listed only."},
    ]
    update_log = [
        {"step": "read_method", "status": "completed", "detail": f"{len(method_rows)} method sections parsed"},
        {"step": "read_workbooks", "status": "completed", "detail": f"{len(sources)} workbooks, {len(inventory_rows)} sheets inventoried"},
        {"step": "scan_files", "status": "completed", "detail": f"{len(file_rows)} source files indexed"},
        {"step": "match_checkpoints", "status": "completed", "detail": f"{len(match_rows)} checkpoints evaluated"},
        {"step": "write_outputs", "status": "completed", "detail": str(out_dir)},
    ]
    output_xlsx = out_dir / "07_updated_master_workbook.xlsx"
    write_updated_workbook(
        primary_master,
        output_xlsx,
        readme_rows,
        match_rows,
        found_rows,
        missing_rows,
        tasks,
        file_rows,
        modules,
        figures,
        update_log,
    )
    qc, report_md, qc_json = make_reports(out_dir, verdict, input_rows, modules, match_rows, tasks, copied_rows)
    verification = verify_outputs(out_dir, output_xlsx)
    (out_dir / "verification_summary.json").write_text(json.dumps(verification, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output_dir": str(out_dir),
        "updated_workbook": str(output_xlsx),
        "overall_verdict": verdict,
        "status_counts": dict(Counter(r["status"] for r in match_rows)),
        "required_files_ok": verification["all_required_files_exist"],
        "required_sheets_ok": verification["all_required_sheets_exist"],
        "qc_report_md": report_md,
        "qc_report_json": qc_json,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
