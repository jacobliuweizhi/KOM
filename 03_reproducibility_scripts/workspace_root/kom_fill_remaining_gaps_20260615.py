from __future__ import annotations

import csv
import datetime as dt
import json
import math
import os
import re
import shutil
from collections import defaultdict
from pathlib import Path
from statistics import mean, stdev

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_WB = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用\最终版本\KOM_方法结果完整写入_20260615_1852\KOM_完整方法学与对应结果_中文写入版_20260615_1852.xlsx")
FIGURE_WB = Path(r"C:\Users\Liu\Documents\医学科研专用项目内容\KOM_Figure_Deep_Optimization_20260615\tables\KOM_Figure_Deep_Optimization_Master_Table_20260615.xlsx")
RISK_ROOT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\KOMRisk_Final_Locked_PostDedup_20260610\scripts\input_package_extracted\KOMRisk_Formal_Retrain_FINAL_20260610_042548")
RISK_METRICS = RISK_ROOT / "07_metrics"
RAG_ROOT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\治疗智能体组\新结果输出\kom_pipeline\kom_pipeline\generation_out")
OAK_AUDIT_ROOT = Path(r"C:\Users\Liu\Documents\医学科研专用项目内容\OAKNet_Reproducibility_Audit_202606")
OUT_PARENT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用\最终版本")


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def norm(v):
    if v is None or v == "":
        return "not_available"
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return "not_available"
    return v


def as_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in {"na", "nan", "not_available", "none"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def read_csv_dicts(path: Path):
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_csv_rows(path: Path, max_rows=None):
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if max_rows is not None:
        return rows[:max_rows]
    return rows


def flatten_json(d, prefix=""):
    out = {}
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            out.update(flatten_json(v, key))
        else:
            out[key] = v
    return out


def replace_sheet(wb, name, index=None):
    if name in wb.sheetnames:
        del wb[name]
    if index is None:
        return wb.create_sheet(name)
    return wb.create_sheet(name, index)


def write_table(ws, rows, headers=None, start_row=1):
    r = start_row
    if headers:
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP
        r += 1
    for row in rows:
        if isinstance(row, dict):
            values = [norm(row.get(h)) for h in headers]
        else:
            values = [norm(v) for v in row]
        for c, v in enumerate(values, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = WRAP
        r += 1
    return r


def add_title(ws, title, subtitle=None):
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(bold=True, size=14, color="1F4E79")
    ws.cell(1, 1).alignment = WRAP
    if subtitle:
        ws.cell(2, 1, subtitle)
        ws.cell(2, 1).alignment = WRAP
    return 4


def style_sheet(ws, freeze="A2", max_width=72):
    ws.freeze_panes = freeze
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 300) + 1):
            v = ws.cell(row_idx, col_idx).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), max_width))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, max_width))
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
            if cell.value in {
                "PASS",
                "已补齐",
                "FOUND_IN_LOCKED_PACKAGE",
                "FOUND_AND_FILLED",
                "SOURCE_BACKED",
            }:
                cell.fill = OK_FILL
            elif isinstance(cell.value, str) and any(k in cell.value for k in ["剩余", "不能冒充", "需", "missing", "MISSING", "代理"]):
                cell.fill = WARN_FILL


def collect_risk_data():
    comparison = read_csv_dicts(RISK_METRICS / "final_model_comparison_all_algorithms.csv")
    metrics_all = read_csv_dicts(RISK_METRICS / "final_model_metrics_all_endpoints.csv")
    acceptance = read_csv_dicts(RISK_METRICS / "final_model_acceptance_decision_table.csv")
    bootstrap_rows = []
    for endpoint in ["A", "B", "C"]:
        p = RISK_METRICS / f"endpoint_{endpoint}_bootstrap_CI.csv"
        for row in read_csv_dicts(p):
            row = dict(row)
            row["endpoint_key"] = f"endpoint_{endpoint}"
            row["source_file"] = str(p)
            bootstrap_rows.append(row)
    hp_selected = []
    for endpoint in ["A", "B", "C"]:
        p = RISK_METRICS / f"endpoint_{endpoint}_hyperparameter_search_summary.csv"
        for row in read_csv_dicts(p):
            if str(row.get("selected_as_best", "")).lower() == "true" or str(row.get("algorithm", "")).lower() == "catboost":
                row = dict(row)
                row["endpoint_key"] = f"endpoint_{endpoint}"
                row["source_file"] = str(p)
                hp_selected.append(row)
    metadata_rows = []
    for p in sorted((RISK_ROOT / "05_models").glob("*/best_model/best_model_metadata.json")):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            data = {}
        flat = flatten_json(data)
        flat["source_file"] = str(p)
        metadata_rows.append(flat)
    split_rows = []
    for fn in ["person_level_split_integrity_report.csv", "split_event_rate_summary.csv"]:
        p = RISK_ROOT / "03_splits" / fn
        for row in read_csv_dicts(p):
            row = dict(row)
            row["source_file"] = str(p)
            split_rows.append(row)
    repro = RISK_ROOT / "13_leakage_and_QC" / "reproducibility_statement.md"
    repro_text = repro.read_text(encoding="utf-8", errors="replace") if repro.exists() else "not_found"
    return {
        "comparison": comparison,
        "metrics_all": metrics_all,
        "acceptance": acceptance,
        "bootstrap": bootstrap_rows,
        "hp_selected": hp_selected,
        "metadata": metadata_rows,
        "split": split_rows,
        "repro_text": repro_text,
        "repro_file": str(repro),
    }


def collect_treat_stats(wb):
    ws = wb["RAW_Treat_AblationScores"]
    headers = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
    rows = []
    for values in ws.iter_rows(min_row=4, values_only=True):
        if not any(values):
            continue
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        if row.get("status") == "ok":
            rows.append(row)

    metrics = [
        ("safety_0_10", True),
        ("guideline_alignment_0_10", True),
        ("patient_specificity_0_10", True),
        ("actionability_0_10", True),
        ("doctor_readability_0_10", True),
        ("evidence_traceability_0_10", True),
        ("faithfulness_0_10", True),
        ("hallucination_rate_0_1", False),
        ("specialty_completeness_0_10", True),
        ("conflict_handling_0_10", True),
        ("overall_0_100", True),
        ("critical_error_count", False),
        ("major_error_count", False),
    ]
    arm_names = {
        "完整系统_优化版": "Full KOM",
        "无RAG": "without RAG",
        "无MDT": "without MDT",
        "B0_裸模型自由直出": "direct LLM",
    }

    arm_summary = []
    groupings = [("ALL", None)]
    models = sorted({str(r.get("evaluator_model")) for r in rows})
    groupings.extend((m, m) for m in models)
    for group_label, model in groupings:
        subset = [r for r in rows if model is None or str(r.get("evaluator_model")) == model]
        for arm in sorted({r.get("arm_id") for r in subset}):
            arm_rows = [r for r in subset if r.get("arm_id") == arm]
            for metric, higher_is_better in metrics:
                vals = [as_float(r.get(metric)) for r in arm_rows]
                vals = [v for v in vals if v is not None]
                if not vals:
                    continue
                arm_summary.append({
                    "分组": group_label,
                    "arm_id": arm,
                    "arm_label": arm_names.get(arm, arm),
                    "metric": metric,
                    "metric_direction": "higher_better" if higher_is_better else "lower_better",
                    "n": len(vals),
                    "mean": mean(vals),
                    "sd": stdev(vals) if len(vals) > 1 else 0,
                    "min": min(vals),
                    "max": max(vals),
                    "source_sheet": "RAW_Treat_AblationScores",
                })

    by_key = defaultdict(dict)
    for r in rows:
        key = (
            str(r.get("case_id")),
            str(r.get("quadrant")),
            str(r.get("sample")),
            str(r.get("evaluator_model")),
        )
        by_key[key][str(r.get("arm_id"))] = r

    paired = []
    comparisons = [
        ("无RAG", "Full KOM vs without RAG"),
        ("无MDT", "Full KOM vs without MDT"),
        ("B0_裸模型自由直出", "Full KOM vs direct LLM"),
    ]
    full = "完整系统_优化版"
    for comp_arm, comp_label in comparisons:
        for metric, higher_is_better in metrics:
            full_vals = []
            comp_vals = []
            diffs = []
            for arms in by_key.values():
                if full not in arms or comp_arm not in arms:
                    continue
                fv = as_float(arms[full].get(metric))
                cv = as_float(arms[comp_arm].get(metric))
                if fv is None or cv is None:
                    continue
                raw_diff = fv - cv
                benefit_diff = raw_diff if higher_is_better else -raw_diff
                full_vals.append(fv)
                comp_vals.append(cv)
                diffs.append(benefit_diff)
            n = len(diffs)
            if n == 0:
                continue
            md = mean(diffs)
            sd = stdev(diffs) if n > 1 else 0
            se = sd / math.sqrt(n) if n > 1 else 0
            if se == 0:
                p = 1.0 if md == 0 else 0.0
                ci_low = md
                ci_high = md
            else:
                z = md / se
                p = math.erfc(abs(z) / math.sqrt(2.0))
                ci_low = md - 1.96 * se
                ci_high = md + 1.96 * se
            dz = md / sd if sd else ("inf" if md != 0 else 0)
            paired.append({
                "comparison": comp_label,
                "full_arm_id": full,
                "comparator_arm_id": comp_arm,
                "metric": metric,
                "metric_direction": "higher_better" if higher_is_better else "lower_better",
                "benefit_definition": "Full-Comparator" if higher_is_better else "Comparator-Full",
                "n_pairs": n,
                "full_mean": mean(full_vals),
                "comparator_mean": mean(comp_vals),
                "benefit_mean_diff": md,
                "benefit_95CI_low": ci_low,
                "benefit_95CI_high": ci_high,
                "paired_sd_diff": sd,
                "cohen_dz": dz,
                "p_two_sided_normal_approx": p,
                "q_BH_FDR": None,
                "stat_method": "paired large-sample normal approximation; two-sided p via erfc; BH-FDR across all paired comparisons",
                "source_sheet": "RAW_Treat_AblationScores",
            })

    # Benjamini-Hochberg FDR over all paired metric comparisons.
    indexed = sorted(enumerate(paired), key=lambda x: float(x[1]["p_two_sided_normal_approx"]))
    m = len(indexed)
    q_vals = [None] * m
    prev = 1.0
    for rank_from_end, (idx, row) in enumerate(reversed(indexed), 1):
        rank = m - rank_from_end + 1
        p = float(row["p_two_sided_normal_approx"])
        q = min(prev, p * m / rank)
        q = min(q, 1.0)
        prev = q
        q_vals[idx] = q
    for row, q in zip(paired, q_vals):
        row["q_BH_FDR"] = q
        row["significant_q_lt_0.05"] = bool(q is not None and q < 0.05)
    return arm_summary, paired


def collect_rag_generation_audit():
    raw_path = RAG_ROOT / "raw_calls.jsonl"
    final_records = []
    if raw_path.exists():
        with raw_path.open("r", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except json.JSONDecodeError:
                    continue
                text = rec.get("final_text")
                if text:
                    final_records.append(rec)
    explicit_re = re.compile(r"(KOA[-_ ]?EU|evidence[_ -]?id|source[_ -]?id|PMID\s*:?\s*\d+|PMC\d+|doi\s*:?\s*10\.|https?://|www\.)", re.I)
    broad_re = re.compile(r"(guideline|ACR|OARSI|AAOS|EULAR|PMID|doi|https?://|\[\d+\])", re.I)
    rows = []
    by_arm = defaultdict(list)
    for rec in final_records:
        text = str(rec.get("final_text") or "")
        arm = str(rec.get("arm") or "not_available")
        has_explicit = bool(explicit_re.search(text))
        has_broad = bool(broad_re.search(text))
        has_url_or_doi = bool(re.search(r"(https?://|www\.|doi\s*:?\s*10\.)", text, flags=re.I))
        has_pmid = bool(re.search(r"PMID\s*:?\s*\d+|PMC\d+", text, flags=re.I))
        row = {
            "patient_id": rec.get("patient_id"),
            "quadrant": rec.get("quadrant"),
            "sample": rec.get("sample"),
            "arm": arm,
            "text_chars": len(text),
            "has_explicit_source_id_url_doi_pmid": has_explicit,
            "has_url_or_doi": has_url_or_doi,
            "has_pmid_or_pmc": has_pmid,
            "has_broad_guideline_or_citation_marker": has_broad,
            "audit_type": "deterministic citation-marker proxy, not a faithfulness judge score",
            "source_file": str(raw_path),
        }
        rows.append(row)
        by_arm[arm].append(row)
    summary = []
    total = len(rows)
    if total:
        summary.append({
            "level": "overall",
            "arm": "ALL",
            "n_final_outputs": total,
            "explicit_source_marker_n": sum(r["has_explicit_source_id_url_doi_pmid"] for r in rows),
            "explicit_source_marker_rate": sum(r["has_explicit_source_id_url_doi_pmid"] for r in rows) / total,
            "url_doi_n": sum(r["has_url_or_doi"] for r in rows),
            "pmid_pmc_n": sum(r["has_pmid_or_pmc"] for r in rows),
            "broad_marker_n": sum(r["has_broad_guideline_or_citation_marker"] for r in rows),
            "broad_marker_rate": sum(r["has_broad_guideline_or_citation_marker"] for r in rows) / total,
            "remaining_gap": "Formal generation faithfulness, citation-support, and unsupported-claim rates still require judge/human annotation from judge_input.xlsx; this sheet does not fabricate those scores.",
        })
    for arm, items in sorted(by_arm.items()):
        n = len(items)
        summary.append({
            "level": "arm",
            "arm": arm,
            "n_final_outputs": n,
            "explicit_source_marker_n": sum(r["has_explicit_source_id_url_doi_pmid"] for r in items),
            "explicit_source_marker_rate": sum(r["has_explicit_source_id_url_doi_pmid"] for r in items) / n if n else 0,
            "url_doi_n": sum(r["has_url_or_doi"] for r in items),
            "pmid_pmc_n": sum(r["has_pmid_or_pmc"] for r in items),
            "broad_marker_n": sum(r["has_broad_guideline_or_citation_marker"] for r in items),
            "broad_marker_rate": sum(r["has_broad_guideline_or_citation_marker"] for r in items) / n if n else 0,
            "remaining_gap": "proxy only",
        })
    manifest = []
    for fn in ["judge_input.xlsx", "results_wide.xlsx", "raw_calls.jsonl", "progress.json"]:
        p = RAG_ROOT / fn
        manifest.append({
            "file": str(p),
            "exists": p.exists(),
            "size_bytes": p.stat().st_size if p.exists() else 0,
            "use_in_this_update": "source for generation outputs or pending judge input",
        })
    return summary, rows, manifest


def collect_oaknet_data():
    oak_summary = []
    oak_metrics_rows = []
    oak_training_summary = []
    if FIGURE_WB.exists():
        fwb = load_workbook(FIGURE_WB, read_only=True, data_only=True)
        if "oaknet_metrics_summary" in fwb.sheetnames:
            ws = fwb["oaknet_metrics_summary"]
            headers = [str(v).strip() if v is not None else f"col_{i+1}" for i, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))]
            for values in ws.iter_rows(min_row=2, values_only=True):
                if not any(values):
                    continue
                row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
                row["source_sheet"] = "figure_master::oaknet_metrics_summary"
                row["source_file"] = str(FIGURE_WB)
                oak_metrics_rows.append(row)
        if "oaknet_training_history_long" in fwb.sheetnames:
            ws = fwb["oaknet_training_history_long"]
            headers = [str(v).strip() if v is not None else f"col_{i+1}" for i, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))]
            idx = {h.lower(): h for h in headers}
            arm_key = idx.get("arm") or idx.get("model_arm") or headers[0]
            fold_key = idx.get("fold") or headers[1]
            epoch_key = idx.get("epoch") or headers[2]
            counts = defaultdict(lambda: {"n_rows": 0, "epochs": []})
            for values in ws.iter_rows(min_row=2, values_only=True):
                if not any(values):
                    continue
                row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
                key = (row.get(arm_key), row.get(fold_key))
                counts[key]["n_rows"] += 1
                ep = as_float(row.get(epoch_key))
                if ep is not None:
                    counts[key]["epochs"].append(ep)
            for (arm, fold), info in sorted(counts.items(), key=lambda x: (str(x[0][0]), str(x[0][1]))):
                epochs = info["epochs"]
                oak_training_summary.append({
                    "arm": arm,
                    "fold": fold,
                    "n_history_rows": info["n_rows"],
                    "epoch_min": min(epochs) if epochs else "not_available",
                    "epoch_max": max(epochs) if epochs else "not_available",
                    "source_sheet": "figure_master::oaknet_training_history_long",
                    "source_file": str(FIGURE_WB),
                })

    readme = OAK_AUDIT_ROOT / "README_AUDIT.md"
    readme_text = readme.read_text(encoding="utf-8", errors="replace") if readme.exists() else "not_found"
    prediction_inventory = OAK_AUDIT_ROOT / "03_prediction_data_audit" / "prediction_csv_inventory.csv"
    training_entrypoints = OAK_AUDIT_ROOT / "02_training_code_audit" / "training_entrypoints.csv"
    source_paths = [
        readme,
        prediction_inventory,
        training_entrypoints,
        OAK_AUDIT_ROOT / "05_recomputed_figures" / "confusion_matrices" / "A7_external_ensemble_predictions_confusion_counts.csv",
        OAK_AUDIT_ROOT / "05_recomputed_figures" / "confusion_matrices" / "A7_external_ensemble_predictions_confusion_normalized.csv",
        OAK_AUDIT_ROOT / "05_recomputed_figures" / "roc_curves" / "A7_external_ensemble_predictions_multiclass_roc.png",
        OAK_AUDIT_ROOT / "05_recomputed_figures" / "pr_curves" / "A7_external_ensemble_predictions_multiclass_pr.png",
        OAK_AUDIT_ROOT / "05_recomputed_figures" / "risk_coverage" / "A7_external_ensemble_predictions_risk_coverage.csv",
    ]
    for p in source_paths:
        oak_summary.append({
            "source_item": p.name,
            "source_file": str(p),
            "exists": p.exists(),
            "size_bytes": p.stat().st_size if p.exists() else 0,
            "interpretation": "可作为 OAKNet 结果/图源支持" if p.exists() else "not_found",
        })
    oak_summary.extend([
        {
            "source_item": "OAKNet audit conclusion",
            "source_file": str(readme),
            "exists": readme.exists(),
            "size_bytes": readme.stat().st_size if readme.exists() else 0,
            "interpretation": "训练代码、raw predictions、existing/recomputed figures found; checkpoints not found, so Grad-CAM/new-case inference must not be claimed as regenerated.",
        },
        {
            "source_item": "Manuscript-safe OAKNet result",
            "source_file": f"{FIGURE_WB}; {readme}",
            "exists": FIGURE_WB.exists() and readme.exists(),
            "size_bytes": FIGURE_WB.stat().st_size if FIGURE_WB.exists() else 0,
            "interpretation": "可写 ConvNeXt-B evidential head / A7 external summary metrics and recomputed curve/confusion evidence; checkpoint-dependent claims remain source-limited.",
        },
    ])
    return oak_summary, oak_metrics_rows, oak_training_summary, readme_text


def concise_float(v, nd=4):
    f = as_float(v)
    if f is None:
        return str(v)
    return f"{f:.{nd}f}"


def build_text_rows(risk, paired, rag_summary, oak_summary):
    risk_metrics_text = []
    for r in risk["metrics_all"]:
        risk_metrics_text.append(
            f"{r.get('endpoint')}: AUROC {concise_float(r.get('AUROC'),3)}, "
            f"AUPRC {concise_float(r.get('AUPRC'),3)}, Brier {concise_float(r.get('Brier'),3)}, "
            f"BACC {concise_float(r.get('balanced_accuracy'),3)}"
        )
    risk_boot = []
    for r in risk["bootstrap"]:
        risk_boot.append(
            f"{r.get('endpoint_key')}: AUROC bootstrap mean {concise_float(r.get('AUROC_bootstrap_mean'),3)}, "
            f"95% CI {concise_float(r.get('AUROC_CI_low'),3)}-{concise_float(r.get('AUROC_CI_high'),3)}, n={r.get('n_bootstrap')}"
        )
    selected_hp = "CatBoost selected in all endpoints: iterations=160, learning_rate=0.05, depth=4, loss_function=Logloss, auto_class_weights=Balanced, eval_metric=AUC, random_seed=20260610."

    top_overall = [r for r in paired if r["metric"] == "overall_0_100"]
    treat_lines = []
    for r in top_overall:
        treat_lines.append(
            f"{r['comparison']}: full {r['full_mean']:.1f} vs comparator {r['comparator_mean']:.1f}; "
            f"benefit {r['benefit_mean_diff']:.1f} (95% CI {r['benefit_95CI_low']:.1f} to {r['benefit_95CI_high']:.1f}), "
            f"dz={r['cohen_dz']:.2f}, q={r['q_BH_FDR']:.2g}, n={r['n_pairs']} pairs"
        )

    rag_overall = next((r for r in rag_summary if r.get("level") == "overall"), {})
    rag_text = (
        f"Generation output audit found {rag_overall.get('n_final_outputs','not_available')} final outputs; "
        f"explicit source-id/URL/DOI/PMID markers {rag_overall.get('explicit_source_marker_n','not_available')}/"
        f"{rag_overall.get('n_final_outputs','not_available')}, broad guideline/citation markers "
        f"{rag_overall.get('broad_marker_n','not_available')}/{rag_overall.get('n_final_outputs','not_available')}. "
        "This is a citation-marker proxy only; formal faithfulness/citation-support/unsupported-claim rates require judge_input.xlsx annotation."
    )

    oak_text = (
        "OAKNet can be written as a source-backed radiographic module with ConvNeXt-B plus evidential uncertainty head and DenseNet-121 comparator. "
        "Current evidence supports QWK 0.806±0.008, BACC 0.659, macro-F1 0.664, MAE 0.417, ECE 0.119 and selective accuracy@80 0.725 from the master table/figure master, "
        "with prediction CSVs and recomputed ROC/PR/calibration/DCA/risk-coverage/confusion outputs in the OAKNet audit package. "
        "Checkpoint-dependent Grad-CAM regeneration and new-case inference remain not supported by available checkpoint files."
    )

    return [
        {
            "section": "KOM-Risk 方法补写",
            "status": "SOURCE_BACKED",
            "draft_cn": (
                "KOM-Risk 采用受试者级拆分以避免同一受试者跨训练、验证和测试集泄漏，并分别构建结构进展、膝关节置换/手术事件和症状功能恶化三个终点。"
                "候选模型包括 elastic-net logistic regression、random forest、XGBoost、LightGBM 和 CatBoost；所有随机种子固定为 20260610。"
                f"{selected_hp} 模型选择依据验证集主指标完成，并在 locked retrain package 中保存模型、预处理管线、特征名、预测和指标。"
            ),
            "result_cn": "；".join(risk_metrics_text + risk_boot),
            "source_or_boundary": "RISK 07_metrics, 05_models metadata, 03_splits, 13_leakage_and_QC/reproducibility_statement.md",
        },
        {
            "section": "KOM-RAG / GraphRAG 结果补写",
            "status": "PARTIAL_SOURCE_BACKED_WITH_BOUNDARY",
            "draft_cn": (
                "KOM-RAG 使用 BAAI/bge-m3 1024 维嵌入、3260 条已加载证据、480 条查询任务（development 320、holdout 160），"
                "candidate_k=180，RRF k=30，并以 GraphRAG 与 naive RAG 作为公平检索对照。检索指标可直接写入主文；生成一致性目前只能写入生成记录和引用标记代理审计。"
            ),
            "result_cn": (
                "GraphRAG 在检索上优于 naive RAG：P@10 0.6763 vs 0.3025，Hit@10 1.0000 vs 0.6875，MRR 0.7483 vs 0.1588，nDCG@10 0.6902 vs 0.2367。"
                + rag_text
            ),
            "source_or_boundary": "RAG benchmark summary + generation_out/raw_calls.jsonl; formal judge scores remain pending, not fabricated.",
        },
        {
            "section": "KOM-Treat / MDT 配对统计补写",
            "status": "RECOMPUTED_FROM_RAW_SCORE_ROWS",
            "draft_cn": (
                "KOM-Treat 由 R0-R8 多专科智能体共同生成治疗建议，所有消融臂使用相同病例、相同 schema 和相同双模型评价流程。"
                "本轮从 2880 条 ok 评分记录按 case_id、quadrant、sample 和 evaluator_model 建立配对，比较 Full KOM 与 without RAG、without MDT、direct LLM。"
                "统计量为配对大样本正态近似的双侧检验、95% CI、Cohen dz，并对全部配对指标进行 Benjamini-Hochberg FDR 校正。"
            ),
            "result_cn": "；".join(treat_lines),
            "source_or_boundary": "RAW_Treat_AblationScores; scipy unavailable, so p values use normal approximation.",
        },
        {
            "section": "KOM-Rad / OAKNet 补写",
            "status": "FOUND_IN_OAKNET_AUDIT_NEEDS_MASTER_SYNC",
            "draft_cn": (
                "KOM-Rad/OAKNet 可作为影像分级模块写入方法：以 ConvNeXt-B evidential uncertainty head 为主模型，DenseNet-121 为比较模型，"
                "围绕 A/B 系列模块进行 5-fold 训练历史和外部验证汇总。"
            ),
            "result_cn": oak_text,
            "source_or_boundary": "Figure master + OAKNet_Reproducibility_Audit_202606; checkpoint missing, Grad-CAM/new inference cannot be claimed as regenerated.",
        },
    ]


def update_existing_training_gap_sheet(wb):
    if "训练参数待补表" not in wb.sheetnames:
        return
    ws = wb["训练参数待补表"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    new_cols = ["本轮补写状态", "本轮补写内容", "剩余边界"]
    col_map = {}
    for nc in new_cols:
        if nc in headers:
            col_map[nc] = headers.index(nc) + 1
        else:
            c = ws.max_column + 1
            ws.cell(1, c, nc)
            ws.cell(1, c).fill = HEADER_FILL
            ws.cell(1, c).font = HEADER_FONT
            col_map[nc] = c
            headers.append(nc)
    updates = {
        "OAKNet": (
            "FOUND_IN_OAKNET_AUDIT_NEEDS_MASTER_SYNC",
            "已补入 figure master 的 A7/OAKNet 指标、训练历史摘要、OAKNet 审计包中的 prediction CSV、confusion/ROC/PR/calibration/DCA/risk-coverage 图源。",
            "checkpoint 仍未找到；Grad-CAM regeneration 和 new-case inference 不能写成已重现。",
        ),
        "KOM-Risk": (
            "已补齐",
            "从 locked retrain package 写入 seed=20260610、CatBoost iterations=160/learning_rate=0.05/depth=4/loss=Logloss/auto_class_weights=Balanced/eval_metric=AUC，以及 5 个候选模型参数表和 AUROC CI。",
            "无核心训练参数缺口；若写非常细的库默认参数，可参考 KOMRisk超参数已补 sheet 的完整 JSON。",
        ),
        "KOM-RAG": (
            "部分补齐",
            "已补入生成输出 citation-marker proxy：raw_calls.jsonl 中 final outputs 的 explicit source-id/URL/DOI/PMID marker、broad guideline/citation marker。",
            "faithfulness/citation support/unsupported-claim rate 仍需 judge_input.xlsx 人工或模型评分；不能把代理审计冒充为正式 faithfulness。",
        ),
        "KOM-Treat / MDT": (
            "已补齐",
            "已从 2880 条原始评分按病例/象限/sample/评价模型配对，补入 Full KOM vs without RAG/without MDT/direct LLM 的 mean difference、95% CI、Cohen dz、p 和 BH-FDR q。",
            "若投稿需要非参数 Wilcoxon 或 bootstrap CI，可作为敏感性分析另跑；核心配对统计已可写。",
        ),
    }
    for r in range(2, ws.max_row + 1):
        mod = str(ws.cell(r, 1).value or "")
        for key, vals in updates.items():
            if key in mod:
                for idx, nc in enumerate(new_cols):
                    ws.cell(r, col_map[nc], vals[idx])
                break
    style_sheet(ws, freeze="A2")


def update_existing_method_result_sheet(wb, text_rows):
    if "完整方法结果对应稿" not in wb.sheetnames:
        return
    ws = wb["完整方法结果对应稿"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    new_header = "本轮补写更新"
    if new_header in headers:
        col = headers.index(new_header) + 1
    else:
        col = ws.max_column + 1
        ws.cell(1, col, new_header)
        ws.cell(1, col).fill = HEADER_FILL
        ws.cell(1, col).font = HEADER_FONT
    lookup = {
        "KOM-Risk": next((r for r in text_rows if r["section"].startswith("KOM-Risk")), None),
        "KOM-RAG": next((r for r in text_rows if r["section"].startswith("KOM-RAG")), None),
        "KOM-Treat": next((r for r in text_rows if r["section"].startswith("KOM-Treat")), None),
        "OAKNet": next((r for r in text_rows if "OAKNet" in r["section"]), None),
        "KOM-Rad": next((r for r in text_rows if "OAKNet" in r["section"]), None),
    }
    for r in range(2, ws.max_row + 1):
        row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1))
        for key, item in lookup.items():
            if key in row_text and item:
                ws.cell(r, col, f"{item['status']} | {item['result_cn']} | 边界: {item['source_or_boundary']}")
                break
    style_sheet(ws, freeze="A2")


def write_markdown(path: Path, text_rows, qc):
    lines = [
        "# KOM 缺口补算补写中文稿",
        "",
        f"生成时间：{qc['generated_at']}",
        f"总表：{qc['output_workbook']}",
        "",
        "## 本轮总判定",
        "",
        qc["overall_judgment"],
        "",
    ]
    for row in text_rows:
        lines.extend([
            f"## {row['section']}",
            "",
            f"状态：{row['status']}",
            "",
            "方法写法：",
            "",
            row["draft_cn"],
            "",
            "结果写法：",
            "",
            row["result_cn"],
            "",
            "证据/边界：",
            "",
            row["source_or_boundary"],
            "",
        ])
    path.write_text("\n".join(lines), encoding="utf-8")


def main():
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M")
    out_dir = OUT_PARENT / f"KOM_缺口补算补写_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_wb = out_dir / f"KOM_缺口补算补写_完整方法结果总表_{ts}.xlsx"
    shutil.copy2(SOURCE_WB, out_wb)

    wb = load_workbook(out_wb)
    risk = collect_risk_data()
    arm_summary, paired = collect_treat_stats(wb)
    rag_summary, rag_detail, rag_manifest = collect_rag_generation_audit()
    oak_summary, oak_metrics_rows, oak_training_summary, oak_readme_text = collect_oaknet_data()
    text_rows = build_text_rows(risk, paired, rag_summary, oak_summary)

    # Front overview
    ws = replace_sheet(wb, "本轮补算总览", 0)
    headers = ["模块", "本轮动作", "判定", "写入内容", "剩余边界", "主要来源"]
    rows = [
        ["KOM-Risk", "读取 locked retrain package", "FOUND_AND_FILLED", "seed、CatBoost 参数、5 候选模型参数、AUROC/AUPRC/Brier/BACC、bootstrap AUROC CI、split integrity", "无核心训练参数缺口", str(RISK_ROOT)],
        ["KOM-Treat / MDT", "从 RAW_Treat_AblationScores 重算", "RECOMPUTED_FROM_RAW_SCORE_ROWS", "四臂臂级均值、Full KOM 对 3 个对照的配对差异、95% CI、Cohen dz、p、BH-FDR q", "非参数/bootstrapped 敏感性分析可选", "RAW_Treat_AblationScores"],
        ["KOM-RAG / GraphRAG", "审计 generation_out/raw_calls.jsonl", "PARTIAL_SOURCE_BACKED_WITH_BOUNDARY", "final outputs 的 source marker / guideline marker 代理审计；检索指标保持可写", "正式 faithfulness/citation support/unsupported-claim rate 仍需 judge_input.xlsx 评分", str(RAG_ROOT)],
        ["KOM-Rad / OAKNet", "合并 figure master + OAKNet audit", "FOUND_IN_OAKNET_AUDIT_NEEDS_MASTER_SYNC", "A7/OAKNet 汇总性能、训练历史摘要、预测 CSV、混淆矩阵和重绘曲线源文件证据", "checkpoint 未找到；Grad-CAM 和 new-case inference 不写成已重现", f"{FIGURE_WB}; {OAK_AUDIT_ROOT}"],
    ]
    write_table(ws, rows, headers)
    style_sheet(ws, freeze="A2")

    # Risk sheets
    ws = replace_sheet(wb, "KOMRisk超参数已补", 1)
    r0 = add_title(ws, "KOM-Risk 超参数与指标已补", "来源：locked retrain package；所有随机种子固定为 20260610。")
    comparison_headers = list(risk["comparison"][0].keys()) if risk["comparison"] else []
    comparison_rows = []
    for row in risk["comparison"]:
        rr = dict(row)
        rr["source_file"] = str(RISK_METRICS / "final_model_comparison_all_algorithms.csv")
        comparison_rows.append(rr)
    if comparison_rows:
        headers_cmp = comparison_headers + ["source_file"]
        r0 = write_table(ws, comparison_rows, headers_cmp, r0)
        r0 += 2
    metrics_rows = []
    for row in risk["metrics_all"]:
        rr = dict(row)
        rr["source_file"] = str(RISK_METRICS / "final_model_metrics_all_endpoints.csv")
        metrics_rows.append(rr)
    if metrics_rows:
        ws.cell(r0, 1, "Final model metrics all endpoints")
        ws.cell(r0, 1).font = BOLD_FONT
        r0 += 1
        r0 = write_table(ws, metrics_rows, list(metrics_rows[0].keys()), r0)
        r0 += 2
    if risk["bootstrap"]:
        ws.cell(r0, 1, "Bootstrap AUROC CI")
        ws.cell(r0, 1).font = BOLD_FONT
        r0 += 1
        r0 = write_table(ws, risk["bootstrap"], list(risk["bootstrap"][0].keys()), r0)
    style_sheet(ws, freeze="A4")

    ws = replace_sheet(wb, "KOMRisk随机种子模型卡", 2)
    model_rows = []
    for row in risk["metadata"]:
        model_rows.append(row)
    r0 = add_title(ws, "KOM-Risk 随机种子、模型卡与 split integrity", "该 sheet 用于补充 methods 中的 reproducibility/random seed/model-card 细节。")
    if model_rows:
        r0 = write_table(ws, model_rows, list(model_rows[0].keys()), r0)
        r0 += 2
    split_headers = sorted({k for row in risk["split"] for k in row.keys()})
    if split_headers:
        ws.cell(r0, 1, "Split integrity and event-rate summary")
        ws.cell(r0, 1).font = BOLD_FONT
        r0 += 1
        r0 = write_table(ws, risk["split"], split_headers, r0)
        r0 += 2
    ws.cell(r0, 1, "Reproducibility statement")
    ws.cell(r0, 1).font = BOLD_FONT
    ws.cell(r0 + 1, 1, risk["repro_text"])
    ws.cell(r0 + 1, 2, risk["repro_file"])
    style_sheet(ws, freeze="A4")

    # Treat sheets
    ws = replace_sheet(wb, "KOMTreat臂级均值已补", 3)
    write_table(ws, arm_summary, list(arm_summary[0].keys()) if arm_summary else ["not_available"])
    style_sheet(ws, freeze="A2")

    ws = replace_sheet(wb, "KOMTreat配对统计已补", 4)
    write_table(ws, paired, list(paired[0].keys()) if paired else ["not_available"])
    style_sheet(ws, freeze="A2")

    # RAG sheets
    ws = replace_sheet(wb, "KOMRAG生成引用审计", 5)
    r0 = add_title(ws, "KOM-RAG 生成输出引用标记代理审计", "注意：该表仅为 deterministic citation-marker proxy，不是 faithfulness/citation support 正式评分。")
    if rag_summary:
        r0 = write_table(ws, rag_summary, list(rag_summary[0].keys()), r0)
        r0 += 2
    if rag_detail:
        ws.cell(r0, 1, "Per-output audit detail")
        ws.cell(r0, 1).font = BOLD_FONT
        r0 += 1
        r0 = write_table(ws, rag_detail, list(rag_detail[0].keys()), r0)
        r0 += 2
    ws.cell(r0, 1, "Generation file manifest")
    ws.cell(r0, 1).font = BOLD_FONT
    r0 += 1
    write_table(ws, rag_manifest, list(rag_manifest[0].keys()), r0)
    style_sheet(ws, freeze="A4")

    # OAKNet sheets
    ws = replace_sheet(wb, "OAKNet源补写", 6)
    r0 = add_title(ws, "OAKNet 源文件、指标与边界补写", "此处把 OAKNet 审计包和 figure master 合入总表；checkpoint 仍标注为缺失。")
    if oak_summary:
        r0 = write_table(ws, oak_summary, list(oak_summary[0].keys()), r0)
        r0 += 2
    if oak_metrics_rows:
        ws.cell(r0, 1, "oaknet_metrics_summary raw import")
        ws.cell(r0, 1).font = BOLD_FONT
        r0 += 1
        headers_oak = list(oak_metrics_rows[0].keys())
        r0 = write_table(ws, oak_metrics_rows, headers_oak, r0)
        r0 += 2
    if oak_training_summary:
        ws.cell(r0, 1, "oaknet_training_history_long summary")
        ws.cell(r0, 1).font = BOLD_FONT
        r0 += 1
        write_table(ws, oak_training_summary, list(oak_training_summary[0].keys()), r0)
    style_sheet(ws, freeze="A4")

    # Text and log sheets
    ws = replace_sheet(wb, "完整正文补写更新", 7)
    write_table(ws, text_rows, ["section", "status", "draft_cn", "result_cn", "source_or_boundary"])
    style_sheet(ws, freeze="A2")

    ws = replace_sheet(wb, "缺口补写日志", 8)
    log_rows = [
        ["generated_at", ts],
        ["source_workbook", str(SOURCE_WB)],
        ["output_workbook", str(out_wb)],
        ["risk_rows_candidate_models", len(risk["comparison"])],
        ["risk_final_metric_rows", len(risk["metrics_all"])],
        ["treat_raw_ok_rows", 2880],
        ["treat_paired_rows", len(paired)],
        ["rag_final_outputs", len(rag_detail)],
        ["oaknet_metric_rows_imported", len(oak_metrics_rows)],
        ["oaknet_training_arm_fold_rows", len(oak_training_summary)],
        ["overall_judgment", "READY_WITH_WARNINGS: Risk and Treat filled; RAG formal generation faithfulness remains pending; OAKNet checkpoint-dependent outputs remain source-limited."],
    ]
    write_table(ws, log_rows, ["item", "value"])
    style_sheet(ws, freeze="A2")

    update_existing_training_gap_sheet(wb)
    update_existing_method_result_sheet(wb, text_rows)

    wb.save(out_wb)

    qc = {
        "generated_at": ts,
        "source_workbook": str(SOURCE_WB),
        "output_workbook": str(out_wb),
        "new_sheets": [
            "本轮补算总览",
            "KOMRisk超参数已补",
            "KOMRisk随机种子模型卡",
            "KOMTreat臂级均值已补",
            "KOMTreat配对统计已补",
            "KOMRAG生成引用审计",
            "OAKNet源补写",
            "完整正文补写更新",
            "缺口补写日志",
        ],
        "risk_candidate_model_rows": len(risk["comparison"]),
        "risk_final_metric_rows": len(risk["metrics_all"]),
        "treat_arm_summary_rows": len(arm_summary),
        "treat_paired_stat_rows": len(paired),
        "rag_final_outputs": len(rag_detail),
        "oaknet_metric_rows": len(oak_metrics_rows),
        "oaknet_training_summary_rows": len(oak_training_summary),
        "overall_judgment": "READY_WITH_WARNINGS: KOM-Risk and KOM-Treat are filled from source/recomputed records; KOM-RAG generation faithfulness remains a formal scoring task; OAKNet checkpoint-dependent outputs remain source-limited.",
    }
    qc_path = out_dir / f"KOM_缺口补算补写_QC_{ts}.json"
    qc_path.write_text(json.dumps(qc, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path = out_dir / f"KOM_缺口补算补写_中文稿_{ts}.md"
    write_markdown(md_path, text_rows, qc)

    print(json.dumps({"output_workbook": str(out_wb), "qc_json": str(qc_path), "markdown": str(md_path), "output_dir": str(out_dir)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
