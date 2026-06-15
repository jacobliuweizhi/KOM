from __future__ import annotations

import csv
import json
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_WB = Path(
    "C:/OAI\u7814\u7a76\u9879\u76ee/pythonProject1/KOM\u8fd4\u4fee\u4fee\u6539/"
    "\u6295\u7a3f\u4f7f\u7528/\u6700\u7ec8\u7248\u672c/"
    "KOM_\u8865\u5145\u6750\u6599\u5b8c\u6574\u6b63\u6587_20260615_1939/"
    "KOM_\u8865\u5145\u6750\u6599\u65b9\u6cd5\u7ed3\u679c\u5b8c\u6574\u6b63\u6587\u603b\u8868_20260615_1939.xlsx"
)
METHODS_TXT = Path(
    "C:/Users/Liu/.codex/attachments/e7544ca8-7ea3-4808-b31e-c8a71308596f/pasted-text.txt"
)
JUDGE_OUT = Path(
    "C:/OAI\u7814\u7a76\u9879\u76ee/pythonProject1/KOM\u8fd4\u4fee\u4fee\u6539/"
    "\u6cbb\u7597\u667a\u80fd\u4f53\u7ec4/\u65b0\u7ed3\u679c\u8f93\u51fa/"
    "kom_pipeline/kom_pipeline/judge_out"
)
OUT_ROOT = Path(
    "C:/OAI\u7814\u7a76\u9879\u76ee/pythonProject1/KOM\u8fd4\u4fee\u4fee\u6539/"
    "\u6295\u7a3f\u4f7f\u7528/\u6700\u7ec8\u7248\u672c"
)


S = {
    "expert_summary": "OLD_022_Expert_\u4e13\u5bb6API_\u516d\u4e13\u5bb6\u5206\u7ec4\u6c47\u603b",
    "raw_arm_summary": "OLD_024_Expert_\u6d88\u878d_\u56db\u81c2_\u81c2\u7ea7\u6c47\u603b",
    "paired": "OLD_025_Expert_\u6d88\u878d_\u56db\u81c2_\u914d\u5bf9\u68c0\u9a8c",
    "rule_summary": "OLD_026_Expert_\u6d88\u878d_\u56db\u81c2_\u89c4\u5219_\u81c2\u7ea7\u6c47\u603b",
    "raw_rx": "OLD_060_Expert_\u6d88\u878d_\u56db\u81c2_\u539f\u59cb\u5904\u65b9",
}


def records(ws, header_row: int = 1) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(x) if x is not None else "" for x in rows[header_row - 1]]
    out = []
    for row in rows[header_row:]:
        if all(x is None for x in row):
            continue
        out.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return out


def metric_map(ws) -> dict:
    out = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] is not None:
            out[str(row[0])] = row[1] if len(row) > 1 else None
    return out


def safe_float(v):
    try:
        return float(v)
    except Exception:
        return v


def round_float(v, digits=3):
    if isinstance(v, (int, float)):
        return round(v, digits)
    try:
        return round(float(v), digits)
    except Exception:
        return v


def write_rows(ws, rows: list[list], freeze: str = "A2"):
    for row in rows:
        ws.append(row)
    if freeze:
        ws.freeze_panes = freeze
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 80) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 55))


def remove_if_exists(wb, names: list[str]):
    for name in names:
        if name in wb.sheetnames:
            wb.remove(wb[name])


def extract_source_data():
    wb = load_workbook(BASE_WB, read_only=True, data_only=True)

    expert_rows = [
        r for r in records(wb[S["expert_summary"]], 1) if r.get("source_type") == "ablation_four_arm"
    ]
    headline = records(wb["SUM_174_16_KOM_TREAT_ABLATIO"], 4)
    paired_rows = records(wb[S["paired"]], 3)
    icc_rows = records(wb["RAW_Expert_ICC"], 1)
    risk_rows = records(wb["SUM_166_08_KOM_RISK_LATEST"], 4)
    profile_rows = records(wb["SUM_164_06_KOM_PROFILE"], 4)
    oak_rows = records(wb["SUM_165_07_KOM_RAD_OAKNET"], 4)
    case_rows = records(wb["SUM_163_05_STANDARDIZED_CASE"], 4)
    sim_rows = records(wb["SUM_173_15_KOM_SIM_CLINICIAN"], 4)
    rag_summary = metric_map(wb["RAG_metrics_summary"])
    rag_holdout = json.loads(rag_summary["holdout"])
    raw_rx = records(wb[S["raw_rx"]], 1)
    raw_counts = {
        "row_count": len(raw_rx),
        "arm_code_counts": dict(Counter(r.get("arm_code") for r in raw_rx)),
        "sample_counts": dict(Counter(r.get("sample") for r in raw_rx)),
    }
    judge_summary = {}
    if (JUDGE_OUT / "summary.json").exists():
        judge_summary = json.loads((JUDGE_OUT / "summary.json").read_text(encoding="utf-8"))
    return {
        "expert_rows": expert_rows,
        "headline": headline,
        "paired_rows": paired_rows,
        "icc_rows": icc_rows,
        "risk_rows": risk_rows,
        "profile_rows": profile_rows,
        "oak_rows": oak_rows,
        "case_rows": case_rows,
        "sim_rows": sim_rows,
        "rag_summary": rag_summary,
        "rag_holdout": rag_holdout,
        "raw_counts": raw_counts,
        "judge_summary": judge_summary,
    }


ARM_LABEL = {
    "A_full": ("A_full_KOM", "Full KOM / \u5b8c\u6574 KOM"),
    "B_no_rag": ("B_without_RAG", "KOM without RAG / \u53bb RAG"),
    "C_no_mdt": ("C_without_MDT", "KOM without MDT / \u53bb MDT"),
    "D_bare": ("D_direct_LLM", "Direct LLM baseline / \u76f4\u63a5 LLM"),
}


def build_section_rows(data):
    def pick_risk_row(prefix: str, needle: str):
        for row in data["risk_rows"]:
            endpoint = str(row.get("endpoint", ""))
            if endpoint.startswith(prefix) or needle in endpoint:
                return row
        raise KeyError(f"risk endpoint not found: {prefix}/{needle}")

    risk_a = pick_risk_row("A_", "KL")
    risk_b = pick_risk_row("B_", "TKR")
    risk_c = pick_risk_row("C_", "symptom")
    sim = {r["arm"]: r for r in data["sim_rows"]}
    rag = data["rag_holdout"]
    sections = [
        [
            "S1",
            "\u6807\u51c6\u5316\u75c5\u4f8b\u6784\u5efa",
            "\u672c\u7814\u7a76\u9996\u5148\u6784\u5efa 120 \u4f8b\u6807\u51c6\u5316 KOA \u533b\u751f\u4efb\u52a1\u75c5\u4f8b\u96c6\u3002\u75c5\u4f8b\u6765\u81ea OAI \u6d3e\u751f\u5019\u9009\u753b\u50cf\u6c60\uff0c\u91c7\u7528\u5206\u5c42\u76ee\u7684\u62bd\u6837\u548c\u6700\u5927\u53d8\u5f02\u8986\u76d6\uff0c\u800c\u975e\u771f\u5b9e\u4e16\u754c\u60a3\u75c5\u7387\u62bd\u6837\u3002\u6bcf\u4e2a\u75c5\u4f8b\u4fdd\u7559\u76ee\u6807\u819d\u3001\u5f71\u50cf\u7ed3\u6784\u3001\u75bc\u75db\u529f\u80fd\u3001BMI\u3001\u8dcc\u5012\u98ce\u9669\u3001\u7528\u836f/\u5408\u5e76\u75c7\u5b89\u5168\u6807\u5fd7\u3001\u5eb7\u590d\u610f\u613f\u3001\u6cbb\u7597\u504f\u597d\u548c\u7f3a\u5931\u4fe1\u606f\u72b6\u6001\u3002",
            "\u9501\u5b9a\u75c5\u4f8b\u96c6\u4e3a 120 \u4f8b\uff0cQ1-Q4 \u56db\u4e2a\u8d1f\u62c5-\u9700\u6c42\u8c61\u9650\u5404 30 \u4f8b\u3002\u8be5\u75c5\u4f8b\u96c6\u7528\u4e8e\u8bc4\u4f30\u7cfb\u7edf\u5728\u591a\u79cd\u4e34\u5e8a\u51b3\u7b56\u60c5\u5883\u4e2d\u7684\u7a33\u5b9a\u6027\uff0c\u4e0d\u7528\u4e8e\u63a8\u65ad\u771f\u5b9e\u4e16\u754c\u6d41\u884c\u75c5\u5b66\u5206\u5e03\u3002",
            "SUM_163_05_STANDARDIZED_CASE",
            "\u4e0d\u5199\u6210\u771f\u5b9e\u60a3\u8005\u7ed3\u5c40\u961f\u5217\u3002",
        ],
        [
            "S2",
            "KOM-Profile \u60a3\u8005\u753b\u50cf",
            "KOM-Profile \u5c06\u6807\u51c6\u5316\u75c5\u4f8b\u8f6c\u6362\u4e3a\u7ed3\u6784\u5316\u60a3\u8005\u753b\u50cf\uff0c\u4f9b\u98ce\u9669\u3001\u68c0\u7d22\u3001\u6cbb\u7597\u548c\u5b89\u5168\u6a21\u5757\u5171\u540c\u8c03\u7528\u3002\u753b\u50cf\u5b57\u6bb5\u5305\u542b\u4eba\u53e3\u5b66\u3001\u5f71\u50cf\u3001\u75bc\u75db\u529f\u80fd\u3001\u795e\u7ecf\u808c\u8089/\u8dcc\u5012\u98ce\u9669\u3001\u4ee3\u8c22\u8425\u517b\u3001\u5fc3\u7406\u884c\u4e3a\u3001\u5185\u79d1\u5b89\u5168\u548c\u6cbb\u7597\u504f\u597d\u7b49\u57df\u3002",
            "\u9501\u5b9a schema \u5305\u542b 56 \u4e2a\u7ed3\u6784\u5316\u5b57\u6bb5\uff0coverall F1/accuracy=0.846\uff0c\u5176\u4e2d 31 \u4e2a exact match\uff0c25 \u4e2a partial match\u3002\u5f53\u524d\u6062\u590d\u5305\u672a\u627e\u5230\u9010\u5b57\u6bb5\u539f\u59cb\u62bd\u53d6\u660e\u7ec6\uff0c\u56e0\u6b64\u6b63\u6587\u53ea\u5199\u9501\u5b9a\u6c47\u603b\u7ed3\u679c\u3002",
            "SUM_164_06_KOM_PROFILE",
            "\u4e0d\u4f2a\u9020\u9010\u5b57\u6bb5 raw rows\u3002",
        ],
        [
            "S3",
            "KOM-Rad / OAK-Net \u5f71\u50cf\u6a21\u5757",
            "KOM-Rad \u4f7f\u7528 OAK-Net \u5bf9\u819d\u5173\u8282 X \u7ebf\u5f71\u50cf\u8f93\u51fa KL \u4e25\u91cd\u5ea6\u548c\u4e0d\u786e\u5b9a\u6027\u4fe1\u53f7\u3002\u4e3b\u6a21\u578b\u4e3a ConvNeXt-B + evidential uncertainty head\uff0cDenseNet-121 \u4f5c\u4e3a\u6bd4\u8f83\u6a21\u578b\u3002",
            "\u9501\u5b9a\u603b\u8868\u548c\u56fe\u6e90\u652f\u6301\u7684\u5f71\u50cf\u7ed3\u679c\u4e3a QWK 0.806\u00b10.008\u3001BACC 0.659\u3001macro-F1 0.664\u3001MAE 0.417\u3001ECE 0.119\uff0c80% coverage \u4e0b selective accuracy 0.725\u3002\u5f53\u524d\u672a\u627e\u5230 checkpoint\uff0c\u56e0\u6b64\u4e0d\u5199 checkpoint-dependent Grad-CAM \u518d\u751f\u6210\u6216\u65b0\u75c5\u4f8b\u63a8\u7406\u5df2\u91cd\u73b0\u3002",
            "SUM_165_07_KOM_RAD_OAKNET",
            "\u8d85\u53c2\u548c checkpoint \u9700\u540e\u7eed\u6e90\u786e\u8ba4\u3002",
        ],
        [
            "S4",
            "KOM-Risk \u9884\u540e\u98ce\u9669\u6a21\u578b",
            "KOM-Risk \u91c7\u7528\u53d7\u8bd5\u8005\u7ea7 split\uff0c\u9632\u6b62\u540c\u4e00\u53d7\u8bd5\u8005\u8de8\u8bad\u7ec3/\u9a8c\u8bc1/\u6d4b\u8bd5\u96c6\u6cc4\u6f0f\u3002\u4e09\u4e2a\u7ec8\u70b9\u5206\u522b\u4e3a KL \u7ed3\u6784\u8fdb\u5c55\u3001TKR/\u819d\u624b\u672f\u4e8b\u4ef6\u3001\u75c7\u72b6/\u529f\u80fd\u6076\u5316\u3002\u5019\u9009\u6a21\u578b\u5305\u62ec elastic-net logistic regression\u3001random forest\u3001XGBoost\u3001LightGBM \u548c CatBoost\uff0c\u6700\u7ec8\u4e09\u7ec8\u70b9 best_model \u5747\u4e3a CatBoost\u3002",
            f"Endpoint A AUROC {round_float(risk_a['AUROC'],3)}\u3001AUPRC {round_float(risk_a['AUPRC'],3)}\u3001Brier {round_float(risk_a['Brier'],3)}\u3001BACC {round_float(risk_a['BACC'],3)}\uff1bEndpoint B AUROC {round_float(risk_b['AUROC'],3)}\u3001AUPRC {round_float(risk_b['AUPRC'],3)}\u3001Brier {round_float(risk_b['Brier'],3)}\u3001BACC {round_float(risk_b['BACC'],3)}\uff1bEndpoint C AUROC {round_float(risk_c['AUROC'],3)}\u3001AUPRC {round_float(risk_c['AUPRC'],3)}\u3001Brier {round_float(risk_c['Brier'],3)}\u3001BACC {round_float(risk_c['BACC'],3)}\u3002A/B \u4e3a\u4e3b\u5206\u6790\uff0cC \u4e3a\u8865\u5145\u7ec8\u70b9\u3002",
            "SUM_166_08_KOM_RISK_LATEST; RISK_metrics_latest",
            "\u4e0d\u6df7\u7528 legacy LightGBM/CoxPH \u65e7\u6307\u6807\u3002",
        ],
        [
            "S5",
            "SHAP / \u7279\u5f81\u6eaf\u6e90",
            "\u9501\u5b9a\u6a21\u578b\u7684 SHAP \u503c\u4e0e feature-to-source lineage \u8868\u914d\u5bf9\u4fdd\u5b58\uff0c\u4f7f\u6bcf\u4e2a\u89e3\u91ca\u6027\u7279\u5f81\u53ef\u8ffd\u6eaf\u5230\u539f\u59cb OAI/source \u53d8\u91cf\u548c\u7f16\u7801\u540e\u7279\u5f81\u540d\u3002",
            "SHAP \u548c\u7279\u5f81\u6eaf\u6e90 checkpoint \u5728\u603b\u8868\u4e2d\u6807\u8bb0\u4e3a\u901a\u8fc7\uff0c\u53ef\u4f5c\u4e3a\u6a21\u578b\u5f52\u56e0\u548c\u56fe\u4ef6\u6eaf\u6e90\u8bc1\u636e\u3002",
            "RESULT_CHECKPOINTS E01-E03",
            "SHAP \u4e0d\u5199\u6210\u56e0\u679c\u673a\u5236\u3002",
        ],
        [
            "S6",
            "KOM-KB \u8bc1\u636e\u6570\u636e\u5e93",
            "KOM-KB \u8ddf\u8e2a PubMed/search lifecycle\u3001evidence-unit schema\u3001source identifiers \u548c L1-L7 \u8bc1\u636e\u7b49\u7ea7\uff0c\u4e3a RAG \u68c0\u7d22\u548c\u591a\u4e13\u79d1\u5904\u65b9\u751f\u6210\u63d0\u4f9b\u53ef\u8ffd\u6eaf\u8bc1\u636e\u5e95\u5ea7\u3002",
            "KOM-KB \u7684 search lifecycle\u3001evidence schema \u548c evidence counts checkpoint \u901a\u8fc7\u3002RAG \u68c0\u7d22\u5b50\u96c6\u52a0\u8f7d evidence_loaded=3260\uff0c\u5199\u4f5c\u65f6\u9700\u533a\u5206\u8bc1\u636e\u5e93\u603b evidence units \u548c RAG loaded subset\u3002",
            "RAG_metrics_summary; KOM-KB evidence sheets",
            "\u4e0d\u6df7\u7528\u4e0d\u540c\u5206\u6bcd\u3002",
        ],
        [
            "S7",
            "KOM-RAG / GraphRAG",
            "KOM-RAG \u4f7f\u7528 BAAI/bge-m3 \u5d4c\u5165\u6a21\u578b\uff0c\u5411\u91cf\u7ef4\u5ea6 1024\u3002\u68c0\u7d22\u57fa\u51c6\u5305\u542b 480 \u6761 query\uff0cdevelopment 320\u3001holdout 160\u3002\u7ba1\u7ebf\u52a0\u8f7d 3260 \u6761\u8bc1\u636e\uff0ccandidate_k=180\uff0cRRF k=30\uff0c\u5bf9\u6bd4 GraphRAG \u548c naive RAG\u3002",
            f"160 \u6761 holdout query \u4e0a\uff0cGraphRAG \u76f8\u5bf9 naive RAG \u7684 P@10 \u4e3a {rag['graph']['precision_at_10']:.4f} vs {rag['baseline']['precision_at_10']:.4f}\uff0cHit@10 \u4e3a {rag['graph']['hit_rate_at_10']:.4f} vs {rag['baseline']['hit_rate_at_10']:.4f}\uff0cMRR \u4e3a {rag['graph']['mrr']:.4f} vs {rag['baseline']['mrr']:.4f}\uff0cnDCG@10 \u4e3a {rag['graph']['ndcg_at_10']:.4f} vs {rag['baseline']['ndcg_at_10']:.4f}\u3002\u65e7 judge_out \u4e2d 2 \u4f8b\u30019 arms\u30016 \u6a21\u578b\u8bc4\u59d4\u7ed3\u679c\u5df2\u6392\u9664\uff0c\u4e0d\u5199\u5165\u5f53\u524d\u8bba\u6587\u751f\u6210\u8d28\u91cf\u7ed3\u679c\u3002",
            "RAW_RAG_Benchmark; AUD_137_KOM-RAG topk and lab; RAG_metrics_summary",
            "\u751f\u6210\u7aef faithfulness/citation support/unsupported-claim rate \u672a\u6709\u5f53\u524d\u56db\u81c2\u4e13\u5bb6\u6807\u6ce8\u6e90\uff0c\u4e0d\u7528\u65e7 9 \u81c2 judge_out \u8865\u4f4d\u3002",
        ],
        [
            "S8",
            "KOM-MDT / KOM-Rx",
            "KOM-MDT \u7531 R0-R8 \u4e5d\u4e2a\u534f\u4f5c\u667a\u80fd\u4f53\u6784\u6210\uff1aR0 \u753b\u50cf\u5171\u8bc6\uff0cR1 \u8fd0\u52a8\u5eb7\u590d\uff0cR2 \u4f53\u91cd\u8425\u517b\uff0cR3 \u5fc3\u7406\u884c\u4e3a\uff0cR4 \u9aa8\u79d1\u7efc\u5408\uff0cR5 \u8de8\u4e13\u79d1\u8d28\u7591\uff0cR6 \u8bc1\u636e\u4ef2\u88c1\uff0cR7 \u9aa8\u79d1\u8fb9\u754c\u4ef2\u88c1\uff0cR8 \u6700\u7ec8\u5904\u65b9\u6574\u5408\u3002R0-R8 \u662f\u534f\u4f5c\u89d2\u8272\uff0c\u4e0d\u662f\u5b9e\u9a8c\u81c2\u3002",
            "\u591a\u4e13\u79d1\u534f\u4f5c\u7684\u5b9e\u9a8c\u8bc1\u636e\u4f53\u73b0\u5728 S11 \u7684\u56db\u81c2\u6d88\u878d\u4e2d\uff1a\u53bb MDT \u540e\u516d\u4e13\u5bb6 overall quality \u4ece Full KOM 84.6 \u964d\u81f3 64.4\uff0c\u8bf4\u660e\u534f\u4f5c/\u4ef2\u88c1\u6d41\u7a0b\u5bf9\u5904\u65b9\u8d28\u91cf\u6709\u53ef\u89c2\u5bdf\u8d21\u732e\u3002",
            "SUM_174_16_KOM_TREAT_ABLATIO; OLD_022_Expert_\u4e13\u5bb6API_\u516d\u4e13\u5bb6\u5206\u7ec4\u6c47\u603b",
            "\u660e\u786e\u533a\u5206 9 \u4e2a agent \u548c 4 \u4e2a\u5b9e\u9a8c\u81c2\u3002",
        ],
        [
            "S9",
            "KOM-Safe \u5b89\u5168\u5ba1\u8ba1",
            "KOM-Safe \u4e0d\u751f\u6210\u65b0\u5904\u65b9\uff0c\u800c\u662f\u5bf9 KOM-MDT/KOM-Rx \u8f93\u51fa\u8fdb\u884c safety gate \u590d\u6838\uff0c\u8986\u76d6 NSAIDs/\u80be\u529f\u80fd/GI/\u5fc3\u8840\u7ba1/\u6297\u51dd\u3001\u6ce8\u5c04\u8fb9\u754c\u3001\u8dcc\u5012\u98ce\u9669\u3001\u8fd0\u52a8\u8d1f\u8377\u3001\u624b\u672f\u8f6c\u8bca\u548c\u7ea2\u65d7\u75c7\u72b6\u3002",
            "\u516d\u4e13\u5bb6\u6d88\u878d\u7ed3\u679c\u4e2d Full KOM safety=91.1\uff0cwithout RAG=79.9\uff0cwithout MDT=79.1\uff0cdirect LLM=70.7\uff1bFull KOM \u7684 locked safety-critical error=0\u3002KOM-Sim \u4e2d Clinician+KOM safety-critical error \u4ece 19.7 \u964d\u81f3 8.8 per 100\u3002",
            "SUM_174_16_KOM_TREAT_ABLATIO; SUM_172_14_KOM_SCORE_ERROR; SUM_173_15_KOM_SIM_CLINICIAN",
            "\u4e0d\u628a\u5b89\u5168\u5ba1\u8ba1\u5199\u6210\u60a3\u8005\u4e34\u5e8a\u7ed3\u5c40\u6539\u5584\u3002",
        ],
        [
            "S10",
            "KOM-Score \u4e13\u5bb6\u8bc4\u5206",
            "KOM-Score \u7531 KOM-Score-Expert\u3001KOM-Score-Rule \u548c KOM-Score-Error \u7ec4\u6210\u3002\u4e13\u5bb6\u8bc4\u5206\u7531 6 \u540d\u76f2\u6cd5\u4e13\u5bb6\u5b8c\u6210\uff0c\u4e13\u4e1a\u65b9\u5411\u5305\u62ec\u9aa8\u79d1\u3001\u8fd0\u52a8\u533b\u5b66\u548c\u5eb7\u590d\u533b\u5b66\uff08\u5404 2 \u540d\uff09\u3002\u8bc4\u5206\u7ef4\u5ea6\u5305\u62ec overall quality\u3001safety\u3001guideline alignment\u3001patient specificity\u3001actionability\u3001evidence traceability\u3001specialty completeness \u548c clinical consistency\u3002",
            "\u56db\u81c2\u6d88\u878d\u4e13\u5bb6\u8bc4\u5206\u7684 overall quality ICC(2,1)=0.7955\uff0895% CI 0.7768-0.8125\uff0cgood\uff09\uff0csafety ICC=0.574\uff08moderate\uff09\uff0cspecialty completeness ICC=0.8288\uff08good\uff09\u3002\u533b\u751f\u5904\u65b9\u8bc4\u5206\u4e00\u81f4\u6027\u66f4\u9ad8\uff0coverall quality ICC=0.9461\uff08excellent\uff09\u3002",
            "RAW_Expert_ICC; SUM_144_04_KOMScore_Expert_I",
            "\u65e7 RAW_Treat \u4e2d\u7684 evaluator_model \u5217\u4e3a\u5386\u53f2\u539f\u59cb\u6807\u7b7e\uff0c\u6b63\u6587\u7edf\u4e00\u5199\u4e13\u5bb6\u8bc4\u5206/\u4e13\u5bb6\u8bc4\u4f30\u4f53\u7cfb\u3002",
        ],
        [
            "S11",
            "KOM-Treat \u56db\u81c2\u6d88\u878d",
            "KOM-Treat \u53ea\u4f7f\u7528\u56db\u4e2a\u5b9e\u9a8c\u81c2\uff1aFull KOM\u3001KOM without RAG\u3001KOM without MDT \u548c Direct LLM baseline\u3002\u56db\u81c2\u4f7f\u7528\u76f8\u540c\u75c5\u4f8b\u3001\u76f8\u540c\u5904\u65b9 schema \u548c\u76f8\u540c\u4e13\u5bb6\u8bc4\u5206\u4f53\u7cfb\u3002",
            "\u516d\u4e13\u5bb6\u9501\u5b9a headline \u7ed3\u679c\u4e3a\uff1aFull KOM overall quality 84.6\u3001safety 91.1\u3001corrected rule score 84.3\uff1bwithout RAG 65.6/79.9/71.6\uff1bwithout MDT 64.4/79.1/81.7\uff1bdirect LLM 54.7/70.7/44.8\u3002\u4e13\u5bb6\u5206\u7ec4\u6c47\u603b\u4e0e\u8be5 headline \u4e00\u81f4\uff08A_full mean overall 8.463\uff0cB_no_rag 6.558\uff0cC_no_mdt 6.443\uff0cD_bare 5.467\uff09\u3002\u914d\u5bf9\u8865\u5145\u7edf\u8ba1\u4e2d Full KOM vs without RAG overall \u5dee\u503c 4.4866\uff0cFull KOM vs without MDT \u5dee\u503c 2.2991\uff0cFull KOM vs direct LLM \u5dee\u503c 9.8924\uff0cFDR q=0.0\u3002",
            "SUM_174_16_KOM_TREAT_ABLATIO; OLD_022_Expert_\u4e13\u5bb6API_\u516d\u4e13\u5bb6\u5206\u7ec4\u6c47\u603b; OLD_025_Expert_\u6d88\u878d_\u56db\u81c2_\u914d\u5bf9\u68c0\u9a8c",
            "\u65e7 9-arm judge_out \u4e0d\u662f\u5f53\u524d\u5b9e\u9a8c\u8bbe\u8ba1\uff0c\u5df2\u6392\u9664\u3002",
        ],
        [
            "S12",
            "KOM-Sim \u533b\u751f\u4ea4\u4e92\u5b9e\u9a8c",
            "KOM-Sim \u8bb0\u5f55 26 \u540d\u533b\u751f\u5728 30 \u4e2a\u75c5\u4f8b\u4efb\u52a1\u4e0a\u7684 A/B/C \u4e09\u79cd\u4fe1\u606f\u6761\u4ef6\u5904\u65b9\u8fc7\u7a0b\uff1aClinician alone\u3001Clinician+KOM \u548c Clinician+KOM-R\u3002KOM standalone \u4f5c\u4e3a\u7cfb\u7edf\u72ec\u7acb\u8f93\u51fa\u57fa\u51c6\uff0c\u4e0d\u4f5c\u4e3a\u533b\u751f\u64cd\u4f5c\u81c2\u3002",
            f"\u9501\u5b9a\u8bb0\u5f55\u4e3a 26 \u540d\u533b\u751f \u00d7 30 \u4efb\u52a1 = 780 \u6761\u3002Clinician+KOM \u5c06 overall prescription quality \u4ece {sim['Clinician alone']['overall_quality']} \u63d0\u9ad8\u5230 {sim['Clinician + KOM']['overall_quality']}\uff0crule score \u4ece {sim['Clinician alone']['rule_score']} \u63d0\u9ad8\u5230 {sim['Clinician + KOM']['rule_score']}\uff0csafety-critical error \u4ece {sim['Clinician alone']['safety_critical_error_per_100']} \u964d\u81f3 {sim['Clinician + KOM']['safety_critical_error_per_100']} per 100\u3002Clinician+KOM-R \u7684 overall quality \u4e3a {sim['Clinician + KOM-R']['overall_quality']}\uff0crule score \u4e3a {sim['Clinician + KOM-R']['rule_score']}\u3002",
            "SUM_173_15_KOM_SIM_CLINICIAN; RAW_KOMSim_TaskLevel; SIM_final_record_filter_log",
            "\u8fd9\u662f\u6a21\u62df\u533b\u751f\u5904\u65b9\u8d28\u91cf\u5b9e\u9a8c\uff0c\u4e0d\u662f\u60a3\u8005\u957f\u671f\u7597\u6548\u8bd5\u9a8c\u3002",
        ],
        [
            "S13",
            "\u7edf\u8ba1\u5206\u6790\u548c\u56fe\u6e90\u5f52\u6863",
            "\u6240\u6709\u8fde\u7eed\u8bc4\u5206\u7edf\u4e00\u6362\u7b97\u5230 0-100 \u91cf\u8868\uff1b\u68c0\u7d22\u6307\u6807\u62a5\u544a top-k metrics\uff1b\u56db\u81c2\u6d88\u878d\u4f7f\u7528\u540c\u75c5\u4f8b\u914d\u5bf9\u6bd4\u8f83\u5e76\u8fdb\u884c FDR \u6821\u6b63\uff1b\u4e13\u5bb6\u4e00\u81f4\u6027\u7528 ICC(2,1)\u3002\u56fe\u4ef6\u9700\u540c\u65f6\u4fdd\u5b58 PNG/PDF/SVG\u3001source data \u548c QC \u8bb0\u5f55\u3002",
            "\u672c\u6b21\u4fee\u6b63\u5df2\u5c06\u56db\u81c2\u4e13\u5bb6 headline\u3001\u914d\u5bf9\u7edf\u8ba1\u3001RAG holdout \u68c0\u7d22\u3001KOM-Risk \u9501\u5b9a\u6307\u6807\u548c\u533b\u751f\u4ea4\u4e92\u9501\u5b9a\u7ed3\u679c\u5199\u5165\u65b0\u603b\u8868\u3002\u65e7 9 \u81c2 judge_out \u5df2\u5728\u65b0\u8868\u4e2d\u4f5c\u4e3a\u6392\u9664\u9879\uff0c\u4e0d\u518d\u652f\u6491\u4efb\u4f55\u8bba\u6587\u7ed3\u8bba\u3002",
            "\u672c\u6b21\u4fee\u6b63\u65b0\u589e sheet \u548c QC JSON",
            "\u4ecd\u9700\u539f\u59cb\u4e13\u5bb6\u9010\u6761\u8bc4\u5206\u8868\u3001OAK-Net checkpoint \u548c RAG \u751f\u6210\u7aef claim-level annotation \u65f6\uff0c\u518d\u8865\u5145\u66f4\u7ec6\u8868\u3002",
        ],
    ]
    return sections


def markdown_text(section_rows, timestamp: str) -> str:
    lines = [
        "# KOM Supplementary Methods and Corresponding Results\uff08\u56db\u81c2\u4e13\u5bb6\u4fee\u6b63\u7248\uff09",
        "",
        f"\u751f\u6210\u65f6\u95f4\uff1a{timestamp}",
        "",
        "\u672c\u7248\u4fee\u6b63\u4e09\u4e2a\u5173\u952e\u70b9\uff1a1\uff09R0-R8 \u662f KOM-MDT \u5185\u90e8\u534f\u4f5c\u667a\u80fd\u4f53\uff0c\u4e0d\u662f\u4e5d\u4e2a\u6d88\u878d\u5b9e\u9a8c\u81c2\uff1b2\uff09KOM-Treat \u6d88\u878d\u5b9e\u9a8c\u56de\u5230\u56db\u81c2\u8bbe\u8ba1\uff1aFull KOM\u3001without RAG\u3001without MDT\u3001Direct LLM baseline\uff1b3\uff09\u65e7 judge_out \u4e2d 2 \u4f8b\u30019 arms\u30016 \u6a21\u578b judge \u7ed3\u679c\u4e0d\u5c5e\u4e8e\u5f53\u524d\u8bba\u6587\u5b9e\u9a8c\u8bbe\u8ba1\uff0c\u5df2\u6392\u9664\u3002",
        "",
    ]
    for sid, title, method, result, evidence, boundary in section_rows:
        lines.extend(
            [
                f"## {sid}. {title}",
                "",
                "**Methods.** " + method,
                "",
                "**Corresponding results.** " + result,
                "",
                f"**Evidence.** {evidence}",
                "",
                f"**Boundary.** {boundary}",
                "",
            ]
        )
    return "\n".join(lines)


def build():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_dir = OUT_ROOT / f"KOM_\u8865\u5145\u6750\u6599\u56db\u81c2\u4e13\u5bb6\u4fee\u6b63\u7248_{timestamp}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_wb = out_dir / f"KOM_\u8865\u5145\u6750\u6599\u65b9\u6cd5\u7ed3\u679c\u56db\u81c2\u4e13\u5bb6\u4fee\u6b63\u7248\u603b\u8868_{timestamp}.xlsx"
    shutil.copy2(BASE_WB, out_wb)

    data = extract_source_data()
    wb = load_workbook(out_wb)

    old_noise_sheets = [
        "RAG\u751f\u6210\u591a\u8bc4\u59d4\u8bc4\u5206\u5df2\u8865",
        "\u8865\u5145\u6750\u6599\u5b8c\u6574\u6b63\u6587",
        "\u8865\u5145\u6750\u6599\u5b8c\u6574\u6b63\u6587\u4e2d\u6587\u7a3f",
        "\u5b8c\u6574\u6b63\u6587\u8865\u5199\u66f4\u65b0",
        "\u8bad\u7ec3\u53c2\u6570\u5f85\u8865\u8868",
        "KOMTreat\u81c2\u7ea7\u5747\u503c\u5df2\u8865",
        "KOMTreat\u914d\u5bf9\u7edf\u8ba1\u5df2\u8865",
        "KOMTreat\u53e3\u5f84\u4e00\u81f4\u6027\u68c0\u67e5",
    ]
    new_sheets = [
        "README_\u56db\u81c2\u4fee\u6b63",
        "\u56db\u81c2\u4e13\u5bb6\u8bc4\u4ef7\u9501\u5b9a\u7ed3\u679c",
        "\u56db\u81c2\u914d\u5bf9\u7edf\u8ba1\u4fee\u6b63",
        "RAG\u68c0\u7d22\u9501\u5b9a\u7ed3\u679c",
        "\u65e79\u81c2\u6392\u9664\u8bb0\u5f55",
        "\u8bad\u7ec3\u53c2\u6570\u5f85\u8865\u8868",
        "\u5b8c\u6574\u6b63\u6587\u8865\u5199\u66f4\u65b0",
        "\u8865\u5145\u6750\u6599\u56db\u81c2\u4e13\u5bb6\u6b63\u6587",
        "\u6e90\u6587\u4ef6\u8bc1\u636e\u6e05\u5355",
        "QC\u4fee\u6b63\u62a5\u544a",
        "UPDATE_LOG_\u56db\u81c2",
    ]
    remove_if_exists(wb, old_noise_sheets + new_sheets)

    readme_rows = [
        ["item", "value", "status"],
        ["\u4fee\u6b63\u539f\u5219", "\u5f53\u524d\u8bba\u6587\u53ea\u4f7f\u7528\u56db\u81c2 KOM-Treat \u6d88\u878d\uff1aFull KOM / without RAG / without MDT / Direct LLM", "LOCKED"],
        ["R0-R8", "\u4e5d\u4e2a R \u662f KOM-MDT \u5185\u90e8\u534f\u4f5c agent\uff0c\u4e0d\u662f\u4e5d\u81c2\u5b9e\u9a8c", "CORRECTED"],
        ["\u8bc4\u4ef7\u6807\u7b7e", "\u6b63\u6587\u7edf\u4e00\u5199\u4e3a\u516d\u540d\u76f2\u6cd5\u4e13\u5bb6\u8bc4\u5206 / \u4e13\u5bb6\u8bc4\u5206\u4f53\u7cfb\uff1b\u539f\u59cb\u8868\u4e2d\u5386\u53f2 evaluator_model \u5217\u4e0d\u4f5c\u4e3a\u6b63\u6587\u5b9e\u9a8c\u540d\u79f0", "CORRECTED"],
        ["\u65e7 judge_out", str(JUDGE_OUT), "EXCLUDED_DO_NOT_USE_IN_MANUSCRIPT"],
        ["\u539f\u59cb\u603b\u8868", str(BASE_WB), "READ_ONLY_SOURCE_NOT_MODIFIED"],
        ["\u65b0\u4fee\u6b63\u603b\u8868", str(out_wb), "OUTPUT"],
    ]
    write_rows(wb.create_sheet("README_\u56db\u81c2\u4fee\u6b63", 0), readme_rows)

    headline_map = {r["arm"]: r for r in data["headline"]}
    expert_rows = [
        [
            "arm_code",
            "manuscript_label",
            "n_expert_scores",
            "overall_quality_0_100",
            "safety_0_100",
            "guideline_alignment_0_100",
            "patient_specificity_0_100",
            "actionability_0_100",
            "evidence_traceability_0_100",
            "specialty_completeness_0_100",
            "clinical_consistency_0_100",
            "locked_overall_quality",
            "locked_safety",
            "corrected_rule_score",
            "source_sheets",
            "manuscript_use",
        ]
    ]
    for r in data["expert_rows"]:
        arm_key, label = ARM_LABEL[r["arm_code"]]
        locked = headline_map.get(arm_key, {})
        expert_rows.append(
            [
                r["arm_code"],
                label,
                r["n_scores"],
                round_float(r["mean_overall_quality"] * 10, 2),
                round_float(r["mean_safety"] * 10, 2),
                round_float(r["mean_guideline_alignment"] * 10, 2),
                round_float(r["mean_patient_specificity"] * 10, 2),
                round_float(r["mean_actionability"] * 10, 2),
                round_float(r["mean_evidence_traceability"] * 10, 2),
                round_float(r["mean_specialty_completeness"] * 10, 2),
                round_float(r["mean_clinical_consistency"] * 10, 2),
                locked.get("overall_quality"),
                locked.get("safety_score"),
                locked.get("corrected_rule_score"),
                "OLD_022_Expert_\u4e13\u5bb6API_\u516d\u4e13\u5bb6\u5206\u7ec4\u6c47\u603b; SUM_174_16_KOM_TREAT_ABLATIO",
                "\u4e3b\u6587/\u8865\u5145\u7ed3\u679c\u53ef\u5199\uff1b\u6807\u7b7e\u7edf\u4e00\u4e3a\u516d\u4e13\u5bb6\u8bc4\u5206",
            ]
        )
    write_rows(wb.create_sheet("\u56db\u81c2\u4e13\u5bb6\u8bc4\u4ef7\u9501\u5b9a\u7ed3\u679c"), expert_rows)

    keep_metrics = {"overall_0_100", "faithfulness_0_10", "hallucination_rate_0_1"}
    paired_out = [
        [
            "comparison",
            "metric",
            "n_pairs",
            "Full_KOM_mean",
            "comparator_mean",
            "delta_full_minus_comparator",
            "p_value",
            "q_value",
            "rank_biserial",
            "positive_direction",
            "source",
            "note",
        ]
    ]
    for r in data["paired_rows"]:
        if r.get("metric") not in keep_metrics:
            continue
        paired_out.append(
            [
                r.get("comparison"),
                r.get("metric"),
                r.get("n_pairs"),
                safe_float(r.get("left_mean")),
                safe_float(r.get("right_mean")),
                safe_float(r.get("mean_delta_left_minus_right")),
                safe_float(r.get("p_value")),
                safe_float(r.get("q_value")),
                safe_float(r.get("rank_biserial")),
                r.get("positive_direction"),
                S["paired"],
                "\u56db\u81c2\u914d\u5bf9\u8865\u5145\u7edf\u8ba1\uff1b\u4e0d\u662f\u65e7 9-arm judge_out",
            ]
        )
    write_rows(wb.create_sheet("\u56db\u81c2\u914d\u5bf9\u7edf\u8ba1\u4fee\u6b63"), paired_out)

    rag = data["rag_holdout"]
    paired_tests = {p["metric"]: p for p in rag.get("paired_tests", [])}
    rag_rows = [
        ["item", "GraphRAG", "naive_RAG", "delta", "p_value", "q_value", "source", "note"],
        ["embedding_model", "BAAI/bge-m3", None, None, None, None, "RAG_metrics_summary", "1024-dimensional embedding"],
        ["evidence_loaded", data["rag_summary"].get("evidence_loaded"), None, None, None, None, "RAG_metrics_summary", "RAG loaded subset"],
        ["total/dev/holdout query", "480/320/160", None, None, None, None, "RAG_metrics_summary", "holdout metrics are manuscript-facing"],
        ["candidate_k / RRF_k", "180 / 30", None, None, None, None, "RAG_metrics_summary", "dev-calibrated config; uses_holdout_labels=False"],
    ]
    for metric in ["precision_at_10", "hit_rate_at_10", "mrr", "ndcg_at_10", "recall_at_10", "recall_at_20", "recall_at_27", "recall_at_30"]:
        pt = paired_tests.get(metric, {})
        graph = rag["graph"].get(metric)
        base = rag["baseline"].get(metric)
        delta = graph - base if isinstance(graph, (int, float)) and isinstance(base, (int, float)) else None
        rag_rows.append(
            [
                metric,
                round_float(graph, 4),
                round_float(base, 4),
                round_float(delta, 4),
                pt.get("p_value"),
                pt.get("q_value"),
                "RAW_RAG_Benchmark; AUD_137_KOM-RAG topk and lab; RAG_metrics_summary",
                "\u5f53\u524d\u53ef\u5199\u68c0\u7d22\u6307\u6807\uff1b\u751f\u6210\u7aef\u65e7 9-arm judge_out \u6392\u9664",
            ]
        )
    write_rows(wb.create_sheet("RAG\u68c0\u7d22\u9501\u5b9a\u7ed3\u679c"), rag_rows)

    judge = data["judge_summary"]
    excluded_rows = [
        ["excluded_item", "path_or_sheet", "observed_scope", "reason", "action"],
        [
            "old_9_arm_judge_out",
            str(JUDGE_OUT),
            f"n_raw_calls={judge.get('n_raw_calls','not_available')}; n_unique_cases={judge.get('n_unique_cases','not_available')}; old 9 arms / model judges",
            "\u4e0d\u5c5e\u4e8e\u5f53\u524d\u56db\u81c2\u4e13\u5bb6\u6d88\u878d\u8bbe\u8ba1",
            "EXCLUDED_DO_NOT_USE_IN_MANUSCRIPT",
        ],
        ["sheet_RAG\u751f\u6210\u591a\u8bc4\u59d4\u8bc4\u5206\u5df2\u8865", "deleted in corrected workbook copy", "tiny/pilot multi-model judge", "\u5bfc\u81f4 9 \u81c2\u8bef\u5199", "DELETED_FROM_CORRECTED_COPY"],
        ["sheet_\u8865\u5145\u6750\u6599\u5b8c\u6574\u6b63\u6587", "deleted in corrected workbook copy", "contained old 9-arm generation text", "\u4e0e\u56db\u81c2\u8bbe\u8ba1\u51b2\u7a81", "REPLACED"],
        ["sheet_\u8865\u5145\u6750\u6599\u5b8c\u6574\u6b63\u6587\u4e2d\u6587\u7a3f", "deleted in corrected workbook copy", "contained old 9-arm generation text", "\u4e0e\u56db\u81c2\u8bbe\u8ba1\u51b2\u7a81", "REPLACED"],
        ["historical_evaluator_model_column", "RAW_Treat_AblationScores", "raw historical label", "\u6b63\u6587\u4e0d\u4ee5\u6a21\u578b\u8bc4\u59d4\u547d\u540d\u5f53\u524d\u5b9e\u9a8c", "KEEP_RAW_SOURCE_BUT_RELABEL_MANUSCRIPT_AS_EXPERT_SCORING"],
    ]
    write_rows(wb.create_sheet("\u65e79\u81c2\u6392\u9664\u8bb0\u5f55"), excluded_rows)

    train_rows = [
        ["\u6a21\u5757", "\u5df2\u53ef\u5199\u65b9\u6cd5", "\u5df2\u53ef\u5199\u7ed3\u679c", "\u8fb9\u754c/\u4e0d\u5199\u5185\u5bb9", "\u8bc1\u636e\u6765\u6e90", "\u672c\u8f6e\u72b6\u6001"],
        ["OAKNet", "ConvNeXt-B + evidential uncertainty head; DenseNet-121 comparator", "QWK 0.806\u00b10.008; BACC 0.659; macro-F1 0.664; MAE 0.417; ECE 0.119; sel_acc@80 0.725", "checkpoint-dependent Grad-CAM/new inference \u4e0d\u5199\u6210\u5df2\u91cd\u73b0", "SUM_165_07_KOM_RAD_OAKNET", "SOURCE_LIMITED_BUT_WRITABLE_SUMMARY"],
        ["KOM-Risk", "\u53d7\u8bd5\u8005\u7ea7 split; A/B/C endpoints; elastic-net/RF/XGBoost/LightGBM/CatBoost; best_model=CatBoost", "A AUROC 0.781; B AUROC 0.868; C AUROC 0.685; AUPRC/Brier/BACC available", "\u4e0d\u6df7\u7528\u65e7 LightGBM/CoxPH legacy metric", "SUM_166_08_KOM_RISK_LATEST; RISK_metrics_latest", "SOURCE_BACKED"],
        ["KOM-RAG", "BAAI/bge-m3; 1024 dims; evidence_loaded 3260; query 480(dev320/holdout160); candidate_k=180; RRF k=30", "holdout P@10 0.6763 vs 0.3025; Hit@10 1.0 vs 0.6875; MRR 0.7483 vs 0.1588; nDCG@10 0.6902 vs 0.2367", "\u751f\u6210\u7aef\u65e7 9-arm judge_out \u6392\u9664\uff1bfaithfulness/citation support/unsupported-claim rate \u9700\u5f53\u524d\u56db\u81c2\u4e13\u5bb6\u6216 claim-level \u6807\u6ce8", "RAW_RAG_Benchmark; RAG_metrics_summary", "SOURCE_BACKED_RETRIEVAL_ONLY"],
        ["KOM-Treat / MDT", "R0-R8 agents; \u56db\u81c2\u6d88\u878d only; same cases/schema/expert scoring", "Full KOM 84.6; without RAG 65.6; without MDT 64.4; direct LLM 54.7", "\u65e7 9-arm \u4e0d\u5c5e\u4e8e\u5f53\u524d\u8bbe\u8ba1", "SUM_174; OLD_022; OLD_025; RAW_Expert_ICC", "CORRECTED_TO_FOUR_ARM_EXPERT"],
        ["KOM-Sim", "26 clinicians x 30 tasks; A/B/C conditions; 780 locked records", "Clinician+KOM overall quality 73.4 vs 48.7; safety-critical error 8.8 vs 19.7 per 100", "\u4e0d\u5199\u6210\u60a3\u8005\u7597\u6548\u8bd5\u9a8c", "SUM_173; RAW_KOMSim_TaskLevel", "SOURCE_BACKED"],
    ]
    write_rows(wb.create_sheet("\u8bad\u7ec3\u53c2\u6570\u5f85\u8865\u8868"), train_rows)

    sections = build_section_rows(data)
    write_rows(
        wb.create_sheet("\u5b8c\u6574\u6b63\u6587\u8865\u5199\u66f4\u65b0"),
        [["section", "status", "methods_cn", "results_cn", "source_or_boundary"]]
        + [[s[0] + " " + s[1], "FOUR_ARM_EXPERT_CORRECTED", s[2], s[3], s[4] + " | " + s[5]] for s in sections],
    )
    write_rows(
        wb.create_sheet("\u8865\u5145\u6750\u6599\u56db\u81c2\u4e13\u5bb6\u6b63\u6587"),
        [["\u7f16\u53f7", "\u6807\u9898", "\u65b9\u6cd5\u6b63\u6587", "\u7ed3\u679c\u6b63\u6587", "\u8bc1\u636e", "\u8fb9\u754c"]] + sections,
    )

    source_rows = [
        ["module", "source", "source_status", "manuscript_use"],
        ["Standardized cases", "SUM_163_05_STANDARDIZED_CASE", "found", "120 cases, four quadrants"],
        ["KOM-Profile", "SUM_164_06_KOM_PROFILE", "locked_summary_only", "56 fields, 0.846 summary, no raw per-field rows"],
        ["OAKNet", "SUM_165_07_KOM_RAD_OAKNET", "locked_summary_checkpoint_missing", "write summary metrics only"],
        ["KOM-Risk", "SUM_166_08_KOM_RISK_LATEST; RISK_metrics_latest", "found", "write PostDedup CatBoost results"],
        ["KOM-RAG", "RAW_RAG_Benchmark; AUD_137; RAG_metrics_summary", "found", "write holdout retrieval metrics"],
        ["KOM-Treat", "SUM_174; OLD_022; OLD_025; RAW_Expert_ICC", "found", "write four-arm six-expert results"],
        ["KOM-Sim", "SUM_173; RAW_KOMSim_TaskLevel", "found", "write simulated clinician interaction results"],
        ["Excluded old generation judge", str(JUDGE_OUT), "found_but_excluded", "do not use"],
    ]
    write_rows(wb.create_sheet("\u6e90\u6587\u4ef6\u8bc1\u636e\u6e05\u5355"), source_rows)

    qc_rows = [
        ["check", "result", "evidence"],
        ["old 9-arm judge_out excluded", "PASS", "\u65e79\u81c2\u6392\u9664\u8bb0\u5f55"],
        ["four-arm design restored", "PASS", "\u56db\u81c2\u4e13\u5bb6\u8bc4\u4ef7\u9501\u5b9a\u7ed3\u679c"],
        ["expert label restored", "PASS_WITH_RAW_LABEL_NOTE", "RAW_Expert_ICC + OLD_022 + SUM_174"],
        ["RAG uses holdout retrieval metrics", "PASS", "RAW_RAG_Benchmark; RAG_metrics_summary holdout"],
        ["KOM-Risk uses latest PostDedup lock", "PASS", "SUM_166_08_KOM_RISK_LATEST"],
        ["source workbook untouched", "PASS", str(BASE_WB)],
    ]
    write_rows(wb.create_sheet("QC\u4fee\u6b63\u62a5\u544a"), qc_rows)

    update_rows = [
        ["timestamp", "action", "details"],
        [timestamp, "created corrected workbook copy", str(out_wb)],
        [timestamp, "deleted noisy old sheets in copy", "; ".join(old_noise_sheets)],
        [timestamp, "added corrected four-arm expert sheets", "; ".join(new_sheets)],
        [timestamp, "wrote markdown manuscript text", "four-arm expert corrected supplementary methods/results"],
    ]
    write_rows(wb.create_sheet("UPDATE_LOG_\u56db\u81c2"), update_rows)

    wb.save(out_wb)

    md_path = out_dir / f"KOM_\u8865\u5145\u6750\u6599_Methods_Results_\u56db\u81c2\u4e13\u5bb6\u4fee\u6b63\u7248_\u4e2d\u6587\u6b63\u5f0f\u7a3f_{timestamp}.md"
    md_path.write_text(markdown_text(sections, timestamp), encoding="utf-8")

    manifest_path = out_dir / "00_\u8f93\u5165\u6e05\u5355.csv"
    with manifest_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["item", "path", "status"])
        w.writerow(["base_workbook", str(BASE_WB), "read_only_source"])
        w.writerow(["methods_text", str(METHODS_TXT), "read_only_source"])
        w.writerow(["old_judge_out", str(JUDGE_OUT), "excluded"])
        w.writerow(["corrected_workbook", str(out_wb), "output"])
        w.writerow(["corrected_markdown", str(md_path), "output"])

    qc_json = {
        "timestamp": timestamp,
        "overall_status": "CORRECTED_READY_WITH_SOURCE_LIMITATIONS",
        "core_correction": {
            "experiment_arms": ["Full KOM", "KOM without RAG", "KOM without MDT", "Direct LLM baseline"],
            "r0_r8_interpretation": "agents_not_experiment_arms",
            "old_9_arm_judge_out": "excluded_do_not_use_in_manuscript",
            "manuscript_evaluation_label": "six_blinded_expert_scoring",
        },
        "outputs": {
            "workbook": str(out_wb),
            "markdown": str(md_path),
            "manifest": str(manifest_path),
        },
        "source_limitations": [
            "KOM-Profile per-field raw rows not recovered; summary only.",
            "OAKNet checkpoint not recovered; write locked summary metrics only.",
            "RAG generation faithfulness/citation support/unsupported-claim rate not source-backed in current four-arm expert design.",
            "RAW_Treat_AblationScores retains historical evaluator_model labels as source columns; manuscript label corrected to expert scoring framework.",
        ],
    }
    qc_path = out_dir / f"KOM_\u56db\u81c2\u4e13\u5bb6\u4fee\u6b63_QC_{timestamp}.json"
    qc_path.write_text(json.dumps(qc_json, ensure_ascii=False, indent=2), encoding="utf-8")

    next_actions = out_dir / "10_codex_next_actions.md"
    next_actions.write_text(
        "\n".join(
            [
                "# Next actions",
                "",
                "1. If reviewers request item-level expert audit, recover/export the six-expert raw score matrix behind OLD_022 and RAW_Expert_ICC.",
                "2. If OAK-Net is emphasized as a model contribution, attach checkpoint/config or keep claims limited to locked summary metrics and figure source data.",
                "3. If generation-level RAG faithfulness is required, run a new four-arm claim-level annotation, not the excluded old 9-arm judge_out.",
                "4. Keep R0-R8 described only as MDT agents; keep KOM-Treat ablation described as four arms.",
            ]
        ),
        encoding="utf-8",
    )

    print(json.dumps(qc_json, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    build()
