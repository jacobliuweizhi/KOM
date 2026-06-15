from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
import os
import re
import json
import shutil
import hashlib
import math
from datetime import datetime
from collections import Counter

import pandas as pd
import openpyxl


ROOT = Path("C:/OAI" + "\u7814\u7a76\u9879\u76ee" + "/pythonProject1/" + "KOM" + "\u8fd4\u4fee\u4fee\u6539")
SUBMISSION = ROOT / "\u6295\u7a3f\u4f7f\u7528"
FINAL = SUBMISSION / "\u6700\u7ec8\u7248\u672c"
LOCAL = ROOT / "\u672c\u5730\u5316" / "koa_mdt_agents"
OUT = ROOT / "KOM_Submission_Audit_Package_202606"
ZIP_OUT = ROOT / "KOM_Submission_Audit_Package_202606.zip"
NOW = datetime.now().strftime("%Y%m%d_%H%M%S")

SOURCE_ZIPS = [
    SUBMISSION / "KOM_All_Data_and_Prediction_Model_Package_202606.zip",
    SUBMISSION / "KOM_Data_Figures_Editable_Package_v2.zip",
    SUBMISSION / "OAKNet_Curves_Replot_202606_complete.zip",
]
STANDALONE_ASSETS = [
    FINAL / ("KOM_" + "\u4e3b\u6587\u603b\u8868_\u7ecf\u9a8c\u83b7\u76ca\u66f2\u7ebf_\u4f18\u5316\u586b\u5145\u7248.xlsx"),
    FINAL / ("KOM_" + "\u7cfb\u7edf\u8bc1\u636e\u77e9\u9635_\u6807\u51c6\u547d\u540d\u7248.xlsx"),
    SUBMISSION / ("KOM_" + "\u7eaf\u51c0\u7248_\u516d\u4e13\u5bb6API\u56de\u586b\u7248_\u6574\u7406\u7248.xlsx"),
]
STAGE4A = LOCAL / "data" / "processed" / "stage_validation" / "stage4a_retrieval_metrics.json"
STAGE3 = LOCAL / "data" / "processed" / "stage_validation" / "stage3_retrieval_dod.json"


def file_hash(path: Path, max_bytes: int | None = None) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        if max_bytes:
            h.update(f.read(max_bytes))
        else:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
    return h.hexdigest()


def write_csv(path: Path, rows, columns=None) -> pd.DataFrame:
    path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    if columns:
        for c in columns:
            if c not in df.columns:
                df[c] = pd.NA
        df = df[columns]
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return df


def write_md(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def read_csv_safe(path: Path) -> pd.DataFrame | None:
    for enc in ["utf-8-sig", "utf-8", "gbk", "latin1"]:
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception:
            pass
    return None


def read_excel_sheet_safe(path: Path, sheet: str) -> pd.DataFrame | None:
    try:
        return pd.read_excel(path, sheet_name=sheet)
    except Exception:
        return None


def rel_to_out(path: Path) -> str:
    try:
        return str(path.relative_to(OUT))
    except Exception:
        return ""


def category_for(path: Path) -> str:
    s = str(path).lower()
    if any(k in s for k in ["komsim", "hci", "hai", "time_workload", "experience", "physician", "workload", "interaction", "materialviews", "adoption"]):
        return "KOM-Sim/HCI"
    if any(k in s for k in ["risk", "oaknet", "prediction", "cox", "lightgbm", "catboost", "calibration", "shap", "feature_importance", "reliability"]):
        return "KOM-Risk/OAKNet"
    if any(k in s for k in ["rag", "retrieval", "evidence", "guideline", "stage4a", "ret_", "graphrag"]):
        return "KOM-RAG/Evidence"
    if path.suffix.lower() in [".xlsx", ".xls", ".csv", ".json", ".md", ".docx", ".pptx", ".png", ".svg", ".pdf", ".zip"]:
        return "KOM source asset"
    return "other"


def source_archive_for(path: Path) -> str:
    parts = path.parts
    try:
        i = parts.index("extracted_archives")
        return parts[i + 1]
    except Exception:
        pass
    try:
        i = parts.index("x")
        return parts[i + 1]
    except Exception:
        pass
    if path in SOURCE_ZIPS:
        return "original_zip"
    if path in STANDALONE_ASSETS:
        return "standalone_key_asset"
    if str(path).startswith(str(LOCAL)):
        return "local_koa_mdt_agents"
    return "project_source"


def first_existing_col(df: pd.DataFrame, terms: list[str]):
    for c in df.columns:
        lc = str(c).lower()
        if any(t in lc for t in terms):
            return c
    return None


def scalar_from_row(row, col):
    if col is None:
        return pd.NA
    try:
        return row.get(col, pd.NA)
    except Exception:
        return pd.NA


def ensure_dirs() -> None:
    if OUT.exists():
        backup = ROOT / f"KOM_Submission_Audit_Package_202606_backup_{NOW}"
        shutil.move(str(OUT), str(backup))
        print("Backed up existing output dir:", backup)
    OUT.mkdir(parents=True, exist_ok=True)
    for d in [
        "00_inventory",
        "01_KOMSim_logs_and_time_definition",
        "02_KOMRisk_reproducible_prediction_package",
        "03_KOMRAG_query_level_evidence_mapping",
        "04_crosscheck_reports",
        "05_submission_ready_tables",
    ]:
        (OUT / d).mkdir(parents=True, exist_ok=True)


def safe_name(text: str, max_len: int = 90) -> str:
    text = re.sub(r"[^A-Za-z0-9_.\-]+", "_", text)
    text = text.strip("._")
    return (text[:max_len] or "file")


def extract_zip_flat(zip_path: Path, dest: Path) -> list[dict]:
    """Extract every file from a zip with flattened unique names.

    The source packages contain deeply nested paths. Preserving the original
    hierarchy exceeds Windows MAX_PATH under the required output root, so the
    manifest stores the original internal path for reviewer traceability.
    """
    dest.mkdir(parents=True, exist_ok=True)
    rows = []
    with ZipFile(zip_path) as z:
        for idx, info in enumerate(z.infolist(), start=1):
            if info.is_dir():
                continue
            internal = info.filename
            base = safe_name(Path(internal).name)
            stem = Path(base).stem[:60]
            suffix = Path(base).suffix[:12]
            digest = hashlib.sha1(internal.encode("utf-8", errors="ignore")).hexdigest()[:10]
            out_name = f"{idx:05d}_{stem}_{digest}{suffix}"
            out_path = dest / out_name
            with z.open(info, "r") as src, out_path.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            rows.append(
                {
                    "source_zip": str(zip_path),
                    "zip_internal_path": internal,
                    "extracted_path": str(out_path),
                    "file_size_bytes": out_path.stat().st_size,
                    "zip_compress_size": info.compress_size,
                    "zip_file_size": info.file_size,
                }
            )
    return rows


def export_matching_sheets(workbook_paths: list[Path], terms: list[str], out_dir: Path, prefix: str):
    exported = []
    for wb_path in workbook_paths:
        try:
            wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
        except Exception:
            continue
        for sh in names:
            low = (sh + " " + wb_path.name).lower()
            if any(t.lower() in low for t in terms):
                df = read_excel_sheet_safe(wb_path, sh)
                if df is not None:
                    safe = re.sub(r"[^A-Za-z0-9_\-]+", "_", sh)[:45]
                    out = out_dir / f"{prefix}_{safe}.csv"
                    df.to_csv(out, index=False, encoding="utf-8-sig")
                    exported.append(
                        {
                            "source_workbook": str(wb_path),
                            "source_sheet": sh,
                            "rows": len(df),
                            "columns": len(df.columns),
                            "exported_csv": str(out),
                        }
                    )
    return exported


def main() -> None:
    ensure_dirs()
    # Keep this short to avoid Windows MAX_PATH failures inside nested archives.
    extract_root = OUT / "00_inventory" / "x"
    extract_root.mkdir(parents=True, exist_ok=True)

    asset_rows = []
    for p in SOURCE_ZIPS + STANDALONE_ASSETS + [STAGE4A, STAGE3]:
        asset_rows.append(
            {
                "path": str(p),
                "exists": p.exists(),
                "size_bytes": p.stat().st_size if p.exists() else None,
                "sha256_first_1mb": file_hash(p, max_bytes=1024 * 1024) if p.exists() else None,
            }
        )
    write_csv(OUT / "00_inventory" / "source_asset_presence.csv", asset_rows, ["path", "exists", "size_bytes", "sha256_first_1mb"])

    short_zip_dirs = {
        "KOM_All_Data_and_Prediction_Model_Package_202606": "all",
        "KOM_Data_Figures_Editable_Package_v2": "fig",
        "OAKNet_Curves_Replot_202606_complete": "oak",
    }
    extraction_rows = []
    for zp in SOURCE_ZIPS:
        if not zp.exists():
            continue
        dest = extract_root / short_zip_dirs.get(zp.stem, zp.stem[:12])
        extraction_rows.extend(extract_zip_flat(zp, dest))
    write_csv(OUT / "00_inventory" / "extraction_manifest.csv", extraction_rows, ["source_zip", "zip_internal_path", "extracted_path", "file_size_bytes", "zip_compress_size", "zip_file_size"])

    for p in STANDALONE_ASSETS:
        if p.exists():
            shutil.copy2(p, OUT / "05_submission_ready_tables" / p.name)

    files = []
    scan_roots = [extract_root, OUT / "05_submission_ready_tables", LOCAL / "data" / "processed", SUBMISSION]
    exclude_dirs = {".venv", "venv", "node_modules", "__pycache__", ".git", "site-packages", "dist", "build"}
    keyword_re = re.compile(r"(KOM|Sim|HCI|log|RAG|evidence|risk|prediction|cox|lightgbm|catboost|calibration|shap|stage4a|OAKNet|guideline|prescription|intervention)", re.I)
    seen = set()
    for sr in scan_roots:
        if not sr.exists():
            continue
        for dirpath, dirnames, filenames in os.walk(sr):
            dirnames[:] = [d for d in dirnames if d not in exclude_dirs and not d.startswith(".")]
            dp = Path(dirpath)
            for fn in filenames:
                p = dp / fn
                if p in seen:
                    continue
                seen.add(p)
                if sr in [SUBMISSION, LOCAL / "data" / "processed"] and not keyword_re.search(str(p)):
                    continue
                try:
                    size = p.stat().st_size
                except Exception:
                    size = None
                files.append(
                    {
                        "absolute_path": str(p),
                        "relative_path": rel_to_out(p),
                        "file_name": p.name,
                        "extension": p.suffix.lower(),
                        "file_size_bytes": size,
                        "source_archive": source_archive_for(p),
                        "detected_category": category_for(p),
                    }
                )
    files_sorted = sorted(files, key=lambda r: (r["detected_category"], r["file_name"], r["absolute_path"]))
    write_csv(OUT / "00_inventory" / "all_files_inventory.csv", files_sorted, ["absolute_path", "relative_path", "file_name", "extension", "file_size_bytes", "source_archive", "detected_category"])

    cat_counts = Counter(r["detected_category"] for r in files_sorted)
    ext_counts = Counter(r["extension"] for r in files_sorted)
    key_text = ["# Key asset summary\n\n", f"Generated: {datetime.now().isoformat(timespec='seconds')}\n\n", "## Source assets\n"]
    for r in asset_rows:
        status = "FOUND" if r["exists"] else "NOT FOUND"
        key_text.append(f"- {status}: `{r['path']}` ({r['size_bytes'] or 'NA'} bytes)\n")
    key_text.append("\n## Detected file categories\n")
    for k, v in cat_counts.items():
        key_text.append(f"- {k}: {v}\n")
    key_text.append("\n## File extensions\n")
    for k, v in ext_counts.items():
        key_text.append(f"- {k or '[none]'}: {v}\n")
    write_md(OUT / "00_inventory" / "key_asset_summary.md", "".join(key_text))

    workbook_paths = [p for p in STANDALONE_ASSETS if p.exists()]
    for r in files_sorted:
        p = Path(r["absolute_path"])
        if p.suffix.lower() == ".xlsx" and ("KOM_" in p.name or "KOM" in str(p)):
            if p not in workbook_paths:
                workbook_paths.append(p)

    sheet_rows = []
    for wb_path in workbook_paths:
        try:
            wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
            for sh in wb.sheetnames:
                ws = wb[sh]
                sheet_rows.append({"workbook": str(wb_path), "sheet": sh, "max_row": ws.max_row, "max_column": ws.max_column})
            wb.close()
        except Exception as e:
            sheet_rows.append({"workbook": str(wb_path), "sheet": "[ERROR]", "max_row": None, "max_column": None, "error": str(e)})
    write_csv(OUT / "00_inventory" / "workbook_sheet_inventory.csv", sheet_rows, ["workbook", "sheet", "max_row", "max_column", "error"])

    sim_sheet_exports = export_matching_sheets(workbook_paths, ["KOMSim", "HAI", "human", "physician", "time", "workload", "experience"], OUT / "01_KOMSim_logs_and_time_definition", "sheet")
    risk_sheet_exports = export_matching_sheets(workbook_paths, ["RISK", "OAKNet", "prediction", "model", "cox", "lightgbm", "catboost"], OUT / "02_KOMRisk_reproducible_prediction_package", "sheet")
    rag_sheet_exports = export_matching_sheets(workbook_paths, ["RAG", "RET", "retrieval", "evidence", "guideline", "GraphRAG"], OUT / "03_KOMRAG_query_level_evidence_mapping", "sheet")

    # KOM-Sim
    sim_files = []
    for r in files_sorted:
        p = Path(r["absolute_path"])
        if category_for(p) == "KOM-Sim/HCI" or any(k in p.name.lower() for k in ["komsim", "time_workload", "experience", "physician", "materialviews", "adoption"]):
            sim_files.append(r)
    for e in sim_sheet_exports:
        p = Path(e["exported_csv"])
        sim_files.append({"absolute_path": str(p), "relative_path": rel_to_out(p), "file_name": p.name, "extension": ".csv", "file_size_bytes": p.stat().st_size, "source_archive": "extracted_workbook_sheet", "detected_category": "KOM-Sim/HCI"})
    write_csv(OUT / "01_KOMSim_logs_and_time_definition" / "KOMSim_raw_log_inventory.csv", sim_files, ["absolute_path", "relative_path", "file_name", "extension", "file_size_bytes", "source_archive", "detected_category"])

    event_candidates, task_candidates, time_frames, physician_frames = [], [], [], []
    for r in sim_files:
        p = Path(r["absolute_path"])
        if p.suffix.lower() != ".csv" or not p.exists():
            continue
        df = read_csv_safe(p)
        if df is None or df.empty:
            continue
        cols = " ".join(map(str, df.columns)).lower()
        if any(t in cols for t in ["timestamp", "event", "click", "page", "action", "session", "log"]):
            event_candidates.append((p, df))
        if any(t in cols for t in ["time", "workload", "duration", "minutes", "completion", "task", "condition", "arm", "physician", "doctor", "user", "case"]):
            task_candidates.append((p, df))
        if any(t in str(p).lower() for t in ["time_workload", "time", "workload"]):
            time_frames.append((p, df))
        if any(t in str(p).lower() for t in ["physician", "hai", "doctor", "distribution"]):
            physician_frames.append((p, df))

    event_cols = ["participant_id", "physician_id", "case_id", "condition", "timestamp", "event_type", "screen_or_module", "action_label", "evidence_id", "recommendation_id", "duration_seconds", "raw_source_file", "cleaning_note"]
    event_rows = []
    for p, df in event_candidates[:5]:
        pid = first_existing_col(df, ["participant", "physician", "doctor", "user", "id"])
        cid = first_existing_col(df, ["case"])
        cond = first_existing_col(df, ["condition", "arm", "group"])
        ts = first_existing_col(df, ["timestamp", "time"])
        ev = first_existing_col(df, ["event", "action", "click"])
        mod = first_existing_col(df, ["screen", "module", "page"])
        dur = first_existing_col(df, ["duration", "seconds", "sec"])
        for _, row in df.head(5000).iterrows():
            event_rows.append(
                {
                    "participant_id": scalar_from_row(row, pid),
                    "physician_id": scalar_from_row(row, pid),
                    "case_id": scalar_from_row(row, cid),
                    "condition": scalar_from_row(row, cond),
                    "timestamp": scalar_from_row(row, ts),
                    "event_type": scalar_from_row(row, ev),
                    "screen_or_module": scalar_from_row(row, mod),
                    "action_label": scalar_from_row(row, ev),
                    "duration_seconds": scalar_from_row(row, dur),
                    "raw_source_file": str(p),
                    "cleaning_note": "standardized from detected event-like log columns",
                }
            )
    event_found = bool(event_rows)
    write_csv(OUT / "01_KOMSim_logs_and_time_definition" / "KOMSim_event_log_standardized.csv", event_rows, event_cols)

    task_cols = ["participant_id", "physician_id", "case_id", "condition", "task_id", "case_quadrant", "time_minutes", "time_seconds", "material_views", "evidence_views", "agent_trace_views", "interaction_count", "perceived_workload", "usability_score", "trust_score", "decision_confidence", "completion_status", "raw_source_file", "standardization_note"]
    task_rows = []
    for p, df in task_candidates[:20]:
        pid = first_existing_col(df, ["participant", "physician", "doctor", "user", "id"])
        cid = first_existing_col(df, ["case"])
        cond = first_existing_col(df, ["condition", "arm", "group", "module"])
        qcol = first_existing_col(df, ["quadrant", "q"])
        tcol = first_existing_col(df, ["time", "minute", "duration"])
        mat = first_existing_col(df, ["material", "view"])
        evid = first_existing_col(df, ["evidence"])
        work = first_existing_col(df, ["workload", "nasa", "tlx"])
        usab = first_existing_col(df, ["usability", "sus"])
        trust = first_existing_col(df, ["trust"])
        conf = first_existing_col(df, ["confidence"])
        for _, row in df.head(2000).iterrows():
            val_time = scalar_from_row(row, tcol)
            sec = pd.NA
            try:
                sec = float(val_time) * 60 if pd.notna(val_time) else pd.NA
            except Exception:
                pass
            task_rows.append(
                {
                    "participant_id": scalar_from_row(row, pid),
                    "physician_id": scalar_from_row(row, pid),
                    "case_id": scalar_from_row(row, cid),
                    "condition": scalar_from_row(row, cond),
                    "task_id": pd.NA,
                    "case_quadrant": scalar_from_row(row, qcol),
                    "time_minutes": val_time,
                    "time_seconds": sec,
                    "material_views": scalar_from_row(row, mat),
                    "evidence_views": scalar_from_row(row, evid),
                    "agent_trace_views": pd.NA,
                    "interaction_count": pd.NA,
                    "perceived_workload": scalar_from_row(row, work),
                    "usability_score": scalar_from_row(row, usab),
                    "trust_score": scalar_from_row(row, trust),
                    "decision_confidence": scalar_from_row(row, conf),
                    "completion_status": pd.NA,
                    "raw_source_file": str(p),
                    "standardization_note": "task-level or summary-level table; not verified as raw event log",
                }
            )
    write_csv(OUT / "01_KOMSim_logs_and_time_definition" / "KOMSim_task_level_log_standardized.csv", task_rows, task_cols)

    def frame_rows(frames):
        rows = []
        for p, df in frames:
            for _, row in df.head(1000).iterrows():
                item = {str(k): v for k, v in row.items()}
                item["raw_source_file"] = str(p)
                rows.append(item)
        return rows

    time_rows = frame_rows(time_frames)
    phys_rows = frame_rows(physician_frames)
    write_csv(OUT / "01_KOMSim_logs_and_time_definition" / "KOMSim_time_and_interaction_summary.csv", time_rows)
    write_csv(OUT / "01_KOMSim_logs_and_time_definition" / "KOMSim_physician_level_summary.csv", phys_rows)
    cond_rows = []
    if task_rows:
        tdf = pd.DataFrame(task_rows)
        if "condition" in tdf.columns and "time_minutes" in tdf.columns:
            for cond, grp in tdf.groupby("condition", dropna=False):
                nums = pd.to_numeric(grp["time_minutes"], errors="coerce").dropna()
                cond_rows.append({"metric": "time_minutes", "condition": cond, "n": len(nums), "mean": nums.mean() if len(nums) else pd.NA, "sd": nums.std() if len(nums) > 1 else pd.NA, "source": "KOMSim_task_level_log_standardized.csv"})
    write_csv(OUT / "01_KOMSim_logs_and_time_definition" / "KOMSim_condition_comparison_statistics.csv", cond_rows, ["metric", "condition", "n", "mean", "sd", "source"])
    write_md(
        OUT / "01_KOMSim_logs_and_time_definition" / "KOM-Sim_log_cleaning_protocol.md",
        f"""# KOM-Sim log cleaning protocol

Generated: {datetime.now().isoformat(timespec='seconds')}

## Source status

- Raw event logs detected: {'FOUND' if event_found else 'NOT FOUND'}
- Task-level or summary-level timing/HCI data detected: {'FOUND' if len(task_rows) > 0 or len(time_rows) > 0 else 'NOT FOUND'}

## Cleaning and standardization rules

1. All source files matching KOM-Sim, HCI, HAI, physician, time, workload, interaction, material views, or adoption keywords were inventoried.
2. Event-level data were only standardized when timestamp/event/action/page/session-like columns were present.
3. If raw event logs were not found, task-level and figure-level summaries were exported separately and marked as summary-level, not raw logs.
4. Missing columns were retained as NA rather than inferred.
5. No task duration, interaction count, physician ID, case ID, or condition label was fabricated.

## Limitation

If the package only contains figure source tables or aggregate HCI summaries, the standardized task table is suitable for audit traceability but not for raw event-log reanalysis.
""",
    )

    # KOM-Risk
    risk_files = [r for r in files_sorted if category_for(Path(r["absolute_path"])) == "KOM-Risk/OAKNet"]
    for e in risk_sheet_exports:
        p = Path(e["exported_csv"])
        risk_files.append({"absolute_path": str(p), "relative_path": rel_to_out(p), "file_name": p.name, "extension": ".csv", "file_size_bytes": p.stat().st_size, "source_archive": "extracted_workbook_sheet", "detected_category": "KOM-Risk/OAKNet"})
    write_csv(OUT / "02_KOMRisk_reproducible_prediction_package" / "KOMRisk_file_inventory.csv", risk_files, ["absolute_path", "relative_path", "file_name", "extension", "file_size_bytes", "source_archive", "detected_category"])
    pred_rows = []
    for r in risk_files:
        p = Path(r["absolute_path"])
        if p.suffix.lower() == ".csv" and "prediction" in p.name.lower() and p.exists():
            df = read_csv_safe(p)
            if df is not None and not df.empty:
                for _, row in df.head(10000).iterrows():
                    rec = {str(k): v for k, v in row.items()}
                    rec["raw_source_file"] = str(p)
                    pred_rows.append(rec)
    write_csv(OUT / "02_KOMRisk_reproducible_prediction_package" / "risk_predictions.csv", pred_rows)
    risk_metrics = [
        {"endpoint": "KL structural progression", "model": "LightGBM", "n": 7855, "event_rate_percent": 13.4, "metric_name": "AUROC", "metric_value": 0.817, "secondary_metric_name": "BACC", "secondary_metric_value": 0.735, "source": "specified main-text result; sample-level endpoint predictions not found in current source package"},
        {"endpoint": "TKR / knee surgery event", "model": "CoxPH", "n": 9014, "event_rate_percent": 5.2, "metric_name": "C-index", "metric_value": 0.862, "secondary_metric_name": pd.NA, "secondary_metric_value": pd.NA, "source": "specified main-text result; sample-level endpoint predictions not found in current source package"},
        {"endpoint": "Symptom/function worsening", "model": "CatBoost", "n": 8962, "event_rate_percent": 31.0, "metric_name": "AUROC", "metric_value": 0.683, "secondary_metric_name": pd.NA, "secondary_metric_value": pd.NA, "source": "specified main-text result; sample-level endpoint predictions not found in current source package"},
    ]
    feat_rows, calib_rows, split_rows = [], [], []
    for r in risk_files:
        p = Path(r["absolute_path"])
        if p.suffix.lower() != ".csv" or not p.exists():
            continue
        df = read_csv_safe(p)
        if df is None or df.empty:
            continue
        low = p.name.lower()
        if any(k in low for k in ["feature", "shap", "importance"]):
            for _, row in df.head(5000).iterrows():
                rec = {str(k): v for k, v in row.items()}
                rec["raw_source_file"] = str(p)
                feat_rows.append(rec)
        if any(k in low for k in ["calibration", "reliability"]):
            for _, row in df.head(5000).iterrows():
                rec = {str(k): v for k, v in row.items()}
                rec["raw_source_file"] = str(p)
                calib_rows.append(rec)
        if any(k in low for k in ["split", "train", "test", "fold"]):
            for _, row in df.head(5000).iterrows():
                rec = {str(k): v for k, v in row.items()}
                rec["raw_source_file"] = str(p)
                split_rows.append(rec)
    write_csv(OUT / "02_KOMRisk_reproducible_prediction_package" / "risk_model_metrics.csv", risk_metrics, ["endpoint", "model", "n", "event_rate_percent", "metric_name", "metric_value", "secondary_metric_name", "secondary_metric_value", "source"])
    write_csv(OUT / "02_KOMRisk_reproducible_prediction_package" / "risk_feature_importance.csv", feat_rows)
    write_csv(OUT / "02_KOMRisk_reproducible_prediction_package" / "risk_calibration_curve.csv", calib_rows)
    write_csv(OUT / "02_KOMRisk_reproducible_prediction_package" / "risk_split_definition.csv", split_rows)
    model_config = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "model_metrics_source": "main-text summary values plus detected package metric files",
        "sample_level_endpoint_prediction_status": "not found for the three specified KOMRisk endpoints" if not pred_rows else "partial prediction CSVs found; verify endpoint identity",
        "raw_prediction_rows_detected": len(pred_rows),
        "feature_importance_rows_detected": len(feat_rows),
        "calibration_rows_detected": len(calib_rows),
        "split_definition_rows_detected": len(split_rows),
    }
    (OUT / "02_KOMRisk_reproducible_prediction_package" / "risk_model_config.json").write_text(json.dumps(model_config, ensure_ascii=False, indent=2), encoding="utf-8")
    missing_risk = []
    if not pred_rows:
        missing_risk.append("Sample-level predictions for the three specified KOMRisk endpoints were not found; do not infer them from summary metrics.")
    if not feat_rows:
        missing_risk.append("Feature importance / SHAP files were not found.")
    if not calib_rows:
        missing_risk.append("Calibration curve files for KOMRisk endpoints were not found; OAKNet calibration may exist separately if listed in inventory.")
    if not split_rows:
        missing_risk.append("Train/validation/test split definition files were not found.")
    write_md(OUT / "02_KOMRisk_reproducible_prediction_package" / "KOMRisk_missing_items_report.md", "# KOMRisk missing items report\n\n" + ("\n".join(f"- {m}" for m in missing_risk) if missing_risk else "- No missing items detected by file-level audit."))

    # KOM-RAG
    rag_files = [r for r in files_sorted if category_for(Path(r["absolute_path"])) == "KOM-RAG/Evidence"]
    for e in rag_sheet_exports:
        p = Path(e["exported_csv"])
        rag_files.append({"absolute_path": str(p), "relative_path": rel_to_out(p), "file_name": p.name, "extension": ".csv", "file_size_bytes": p.stat().st_size, "source_archive": "extracted_workbook_sheet", "detected_category": "KOM-RAG/Evidence"})
    write_csv(OUT / "03_KOMRAG_query_level_evidence_mapping" / "KOMRAG_file_inventory.csv", rag_files, ["absolute_path", "relative_path", "file_name", "extension", "file_size_bytes", "source_archive", "detected_category"])
    stage4a_data = {}
    stage3_data = {}
    if STAGE4A.exists():
        try:
            stage4a_data = json.loads(STAGE4A.read_text(encoding="utf-8"))
        except Exception:
            stage4a_data = {}
    if STAGE3.exists():
        try:
            stage3_data = json.loads(STAGE3.read_text(encoding="utf-8"))
        except Exception:
            stage3_data = {}
    (OUT / "03_KOMRAG_query_level_evidence_mapping" / "stage4a_retrieval_metrics.source.json").write_text(json.dumps(stage4a_data, ensure_ascii=False, indent=2), encoding="utf-8")
    if stage3_data:
        (OUT / "03_KOMRAG_query_level_evidence_mapping" / "stage3_retrieval_dod.source.json").write_text(json.dumps(stage3_data, ensure_ascii=False, indent=2), encoding="utf-8")
    metric_rows, query_rows, retrieval_rows, relevance_rows, anchor_rows = [], [], [], [], []
    if stage4a_data:
        metrics = stage4a_data.get("holdout_metrics") or stage4a_data.get("metrics") or stage4a_data
        metric_rows.append(
            {
                "query_id": "[summary_holdout]",
                "case_id": pd.NA,
                "agent": pd.NA,
                "precision_at_10": metrics.get("precision_at_10") or metrics.get("Precision@10") or metrics.get("precision@10"),
                "recall_at_10": metrics.get("recall_at_10") or metrics.get("Recall@10") or metrics.get("recall@10"),
                "hit_rate_at_10": metrics.get("hit_rate_at_10") or metrics.get("Hit@10") or metrics.get("hit@10"),
                "mrr": metrics.get("mrr") or metrics.get("MRR"),
                "ndcg_at_10": metrics.get("ndcg_at_10") or metrics.get("nDCG@10") or metrics.get("ndcg@10"),
                "source_file": str(STAGE4A),
                "note": "summary row; query-level records only if separately available",
            }
        )
    for r in rag_files:
        p = Path(r["absolute_path"])
        if p.suffix.lower() != ".csv" or not p.exists():
            continue
        df = read_csv_safe(p)
        if df is None or df.empty:
            continue
        cols = " ".join(map(str, df.columns)).lower()
        if any(k in cols for k in ["query", "clinical question"]):
            qcol = first_existing_col(df, ["query", "clinical question", "question"])
            cid = first_existing_col(df, ["case"])
            ag = first_existing_col(df, ["agent", "domain"])
            if qcol:
                for _, row in df.head(2000).iterrows():
                    query_rows.append({"query_id": scalar_from_row(row, first_existing_col(df, ["query_id", "id"])), "case_id": scalar_from_row(row, cid), "agent": scalar_from_row(row, ag), "query_text": scalar_from_row(row, qcol), "source_file": str(p)})
        if any(k in cols for k in ["precision", "recall", "mrr", "ndcg", "hit"]):
            for _, row in df.head(2000).iterrows():
                metric_rows.append({"query_id": scalar_from_row(row, first_existing_col(df, ["query_id", "query"])), "case_id": scalar_from_row(row, first_existing_col(df, ["case"])), "agent": scalar_from_row(row, first_existing_col(df, ["agent", "domain"])), "precision_at_10": scalar_from_row(row, first_existing_col(df, ["precision"])), "recall_at_10": scalar_from_row(row, first_existing_col(df, ["recall"])), "hit_rate_at_10": scalar_from_row(row, first_existing_col(df, ["hit"])), "mrr": scalar_from_row(row, first_existing_col(df, ["mrr"])), "ndcg_at_10": scalar_from_row(row, first_existing_col(df, ["ndcg"])), "source_file": str(p), "note": "extracted from detected metric-like table"})
        if any(k in cols for k in ["evidence", "eu_id", "direct"]) and any(k in cols for k in ["query", "case", "agent"]):
            evid = first_existing_col(df, ["eu_id", "evidence_id", "evidence"])
            if evid:
                for _, row in df.head(3000).iterrows():
                    retrieval_rows.append({"query_id": scalar_from_row(row, first_existing_col(df, ["query_id", "query"])), "case_id": scalar_from_row(row, first_existing_col(df, ["case"])), "agent": scalar_from_row(row, first_existing_col(df, ["agent", "domain"])), "rank": scalar_from_row(row, first_existing_col(df, ["rank"])), "evidence_id": scalar_from_row(row, evid), "score": scalar_from_row(row, first_existing_col(df, ["score"])), "evidence_role": scalar_from_row(row, first_existing_col(df, ["role", "direct", "context"])), "source_file": str(p)})
        if any(k in cols for k in ["gold", "label", "relevance", "strict", "weak"]):
            evid = first_existing_col(df, ["eu_id", "evidence_id", "evidence"])
            if evid:
                for _, row in df.head(3000).iterrows():
                    relevance_rows.append({"query_id": scalar_from_row(row, first_existing_col(df, ["query_id", "query"])), "evidence_id": scalar_from_row(row, evid), "label": scalar_from_row(row, first_existing_col(df, ["label", "relevance", "gold", "strict", "weak"])), "source_file": str(p)})
        if any(k in cols for k in ["guideline", "anchor", "acr", "eular", "nice", "oarsi", "aaos"]):
            evid = first_existing_col(df, ["eu_id", "evidence_id", "evidence"])
            title = first_existing_col(df, ["title", "guideline", "source"])
            year = first_existing_col(df, ["year"])
            for _, row in df.head(2000).iterrows():
                text = " ".join(str(row.get(c, "")) for c in df.columns[:8]).lower()
                if any(k in text for k in ["nice", "eular", "aaos", "oarsi", "american college", "acr", "arthritis foundation", "guideline"]):
                    anchor_rows.append({"evidence_id": scalar_from_row(row, evid), "guideline_family": pd.NA, "title_or_source": scalar_from_row(row, title) if title else text[:300], "year": scalar_from_row(row, year), "source_file": str(p), "source_note": "detected guideline-like row; verify family manually if needed"})
    write_csv(OUT / "03_KOMRAG_query_level_evidence_mapping" / "rag_query_set.csv", query_rows, ["query_id", "case_id", "agent", "query_text", "source_file"])
    write_csv(OUT / "03_KOMRAG_query_level_evidence_mapping" / "guideline_anchor_mapping.csv", anchor_rows, ["evidence_id", "guideline_family", "title_or_source", "year", "source_file", "source_note"])
    write_csv(OUT / "03_KOMRAG_query_level_evidence_mapping" / "rag_retrieval_results.csv", retrieval_rows, ["query_id", "case_id", "agent", "rank", "evidence_id", "score", "evidence_role", "source_file"])
    write_csv(OUT / "03_KOMRAG_query_level_evidence_mapping" / "rag_relevance_labels.csv", relevance_rows, ["query_id", "evidence_id", "label", "source_file"])
    write_csv(OUT / "03_KOMRAG_query_level_evidence_mapping" / "rag_metric_by_query.csv", metric_rows, ["query_id", "case_id", "agent", "precision_at_10", "recall_at_10", "hit_rate_at_10", "mrr", "ndcg_at_10", "source_file", "note"])
    err = []
    for r in metric_rows:
        try:
            p10 = float(r.get("precision_at_10"))
        except Exception:
            p10 = math.nan
        try:
            nd = float(r.get("ndcg_at_10"))
        except Exception:
            nd = math.nan
        if (not math.isnan(p10) and p10 < 0.5) or (not math.isnan(nd) and nd < 0.65):
            e = dict(r)
            e["error_reason"] = "below target threshold or summary indicates review needed"
            err.append(e)
    write_csv(OUT / "03_KOMRAG_query_level_evidence_mapping" / "rag_error_cases.csv", err)
    rag_report = [
        "# KOMRAG query-level audit report\n\n",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}\n\n",
        f"- Stage4A metrics file: {'FOUND' if STAGE4A.exists() else 'NOT FOUND'}\n",
        f"- Stage3 comparison file: {'FOUND' if STAGE3.exists() else 'NOT FOUND'}\n",
        f"- Query set rows extracted: {len(query_rows)}\n",
        f"- Retrieval result rows extracted: {len(retrieval_rows)}\n",
        f"- Relevance label rows extracted: {len(relevance_rows)}\n",
        f"- Guideline anchor mapping rows extracted: {len(anchor_rows)}\n",
    ]
    if not relevance_rows:
        rag_report.append("\n## Missing item\n\n- Query-level human/gold relevance labels were not located in the current source package. Labels were not invented.\n")
    if not retrieval_rows:
        rag_report.append("\n- Query-level retrieval result rows were not located or could not be mapped from available tables. Summary metrics are preserved from source JSON only.\n")
    write_md(OUT / "03_KOMRAG_query_level_evidence_mapping" / "KOMRAG_query_level_audit_report.md", "".join(rag_report))

    # Submission tables and crosscheck workbook
    for src in [
        OUT / "01_KOMSim_logs_and_time_definition" / "KOMSim_time_and_interaction_summary.csv",
        OUT / "02_KOMRisk_reproducible_prediction_package" / "risk_model_metrics.csv",
        OUT / "03_KOMRAG_query_level_evidence_mapping" / "rag_metric_by_query.csv",
        OUT / "03_KOMRAG_query_level_evidence_mapping" / "guideline_anchor_mapping.csv",
    ]:
        if src.exists():
            shutil.copy2(src, OUT / "05_submission_ready_tables" / src.name)
    checklist = []
    def add_check(item, status, evidence, next_action):
        checklist.append({"item": item, "status": status, "evidence": evidence, "recommended_next_action": next_action})

    add_check("Source archives located", "COMPLETE" if all(p.exists() for p in SOURCE_ZIPS) else "PARTIAL", "; ".join(str(p) for p in SOURCE_ZIPS if p.exists()), "Add any missing archive before submission.")
    add_check("KOM-Sim raw event logs", "COMPLETE" if event_found else "MISSING", str(OUT / "01_KOMSim_logs_and_time_definition" / "KOMSim_event_log_standardized.csv"), "Provide raw browser/session event logs if event-level reanalysis is required.")
    add_check("KOM-Sim task/time summaries", "COMPLETE" if (len(task_rows) > 0 or len(time_rows) > 0) else "MISSING", str(OUT / "01_KOMSim_logs_and_time_definition"), "If only summaries exist, state this in methods.")
    add_check("KOMRisk headline metrics", "COMPLETE", str(OUT / "02_KOMRisk_reproducible_prediction_package" / "risk_model_metrics.csv"), "Verify these against manuscript text before submission.")
    add_check("KOMRisk sample-level endpoint predictions", "COMPLETE" if pred_rows else "MISSING/PARTIAL", str(OUT / "02_KOMRisk_reproducible_prediction_package" / "risk_predictions.csv"), "Add endpoint-specific sample predictions if reproducibility requires recalculation.")
    add_check("KOM-RAG Stage4A metrics", "COMPLETE" if STAGE4A.exists() else "MISSING", str(STAGE4A), "Keep source JSON with final workbook.")
    add_check("KOM-RAG query-level retrieval rows", "COMPLETE" if retrieval_rows else "MISSING/PARTIAL", str(OUT / "03_KOMRAG_query_level_evidence_mapping" / "rag_retrieval_results.csv"), "Add per-query topK result table if reviewer asks for query-level audit.")
    add_check("KOM-RAG relevance labels", "COMPLETE" if relevance_rows else "MISSING", str(OUT / "03_KOMRAG_query_level_evidence_mapping" / "rag_relevance_labels.csv"), "Provide strict/wide gold labels table; do not infer from metrics.")
    add_check("Crosscheck workbook", "COMPLETE", str(OUT / "04_crosscheck_reports" / "submission_audit_crosscheck.xlsx"), "Use as reviewer-facing index.")
    write_csv(OUT / "04_crosscheck_reports" / "submission_readiness_checklist.csv", checklist, ["item", "status", "evidence", "recommended_next_action"])
    wb_path = OUT / "04_crosscheck_reports" / "submission_audit_crosscheck.xlsx"
    with pd.ExcelWriter(wb_path, engine="openpyxl") as writer:
        pd.DataFrame([{"package": "KOM_Submission_Audit_Package_202606", "generated": datetime.now().isoformat(timespec="seconds"), "principle": "Real source files only; missing items remain marked missing."}]).to_excel(writer, sheet_name="README", index=False)
        pd.DataFrame(sim_files).to_excel(writer, sheet_name="KOM-Sim log status", index=False)
        pd.DataFrame([{"raw_event_logs_found": event_found, "task_level_rows": len(task_rows), "time_summary_rows": len(time_rows), "physician_summary_rows": len(phys_rows), "condition_comparison_rows": len(cond_rows)}]).to_excel(writer, sheet_name="KOM-Sim time definition", index=False)
        pd.DataFrame(risk_files).to_excel(writer, sheet_name="KOM-Risk files", index=False)
        pd.DataFrame(risk_metrics).to_excel(writer, sheet_name="KOM-Risk metrics", index=False)
        pd.DataFrame(query_rows).to_excel(writer, sheet_name="KOM-RAG query set", index=False)
        pd.DataFrame(anchor_rows).to_excel(writer, sheet_name="KOM-RAG evidence mapping", index=False)
        pd.DataFrame(metric_rows).to_excel(writer, sheet_name="KOM-RAG query metrics", index=False)
        missing_rows = []
        for m in missing_risk:
            missing_rows.append({"module": "KOM-Risk", "missing_item": m})
        if not event_found:
            missing_rows.append({"module": "KOM-Sim", "missing_item": "Raw event-level logs not found."})
        if not relevance_rows:
            missing_rows.append({"module": "KOM-RAG", "missing_item": "Query-level relevance labels not found."})
        if not retrieval_rows:
            missing_rows.append({"module": "KOM-RAG", "missing_item": "Query-level retrieval result table not found or not mappable."})
        pd.DataFrame(missing_rows).to_excel(writer, sheet_name="Missing items", index=False)
        pd.DataFrame(checklist).to_excel(writer, sheet_name="Ready checklist", index=False)

    readme = f"""# KOM Submission Audit Package 202606

Generated: {datetime.now().isoformat(timespec='seconds')}

## Purpose

This package consolidates reviewer-auditable source materials for KOM-Sim, KOMRisk, and KOMRAG. It uses real source files only. Missing raw logs, sample-level predictions, query-level retrieval rows, or relevance labels are explicitly marked as missing or partial.

## Package structure

- `00_inventory/`: source archive extraction, full file inventory, workbook/sheet inventory, and key asset summary.
- `01_KOMSim_logs_and_time_definition/`: KOM-Sim/HCI log inventory, standardized event/task tables when available, time summaries, physician summaries, and cleaning protocol.
- `02_KOMRisk_reproducible_prediction_package/`: risk/OAKNet file inventory, detected predictions, headline risk metrics, calibration/feature/split tables if available, and missing item report.
- `03_KOMRAG_query_level_evidence_mapping/`: RAG/evidence file inventory, query set, guideline anchors, retrieval rows, relevance labels, metric rows, error cases, and audit report.
- `04_crosscheck_reports/`: reviewer-facing crosscheck workbook and readiness checklist.
- `05_submission_ready_tables/`: copied master tables and selected generated audit tables.

## Status summary

- KOM-Sim raw event logs: {'FOUND' if event_found else 'NOT FOUND'}
- KOM-Sim task/time summaries: {'FOUND' if (len(task_rows) > 0 or len(time_rows) > 0) else 'NOT FOUND'}
- KOMRisk sample-level prediction rows detected: {len(pred_rows)}
- KOMRisk feature-importance rows detected: {len(feat_rows)}
- KOMRisk calibration rows detected: {len(calib_rows)}
- KOMRAG Stage4A metrics: {'FOUND' if STAGE4A.exists() else 'NOT FOUND'}
- KOMRAG query rows extracted: {len(query_rows)}
- KOMRAG retrieval result rows extracted: {len(retrieval_rows)}
- KOMRAG relevance label rows extracted: {len(relevance_rows)}

## Recommended next action

1. If reviewer asks for raw KOM-Sim behavior-log reanalysis, provide the original event/session logs.
2. If reviewer asks to recompute KOMRisk endpoints, provide endpoint-specific sample-level predictions and split definitions.
3. If reviewer asks for query-level KOMRAG auditing, add the strict/wide relevance label table and per-query TopK retrieval table if these are stored outside the current package.

## Non-fabrication note

No missing raw log, prediction, guideline-anchor, query-label, or retrieval-result rows were generated synthetically. Empty or partial files preserve the required schemas where possible and document the missing source evidence.
"""
    write_md(OUT / "README_KOM_SUBMISSION_AUDIT_PACKAGE.md", readme)
    audit_ok = True
    audit_rows = []
    for p in [wb_path, OUT / "00_inventory" / "all_files_inventory.csv", OUT / "04_crosscheck_reports" / "submission_readiness_checklist.csv"]:
        audit_rows.append({"path": str(p), "exists": p.exists(), "size_bytes": p.stat().st_size if p.exists() else None})
    try:
        wb = openpyxl.load_workbook(wb_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        audit_rows.append({"path": str(wb_path), "exists": True, "size_bytes": wb_path.stat().st_size, "sheet_count": len(sheets), "sheets": "; ".join(sheets)})
    except Exception as e:
        audit_rows.append({"path": str(wb_path), "exists": False, "error": str(e)})
        audit_ok = False
    write_csv(OUT / "04_crosscheck_reports" / "workbook_audit.csv", audit_rows)
    write_md(OUT / "04_crosscheck_reports" / "workbook_audit_result.md", f"workbook_audit exit {0 if audit_ok else 1}\n")

    if ZIP_OUT.exists():
        backup_zip = ROOT / f"KOM_Submission_Audit_Package_202606_backup_{NOW}.zip"
        shutil.move(str(ZIP_OUT), str(backup_zip))
        print("Backed up existing zip:", backup_zip)
    with ZipFile(ZIP_OUT, "w", ZIP_DEFLATED) as z:
        for dirpath, dirnames, filenames in os.walk(OUT):
            dirnames[:] = [d for d in dirnames if d != "__pycache__"]
            for fn in filenames:
                p = Path(dirpath) / fn
                z.write(p, p.relative_to(ROOT))

    status = {
        "output_dir": str(OUT),
        "zip_path": str(ZIP_OUT),
        "source_archives_found": sum(1 for p in SOURCE_ZIPS if p.exists()),
        "inventory_rows": len(files_sorted),
        "komsim_raw_event_logs_found": event_found,
        "komsim_task_rows": len(task_rows),
        "komsim_time_summary_rows": len(time_rows),
        "komsim_physician_summary_rows": len(phys_rows),
        "komrisk_prediction_rows": len(pred_rows),
        "komrisk_feature_importance_rows": len(feat_rows),
        "komrisk_calibration_rows": len(calib_rows),
        "komrisk_split_rows": len(split_rows),
        "komrag_stage4a_found": STAGE4A.exists(),
        "komrag_query_rows": len(query_rows),
        "komrag_retrieval_rows": len(retrieval_rows),
        "komrag_relevance_rows": len(relevance_rows),
        "komrag_metric_rows": len(metric_rows),
        "workbook_audit_exit": 0 if audit_ok else 1,
        "zip_size_bytes": ZIP_OUT.stat().st_size if ZIP_OUT.exists() else None,
    }
    (OUT / "00_inventory" / "package_build_status.json").write_text(json.dumps(status, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(status, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
