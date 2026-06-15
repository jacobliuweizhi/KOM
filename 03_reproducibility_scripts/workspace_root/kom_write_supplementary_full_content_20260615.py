from __future__ import annotations

import csv
import datetime as dt
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_WB = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用\最终版本\KOM_缺口补算补写_20260615_1910\KOM_缺口补算补写_完整方法结果总表_20260615_1910.xlsx")
OUT_PARENT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用\最终版本")
JUDGE_OUT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\治疗智能体组\新结果输出\kom_pipeline\kom_pipeline\judge_out")
GEN_OUT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\治疗智能体组\新结果输出\kom_pipeline\kom_pipeline\generation_out")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
WRAP = Alignment(wrap_text=True, vertical="top")


def csv_dicts(path: Path):
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def replace_sheet(wb, name, idx=None):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name, idx) if idx is not None else wb.create_sheet(name)


def write_table(ws, rows, headers=None, start_row=1):
    r = start_row
    if headers:
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP
        r += 1
    for row in rows:
        vals = [row.get(h, "not_available") for h in headers] if isinstance(row, dict) else row
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)
            ws.cell(r, c).alignment = WRAP
        r += 1
    return r


def style(ws):
    if ws.max_row and ws.max_column:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    for c in range(1, ws.max_column + 1):
        width = 10
        for r in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                width = max(width, min(len(str(v)) + 2, 78))
        ws.column_dimensions[get_column_letter(c)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
            if isinstance(cell.value, str) and any(k in cell.value for k in ["不能", "边界", "pilot", "不能写", "仍需"]):
                cell.fill = WARN_FILL
            elif isinstance(cell.value, str) and any(k in cell.value for k in ["已补", "可写", "PASS", "SOURCE"]):
                cell.fill = OK_FILL


def read_judge_summary():
    summary = json.loads((JUDGE_OUT / "summary.json").read_text(encoding="utf-8"))
    arm = csv_dicts(JUDGE_OUT / "arm_ranking.csv")
    icc = csv_dicts(JUDGE_OUT / "inter_judge_icc.csv")
    health = json.loads((JUDGE_OUT / "model_health.json").read_text(encoding="utf-8"))
    return summary, arm, icc, health


def arm_line(arm_rows, name):
    row = next((r for r in arm_rows if r.get("arm") == name), None)
    if not row:
        return "not_available"
    return (
        f"{name}: appropriateness {float(row['appropriateness']):.3f}, "
        f"completeness {float(row['completeness']):.3f}, safety {float(row['safety']):.3f}, "
        f"personalization {float(row['personalization']):.3f}, actionability {float(row['actionability']):.3f}, "
        f"overall {float(row['overall']):.3f}±{float(row['overall_sd']):.3f}, n={row['n']}"
    )


def build_sections(arm_rows, icc_rows, summary):
    overall_icc = next((r for r in icc_rows if r.get("dimension") == "overall"), {})
    action_icc = next((r for r in icc_rows if r.get("dimension") == "actionability"), {})
    rag_generation = (
        f"Generation-quality pilot audit used {summary['n_unique_cases']} cases, 9 arms, "
        f"{len(summary['judges'])} judge models and three repeated judge runs, yielding "
        f"{summary['n_raw_calls']} raw judge calls and {summary['n_aggregated_rows']} aggregated judge rows. "
        f"The best overall arms were: {arm_line(arm_rows, '单智能体强基线')}; "
        f"{arm_line(arm_rows, '完整系统_优化版')}; {arm_line(arm_rows, '无结构化')}. "
        f"Removing RAG produced the lowest score: {arm_line(arm_rows, '无RAG')}. "
        f"Inter-judge reliability was good for overall score, ICC(2,k)={overall_icc.get('icc_2k')} "
        f"({overall_icc.get('ci_low')}-{overall_icc.get('ci_high')}), and excellent for actionability, "
        f"ICC(2,k)={action_icc.get('icc_2k')} ({action_icc.get('ci_low')}-{action_icc.get('ci_high')}). "
        "Because this is a tiny/pilot generation audit, it is written as a generation-quality audit, "
        "not as the final full-scale faithfulness or unsupported-claim rate."
    )

    sections = [
        {
            "id": "S1",
            "title": "Standardized Case Construction",
            "methods": (
                "We constructed a locked standardized KOA evaluation set from an OAI-derived candidate clinical profile pool. "
                "Cases were selected by stratified purposive sampling rather than prevalence sampling, so that the test set covered "
                "orthogonal clinical burdens and decision needs. The final set contained four quadrants: low burden/low need, "
                "low burden/high need, high burden/low need and high burden/high need. Each case template preserved the target knee, "
                "visit context, radiographic severity, symptom/function burden, BMI, fall risk, comorbidity/safety flags, rehabilitation "
                "willingness, treatment preference and missing-information state. Duplicate subject, duplicate knee and high-similarity "
                "case checks were archived before downstream model evaluation."
            ),
            "results": (
                "The locked standardized set contained 120 cases, with 30 cases in each of the four quadrants. The complexity axes "
                "included KL grade, pain, BMI, fall risk, patient demand, comorbidity/risk flags and missing information. The case set "
                "should be interpreted as a structured research evaluation set rather than a real-world prevalence sample."
            ),
            "evidence": "SUM_163_05_STANDARDIZED_CASE; RAW_Case_Basic; RESULT_CHECKPOINTS A01-A07",
            "boundary": "Do not claim real-world KOA prevalence or clinical outcome distribution from this standardized set.",
        },
        {
            "id": "S2",
            "title": "KOM-Profile Patient Representation",
            "methods": (
                "KOM-Profile converted each standardized case into a structured patient representation used by the risk, retrieval, "
                "treatment and safety modules. The representation included demographic fields, radiographic fields, pain and function "
                "variables, neuromuscular/fall-risk variables, metabolic/nutritional variables, psychological and behavioral variables, "
                "medical safety variables and treatment preference variables. Anchor fields included age, sex, BMI, KL grade, NRS, WOMAC, "
                "fall risk, lower-limb strength, renal/GI/anticoagulant status, surgery preference and rehabilitation willingness."
            ),
            "results": (
                "The locked profile schema contained 56 structured fields. The locked summary performance was 0.846 for overall F1/accuracy, "
                "with 31 exact fields and 25 partial fields. This supports use of KOM-Profile as the structured intake layer, while the "
                "field-level raw extraction table remains source-limited in the currently recovered package."
            ),
            "evidence": "SUM_164_06_KOM_PROFILE; RESULT_CHECKPOINTS B01-B05",
            "boundary": "Write the locked summary, but do not fabricate field-level extraction rows.",
        },
        {
            "id": "S3",
            "title": "KOM-Rad / OAKNet Radiographic Module",
            "methods": (
                "KOM-Rad used an OAKNet radiographic grading module to provide image-derived KL severity and uncertainty signals. "
                "The primary model was a ConvNeXt-B backbone with an evidential uncertainty head, with DenseNet-121 retained as a comparator. "
                "The available figure master and audit package document 12 model/ablation arms, five folds, epoch 0-59 training histories "
                "and validation/internal/external performance summaries. Prediction CSVs and figure source data were used for recomputable "
                "confusion matrix, ROC, PR, calibration, DCA and risk-coverage visualizations."
            ),
            "results": (
                "The source-backed primary OAKNet result was QWK 0.806±0.008, balanced accuracy 0.659, macro-F1 0.664, MAE 0.417, "
                "ECE 0.119 and selective accuracy at 80% coverage of 0.725. The audit package found training code, raw predictions and "
                "existing/recomputed figures, but no model checkpoint files."
            ),
            "evidence": "SUM_165_07_KOM_RAD_OAKNET; OAKNet源补写; OAKNet_Reproducibility_Audit_202606",
            "boundary": "Do not write checkpoint-dependent Grad-CAM regeneration or new-case inference as reproduced unless checkpoints are recovered.",
        },
        {
            "id": "S4",
            "title": "KOM-Risk Prognostic Modeling",
            "methods": (
                "KOM-Risk modeled three patient-level endpoints: structural KL progression, TKR/knee surgery event and symptom/function worsening. "
                "A person-level split was used to prevent cross-split subject leakage. Candidate models included elastic-net logistic regression, "
                "random forest, XGBoost, LightGBM and CatBoost. The locked retrain package fixed all random seeds at 20260610. CatBoost was selected "
                "for all endpoints, with iterations=160, learning_rate=0.05, depth=4, loss_function=Logloss, auto_class_weights=Balanced, "
                "eval_metric=AUC and random_seed=20260610. The locked package saved preprocessing pipelines, feature names, predictions and metrics."
            ),
            "results": (
                "For KL structural progression, CatBoost achieved AUROC 0.781, AUPRC 0.349, Brier 0.191 and balanced accuracy 0.715; "
                "bootstrap AUROC mean was 0.782 with 95% CI 0.742-0.817. For TKR/knee surgery event, AUROC was 0.868, AUPRC 0.362, "
                "Brier 0.128 and balanced accuracy 0.780; bootstrap AUROC mean was 0.868 with 95% CI 0.829-0.902. For symptom/function worsening, "
                "AUROC was 0.685, AUPRC 0.348, Brier 0.222 and balanced accuracy 0.609; bootstrap AUROC mean was 0.686 with 95% CI 0.653-0.720."
            ),
            "evidence": "KOMRisk超参数已补; KOMRisk随机种子模型卡; RISK 07_metrics; 03_splits",
            "boundary": "Core training and seed details are complete; library-default parameters are available in the JSON hyperparameter sheet if needed.",
        },
        {
            "id": "S5",
            "title": "SHAP and Feature Lineage",
            "methods": (
                "For model interpretability, locked-model SHAP values were paired with a feature-to-source lineage table. This linked model-attribution "
                "features back to raw OAI/source variables and encoded feature names, allowing each reported explanatory feature to be traced to an "
                "auditable data origin. The final SHAP figures were checked against the locked feature-lineage files."
            ),
            "results": (
                "All SHAP/feature-lineage checkpoints passed: locked-model SHAP values, SHAP feature-to-source mapping and final SHAP figure validity "
                "were present in the recovered workbook/source package."
            ),
            "evidence": "RESULT_CHECKPOINTS E01-E03; SHAP / feature lineage sheets",
            "boundary": "SHAP is reported as model attribution, not as causal mechanism.",
        },
        {
            "id": "S6",
            "title": "KOM-KB Evidence Database",
            "methods": (
                "KOM-KB provided the evidence substrate for retrieval and treatment generation. The database construction tracked PubMed/search lifecycle, "
                "evidence-unit schema, source identifiers and L1-L7 evidence hierarchy. Evidence units were organized so that guideline anchors, systematic "
                "reviews, randomized trials and internal study-level hits could be routed into specialty modules."
            ),
            "results": (
                "The evidence database checkpoints passed for PubMed/search lifecycle audit, evidence-unit database/schema and L1-L7 evidence counts. "
                "The RAG retrieval subset loaded 3260 evidence items; this should be distinguished from database-total evidence-unit counts when both are reported."
            ),
            "evidence": "RESULT_CHECKPOINTS F01-F03; KOM-KB evidence/source sheets",
            "boundary": "Do not mix database-total evidence-unit counts with RAG loaded-subset counts without naming the denominator.",
        },
        {
            "id": "S7",
            "title": "KOM-RAG / GraphRAG Retrieval and Generation Audit",
            "methods": (
                "KOM-RAG used BAAI/bge-m3 embeddings with 1024-dimensional vectors. The retrieval benchmark included 480 queries, split into 320 development "
                "and 160 holdout queries. The retrieval pipeline loaded 3260 evidence items, used candidate_k=180 and RRF k=30, and compared GraphRAG against "
                "a naive RAG baseline. Generation outputs were archived in generation_out, and a multi-LLM judge pipeline evaluated generation quality using "
                "appropriateness, completeness, safety, personalization, actionability and overall dimensions."
            ),
            "results": (
                "GraphRAG outperformed naive RAG in retrieval: P@10 was 0.6763 versus 0.3025, Hit@10 was 1.0000 versus 0.6875, MRR was 0.7483 versus 0.1588, "
                "and nDCG@10 was 0.6902 versus 0.2367. " + rag_generation +
                " A deterministic citation-marker audit of 18 final outputs found explicit source-id/URL/DOI/PMID markers in 0/18 outputs and broad guideline/"
                "citation markers in 1/18 outputs."
            ),
            "evidence": "SUM_168_10_KOM_RAG_GRAPHRA; RAG生成多评委评分已补; KOMRAG生成引用审计; judge_out; generation_out",
            "boundary": "The judge_out result is a tiny/pilot generation-quality audit; exact full-scale faithfulness/citation-support/unsupported-claim rates still need claim-level annotation.",
        },
        {
            "id": "S8",
            "title": "KOM-MDT / KOM-Treat Treatment Recommendation System",
            "methods": (
                "KOM-Treat implemented a multi-specialty decision pipeline with R0-R8 agents. R0 standardized the case and missing-information gate; R1-R4 generated "
                "specialty drafts across medication, surgery, exercise, nutrition and psychology/behavior; R5-R8 performed self-audit, central audit, cross-review "
                "and final consensus. Ablations used the same cases, schema and scoring process across four arms: Full KOM, without RAG, without MDT and direct LLM. "
                "The raw scoring file contained 2880 ok rows across four arms and two evaluator models. Paired statistics were recomputed by matching case_id, quadrant, "
                "sample and evaluator_model, using paired large-sample normal approximation, 95% CI, Cohen dz and Benjamini-Hochberg FDR."
            ),
            "results": (
                "The locked headline treatment-ablation score from SUM_174 was: Full KOM overall quality 84.6, safety 91.1, corrected rule score 84.3 and safety-critical "
                "error 0; without RAG 65.6/79.9/71.6; without MDT 64.4/79.1/81.7; direct LLM 54.7/70.7/44.8. In the raw evaluator overall_0_100 paired-statistic口径, "
                "Full KOM scored 72.3 versus 67.8 without RAG, benefit 4.5 (95% CI 4.0-5.0), dz=0.65, q=4.8e-67, n=707; 72.3 versus 70.0 without MDT, benefit 2.3 "
                "(95% CI 1.8-2.8), dz=0.34, q=1.1e-19, n=702; and 72.3 versus 62.4 for direct LLM, benefit 9.9 (95% CI 9.3-10.5), dz=1.21, q=3.2e-224, n=706."
            ),
            "evidence": "SUM_174_16_KOM_TREAT_ABLATIO; RAW_Treat_AblationScores; KOMTreat配对统计已补; KOMTreat口径一致性检查",
            "boundary": "The 84.6 headline score and 72.3 raw paired-statistic score are different endpoints and must not be mixed.",
        },
        {
            "id": "S9",
            "title": "KOM-Safe Safety Audit",
            "methods": (
                "KOM-Safe applied rule-based and reviewer-facing safety checks to treatment recommendations. Required gates included missing-information-first logic, "
                "renal/eGFR, GI bleeding, anticoagulant/current medication and cardiovascular risk checks before oral NSAID recommendation, injection boundaries, "
                "surgery referral boundaries, exercise stop rules and human review of safety-critical outputs."
            ),
            "results": (
                "The safety-rule library, PASS/WARN/FAIL safety records and safety-critical repair/human-review checkpoints were source-confirmed. The locked MDT safety "
                "sheet documents explicit gates for missing information, oral NSAIDs, injections, surgery escalation and exercise boundaries."
            ),
            "evidence": "SUM_169_11_KOM_MDT_RX_SAFE; RESULT_CHECKPOINTS I01-I03",
            "boundary": "Report recommendation safety and audit behavior; do not claim real patient safety outcomes.",
        },
        {
            "id": "S10",
            "title": "KOM-Score Evaluation Framework",
            "methods": (
                "KOM-Score combined expert raw scores, rule-score outputs, reliability analysis, error taxonomy, agreement/kappa evaluation and training/arbitration "
                "materials. This provided the evaluation spine for recommendation quality, safety, readability, evidence traceability and error classification."
            ),
            "results": (
                "Expert scoring, reliability, rule scores, error taxonomy/agreement and arbitration materials all passed checkpoint review. KOM-Score therefore supports "
                "both treatment-ablation and clinician-simulation evaluation."
            ),
            "evidence": "SUM_170_12_KOM_SCORE_EVAL; RESULT_CHECKPOINTS J01-J05",
            "boundary": "KOM-Score measures recommendation quality, not clinical efficacy.",
        },
        {
            "id": "S11",
            "title": "KOM-Sim Human Interaction Evaluation",
            "methods": (
                "KOM-Sim evaluated doctor-facing interaction using 26 clinicians and 30 tasks, generating 780 final locked records. Conditions compared clinician alone, "
                "clinician plus KOM, clinician plus KOM-R and KOM standalone. Task assignment, final record filtering, primary/secondary comparisons and experience/LOWESS "
                "analysis were archived."
            ),
            "results": (
                "Clinician alone achieved overall quality 48.7, rule score 30.1, safety-critical errors 19.7 per 100 and high-quality rate 4.6%. Clinician + KOM achieved "
                "overall quality 73.4, rule score 63.7, safety-critical errors 8.8 per 100 and high-quality rate 51.5%. Clinician + KOM-R achieved overall quality 70.1, "
                "rule score 61.0 and safety-critical errors 14.5 per 100. KOM standalone had overall quality 84.6, rule score 70 and safety-critical error 0."
            ),
            "evidence": "SUM_173_15_KOM_SIM_CLINICIAN; SIM_final_record_filter_log; RESULT_CHECKPOINTS K01-K06",
            "boundary": "This is simulated clinician interaction and prescription-quality evaluation, not a real-world patient outcome trial.",
        },
        {
            "id": "S12",
            "title": "Figure Source Data and Audit Trail",
            "methods": (
                "All manuscript-critical numerical values and figure source data were indexed in the master workbook. The figure package tracked source-data manifests, "
                "figure denominators, supplementary method figures and the rule that manuscript figures should not rely on bitmap-only evidence when source data are available."
            ),
            "results": (
                "Main figure/source-data checkpoints passed. The latest workbook contains module-specific source sheets for cases, profile, radiography, risk, RAG, treatment, "
                "safety, scoring, clinician simulation and final performance summaries, with numeric traceability for manuscript-critical values."
            ),
            "evidence": "RESULT_CHECKPOINTS L01-L04; SUM_175_17_FINAL_MODEL_PERFO; SUM_181_23_NUMERIC_TRACEABIL; figure master",
            "boundary": "Figures depending on unavailable checkpoints or unscored generation faithfulness should be labeled as pending or excluded from final claims.",
        },
    ]
    return sections


def write_markdown(path: Path, sections):
    lines = [
        "# Supplementary Materials: Methods and Linked Results for the KOM Knee Osteoarthritis Decision-Support System",
        "",
        "## Supplementary Methods Overview",
        "",
        (
            "KOM was organized as a modular doctor-facing knee osteoarthritis decision-support workflow. "
            "The supplementary methods below describe how standardized cases were constructed, how patient profiles and radiographic/risk representations were generated, "
            "how evidence was retrieved and routed, how multi-specialty treatment recommendations were produced, and how safety, scoring, human interaction and figure-source "
            "audits were performed. Each methods subsection is followed by the directly corresponding result paragraph so that every methodological promise is paired with "
            "an auditable output."
        ),
        "",
    ]
    for s in sections:
        lines.extend([
            f"## {s['id']}. {s['title']}",
            "",
            "### Methods",
            "",
            s["methods"],
            "",
            "### Linked Results",
            "",
            s["results"],
            "",
            "### Source Boundary",
            "",
            f"Evidence: {s['evidence']}. Boundary: {s['boundary']}",
            "",
        ])
    lines.extend([
        "## Manuscript Boundary Statement",
        "",
        (
            "The current supplementary methods and linked results are source-backed for standardized cases, KOM-Profile locked summaries, OAKNet performance summaries, "
            "KOM-Risk locked model training and metrics, SHAP/feature lineage, KOM-KB, GraphRAG retrieval, KOM-Treat ablations, KOM-Safe, KOM-Score, KOM-Sim and figure-source manifests. "
            "The remaining boundaries are narrow and explicit: RAG generation faithfulness/citation-support/unsupported-claim rates require claim-level annotation beyond the current "
            "tiny/pilot judge audit; OAKNet checkpoint-dependent Grad-CAM regeneration and new-case inference require checkpoint recovery or retraining; and treatment headline scores "
            "must be kept distinct from raw evaluator paired-statistic endpoints."
        ),
        "",
    ])
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def main():
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M")
    out_dir = OUT_PARENT / f"KOM_补充材料完整正文_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_wb = out_dir / f"KOM_补充材料方法结果完整正文总表_{ts}.xlsx"
    shutil.copy2(BASE_WB, out_wb)

    summary, arm_rows, icc_rows, health = read_judge_summary()
    sections = build_sections(arm_rows, icc_rows, summary)

    wb = load_workbook(out_wb)

    ws = replace_sheet(wb, "RAG生成多评委评分已补", 0)
    write_table(ws, [
        ["judge_out summary", "n_raw_calls", summary["n_raw_calls"], str(JUDGE_OUT / "summary.json")],
        ["judge_out summary", "n_aggregated_rows", summary["n_aggregated_rows"], str(JUDGE_OUT / "results_agg.csv")],
        ["judge_out summary", "n_unique_cases", summary["n_unique_cases"], str(JUDGE_OUT / "summary.json")],
        ["judge_out summary", "judges", "; ".join(summary["judges"]), str(JUDGE_OUT / "summary.json")],
        ["scope", "interpretation", "tiny/pilot generation-quality audit; not full-scale claim-level faithfulness rate", str(JUDGE_OUT)],
    ], ["section", "item", "value", "source"], 1)
    r = ws.max_row + 2
    ws.cell(r, 1, "Arm ranking")
    ws.cell(r, 1).font = Font(bold=True)
    r += 1
    write_table(ws, arm_rows, list(arm_rows[0].keys()), r)
    r = ws.max_row + 2
    ws.cell(r, 1, "Inter-judge ICC")
    ws.cell(r, 1).font = Font(bold=True)
    r += 1
    write_table(ws, icc_rows, list(icc_rows[0].keys()), r)
    r = ws.max_row + 2
    ws.cell(r, 1, "Model health")
    ws.cell(r, 1).font = Font(bold=True)
    r += 1
    health_rows = [{"judge": k, **v} for k, v in health.items()]
    write_table(ws, health_rows, list(health_rows[0].keys()), r)
    style(ws)

    ws = replace_sheet(wb, "补充材料完整正文", 1)
    content_rows = []
    for s in sections:
        content_rows.append({
            "section_id": s["id"],
            "section_title": s["title"],
            "methods_text": s["methods"],
            "linked_results_text": s["results"],
            "evidence": s["evidence"],
            "boundary": s["boundary"],
        })
    write_table(ws, content_rows, ["section_id", "section_title", "methods_text", "linked_results_text", "evidence", "boundary"])
    style(ws)

    # Update existing RAG gap/status language if present.
    if "训练参数待补表" in wb.sheetnames:
        ws = wb["训练参数待补表"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        status_col = headers.index("本轮补写状态") + 1 if "本轮补写状态" in headers else None
        fill_col = headers.index("本轮补写内容") + 1 if "本轮补写内容" in headers else None
        boundary_col = headers.index("剩余边界") + 1 if "剩余边界" in headers else None
        for rr in range(2, ws.max_row + 1):
            if str(ws.cell(rr, 1).value).startswith("KOM-RAG"):
                if status_col:
                    ws.cell(rr, status_col, "已补 tiny/pilot 多评委生成质量评分")
                if fill_col:
                    ws.cell(rr, fill_col, "已纳入 judge_out：330 raw judge calls、108 aggregated rows、6 judges、9 arms；overall ICC(2,k)=0.7653；Full KOM overall=5.694±0.979；无RAG overall=3.694±0.771。")
                if boundary_col:
                    ws.cell(rr, boundary_col, "该结果为 tiny/pilot generation-quality audit；claim-level faithfulness/citation-support/unsupported-claim rate 仍需单独标注。")
        style(ws)

    if "完整正文补写更新" in wb.sheetnames:
        ws = wb["完整正文补写更新"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col = {h: i + 1 for i, h in enumerate(headers)}
        for rr in range(2, ws.max_row + 1):
            if str(ws.cell(rr, col.get("section", 1)).value).startswith("KOM-RAG"):
                ws.cell(rr, col["status"], "SOURCE_BACKED_RETRIEVAL_AND_TINY_GENERATION_JUDGE")
                ws.cell(rr, col["result_cn"], next(s for s in sections if s["id"] == "S7")["results"])
                ws.cell(rr, col["source_or_boundary"], "RAG生成多评委评分已补 + KOMRAG生成引用审计 + judge_out + generation_out")
        style(ws)

    md_path = out_dir / f"KOM_Supplementary_Methods_and_Linked_Results_完整正文_中文稿_{ts}.md"
    write_markdown(md_path, sections)

    wb.save(out_wb)

    qc = {
        "generated_at": ts,
        "output_dir": str(out_dir),
        "output_workbook": str(out_wb),
        "markdown": str(md_path),
        "judge_out_used": str(JUDGE_OUT),
        "rag_generation_judge_scope": {
            "n_raw_calls": summary["n_raw_calls"],
            "n_aggregated_rows": summary["n_aggregated_rows"],
            "n_unique_cases": summary["n_unique_cases"],
            "n_judges": len(summary["judges"]),
            "arms": summary["arms_evaluated"],
        },
    }
    qc_path = out_dir / f"KOM_补充材料完整正文_QC_{ts}.json"
    qc_path.write_text(json.dumps(qc, ensure_ascii=False, indent=2), encoding="utf-8-sig")
    print(json.dumps(qc, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
