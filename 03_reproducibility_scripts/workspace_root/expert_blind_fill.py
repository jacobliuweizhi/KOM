from __future__ import annotations

import csv
import io
import json
import shutil
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用")
WORKBOOK = ROOT / "KOM_纯净版.xlsx"
ZIP_PATH = ROOT / "files.zip"


DETAIL_SHEET = "专家盲评_五模块_明细"
STAT_SHEET = "专家盲评_五模块_统计"
INFO_SHEET = "专家盲评_专家信息"
STATUS_SHEET = "专家盲评_执行状态"
METHOD_SHEET = "专家盲评_方法说明"


def read_text_from_zip(zf: zipfile.ZipFile, name: str) -> str:
    data = zf.read(name)
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace")


def read_expert_package():
    with zipfile.ZipFile(ZIP_PATH, "r") as outer:
        pilot_report = read_text_from_zip(outer, "PILOT_REPORT_zh.md")
        nested_bytes = outer.read("KOM_multiexpert_eval.zip")

    nested = zipfile.ZipFile(io.BytesIO(nested_bytes), "r")
    instruction = read_text_from_zip(nested, "KOM_multiexpert_eval/CODEX_INSTRUCTION_zh.md")
    reference = read_text_from_zip(nested, "KOM_multiexpert_eval/REFERENCE_ANALYSIS_zh.md")
    expert_info = json.loads(read_text_from_zip(nested, "KOM_multiexpert_eval/expert_ref/expert_info.json"))

    module_rows = {}
    for module in ["A", "B", "C", "D", "E"]:
        raw = read_text_from_zip(nested, f"KOM_multiexpert_eval/expert_ref/expert_ratings_{module}.csv")
        rows = list(csv.DictReader(io.StringIO(raw)))
        module_rows[module] = rows
    return pilot_report, instruction, reference, expert_info, module_rows


def clean_sheet(wb, title):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = border
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(v), 60))
        ws.column_dimensions[letter].width = max(12, min(max_len + 2, 42))
    ws.freeze_panes = "A2"


def get_field(row, key, default=None):
    aliases = {
        "item_id": ["item_id", "blind_id", "盲评编号"],
        "timestamp": ["timestamp", "时间"],
        "overall_score": ["overall_score", "overall_quality", "整体质量"],
        "safety_score": ["safety_score", "safety", "安全性"],
        "guideline_score": ["guideline_score", "guideline", "指南一致性"],
        "individualization_score": ["individualization_score", "individualization", "个体化"],
        "actionability_score": ["actionability_score", "actionability", "可执行性"],
        "critical_error": ["critical_error", "严重错误"],
        "major_error": ["major_error", "重要错误"],
        "minor_error": ["minor_error", "轻微错误"],
        "staging_judgement": ["staging_judgement", "staging_correct", "分期判断"],
        "phenotype_judgement": ["phenotype_judgement", "phenotype_correct", "主表型判断"],
        "risk_judgement": ["risk_judgement", "risk_factor_completeness", "风险因素判断"],
        "pair_choice": ["pair_choice", "choice", "盲选结果"],
    }
    for actual_key, value in row.items():
        actual = str(actual_key)
        for candidate in aliases.get(key, [key]):
            if candidate == actual or candidate in actual:
                return value
    return default


def score_mean(rows, key):
    vals = []
    for row in rows:
        v = get_field(row, key)
        if v not in (None, ""):
            vals.append(float(v))
    return round(sum(vals) / len(vals), 2) if vals else None


def count_yes(rows, key):
    return sum(1 for row in rows if str(get_field(row, key, "")).strip() in {"是", "Yes", "yes", "Y"})


def build_stats(module_rows):
    stats = {}
    for mod in ["A", "E"]:
        rows = module_rows[mod]
        stats[mod] = {
            "n": len(rows),
            "overall_mean": score_mean(rows, "overall_score"),
            "safety_mean": score_mean(rows, "safety_score"),
            "guideline_mean": score_mean(rows, "guideline_score"),
            "individualization_mean": score_mean(rows, "individualization_score"),
            "actionability_mean": score_mean(rows, "actionability_score"),
            "critical_error_count": count_yes(rows, "critical_error"),
            "major_error_count": count_yes(rows, "major_error"),
            "minor_error_count": count_yes(rows, "minor_error"),
        }
    for mod in ["B", "C", "D"]:
        rows = module_rows[mod]
        counters = defaultdict(Counter)
        for row in rows:
            for k, v in row.items():
                if k not in {"item_id", "timestamp"} and v not in (None, ""):
                    counters[k][v] += 1
        stats[mod] = {"n": len(rows), "counts": {k: dict(v) for k, v in counters.items()}}
    return stats


def write_detail_sheet(ws, module_rows):
    headers = [
        "盲评编号",
        "模块",
        "评分时间",
        "整体质量",
        "安全性",
        "指南一致性",
        "个体化",
        "可执行性",
        "严重错误",
        "重要错误",
        "轻微错误",
        "分期判断",
        "主表型判断",
        "风险因素判断",
        "盲选结果",
        "数据来源",
        "导入状态",
        "说明",
    ]
    ws.append(headers)
    module_name = {
        "A": "系统处方质量打分",
        "B": "评估准确性",
        "C": "系统消融盲选",
        "D": "人机交互盲选",
        "E": "医生处方打分",
    }
    for mod in ["A", "B", "C", "D", "E"]:
        for row in module_rows[mod]:
            ws.append(
                [
                    get_field(row, "item_id"),
                    module_name[mod],
                    get_field(row, "timestamp"),
                    get_field(row, "overall_score"),
                    get_field(row, "safety_score"),
                    get_field(row, "guideline_score"),
                    get_field(row, "individualization_score"),
                    get_field(row, "actionability_score"),
                    get_field(row, "critical_error"),
                    get_field(row, "major_error"),
                    get_field(row, "minor_error"),
                    get_field(row, "staging_judgement"),
                    get_field(row, "phenotype_judgement"),
                    get_field(row, "risk_judgement"),
                    get_field(row, "pair_choice"),
                    "files.zip / KOM_multiexpert_eval / expert_ref",
                    "Imported",
                    "真实专家参考评分；用于五模块盲评pilot/锚点，不冒充780处方全量多专家评分。",
                ]
            )
    style_sheet(ws)


def write_stats_sheet(ws, stats):
    ws.append(["统计对象", "样本数", "指标", "数值", "说明"])
    for mod, label in [("A", "系统处方质量打分"), ("E", "医生处方打分")]:
        st = stats[mod]
        for key, cn in [
            ("overall_mean", "整体质量均值"),
            ("safety_mean", "安全性均值"),
            ("guideline_mean", "指南一致性均值"),
            ("individualization_mean", "个体化均值"),
            ("actionability_mean", "可执行性均值"),
            ("critical_error_count", "严重错误数"),
            ("major_error_count", "重要错误数"),
            ("minor_error_count", "轻微错误数"),
        ]:
            ws.append([label, st["n"], cn, st[key], "来自 expert_ref 真实专家CSV"])
    for mod, label in [("B", "评估准确性"), ("C", "系统消融盲选"), ("D", "人机交互盲选")]:
        for metric, counts in stats[mod]["counts"].items():
            for value, count in counts.items():
                ws.append([label, stats[mod]["n"], f"{metric}={value}", count, "计数"])
    style_sheet(ws)


def write_info_sheet(ws, expert_info):
    ws.append(["字段", "值"])
    keys = [
        ("specialty", "专科"),
        ("experience", "年资"),
        ("title", "职称"),
        ("started_at", "开始时间"),
        ("finished_at", "完成时间"),
        ("duration_seconds", "用时秒数"),
    ]
    for key, label in keys:
        ws.append([label, expert_info.get(key)])
    ws.append(["导入时间", datetime.now().isoformat(timespec="seconds")])
    ws.append(["来源", "files.zip / KOM_multiexpert_eval / expert_ref / expert_info.json"])
    style_sheet(ws)


def write_status_sheet(ws):
    rows = [
        ["任务", "状态", "证据", "是否可作为正式论文主结果", "说明"],
        [
            "五模块真实专家盲评参考评分",
            "Completed",
            "expert_ratings_A.csv 至 expert_ratings_E.csv，共15条",
            "Limited",
            "可作为pilot/锚点；单专家不能替代多专家ICC或κ一致性分析。",
        ],
        [
            "多persona×多模型API评委",
            "Not completed in supplied package",
            "PILOT_REPORT_zh.md 明确说明上一轮为本地stub/机制验证",
            "No",
            "不得作为真实多模型judge结论。",
        ],
        [
            "处方_专家评价 780医生处方",
            "Pending expert review",
            "本次15条五模块盲评不是780处方全量专家评分",
            "No",
            "继续保留待真人多专家评分，不伪造。",
        ],
        [
            "A/C真实处方来源",
            "Verified by package provenance",
            "压缩包说明为真实KOM内容；本次导入专家评分，不重新编写处方",
            "Yes for pilot",
            "若需处方全文核验，应对照盲评HTML和case_id映射。",
        ],
        [
            "长度/冗长偏倚",
            "Flagged",
            "PILOT_REPORT_zh.md 提示优化臂约为baseline 9倍长度",
            "Important limitation",
            "盲评和统计解释时需控制长度偏倚。",
        ],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def write_method_sheet(ws, pilot_report, instruction, reference):
    ws.append(["项目", "内容"])
    ws.append(["输入包", "files.zip；内部包含 PILOT_REPORT_zh.md 与 KOM_multiexpert_eval.zip"])
    ws.append(["本次实际导入", "expert_ref 下 A-E 五模块真实专家参考评分，共15条；expert_info.json专家信息。"])
    ws.append(["未导入为正式结果", "pilot_outputs 中的AI多persona结果，因为PILOT_REPORT_zh.md说明其为本地stub，并非真实API评委。"])
    ws.append(["统计口径", "A/E计算5维均值和错误数；B计算分期/表型/风险因素判断；C/D计算盲选结果分布。"])
    ws.append(["边界", "本次不伪造780医生处方专家评分、不伪造多专家一致性；这些仍需真实专家补齐。"])
    ws.append(["指令摘要", instruction[:1200]])
    ws.append(["参考分析摘要", reference[:1200]])
    ws.append(["pilot报告摘要", pilot_report[:1200]])
    style_sheet(ws)


def update_existing_sheets(wb):
    # 专家组
    if "专家组" in wb.sheetnames:
        ws = wb["专家组"]
        found = any(str(cell.value).strip() == "E_REF01" for row in ws.iter_rows() for cell in row if cell.value is not None)
        if not found:
            ws.append(
                [
                    "E_REF01",
                    "运动医学科",
                    "副主任医师; 5–10年",
                    "五模块专家盲评参考评分",
                    15,
                    "A-E five-module blind review",
                    "是",
                    "否-单专家锚点",
                    "Imported from files.zip expert_ref",
                ]
            )

    # 局限与缺口
    if "局限与缺口" in wb.sheetnames:
        ws = wb["局限与缺口"]
        updated = False
        for row in ws.iter_rows(min_row=1):
            values = [str(c.value or "") for c in row]
            joined = " ".join(values)
            if "专家评分" in joined and "一致性" in joined:
                if len(row) >= 7:
                    row[6].value = "部分完成：五模块15条真实专家盲评已导入；多专家ICC/κ和780处方专家评分仍待补"
                    updated = True
        if not updated:
            ws.append(
                [
                    "专家盲评",
                    "五模块15条真实专家参考评分已导入，但多专家一致性与780处方专家评分仍未完成",
                    "可作为pilot/锚点，不能作为全量多专家结论",
                    "是",
                    "必须",
                    "继续招募/导入多专家评分并计算ICC/κ",
                    "部分完成",
                ]
            )

    # 导航索引
    if "导航索引" in wb.sheetnames:
        ws = wb["导航索引"]
        existing = {str(row[0].value) for row in ws.iter_rows(min_row=1, max_col=1) if row[0].value}
        for sheet_name in [DETAIL_SHEET, STAT_SHEET, INFO_SHEET, STATUS_SHEET, METHOD_SHEET]:
            if sheet_name not in existing:
                ws.append([sheet_name, "Expert blind review", "Imported from files.zip expert_ref", "Five-module pilot/reference expert rating"])


def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)
    if not ZIP_PATH.exists():
        raise FileNotFoundError(ZIP_PATH)

    pilot_report, instruction, reference, expert_info, module_rows = read_expert_package()
    stats = build_stats(module_rows)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ROOT / f"KOM_纯净版.before_expert_blind_fill_{ts}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK)
    write_detail_sheet(clean_sheet(wb, DETAIL_SHEET), module_rows)
    write_stats_sheet(clean_sheet(wb, STAT_SHEET), stats)
    write_info_sheet(clean_sheet(wb, INFO_SHEET), expert_info)
    write_status_sheet(clean_sheet(wb, STATUS_SHEET))
    write_method_sheet(clean_sheet(wb, METHOD_SHEET), pilot_report, instruction, reference)
    update_existing_sheets(wb)

    wb.save(WORKBOOK)

    # Reopen validation.
    wb2 = load_workbook(WORKBOOK, data_only=False)
    required = [DETAIL_SHEET, STAT_SHEET, INFO_SHEET, STATUS_SHEET, METHOD_SHEET]
    missing = [s for s in required if s not in wb2.sheetnames]
    detail_rows = wb2[DETAIL_SHEET].max_row - 1 if DETAIL_SHEET in wb2.sheetnames else 0
    stats_rows = wb2[STAT_SHEET].max_row - 1 if STAT_SHEET in wb2.sheetnames else 0
    expert_ref_present = False
    if "专家组" in wb2.sheetnames:
        expert_ref_present = any(
            str(cell.value).strip() == "E_REF01"
            for row in wb2["专家组"].iter_rows()
            for cell in row
            if cell.value is not None
        )

    validation = {
        "status": "PASS" if not missing and detail_rows == 15 and expert_ref_present else "FAIL",
        "workbook": str(WORKBOOK),
        "backup": str(backup),
        "required_sheets_missing": missing,
        "detail_data_rows": detail_rows,
        "stats_data_rows": stats_rows,
        "expert_group_E_REF01_present": expert_ref_present,
        "module_counts": {m: len(rows) for m, rows in module_rows.items()},
        "stats": stats,
        "expert_info": expert_info,
        "notes": [
            "15 real expert reference rows imported from files.zip expert_ref.",
            "Pilot AI persona outputs were not imported as formal results because the package report labels them as local stub/mechanism validation.",
            "780 doctor-prescription expert scoring remains pending and was not fabricated.",
        ],
    }
    report_path = ROOT / f"专家盲评回填统计_{ts}.json"
    report_path.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(validation, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
