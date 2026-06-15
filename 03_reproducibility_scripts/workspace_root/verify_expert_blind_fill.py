from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用")
WORKBOOK = ROOT / "KOM_纯净版.xlsx"

REQUIRED = [
    "专家盲评_五模块_明细",
    "专家盲评_五模块_统计",
    "专家盲评_专家信息",
    "专家盲评_执行状态",
    "专家盲评_方法说明",
]


def main():
    wb = load_workbook(WORKBOOK, data_only=False, read_only=True)
    sheet_shapes = {name: [wb[name].max_row, wb[name].max_column] for name in REQUIRED if name in wb.sheetnames}
    missing = [name for name in REQUIRED if name not in wb.sheetnames]

    e_ref_present = False
    expert_row = None
    if "专家组" in wb.sheetnames:
        for row in wb["专家组"].iter_rows(values_only=True):
            if any(str(v).strip() == "E_REF01" for v in row if v is not None):
                e_ref_present = True
                expert_row = list(row)
                break

    limit_row = None
    if "局限与缺口" in wb.sheetnames:
        for row in wb["局限与缺口"].iter_rows(values_only=True):
            text = " ".join(str(v or "") for v in row)
            if "五模块15条真实专家盲评已导入" in text:
                limit_row = list(row)
                break

    prescription_expert_pending_markers = 0
    prescription_expert_fivemodule_markers = 0
    if "处方_专家评价" in wb.sheetnames:
        for row in wb["处方_专家评价"].iter_rows(values_only=True):
            text = " ".join(str(v or "") for v in row)
            if "Pending expert review" in text or "待" in text:
                prescription_expert_pending_markers += 1
            if "五模块" in text or "E_REF01" in text:
                prescription_expert_fivemodule_markers += 1

    detail_rows = sheet_shapes.get("专家盲评_五模块_明细", [0, 0])[0] - 1
    status = (
        "PASS"
        if not missing
        and detail_rows == 15
        and e_ref_present
        and limit_row is not None
        and prescription_expert_fivemodule_markers == 0
        else "FAIL"
    )

    print(
        json.dumps(
            {
                "status": status,
                "workbook": str(WORKBOOK),
                "missing_required_sheets": missing,
                "sheet_shapes": sheet_shapes,
                "detail_data_rows": detail_rows,
                "expert_group_E_REF01_present": e_ref_present,
                "expert_group_row": expert_row,
                "limitation_row_updated": limit_row is not None,
                "limitation_row": limit_row,
                "prescription_expert_pending_markers": prescription_expert_pending_markers,
                "prescription_expert_fivemodule_markers": prescription_expert_fivemodule_markers,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
