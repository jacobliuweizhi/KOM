from __future__ import annotations

import datetime as dt
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用\最终版本\KOM_缺口补算补写_20260615_1910\KOM_缺口补算补写_完整方法结果总表_20260615_1910.xlsx")
OUT_DIR = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用\最终版本\KOM_缺口补算补写_20260615_1910")


def get_status_counts(wb):
    ws = wb["RESULT_CHECKPOINTS"]
    headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers)}
    counts = defaultdict(Counter)
    total = Counter()
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        module = row[idx["module"]]
        status = row[idx["status"]]
        counts[module][status] += 1
        total[status] += 1
        rows.append({
            "checkpoint_id": row[idx["checkpoint_id"]],
            "module": module,
            "title": row[idx["checkpoint_title"]],
            "status": status,
            "rationale": row[idx["rationale"]],
        })
    return counts, total, rows


def sheet_rows(wb, sheet, min_row=1, max_row=20):
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    return [[v for v in row] for row in ws.iter_rows(min_row=min_row, max_row=min(max_row, ws.max_row), values_only=True)]


def fmt_counts(counter):
    return "; ".join(f"{k}={v}" for k, v in counter.items())


def main():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    counts, total, checkpoint_rows = get_status_counts(wb)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M")
    out_path = OUT_DIR / f"KOM_Supplementary_Methods_Results_现状详细报告_{ts}.md"

    modules = [
        {
            "id": "A",
            "name": "Standardized cases",
            "status": "可直接写入 Supplementary Methods，结果链完整。",
            "method": "写 120 例标准化 KOA 病例的构建逻辑：OAI-derived candidate clinical profile pool，stratified purposive sampling，Q1-Q4 四象限各 30 例，覆盖 KL grade、疼痛、BMI、跌倒风险、患者需求、合并风险和缺失信息。",
            "results": "120 cases；Q1 low burden/low need 30，Q2 low burden/high need 30，Q3 high burden/low need 30，Q4 high burden/high need 30。需说明不是流行病学患病率样本，也不是直接真实患者记录。",
            "evidence": "SUM_163_05_STANDARDIZED_CASE；RAW_Case_Basic；RESULT_CHECKPOINTS A01-A07。",
            "boundary": "不要写成真实世界患病率分布；external clinical distribution 仍是 pending_manual_confirmation。",
        },
        {
            "id": "B",
            "name": "KOM-Profile",
            "status": "可写主结果，但字段级原始逐行表存在来源限制。",
            "method": "写 56 个结构化字段的患者画像抽取 schema，锚定 age、sex、BMI、KL、NRS、WOMAC、fall risk、strength、renal/GI/anticoagulant status、surgery preference、rehab willingness 等下游个体化字段。",
            "results": "locked summary 可写：structured fields=56；overall F1/accuracy=0.846；exact fields=31；partial fields=25。",
            "evidence": "SUM_164_06_KOM_PROFILE；RESULT_CHECKPOINTS B01-B05；B02 为 PASS_USER_CONFIRMED_WITH_SOURCE_LIMITATION。",
            "boundary": "field-level raw extraction table 当前文件中 not_found；不能编造逐字段原始行。正文可写锁定汇总，补充材料应显式标注 source limitation。",
        },
        {
            "id": "C",
            "name": "KOM-Rad / OAKNet",
            "status": "性能和图源可写；checkpoint 相关解释必须有边界。",
            "method": "写 ConvNeXt-B with evidential uncertainty head 为主影像分级模型，DenseNet-121 为 comparator；12 arms、5 folds、epoch 0-59、7200 training-history rows；用于 KL radiographic grading、calibration、selective prediction 和风险覆盖曲线。",
            "results": "ConvNeXt-B/OAKNet 主结果可写：QWK 0.806±0.008，BACC 0.659，macro-F1 0.664，MAE 0.417，ECE 0.119，selective accuracy@80 0.725。OAKNet audit 中已找到 prediction CSV、confusion matrix、ROC、PR、calibration、DCA/risk-coverage 等图源。",
            "evidence": "SUM_165_07_KOM_RAD_OAKNET；OAKNet源补写；KOM_Figure_Deep_Optimization_Master_Table_20260615.xlsx；OAKNet_Reproducibility_Audit_202606。",
            "boundary": "RESULT_CHECKPOINTS 中 C05/C06 仍显示 DEFERRED_EXTERNAL_TRAINING_USER_INSTRUCTION，与新加入的 OAKNet源补写存在状态同步问题；论文安全写法是：核心性能和已重绘曲线可写，checkpoint-dependent Grad-CAM regeneration/new-case inference 不写成已重现。",
        },
        {
            "id": "D",
            "name": "KOM-Risk",
            "status": "现在已经是最完整、最适合直接写的模块之一。",
            "method": "写受试者级 split、防泄漏检查、Endpoint A/B/C、raw/encoded feature counts、elastic-net/RF/XGBoost/LightGBM/CatBoost 候选模型，最终 best_model=CatBoost。所有 seed 固定为 20260610。CatBoost 参数：iterations=160，learning_rate=0.05，depth=4，loss_function=Logloss，auto_class_weights=Balanced，eval_metric=AUC，random_seed=20260610。",
            "results": "Endpoint A KL structural progression：AUROC 0.781，AUPRC 0.349，Brier 0.191，BACC 0.715，AUROC 95% CI 0.742-0.817。Endpoint B TKR/surgery：AUROC 0.868，AUPRC 0.362，Brier 0.128，BACC 0.780，AUROC 95% CI 0.829-0.902。Endpoint C symptom/function worsening：AUROC 0.685，AUPRC 0.348，Brier 0.222，BACC 0.609，AUROC 95% CI 0.653-0.720。",
            "evidence": "KOMRisk超参数已补；KOMRisk随机种子模型卡；RISK 07_metrics；05_models metadata；03_splits；13_leakage_and_QC/reproducibility_statement.md。",
            "boundary": "无核心训练参数缺口；如果写极细库默认参数，引用 KOMRisk超参数已补 中完整 JSON。",
        },
        {
            "id": "E",
            "name": "SHAP / Feature lineage",
            "status": "可直接作为模型解释与特征溯源补充方法。",
            "method": "写 locked model SHAP values、feature-to-source mapping 和最终 SHAP figure validity 的三层解释链。",
            "results": "RESULT_CHECKPOINTS E01-E03 全 PASS；可作为 Risk/Score/feature lineage 的解释证据。",
            "evidence": "RESULT_CHECKPOINTS E01-E03；SHAP / feature lineage sheets；numeric traceability/source index。",
            "boundary": "不要把 SHAP 解释写成因果机制；只写 model-attribution / feature-contribution。",
        },
        {
            "id": "F",
            "name": "KOM-KB evidence database",
            "status": "检索数据库构建可以写；与 RAG 检索结果形成闭环。",
            "method": "写 PubMed/search lifecycle audit、evidence unit database schema、L1-L7 evidence hierarchy/counts。",
            "results": "checkpoint F01-F03 全 PASS；与 RAG 使用的 evidence_loaded 3260/3266 左右证据单元和 source-traceability 共同支撑证据库。",
            "evidence": "RESULT_CHECKPOINTS F01-F03；KOM-KB source/evidence sheets；figure/source data manifest。",
            "boundary": "注意 evidence_loaded 3260 与 KB 总 evidence units 3266 可能是不同口径，报告时要标注 database total vs RAG loaded subset。",
        },
        {
            "id": "G",
            "name": "KOM-RAG / GraphRAG",
            "status": "检索性能可直接写；生成 faithfulness 只能写记录和代理审计。",
            "method": "写 BAAI/bge-m3，1024 维 embedding，evidence_loaded 3260，480 queries（dev320/holdout160），candidate_k=180，RRF k=30，GraphRAG vs naive RAG。生成端可写 generation_out 已归档、raw_calls/judge_input/results_wide 存在。",
            "results": "GraphRAG vs naive RAG：P@10 0.6763 vs 0.3025；Hit@10 1.0000 vs 0.6875；MRR 0.7483 vs 0.1588；nDCG@10 0.6902 vs 0.2367。生成输出审计：18 final outputs；explicit source-id/URL/DOI/PMID marker 0/18；broad guideline/citation marker 1/18；Full KOM arm 0/2。",
            "evidence": "SUM_168_10_KOM_RAG_GRAPHRA；KOMRAG生成引用审计；generation_out/raw_calls.jsonl；RESULT_CHECKPOINTS G01-G06。",
            "boundary": "不能把 citation-marker proxy 写成 faithfulness/citation support/unsupported-claim rate。正式生成一致性需要 judge_input.xlsx 人工或模型评分后才能给百分比。",
        },
        {
            "id": "H",
            "name": "KOM-MDT / KOM-Treat",
            "status": "主四臂消融和配对统计均可写，但必须区分两个评分口径。",
            "method": "写 R0-R8 agents、统一病例/schema/scoring、4 arms、2 evaluator models、2880 ok scoring rows。统计方法写 paired large-sample normal approximation、95% CI、Cohen dz、Benjamini-Hochberg FDR。",
            "results": "SUM_174 locked headline score：Full KOM overall quality 84.6，safety 91.1，corrected rule score 84.3，safety-critical error 0；without RAG 65.6/79.9/71.6；without MDT 64.4/79.1/81.7；direct LLM 54.7/70.7/44.8。RAW evaluator paired statistics：Full vs without RAG raw overall 72.3 vs 67.8，benefit 4.5，95% CI 4.0-5.0，dz=0.65，q=4.8e-67，n=707；Full vs without MDT 72.3 vs 70.0，benefit 2.3，95% CI 1.8-2.8，dz=0.34，q=1.1e-19，n=702；Full vs direct LLM 72.3 vs 62.4，benefit 9.9，95% CI 9.3-10.5，dz=1.21，q=3.2e-224，n=706。",
            "evidence": "SUM_174_16_KOM_TREAT_ABLATIO；RAW_Treat_AblationScores；KOMTreat臂级均值已补；KOMTreat配对统计已补；KOMTreat口径一致性检查。",
            "boundary": "84.6 是 locked headline score，72.3 是 raw evaluator overall_0_100；二者不能混写为同一指标。",
        },
        {
            "id": "I",
            "name": "KOM-Safe",
            "status": "安全规则库和 PASS/WARN/FAIL 记录可直接写。",
            "method": "写 safety-rule library、PASS/WARN/FAIL safety records、安全关键错误 repair/human review。包括 missing-information-first、oral NSAID gate、injection boundary、surgery boundary、exercise stop rules。",
            "results": "RESULT_CHECKPOINTS I01-I03 均已 source confirmed/pass；SUM_169 给出 R0-R8 audit/consensus safety gates。",
            "evidence": "SUM_169_11_KOM_MDT_RX_SAFE；RESULT_CHECKPOINTS I01-I03；规则表、错误日志、病例级安全审计文件。",
            "boundary": "只能写安全审计和处方建议质量，不能写真实患者安全结局。",
        },
        {
            "id": "J",
            "name": "KOM-Score",
            "status": "专家评分、规则评分、错误分类和仲裁材料完整。",
            "method": "写 expert raw scores、expert reliability、rule score outputs、error taxonomy/kappa/agreement、training/arbitration materials。",
            "results": "RESULT_CHECKPOINTS J01-J05 全 PASS；可作为 KOM-Treat、KOM-Sim 的评分体系依据。",
            "evidence": "SUM_170_12_KOM_SCORE_EVAL；RESULT_CHECKPOINTS J01-J05；RAW scoring sheets。",
            "boundary": "不要把评分体系写成临床疗效终点；它是 recommendation quality/safety/readability/evidence scoring。",
        },
        {
            "id": "K",
            "name": "KOM-Sim human interaction",
            "status": "人机交互模拟结果可直接写，边界是模拟处方质量而不是真实疗效。",
            "method": "写 26 clinicians × 30 tasks，conditions A/B/C，final locked records=780，任务分配、记录过滤、primary/secondary comparisons、experience/LOWESS analysis。",
            "results": "Clinician alone overall quality 48.7，rule score 30.1，safety-critical error 19.7 per 100，high-quality 4.6%。Clinician + KOM overall quality 73.4，rule score 63.7，safety-critical error 8.8 per 100，high-quality 51.5%。Clinician + KOM-R overall quality 70.1，rule score 61.0，safety-critical error 14.5 per 100，high-quality 50.8%。KOM standalone overall 84.6，rule score 70，safety-critical error 0。",
            "evidence": "SUM_173_15_KOM_SIM_CLINICIAN；SIM_final_record_filter_log；RESULT_CHECKPOINTS K01-K06。",
            "boundary": "写成 simulated clinician interaction / prescription quality improvement；不能写成 patient outcome or long-term adherence improvement。",
        },
        {
            "id": "L",
            "name": "Main figures and source data",
            "status": "主图/补图源数据总体可作为投稿整理入口。",
            "method": "写 main/figure source-data manifest、figure denominators、Supplementary method figures S1-S13、no bitmap-only figure rule。",
            "results": "RESULT_CHECKPOINTS L01-L04 全 PASS；figure deep optimization master 已合并 OAKNet/Risk/RAG/Treat 等结果源。",
            "evidence": "RESULT_CHECKPOINTS L01-L04；KOM_Figure_Deep_Optimization_Master_Table_20260615.xlsx；OAKNet源补写；SOURCE_FILE_MANIFEST。",
            "boundary": "OAKNet checkpoint-dependent 图和 RAG generation faithfulness 图若要定量呈现，还需补充评分/训练源确认。",
        },
    ]

    lines = []
    lines.append("# KOM Supplementary Methods 与对应结果现状详细报告")
    lines.append("")
    lines.append(f"生成时间：{dt.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"当前总表：`{WORKBOOK}`")
    lines.append(f"当前工作簿 sheet 数：{len(wb.sheetnames)}")
    lines.append("")
    lines.append("## 1. 总体判定")
    lines.append("")
    lines.append("当前版本已经达到 **Supplementary Methods 可供合作者/审稿前内部直接观看的 draft-complete 状态**：方法学承诺被拆成 62 个 checkpoint，均已有 PASS、source-confirmed、user-confirmed-with-source-limitation、generation-record-pass 或 deferred-boundary 标签。")
    lines.append("")
    lines.append("但它还不是“所有潜在审稿追问都完全归零”的状态。总判定应写为 **READY_WITH_WARNINGS**，而不是无条件 READY。主要 warning 有三类：")
    lines.append("")
    lines.append("1. KOM-RAG 生成端已有生成记录和引用标记代理审计，但没有正式 faithfulness/citation-support/unsupported-claim rate。")
    lines.append("2. OAKNet 的性能、预测 CSV 和重绘图源已找到，但 checkpoint-dependent Grad-CAM regeneration/new-case inference 不能写成已重现。")
    lines.append("3. KOM-Treat 存在两个评分口径：SUM_174 的 locked headline score 与 RAW_Treat_AblationScores 的 raw evaluator paired-stat 口径，报告中必须分开命名。")
    lines.append("")
    lines.append("## 2. Checkpoint 总览")
    lines.append("")
    lines.append(f"总 checkpoint 数：{sum(total.values())}。状态分布：{fmt_counts(total)}。")
    lines.append("")
    lines.append("| 模块 | checkpoint 状态 | 报告判定 |")
    lines.append("|---|---:|---|")
    for module, counter in counts.items():
        warning = "PASS / 可写" if set(counter.keys()) == {"PASS"} else "可写，但需保留边界"
        lines.append(f"| {module} | {fmt_counts(counter)} | {warning} |")
    lines.append("")
    lines.append("## 3. Supplementary Methods 建议结构")
    lines.append("")
    lines.append("建议 Supplementary Materials 的 Methods 按以下顺序组织，和当前总表 sheet/模块完全对齐：")
    lines.append("")
    for number, m in enumerate(modules, 1):
        lines.append(f"- **S{number}. {m['name']}**：方法设计 -> 评估/审计流程 -> 对应结果 -> 源文件/边界。")
    lines.append("")
    lines.append("## 4. 模块级 Methods-Results 现状")
    lines.append("")
    for m in modules:
        lines.append(f"### {m['id']}. {m['name']}")
        lines.append("")
        lines.append(f"**当前状态：** {m['status']}")
        lines.append("")
        lines.append(f"**Supplementary Methods 现在可以写：** {m['method']}")
        lines.append("")
        lines.append(f"**对应结果现在可以写：** {m['results']}")
        lines.append("")
        lines.append(f"**证据位置：** {m['evidence']}")
        lines.append("")
        lines.append(f"**不能越过的边界：** {m['boundary']}")
        lines.append("")
    lines.append("## 5. 可以直接进入论文/补充材料的结果语句")
    lines.append("")
    lines.append("下面这些是当前最稳的、可直接进入 Supplementary Results 或主文结果支撑表的语句：")
    lines.append("")
    lines.append("- Standardized cases：The standardized evaluation set included 120 OAI-derived KOA cases, evenly stratified into four burden/need quadrants of 30 cases each.")
    lines.append("- KOM-Profile：The profile module used a 56-field structured schema and achieved a locked overall F1/accuracy of 0.846, with 31 exact and 25 partial field matches; field-level raw extraction rows remain source-limited.")
    lines.append("- KOM-Rad/OAKNet：The radiographic model achieved QWK 0.806±0.008, BACC 0.659, macro-F1 0.664, MAE 0.417 and ECE 0.119, with selective accuracy@80 of 0.725.")
    lines.append("- KOM-Risk：CatBoost was selected for all three endpoints under seed 20260610; AUROC values were 0.781, 0.868 and 0.685 for structural progression, surgery event and symptom/function worsening.")
    lines.append("- KOM-RAG：GraphRAG improved retrieval over naive RAG across P@10, Hit@10, MRR and nDCG@10; generation faithfulness remains pending formal scoring.")
    lines.append("- KOM-Treat：Full KOM had SUM_174 locked overall quality 84.6 versus 65.6 without RAG, 64.4 without MDT and 54.7 for direct LLM; raw paired evaluator statistics also favored Full KOM, but under a separate endpoint label.")
    lines.append("- KOM-Sim：Clinician + KOM improved simulated prescription quality from 48.7 to 73.4 and reduced safety-critical errors from 19.7 to 8.8 per 100.")
    lines.append("")
    lines.append("## 6. 当前不能写成最终定量结论的内容")
    lines.append("")
    lines.append("1. **RAG generation faithfulness/citation support/unsupported-claim rate**：当前只能写 generation records exist 和 citation-marker proxy。正式百分比必须等 judge_input.xlsx 完成评分。")
    lines.append("2. **OAKNet checkpoint-dependent outputs**：checkpoint 未找到，因此 Grad-CAM regeneration、新病例 inference、模型权重复现实验不能写成已完成。")
    lines.append("3. **KOM-Profile field-level raw table**：可写 locked summary，不可伪造逐字段原始行。")
    lines.append("4. **临床疗效/长期依从性/真实世界安全结局**：KOM-Sim 和 KOM-Treat 是模拟医生交互与处方质量评价，不是患者临床结局研究。")
    lines.append("")
    lines.append("## 7. 对当前 Supplementary Materials 的实用结论")
    lines.append("")
    lines.append("现在的 Supplementary Methods 适合整理成一份“方法-结果一一对应”的长补充材料，而不是单纯的方法说明。最稳的写法是：每个模块先写设计和流程，再紧接一个 evidence-backed result paragraph，并在段末用括号标出 source sheet 或 audit package。")
    lines.append("")
    lines.append("当前版本已经足够支持论文内部审阅、图表拆分和补充材料初稿撰写。正式投稿前还建议做三件小收口：")
    lines.append("")
    lines.append("1. 将 RESULT_CHECKPOINTS 中 OAKNet C05/C06 与 `OAKNet源补写` 的新发现状态同步，避免同一工作簿里同时出现 deferred 和 found evidence。")
    lines.append("2. 对 `judge_input.xlsx` 完成一次正式生成质量评分，补出 RAG generation faithfulness/citation-support/unsupported-claim rate。")
    lines.append("3. 如果目标期刊非常重视统计检验，把 KOM-Treat 当前 normal-approx 配对统计再补一个 bootstrap 或 Wilcoxon sensitivity analysis。")
    lines.append("")

    # UTF-8 with BOM is friendlier for Windows Notepad/PowerShell previews.
    out_path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(out_path)


if __name__ == "__main__":
    main()
