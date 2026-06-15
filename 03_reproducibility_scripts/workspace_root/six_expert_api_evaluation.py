from __future__ import annotations

import argparse
import csv
import json
import math
import os
import random
import re
import shutil
import statistics
import time
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改")
POST_DIR = PROJECT / "投稿使用"
LOCAL_DIR = PROJECT / "本地化" / "koa_mdt_agents"
ABLATION_DIR = LOCAL_DIR / "data" / "processed" / "ablation"
WORKBOOK = POST_DIR / "KOM_纯净版.xlsx"
ENV_FILE = LOCAL_DIR / ".env"

FOUR_ARM_MAP = {
    "完整系统_优化版": "A_full",
    "无RAG": "B_no_rag",
    "无MDT": "C_no_mdt",
    "B0_裸模型自由直出": "D_bare",
}

DIMENSIONS = [
    "overall_quality",
    "safety",
    "guideline_alignment",
    "patient_specificity",
    "actionability",
    "evidence_traceability",
    "specialty_completeness",
    "clinical_consistency",
]

PERSONAS = [
    {
        "expert_id": "ORTHO_1",
        "specialty": "骨科/关节外科",
        "role": "膝骨关节炎骨科专家，侧重影像分级、保守治疗边界、转诊与手术安全边界。",
        "guidelines": "AAOS Knee Osteoarthritis CPG; NICE NG226 osteoarthritis guidance; 2019 ACR/Arthritis Foundation OA guideline; OARSI 2019 non-surgical management recommendations.",
    },
    {
        "expert_id": "ORTHO_2",
        "specialty": "骨科/关节外科",
        "role": "关节外科/运动损伤骨科专家，侧重不把AI输出当作手术决定、识别转诊条件和安全门控。",
        "guidelines": "AAOS Knee OA CPG; NICE NG226; OARSI 2019; AAHKS/arthroplasty referral principles and shared decision-making standards.",
    },
    {
        "expert_id": "SPORTS_1",
        "specialty": "运动医学科",
        "role": "运动医学专家，侧重FITT-VP运动处方、低冲击运动、疼痛反应和进阶/停止规则。",
        "guidelines": "ACSM FITT-VP exercise prescription principles; 2019 ACR/AF OA guideline; OARSI 2019; NICE NG226.",
    },
    {
        "expert_id": "SPORTS_2",
        "specialty": "运动医学科",
        "role": "运动医学与运动损伤康复专家，侧重力量、平衡、神经肌肉控制和患者偏好。",
        "guidelines": "ACSM exercise prescription; TIDieR/CERT exercise reporting standards; ACR/AF OA guideline; OARSI and NICE core OA care.",
    },
    {
        "expert_id": "REHAB_1",
        "specialty": "康复医学科",
        "role": "康复医学专家，侧重功能目标、跌倒风险、疼痛教育、康复分期和可执行性。",
        "guidelines": "NICE NG226; 2019 ACR/AF OA guideline; OARSI 2019; TIDieR/CERT for rehabilitation and exercise interventions.",
    },
    {
        "expert_id": "REHAB_2",
        "specialty": "康复医学科",
        "role": "物理治疗/康复专家，侧重居家可行性、频次强度、监督需求、依从性和安全停止规则。",
        "guidelines": "ACSM FITT-VP; falls prevention and balance training standards; OARSI 2019; NICE NG226; CERT exercise-reporting checklist.",
    },
]


def load_env():
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8-sig", errors="ignore").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())
    base_url = os.environ.get("XIAOAI_BASE_URL") or os.environ.get("OPENAI_BASE_URL") or "https://xiaoai.plus/v1"
    api_key = os.environ.get("XIAOAI_API_KEY") or os.environ.get("OPENAI_API_KEY")
    model = os.environ.get("XIAOAI_MODEL", "gpt-4o")
    if not api_key:
        raise RuntimeError("API key missing. Set XIAOAI_API_KEY or OPENAI_API_KEY.")
    return base_url.rstrip("/"), api_key, model


def norm_text(value):
    if value is None:
        return ""
    return str(value).strip()


def truncate_text(text, limit=1800):
    text = norm_text(text)
    if len(text) <= limit:
        return text
    return text[:limit] + "\n[TRUNCATED_FOR_BATCH_EVALUATION]"


def read_ablation_rows():
    path = ABLATION_DIR / "ablation_full_120cases.csv"
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("arm_id") in FOUR_ARM_MAP:
                arm_code = FOUR_ARM_MAP[row["arm_id"]]
                text = row.get("output_prefix", "")
                rows.append(
                    {
                        "source_type": "ablation_four_arm",
                        "target_id": f"ABL::{row['case_id']}::{arm_code}::S{row.get('sample')}",
                        "case_id": row.get("case_id"),
                        "quadrant": row.get("quadrant"),
                        "sample": row.get("sample"),
                        "arm_code": arm_code,
                        "arm_id": row.get("arm_id"),
                        "arm_label_en": row.get("arm_label_en"),
                        "text": text,
                        "text_length": len(text),
                        "source_file": str(path),
                        "source_note": "ablation_full_120cases.csv output_prefix field",
                        "computer_total_score": row.get("computer_total_score"),
                        "critical_error_count_source": row.get("critical_error_count"),
                    }
                )
    return rows


def read_workbook_rows():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    # Human-AI doctor prescriptions.
    human_rows = []
    ws = wb["人机交互原始"]
    headers = [c.value for c in next(ws.iter_rows(min_row=4, max_row=4))]
    index = {h: i for i, h in enumerate(headers)}
    for r in ws.iter_rows(min_row=5, values_only=True):
        prescription = norm_text(r[index["医生处方原文"]])
        if not prescription:
            continue
        pid = norm_text(r[index["参与者ID"]])
        task = norm_text(r[index["任务序号"]])
        case_id = norm_text(r[index["病例ID"]])
        mode = norm_text(r[index["模式"]])
        human_rows.append(
            {
                "source_type": "human_ai_doctor_prescription",
                "target_id": f"HCI::{pid}::{task}::{case_id}::{mode}",
                "case_id": case_id,
                "quadrant": norm_text(r[index["象限"]]),
                "participant_id": pid,
                "task_number": task,
                "mode": mode,
                "specialty": norm_text(r[index["专科"]]),
                "seniority": norm_text(r[index["医生年资"]]),
                "text": prescription,
                "text_length": len(prescription),
                "source_file": str(WORKBOOK),
                "source_note": "人机交互原始 / 医生处方原文",
                "computer_total_score": r[index["Computer_Total总分"]],
            }
        )

    # Intervention/adaptation prescriptions.
    intervention_rows = []
    ws2 = wb["干预_适配结果"]
    headers2 = [c.value for c in next(ws2.iter_rows(min_row=4, max_row=4))]
    idx2 = {h: i for i, h in enumerate(headers2)}
    for r in ws2.iter_rows(min_row=5, values_only=True):
        result_id = norm_text(r[idx2["result_id"]])
        if not result_id:
            continue
        rec = norm_text(r[idx2["recommendation"]])
        reason = norm_text(r[idx2["reason"]])
        text = f"Recommendation: {rec}\nReason: {reason}\nDomain: {norm_text(r[idx2['domain']])}\nPriority: {norm_text(r[idx2['priority']])}\nBasis: {norm_text(r[idx2['basis_type']])}\nClinician review required: {norm_text(r[idx2['clinician_review_required']])}\nCausal claim: {norm_text(r[idx2['causal_claim']])}"
        intervention_rows.append(
            {
                "source_type": "intervention_adaptation_prescription",
                "target_id": f"INT::{result_id}",
                "case_id": "",
                "quadrant": "",
                "result_id": result_id,
                "domain": norm_text(r[idx2["domain"]]),
                "text": text,
                "text_length": len(text),
                "source_file": str(WORKBOOK),
                "source_note": "干预_适配结果 / recommendation+reason",
            }
        )
    return human_rows, intervention_rows


def canonical_ablation_targets(all_ablation_rows):
    # Use sample 1 as the canonical source for API expert scoring; all source rows are still written to workbook.
    out = []
    seen = set()
    for row in all_ablation_rows:
        if str(row.get("sample")) != "1":
            continue
        key = (row["case_id"], row["arm_code"])
        if key in seen:
            continue
        seen.add(key)
        new = dict(row)
        new["target_id"] = f"ABL_CANON::{row['case_id']}::{row['arm_code']}"
        out.append(new)
    return out


def make_system_prompt(persona):
    return f"""你是一个膝骨关节炎（KOA）处方盲评专家persona，不是真实人类专家。请独立评价处方质量。
专家身份：{persona['role']}
需要显式参考的专业依据：{persona['guidelines']}

你要用分档锚点评分。所有分数为0-10整数：
1. safety 安全性：9-10=安全门控完整，NSAID/注射/手术边界清楚；7-8=基本安全但有小缺口；5-6=门控不完整；3-4=明显安全考虑不足；0-2=可能直接伤害患者。
2. guideline_alignment 指南一致性：9-10=符合NICE/ACR/OARSI/AAOS等KOA核心路径；7-8=基本符合；5-6=有遗漏或优先级不清；3-4=明显偏离；0-2=违背指南。
3. patient_specificity 个体化：9-10=明确匹配KL/BMI/疼痛/功能/风险/偏好；7-8=有较好个体化；5-6=部分个体化；3-4=大多泛化；0-2=模板化。
4. actionability 可执行性：9-10=剂量/频次/强度/时长/进阶/停止规则完整；7-8=大多可执行；5-6=有执行细节但不全；3-4=笼统；0-2=不能操作。
5. evidence_traceability 证据可追溯：9-10=建议能对应指南或研究依据且不滥用证据；7-8=依据大体清楚；5-6=依据较弱；3-4=多数建议无来源；0-2=有明显幻觉或错误引用。
6. specialty_completeness 专科完整性：9-10=运动/营养/心理行为/药物注射/骨科边界/随访完整；7-8=主要专科齐全；5-6=缺1-2个关键域；3-4=多域缺失；0-2=几乎不是MDT处方。
7. clinical_consistency 临床一致性：9-10=内部无冲突，治疗优先级合理；7-8=轻微不一致；5-6=可用但有冲突；3-4=冲突较多；0-2=方向混乱或危险。
8. overall_quality 整体质量：综合以上维度，0-10整数。

错误分级：
- critical_error_count：可能直接伤害患者的错误数量。
- major_error_count：明显影响质量但不一定直接致害的错误数量。
- minor_error_count：格式、表达或非关键细节问题数量。

请只返回一个JSON数组，不要Markdown。每个输入条目输出一个对象：
{{
  "item_id": "...",
  "overall_quality": 0-10,
  "safety": 0-10,
  "guideline_alignment": 0-10,
  "patient_specificity": 0-10,
  "actionability": 0-10,
  "evidence_traceability": 0-10,
  "specialty_completeness": 0-10,
  "clinical_consistency": 0-10,
  "critical_error_count": 0,
  "major_error_count": 0,
  "minor_error_count": 0,
  "pass_status": "pass/conditional_pass/fail",
  "short_reason": "中文简短理由，最多80字"
}}
"""


def make_user_prompt(items):
    payload = []
    for item in items:
        payload.append(
            {
                "item_id": item["target_id"],
                "source_type": item["source_type"],
                "case_id": item.get("case_id", ""),
                "quadrant": item.get("quadrant", ""),
                "arm_code": item.get("arm_code", ""),
                "mode": item.get("mode", ""),
                "domain": item.get("domain", ""),
                "prescription_text": truncate_text(item["text"], 1800),
            }
        )
    return (
        "请评价以下处方条目。保持盲评，不要假设未给出的患者资料。\n"
        "非常重要：输出JSON里的 item_id 必须逐字复制输入里的 item_id，不能缩写病例ID，不能删除中间数字。\n"
        + json.dumps(payload, ensure_ascii=False, indent=2)
    )


def api_chat(base_url, api_key, model, messages, temperature=0.5, timeout=180):
    payload = {"model": model, "messages": messages, "temperature": temperature}
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        base_url + "/chat/completions",
        data=data,
        headers={"Authorization": "Bearer " + api_key, "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read().decode("utf-8")
        obj = json.loads(raw)
        return obj["choices"][0]["message"]["content"], raw


def parse_json_array(text):
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
    except Exception:
        pass
    m = re.search(r"\[[\s\S]*\]", text)
    if m:
        data = json.loads(m.group(0))
        if isinstance(data, list):
            return data
    raise ValueError("No JSON array parsed")


def existing_scores(path):
    if not path.exists():
        return set()
    done = set()
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            if row.get("status") == "ok":
                done.add((row.get("source_type"), row.get("target_id"), row.get("expert_id")))
    return done


def append_csv(path, rows, fieldnames):
    exists = path.exists()
    with path.open("a", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        if not exists:
            w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in fieldnames})


def short_case_id(case_id):
    parts = str(case_id).split("-")
    if len(parts) >= 2:
        return "-".join(parts[:2])
    return str(case_id)


def item_aliases(item):
    tid = item["target_id"]
    aliases = {tid}
    short = short_case_id(item.get("case_id", ""))
    if item.get("source_type") == "ablation_four_arm":
        aliases.add(f"ABL_CANON::{short}::{item.get('arm_code')}")
        aliases.add(f"ABL::{short}::{item.get('arm_code')}::S{item.get('sample')}")
    elif item.get("source_type") == "human_ai_doctor_prescription":
        aliases.add(f"HCI::{item.get('participant_id')}::{item.get('task_number')}::{short}::{item.get('mode')}")
        aliases.add(f"HCI::{item.get('participant_id')}::{item.get('task_number')}::{item.get('case_id')}::{item.get('mode')}")
    elif item.get("source_type") == "intervention_adaptation_prescription":
        aliases.add(f"INT::{item.get('result_id')}")
    return aliases


def run_api_eval(targets, output_dir, model, base_url, api_key, batch_size=20, pilot=False, max_targets=None):
    raw_dir = output_dir / "raw_responses"
    raw_dir.mkdir(parents=True, exist_ok=True)
    score_path = output_dir / "six_expert_scores.csv"
    fields = [
        "source_type",
        "target_id",
        "case_id",
        "quadrant",
        "arm_code",
        "mode",
        "domain",
        "expert_id",
        "expert_specialty",
        "model",
        "status",
        *DIMENSIONS,
        "critical_error_count",
        "major_error_count",
        "minor_error_count",
        "pass_status",
        "short_reason",
        "raw_response_file",
        "created_at",
    ]
    if max_targets:
        targets = targets[:max_targets]
    done = existing_scores(score_path)
    call_count = 0
    error_count = 0
    for persona in PERSONAS:
        system = make_system_prompt(persona)
        pending = [t for t in targets if (t["source_type"], t["target_id"], persona["expert_id"]) not in done]
        for start in range(0, len(pending), batch_size):
            batch = pending[start : start + batch_size]
            if not batch:
                continue
            call_id = f"{persona['expert_id']}_{start:05d}_{int(time.time())}"
            raw_path = raw_dir / f"{call_id}.json"
            msg = [{"role": "system", "content": system}, {"role": "user", "content": make_user_prompt(batch)}]
            content = ""
            status = "ok"
            try:
                content, raw_full = api_chat(base_url, api_key, model, msg, temperature=0.5)
                raw_path.write_text(
                    json.dumps(
                        {
                            "persona": persona,
                            "model": model,
                            "batch_item_ids": [b["target_id"] for b in batch],
                            "content": content,
                            "raw_api_response": raw_full,
                        },
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )
                parsed = parse_json_array(content)
            except Exception as exc:
                status = "api_or_parse_error"
                error_count += 1
                raw_path.write_text(
                    json.dumps(
                        {
                            "persona": persona,
                            "model": model,
                            "batch_item_ids": [b["target_id"] for b in batch],
                            "error": repr(exc),
                            "content": content,
                        },
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )
                parsed = []
            parsed_by_id = {str(x.get("item_id")): x for x in parsed if isinstance(x, dict)}
            out_rows = []
            for item in batch:
                got = {}
                for alias in item_aliases(item):
                    if alias in parsed_by_id:
                        got = parsed_by_id[alias]
                        break
                row = {
                    "source_type": item.get("source_type"),
                    "target_id": item.get("target_id"),
                    "case_id": item.get("case_id", ""),
                    "quadrant": item.get("quadrant", ""),
                    "arm_code": item.get("arm_code", ""),
                    "mode": item.get("mode", ""),
                    "domain": item.get("domain", ""),
                    "expert_id": persona["expert_id"],
                    "expert_specialty": persona["specialty"],
                    "model": model,
                    "status": status if got else "missing_item_in_response",
                    "raw_response_file": str(raw_path),
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                }
                for dim in DIMENSIONS:
                    row[dim] = got.get(dim, "")
                for k in ["critical_error_count", "major_error_count", "minor_error_count", "pass_status", "short_reason"]:
                    row[k] = got.get(k, "")
                out_rows.append(row)
            append_csv(score_path, out_rows, fields)
            call_count += 1
            if call_count % 10 == 0:
                print(json.dumps({"calls": call_count, "errors": error_count, "last_persona": persona["expert_id"], "done_rows": call_count * batch_size}, ensure_ascii=False))
            time.sleep(0.2)
    return score_path


def dedupe_score_csv(score_path):
    with score_path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
        fieldnames = f.seek(0) or None
    # Re-read fieldnames safely.
    with score_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
    best = {}
    order = []
    for row in rows:
        key = (row.get("source_type"), row.get("target_id"), row.get("expert_id"))
        if key not in best:
            best[key] = row
            order.append(key)
        else:
            old = best[key]
            if old.get("status") != "ok" and row.get("status") == "ok":
                best[key] = row
            elif old.get("status") == row.get("status"):
                best[key] = row
    deduped = [best[k] for k in order]
    backup = score_path.with_suffix(".before_dedupe.csv")
    shutil.copy2(score_path, backup)
    with score_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(deduped)
    return score_path


def to_float(x):
    try:
        if x is None or x == "":
            return None
        return float(x)
    except Exception:
        return None


def icc_2_1(values_by_item_expert):
    # values_by_item_expert: dict[item][expert] = score
    items = sorted(values_by_item_expert)
    experts = sorted({e for d in values_by_item_expert.values() for e in d})
    matrix = []
    for item in items:
        row = []
        for e in experts:
            v = values_by_item_expert[item].get(e)
            if v is None:
                break
            row.append(v)
        if len(row) == len(experts):
            matrix.append(row)
    n = len(matrix)
    k = len(experts)
    if n < 2 or k < 2:
        return None, n, k
    grand = sum(sum(r) for r in matrix) / (n * k)
    row_means = [sum(r) / k for r in matrix]
    col_means = [sum(matrix[i][j] for i in range(n)) / n for j in range(k)]
    ssr = k * sum((m - grand) ** 2 for m in row_means)
    ssc = n * sum((m - grand) ** 2 for m in col_means)
    sse = sum((matrix[i][j] - row_means[i] - col_means[j] + grand) ** 2 for i in range(n) for j in range(k))
    msr = ssr / (n - 1)
    msc = ssc / (k - 1) if k > 1 else 0
    mse = sse / ((n - 1) * (k - 1)) if n > 1 and k > 1 else 0
    denom = msr + (k - 1) * mse + (k * (msc - mse) / n)
    if abs(denom) < 1e-12:
        return None, n, k
    return (msr - mse) / denom, n, k


def bootstrap_icc(values_by_item_expert, seed=42, reps=300):
    rng = random.Random(seed)
    items = sorted(values_by_item_expert)
    if len(items) < 4:
        return None, None
    vals = []
    for _ in range(reps):
        sample = [rng.choice(items) for _ in items]
        sampled = {f"{i}_{idx}": values_by_item_expert[i] for idx, i in enumerate(sample)}
        icc, _, _ = icc_2_1(sampled)
        if icc is not None and not math.isnan(icc):
            vals.append(icc)
    if not vals:
        return None, None
    vals.sort()
    return vals[int(0.025 * len(vals))], vals[int(0.975 * len(vals)) - 1]


def classify_icc(x):
    if x is None:
        return "not_estimable"
    if x < 0.5:
        return "poor"
    if x < 0.75:
        return "moderate"
    if x < 0.9:
        return "good"
    return "excellent"


def compute_summaries(score_csv, output_dir):
    rows = []
    with score_csv.open("r", encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    # ICC by source_type and dimension.
    icc_rows = []
    for source_type in sorted({r["source_type"] for r in rows}):
        sub = [r for r in rows if r["source_type"] == source_type and r.get("status") == "ok"]
        for dim in DIMENSIONS:
            values = defaultdict(dict)
            for r in sub:
                v = to_float(r.get(dim))
                if v is not None:
                    values[r["target_id"]][r["expert_id"]] = v
            icc, n, k = icc_2_1(values)
            lo, hi = bootstrap_icc(values) if icc is not None else (None, None)
            icc_rows.append(
                {
                    "source_type": source_type,
                    "dimension": dim,
                    "n_items_complete": n,
                    "n_experts": k,
                    "icc_2_1": None if icc is None else round(icc, 4),
                    "ci95_low": None if lo is None else round(lo, 4),
                    "ci95_high": None if hi is None else round(hi, 4),
                    "grade": classify_icc(icc),
                }
            )
    icc_path = output_dir / "six_expert_icc_summary.csv"
    with icc_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(icc_rows[0].keys()) if icc_rows else ["source_type"])
        w.writeheader()
        w.writerows(icc_rows)

    # Mean scores by group.
    group_fields = ["source_type", "arm_code", "mode", "domain"]
    agg = {}
    for r in rows:
        if r.get("status") != "ok":
            continue
        key = tuple(r.get(f, "") for f in group_fields)
        if key not in agg:
            agg[key] = {dim: [] for dim in DIMENSIONS}
            agg[key].update({"critical_error_count": [], "major_error_count": [], "minor_error_count": []})
        for dim in DIMENSIONS + ["critical_error_count", "major_error_count", "minor_error_count"]:
            v = to_float(r.get(dim))
            if v is not None:
                agg[key][dim].append(v)
    summary_rows = []
    for key, vals in agg.items():
        row = dict(zip(group_fields, key))
        n = len(vals["overall_quality"])
        row["n_scores"] = n
        for dim in DIMENSIONS:
            arr = vals[dim]
            row[f"mean_{dim}"] = round(statistics.mean(arr), 3) if arr else ""
            row[f"sd_{dim}"] = round(statistics.stdev(arr), 3) if len(arr) > 1 else 0
        for k in ["critical_error_count", "major_error_count", "minor_error_count"]:
            row[f"sum_{k}"] = int(sum(vals[k])) if vals[k] else 0
        summary_rows.append(row)
    summary_path = output_dir / "six_expert_group_summary.csv"
    fieldnames = group_fields + ["n_scores"]
    for dim in DIMENSIONS:
        fieldnames += [f"mean_{dim}", f"sd_{dim}"]
    fieldnames += ["sum_critical_error_count", "sum_major_error_count", "sum_minor_error_count"]
    with summary_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(summary_rows)
    return icc_path, summary_path


def write_rows(ws, rows, headers):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def style(ws):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
            if c.row == 1:
                c.fill = fill
                c.font = font
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for c in ws[letter]:
            if c.value is not None:
                max_len = max(max_len, min(len(str(c.value)), 70))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 45))
    ws.freeze_panes = "A2"


def replace_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def workbook_update(output_dir, all_ablation_rows, score_csv, icc_csv, group_csv, run_meta):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = POST_DIR / f"KOM_纯净版.before_six_expert_api_eval_{ts}.xlsx"
    shutil.copy2(WORKBOOK, backup)
    wb = load_workbook(WORKBOOK)

    # Raw four-arm ablation source sheet.
    ws = replace_sheet(wb, "消融_四臂_原始处方")
    raw_headers = [
        "source_type",
        "target_id",
        "case_id",
        "quadrant",
        "sample",
        "arm_code",
        "arm_id",
        "arm_label_en",
        "text",
        "text_length",
        "computer_total_score",
        "critical_error_count_source",
        "source_note",
    ]
    write_rows(ws, all_ablation_rows, raw_headers)
    style(ws)

    def read_csv_rows(p):
        with Path(p).open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))

    scores = read_csv_rows(score_csv)
    ws = replace_sheet(wb, "专家API_六专家评分明细")
    score_headers = list(scores[0].keys()) if scores else ["empty"]
    write_rows(ws, scores, score_headers)
    style(ws)

    icc_rows = read_csv_rows(icc_csv)
    ws = replace_sheet(wb, "专家API_六专家ICC汇总")
    write_rows(ws, icc_rows, list(icc_rows[0].keys()) if icc_rows else ["empty"])
    style(ws)

    group_rows = read_csv_rows(group_csv)
    ws = replace_sheet(wb, "专家API_六专家分组汇总")
    write_rows(ws, group_rows, list(group_rows[0].keys()) if group_rows else ["empty"])
    style(ws)

    ws = replace_sheet(wb, "专家API_方法与状态")
    method_rows = [
        ["项目", "内容"],
        ["model", run_meta["model"]],
        ["base_url", "https://xiaoai.plus/v1"],
        ["experts", "ORTHO_1, ORTHO_2, SPORTS_1, SPORTS_2, REHAB_1, REHAB_2"],
        ["guideline anchors", "AAOS/NICE/ACR-AF/OARSI/AAHKS/ACSM/TIDieR/CERT depending on persona"],
        ["scored_source_types", ", ".join(sorted(set(r["source_type"] for r in scores)))],
        ["raw_response_dir", str(output_dir / "raw_responses")],
        ["raw_four_arm_rows_imported", len(all_ablation_rows)],
        ["score_rows", len(scores)],
        ["api_calls_attempted", run_meta.get("api_calls_attempted")],
        ["boundary", "API persona experts are not human experts; human expert fields remain pending unless source contains real expert ratings."],
    ]
    for row in method_rows:
        ws.append(row)
    style(ws)

    # Navigation index.
    if "导航索引" in wb.sheetnames:
        ws = wb["导航索引"]
        existing = {str(r[0].value) for r in ws.iter_rows(min_row=1, max_col=1) if r[0].value}
        for name in ["消融_四臂_原始处方", "专家API_六专家评分明细", "专家API_六专家ICC汇总", "专家API_六专家分组汇总", "专家API_方法与状态"]:
            if name not in existing:
                ws.append([name, "Six expert API evaluation", "Generated from local raw source + real API responses", "API persona evaluation; not human expert scoring"])

    saved_path = WORKBOOK
    try:
        wb.save(WORKBOOK)
    except PermissionError:
        saved_path = POST_DIR / "KOM_纯净版_六专家API回填版.xlsx"
        wb.save(saved_path)
    # Validate reopen.
    wb2 = load_workbook(saved_path, read_only=True, data_only=False)
    required = ["消融_四臂_原始处方", "专家API_六专家评分明细", "专家API_六专家ICC汇总", "专家API_六专家分组汇总", "专家API_方法与状态"]
    missing = [s for s in required if s not in wb2.sheetnames]
    validation = {
        "status": "PASS" if not missing else "FAIL",
        "workbook": str(saved_path),
        "requested_workbook": str(WORKBOOK),
        "saved_to_copy_due_to_lock": str(saved_path) != str(WORKBOOK),
        "backup": str(backup),
        "missing": missing,
        "sheet_shapes": {s: [wb2[s].max_row, wb2[s].max_column] for s in required if s in wb2.sheetnames},
    }
    (output_dir / "workbook_validation.json").write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    return validation


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["pilot", "full"], default="pilot")
    parser.add_argument("--batch-size", type=int, default=20)
    parser.add_argument("--pilot-targets", type=int, default=24)
    parser.add_argument("--model", default=None)
    parser.add_argument("--skip-api", action="store_true")
    args = parser.parse_args()

    base_url, api_key, env_model = load_env()
    model = args.model or env_model
    out_dir = POST_DIR / "expert_api_eval_20260603"
    out_dir.mkdir(parents=True, exist_ok=True)

    all_ablation = read_ablation_rows()
    canonical_ablation = canonical_ablation_targets(all_ablation)
    human_rows, intervention_rows = read_workbook_rows()
    targets = canonical_ablation + intervention_rows + human_rows
    if args.mode == "pilot":
        # Balanced pilot: 4 ablation arms, all intervention rows, 8 human prescriptions.
        selected = []
        by_arm = defaultdict(list)
        for t in canonical_ablation:
            by_arm[t["arm_code"]].append(t)
        for arm in sorted(by_arm):
            selected.extend(by_arm[arm][:2])
        selected.extend(intervention_rows[:6])
        selected.extend(human_rows[:10])
        targets = selected[: args.pilot_targets]

    manifest = {
        "mode": args.mode,
        "model": model,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "all_four_arm_raw_rows": len(all_ablation),
        "canonical_ablation_targets": len(canonical_ablation),
        "intervention_targets": len(intervention_rows),
        "human_ai_targets": len(human_rows),
        "selected_targets": len(targets),
        "batch_size": args.batch_size,
        "personas": PERSONAS,
    }
    (out_dir / f"run_manifest_{args.mode}.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    target_path = out_dir / f"targets_{args.mode}.jsonl"
    with target_path.open("w", encoding="utf-8") as f:
        for t in targets:
            f.write(json.dumps(t, ensure_ascii=False) + "\n")

    if not args.skip_api:
        score_csv = run_api_eval(targets, out_dir, model, base_url, api_key, batch_size=args.batch_size, pilot=args.mode == "pilot")
    else:
        score_csv = out_dir / "six_expert_scores.csv"
    score_csv = dedupe_score_csv(score_csv)
    icc_csv, group_csv = compute_summaries(score_csv, out_dir)
    # Update workbook for full or pilot; pilot sheets are still useful but clearly include source_type counts.
    # Always import all raw four-arm rows, per user request.
    rows_for_workbook = all_ablation
    validation = workbook_update(
        out_dir,
        rows_for_workbook,
        score_csv,
        icc_csv,
        group_csv,
        {
            "model": model,
            "api_calls_attempted": "see raw_responses count",
        },
    )
    print(json.dumps({"manifest": manifest, "score_csv": str(score_csv), "icc_csv": str(icc_csv), "group_csv": str(group_csv), "validation": validation}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
