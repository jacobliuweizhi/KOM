from __future__ import annotations

import argparse
import concurrent.futures as cf
import json
import math
import os
import re
import statistics
import threading
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DETAIL_SHEET = "\u5904\u65b9_LLM\u8d28\u91cf\u8bc4\u4ef7_\u660e\u7ec6"
SUMMARY_SHEET = "\u5904\u65b9_LLM\u8d28\u91cf\u8bc4\u4ef7_\u6c47\u603b"
AGREEMENT_SHEET = "\u5904\u65b9_\u8bc4\u4ef7\u4e00\u81f4\u6027"
DOCTOR_RX_TOKEN = "Doctor_Prescriptions_780"

DIM_KEYS = [
    "safety_0_10",
    "guideline_alignment_0_10",
    "patient_specificity_0_10",
    "actionability_0_10",
    "evidence_traceability_0_10",
    "specialty_completeness_0_10",
    "conflict_handling_0_10",
    "doctor_readability_0_10",
    "brevity_0_10",
]


SYSTEM_PROMPT = """You are a clinical AI evaluation methodologist scoring a knee osteoarthritis prescription.
You are not the treating clinician. Score only the prescription text supplied.
Reward patient-specific, safe, actionable, multidisciplinary KOA management.
Penalize vague plans, unsafe oral NSAID or injection escalation, surgery decided without clinician review, missing FITT exercise detail, missing fall prevention, missing weight plus muscle preservation, stigmatizing language, and lack of evidence traceability.
Return only strict JSON. Comments must be in English and <=50 words."""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--outdir", required=True)
    parser.add_argument("--models", nargs="+", default=["gpt-4o-mini", "claude-haiku-4-5-20251001"])
    parser.add_argument("--repeats", type=int, default=3)
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--temperature", type=float, default=0.5)
    return parser.parse_args()


def find_sheet(wb: Any, token: str) -> str:
    matches = [s for s in wb.sheetnames if token in s]
    if not matches:
        raise RuntimeError(f"Sheet containing {token!r} not found")
    return matches[0]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").strip()
    return re.sub(r"\n{3,}", "\n\n", text)


def load_prescriptions(path: Path, limit: int = 0) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[find_sheet(wb, DOCTOR_RX_TOKEN)]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(x) if x is not None else "" for x in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    required = ["Participant", "Case order", "Case ID", "Stage", "Arm", "Condition label", "Doctor final prescription (verbatim source)"]
    missing = [x for x in required if x not in idx]
    if missing:
        raise RuntimeError(f"Missing prescription columns: {missing}")
    records: list[dict[str, Any]] = []
    for excel_row, row in enumerate(rows[1:], start=2):
        participant = row[idx["Participant"]]
        case_order = row[idx["Case order"]]
        case_id = row[idx["Case ID"]]
        stage = row[idx["Stage"]]
        arm = row[idx["Arm"]]
        condition = row[idx["Condition label"]]
        rx = clean_text(row[idx["Doctor final prescription (verbatim source)"]])
        prescription_id = f"{participant}|{case_order}|{case_id}|{stage}|{arm}"
        records.append(
            {
                "Prescription_ID": prescription_id,
                "Workbook_Row": excel_row,
                "Participant": participant,
                "Case_Order": case_order,
                "Case_ID": case_id,
                "Stage": stage,
                "Arm": arm,
                "Condition_Label": condition,
                "Prescription_Text": rx,
                "Prescription_Length": len(rx),
            }
        )
    if limit:
        records = records[:limit]
    return records


def extract_json(text: str) -> dict[str, Any]:
    text = text.strip()
    text = re.sub(r"^```(?:json)?", "", text, flags=re.I).strip()
    text = re.sub(r"```$", "", text).strip()
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        text = text[start : end + 1]
    return json.loads(text)


def make_user_prompt(record: dict[str, Any]) -> str:
    schema = {
        **{key: "number 0-10" for key in DIM_KEYS},
        "overall_0_100": "number 0-100",
        "critical_error_count": "integer",
        "major_error_count": "integer",
        "pass_status": "Pass / Warning / Fail",
        "comments": "English, <=50 words",
    }
    meta = {k: record[k] for k in ["Prescription_ID", "Participant", "Case_ID", "Stage", "Arm", "Condition_Label", "Prescription_Length"]}
    return (
        "Prescription metadata:\n"
        + json.dumps(meta, ensure_ascii=False)
        + "\n\nPrescription text:\n"
        + record["Prescription_Text"]
        + "\n\nReturn JSON with exactly these keys:\n"
        + json.dumps(schema, ensure_ascii=False, indent=2)
    )


def call_model(record: dict[str, Any], model: str, repeat: int, temperature: float, base_url: str, api_key: str) -> dict[str, Any]:
    if not record["Prescription_Text"].strip():
        return base_result(record, model, repeat, "no_input")
    url = base_url.rstrip("/") + "/chat/completions"
    payload = {
        "model": model,
        "temperature": temperature,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": make_user_prompt(record)},
        ],
    }
    headers = {"Authorization": "Bearer " + api_key, "Content-Type": "application/json"}
    last_error = ""
    for attempt in range(3):
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=120)
            if response.status_code != 200:
                last_error = f"http_{response.status_code}: {response.text[:240]}"
                time.sleep(1.5 * (attempt + 1))
                continue
            body = response.json()
            raw = body["choices"][0]["message"]["content"]
            data = extract_json(raw)
            result = base_result(record, model, repeat, "ok")
            for key in DIM_KEYS:
                result[key] = to_float(data.get(key))
            result["Overall_0_100"] = to_float(data.get("overall_0_100"))
            result["Critical_Error_Count"] = to_int(data.get("critical_error_count"))
            result["Major_Error_Count"] = to_int(data.get("major_error_count"))
            result["Pass_Status"] = normalize_pass(data.get("pass_status"))
            result["Comments"] = clean_comment(data.get("comments"))
            result["Raw_Chars"] = len(raw or "")
            return result
        except Exception as exc:  # noqa: BLE001
            last_error = f"{type(exc).__name__}: {str(exc)[:240]}"
            time.sleep(1.5 * (attempt + 1))
    result = base_result(record, model, repeat, "api_error")
    result["Comments"] = last_error
    return result


def base_result(record: dict[str, Any], model: str, repeat: int, status: str) -> dict[str, Any]:
    return {
        "Prescription_ID": record["Prescription_ID"],
        "Workbook_Row": record["Workbook_Row"],
        "Participant": record["Participant"],
        "Case_Order": record["Case_Order"],
        "Case_ID": record["Case_ID"],
        "Stage": record["Stage"],
        "Arm": record["Arm"],
        "Condition_Label": record["Condition_Label"],
        "Prescription_Length": record["Prescription_Length"],
        "Evaluator_Type": "LLM evaluator",
        "Evaluator_Model": "llm:" + model,
        "Repeat": repeat,
        "Status": status,
        "Included_In_Final_Analysis": status == "ok",
        **{key: None for key in DIM_KEYS},
        "Overall_0_100": None,
        "Critical_Error_Count": None,
        "Major_Error_Count": None,
        "Pass_Status": None,
        "Comments": None,
        "Raw_Chars": None,
    }


def to_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return round(float(value), 4)
    except Exception:
        return None


def to_int(value: Any) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(float(value))
    except Exception:
        return None


def normalize_pass(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if "pass" in text:
        return "Pass"
    if "fail" in text:
        return "Fail"
    if "warn" in text:
        return "Warning"
    return str(value).strip()[:40]


def clean_comment(value: Any) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text[:400]


def load_completed(detail_path: Path) -> dict[tuple[str, str, int], dict[str, Any]]:
    done: dict[tuple[str, str, int], dict[str, Any]] = {}
    if not detail_path.exists():
        return done
    with detail_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            row = json.loads(line)
            key = (row["Prescription_ID"], row["Evaluator_Model"].replace("llm:", ""), int(row["Repeat"]))
            done[key] = row
    return done


def run_evaluation(args: argparse.Namespace, records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    detail_path = outdir / "prescription_llm_quality_detail.jsonl"
    progress_path = outdir / "prescription_llm_quality_progress.json"
    api_key = os.environ.get("XIAOAI_API_KEY", "")
    base_url = os.environ.get("XIAOAI_BASE_URL", "https://xiaoai.plus/v1")
    if not api_key:
        raise RuntimeError("XIAOAI_API_KEY is not set")
    completed = load_completed(detail_path) if args.resume else {}
    if not args.resume and detail_path.exists():
        detail_path.unlink()
        completed = {}
    tasks: list[tuple[dict[str, Any], str, int]] = []
    for rec in records:
        for model in args.models:
            for repeat in range(1, args.repeats + 1):
                key = (rec["Prescription_ID"], model, repeat)
                if key not in completed:
                    tasks.append((rec, model, repeat))
    lock = threading.Lock()
    started = time.time()
    ok_counter = 0
    error_counter = 0
    total_expected = len(records) * len(args.models) * args.repeats
    completed_count = len(completed)

    def save_result(result: dict[str, Any]) -> None:
        nonlocal ok_counter, error_counter, completed_count
        with lock:
            with detail_path.open("a", encoding="utf-8") as handle:
                handle.write(json.dumps(result, ensure_ascii=False) + "\n")
            completed_count += 1
            if result["Status"] == "ok":
                ok_counter += 1
            else:
                error_counter += 1
            progress = {
                "started_at_unix": started,
                "updated_at_unix": time.time(),
                "records": len(records),
                "models": args.models,
                "repeats": args.repeats,
                "total_expected_calls": total_expected,
                "completed_calls": completed_count,
                "remaining_calls": max(total_expected - completed_count, 0),
                "ok_in_current_run": ok_counter,
                "non_ok_in_current_run": error_counter,
            }
            progress_path.write_text(json.dumps(progress, ensure_ascii=False, indent=2), encoding="utf-8")

    if tasks:
        with cf.ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = [executor.submit(call_model, rec, model, repeat, args.temperature, base_url, api_key) for rec, model, repeat in tasks]
            for fut in cf.as_completed(futures):
                save_result(fut.result())
    all_rows = list(load_completed(detail_path).values())
    return all_rows, {
        "detail_path": str(detail_path),
        "progress_path": str(progress_path),
        "total_expected_calls": total_expected,
        "completed_calls": len(all_rows),
    }


def mean_sd(values: list[float]) -> tuple[float | None, float | None]:
    vals = [float(v) for v in values if v is not None]
    if not vals:
        return None, None
    mean = statistics.mean(vals)
    sd = statistics.pstdev(vals) if len(vals) > 1 else 0.0
    return round(mean, 3), round(sd, 3)


def correlation(xs: list[float], ys: list[float]) -> float | None:
    pairs = [(x, y) for x, y in zip(xs, ys) if x is not None and y is not None]
    if len(pairs) < 3:
        return None
    xvals = [p[0] for p in pairs]
    yvals = [p[1] for p in pairs]
    mx = statistics.mean(xvals)
    my = statistics.mean(yvals)
    num = sum((x - mx) * (y - my) for x, y in pairs)
    denx = math.sqrt(sum((x - mx) ** 2 for x in xvals))
    deny = math.sqrt(sum((y - my) ** 2 for y in yvals))
    if denx == 0 or deny == 0:
        return None
    return round(num / (denx * deny), 4)


def summarize(records: list[dict[str, Any]], rows: list[dict[str, Any]], models: list[str]) -> tuple[list[dict[str, Any]], list[list[Any]]]:
    by_rx_model: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_rx_model[(row["Prescription_ID"], row["Evaluator_Model"].replace("llm:", ""))].append(row)
    summaries: list[dict[str, Any]] = []
    for rec in records:
        summary: dict[str, Any] = {k: rec[k] for k in ["Prescription_ID", "Workbook_Row", "Participant", "Case_Order", "Case_ID", "Stage", "Arm", "Condition_Label", "Prescription_Length"]}
        pass_sets = []
        for model in models:
            model_rows = [r for r in by_rx_model.get((rec["Prescription_ID"], model), []) if r["Status"] == "ok"]
            vals = [r["Overall_0_100"] for r in model_rows if r.get("Overall_0_100") is not None]
            mean, sd = mean_sd(vals)
            label = model_label(model)
            summary[f"{label}_mean"] = mean
            summary[f"{label}_sd"] = sd
            summary[f"{label}_n_ok"] = len(vals)
            statuses = {r.get("Pass_Status") for r in model_rows if r.get("Pass_Status")}
            pass_sets.append(statuses)
        means = [summary.get(f"{model_label(m)}_mean") for m in models]
        summary["cross_model_abs_diff"] = round(abs(means[0] - means[1]), 3) if len(means) >= 2 and means[0] is not None and means[1] is not None else None
        summary["all_pass_agree"] = len(pass_sets) >= 2 and all(s == pass_sets[0] for s in pass_sets[1:]) if all(pass_sets) else None
        summary["Summary_Status"] = "ok" if all(summary.get(f"{model_label(m)}_n_ok", 0) > 0 for m in models) else "partial_or_no_input"
        summaries.append(summary)
    xs = [s.get(f"{model_label(models[0])}_mean") for s in summaries] if models else []
    ys = [s.get(f"{model_label(models[1])}_mean") for s in summaries] if len(models) > 1 else []
    diffs = [s["cross_model_abs_diff"] for s in summaries if s.get("cross_model_abs_diff") is not None]
    agreements = [s["all_pass_agree"] for s in summaries if s.get("all_pass_agree") is not None]
    agreement_rows = [
        ["Metric", "Value", "Notes"],
        ["Prescription rows evaluated", len(summaries), "One row per doctor prescription task"],
        ["Successful detail rows", sum(1 for r in rows if r.get("Status") == "ok"), "Model-repeat level rows with parsed JSON scores"],
        ["Model 1", models[0] if models else None, ""],
        ["Model 2", models[1] if len(models) > 1 else None, ""],
        ["Pearson correlation of model means", correlation(xs, ys), "Across prescription-level mean scores"],
        ["Mean absolute model difference", round(statistics.mean(diffs), 3) if diffs else None, "Absolute difference between model means"],
        ["Median absolute model difference", round(statistics.median(diffs), 3) if diffs else None, "Absolute difference between model means"],
        ["Pass-status agreement rate", round(sum(1 for x in agreements if x) / len(agreements), 4) if agreements else None, "Exact agreement of available pass-status sets"],
        ["No-input summary rows", sum(1 for s in summaries if s["Summary_Status"] != "ok"), "Should be 0 if all prescriptions had text and parsed results"],
    ]
    return summaries, agreement_rows


def model_label(model: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", model).lower()


def write_sheet(ws: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells[:200])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)


def update_workbook(input_path: Path, output_path: Path, detail_rows: list[dict[str, Any]], summary_rows: list[dict[str, Any]], agreement_rows: list[list[Any]]) -> dict[str, Any]:
    wb = load_workbook(input_path)
    for name in [DETAIL_SHEET, SUMMARY_SHEET, AGREEMENT_SHEET]:
        if name in wb.sheetnames:
            del wb[name]
    summary_ws = wb.create_sheet(SUMMARY_SHEET, 10)
    detail_ws = wb.create_sheet(DETAIL_SHEET, 30)
    agreement_ws = wb.create_sheet(AGREEMENT_SHEET, 11)
    summary_headers = list(summary_rows[0].keys()) if summary_rows else []
    write_sheet(summary_ws, [summary_headers] + [[r.get(h) for h in summary_headers] for r in summary_rows])
    detail_headers = list(detail_rows[0].keys()) if detail_rows else []
    write_sheet(detail_ws, [detail_headers] + [[r.get(h) for h in detail_headers] for r in detail_rows])
    write_sheet(agreement_ws, agreement_rows)
    if "\u5bfc\u822a\u7d22\u5f15" in wb.sheetnames:
        nav = wb["\u5bfc\u822a\u7d22\u5f15"]
        existing = {str(nav.cell(r, 1).value) for r in range(1, nav.max_row + 1)}
        for sheet_name, group, purpose in [
            (SUMMARY_SHEET, "\u2460 \u5bfc\u822a\u4e0e\u56fe\u8868\u6570\u636e", "Prescription-side real LLM evaluator summary for manuscript figures and agreement reporting."),
            (AGREEMENT_SHEET, "\u2460 \u5bfc\u822a\u4e0e\u56fe\u8868\u6570\u636e", "Cross-model agreement statistics for prescription LLM evaluation."),
            (DETAIL_SHEET, "\u2463 \u539f\u59cb\u660e\u7ec6\u6570\u636e", "Model-repeat-level prescription LLM evaluation detail rows."),
        ]:
            if sheet_name not in existing:
                nav.append([sheet_name, group, purpose, "Added by prescription-side real LLM evaluation completion."])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    reopened = load_workbook(output_path, read_only=True, data_only=True)
    return {
        "output": str(output_path),
        "sheet_count": len(reopened.sheetnames),
        "detail_rows": reopened[DETAIL_SHEET].max_row - 1,
        "summary_rows": reopened[SUMMARY_SHEET].max_row - 1,
        "agreement_rows": reopened[AGREEMENT_SHEET].max_row - 1,
        "opens": True,
    }


def scan_for_leaks(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    api_hits = 0
    local_hits = 0
    formulas_errors = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                text = str(value)
                if "sk-" in text or "api_key" in text.lower():
                    api_hits += 1
                if "C:\\" in text or "/home/" in text:
                    local_hits += 1
                if text in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"):
                    formulas_errors += 1
    return {"api_key_hits": api_hits, "local_path_hits": local_hits, "formula_error_like_cells": formulas_errors}


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    outdir = Path(args.outdir)
    records = load_prescriptions(input_path, args.limit)
    rows, run_info = run_evaluation(args, records)
    summary_rows, agreement_rows = summarize(records, rows, args.models)
    update_info = update_workbook(input_path, output_path, rows, summary_rows, agreement_rows)
    leak_info = scan_for_leaks(output_path)
    report = {
        "input": str(input_path),
        "records_loaded": len(records),
        "models": args.models,
        "repeats": args.repeats,
        "workers": args.workers,
        "run_info": run_info,
        "workbook": update_info,
        "leak_scan": leak_info,
        "final_ready": update_info["opens"] and update_info["summary_rows"] == len(records) and leak_info["api_key_hits"] == 0 and leak_info["local_path_hits"] == 0,
    }
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "prescription_llm_completion_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
