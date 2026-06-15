from __future__ import annotations

import csv
import datetime as dt
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WB_PATH = Path(r"C:\OAI研究项目\pythonProject1\KOM返修修改\投稿使用\最终版本\KOM_补充材料完整正文_20260615_1939\KOM_补充材料方法结果完整正文总表_20260615_1939.xlsx")
OUT_DIR = WB_PATH.parent

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def style(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for c in range(1, ws.max_column + 1):
        width = 10
        for r in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                width = max(width, min(len(str(v)) + 2, 82))
        ws.column_dimensions[get_column_letter(c)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP


def write_table(ws, rows, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, row.get(h, ""))
            ws.cell(r, c).alignment = WRAP


def csv_dicts(path):
    with Path(path).open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def build_sections():
    return [
        {
            "编号": "S1",
            "标题": "标准化病例构建",
            "方法正文": "我们首先构建用于系统评估的标准化膝骨关节炎病例集。病例来源于 OAI 派生的候选临床画像池，并采用分层目的抽样，而不是按真实世界患病率抽样。抽样目标是让评估集覆盖影像结构负担、症状功能负担、治疗需求、安全风险、缺失信息和患者目标等相互独立但临床上常同时出现的决策维度。最终病例被固定到四个象限：低负担/低需求、低负担/高需求、高负担/低需求、高负担/高需求。每个病例模板均保留目标膝、随访访视、影像分级、疼痛与功能、BMI、跌倒风险、合并症/用药安全标志、康复意愿、治疗偏好和关键缺失信息。进入下游模型前，项目保留了受试者重复、膝侧重复和高相似病例检查记录。",
            "结果正文": "最终锁定的标准化病例集包含 120 例，四个象限各 30 例。病例复杂度轴包括 KL 分级、疼痛、BMI、跌倒风险、患者需求、合并症/安全风险和缺失信息。该病例集用于评价模型在不同临床情境中的稳定性和可解释性，不用于估计真实世界 KOA 患病率或临床结局分布。",
            "证据": "SUM_163_05_STANDARDIZED_CASE；RAW_Case_Basic；RESULT_CHECKPOINTS A01-A07",
            "边界": "不能写成真实世界患病率样本，也不能把标准化病例等同于真实患者结局队列。",
        },
        {
            "编号": "S2",
            "标题": "KOM-Profile 患者画像结构化",
            "方法正文": "KOM-Profile 将每个标准化病例转换为可供影像、风险、检索、治疗和安全模块共同调用的结构化患者画像。画像包含人口学变量、影像变量、疼痛和功能变量、神经肌肉与跌倒风险变量、代谢/营养变量、心理行为变量、内科安全变量和治疗偏好变量。核心锚点字段包括年龄、性别、BMI、KL 分级、NRS、WOMAC、跌倒风险、下肢肌力、肾功能/GI 风险/抗凝或当前用药状态、手术偏好和康复意愿。该模块的设计目的不是简单抽取文本，而是把后续个体化处方所需的临床约束提前标准化。",
            "结果正文": "锁定的患者画像 schema 包含 56 个结构化字段。当前锁定汇总性能为 overall F1/accuracy=0.846，其中 31 个字段达到 exact match，25 个字段为 partial match。该结果支持 KOM-Profile 作为后续个体化推理和安全审计的结构化入口。当前恢复包中未找到逐字段原始抽取明细表，因此补充材料中应写锁定汇总结果，不写伪造的逐字段原始行。",
            "证据": "SUM_164_06_KOM_PROFILE；RESULT_CHECKPOINTS B01-B05",
            "边界": "field-level raw extraction table 当前不完整；只能写 locked summary。",
        },
        {
            "编号": "S3",
            "标题": "KOM-Rad / OAKNet 影像分级模块",
            "方法正文": "KOM-Rad 使用 OAKNet 影像分级模块提供 KL 严重度和不确定性信号。主模型为带 evidential uncertainty head 的 ConvNeXt-B，DenseNet-121 作为比较模型。当前 figure master 和 OAKNet 审计包记录了 12 个模型/消融 arm、5-fold 训练历史、epoch 0-59 训练过程，以及 validation/internal/external 性能汇总。可追溯的预测 CSV 和图源数据用于重绘混淆矩阵、ROC、PR、校准曲线、DCA 和 risk-coverage 图。",
            "结果正文": "主影像模型的可写结果为 QWK 0.806±0.008、balanced accuracy 0.659、macro-F1 0.664、MAE 0.417、ECE 0.119，80% coverage 下 selective accuracy 为 0.725。OAKNet 审计包已找到训练代码、raw predictions、existing figures 和 recomputed figures；但未找到模型 checkpoint 文件。",
            "证据": "SUM_165_07_KOM_RAD_OAKNET；OAKNet源补写；OAKNet_Reproducibility_Audit_202606",
            "边界": "性能、预测 CSV 和重绘图源可写；checkpoint-dependent Grad-CAM regeneration 和 new-case inference 不能写成已重现。",
        },
        {
            "编号": "S4",
            "标题": "KOM-Risk 预后风险模型",
            "方法正文": "KOM-Risk 针对三个受试者级终点建模：结构性 KL 进展、TKR/膝关节手术事件、症状或功能恶化。为了避免同一受试者跨训练、验证和测试集泄漏，模型使用 person-level split。候选模型包括 elastic-net logistic regression、random forest、XGBoost、LightGBM 和 CatBoost。locked retrain package 中所有随机种子固定为 20260610。三个终点的最终 best_model 均为 CatBoost，参数为 iterations=160、learning_rate=0.05、depth=4、loss_function=Logloss、auto_class_weights=Balanced、eval_metric=AUC、random_seed=20260610。locked package 保存了预处理管线、特征名、预测和指标。",
            "结果正文": "Endpoint A（KL structural progression）AUROC 为 0.781，AUPRC 为 0.349，Brier 为 0.191，balanced accuracy 为 0.715；bootstrap AUROC mean 为 0.782，95% CI 0.742-0.817。Endpoint B（TKR/knee surgery event）AUROC 为 0.868，AUPRC 为 0.362，Brier 为 0.128，balanced accuracy 为 0.780；bootstrap AUROC mean 为 0.868，95% CI 0.829-0.902。Endpoint C（symptom/function worsening）AUROC 为 0.685，AUPRC 为 0.348，Brier 为 0.222，balanced accuracy 为 0.609；bootstrap AUROC mean 为 0.686，95% CI 0.653-0.720。",
            "证据": "KOMRisk超参数已补；KOMRisk随机种子模型卡；RISK 07_metrics；03_splits",
            "边界": "核心训练参数、seed、split 和指标已补齐；更细库默认参数见完整 JSON。",
        },
        {
            "编号": "S5",
            "标题": "SHAP 与特征溯源",
            "方法正文": "为了保证风险模型解释可审计，项目将 locked-model SHAP values 与 feature-to-source lineage 表配对保存。每个解释性特征均追溯到原始 OAI/source 变量、编码变量名和最终模型特征名。该设计使得模型解释不只停留在特征重要性排序，而能回到数据源、变量转换和图件来源。",
            "结果正文": "SHAP/feature lineage 的三个 checkpoint 均通过：locked-model SHAP values、SHAP feature-to-source mapping 和 final SHAP figure validity 均存在。该结果支持在补充材料中将 SHAP 写成模型归因和特征溯源证据。",
            "证据": "RESULT_CHECKPOINTS E01-E03；SHAP / feature lineage sheets",
            "边界": "SHAP 只能写成 model attribution，不能写成因果机制。",
        },
        {
            "编号": "S6",
            "标题": "KOM-KB 证据数据库",
            "方法正文": "KOM-KB 是证据检索和治疗生成的底层证据库。数据库构建记录 PubMed/search lifecycle、证据单元 schema、source identifiers 和 L1-L7 证据等级。证据单元被组织为可以被专科模块路由的结构：包括 guideline anchors、systematic reviews、RCT 和内部 study-level hits。",
            "结果正文": "KOM-KB 的 PubMed/search lifecycle audit、evidence-unit database/schema 和 L1-L7 evidence counts checkpoint 均通过。RAG 检索子集加载 evidence_loaded=3260。写作时需区分证据库总 evidence units 与 RAG loaded subset，避免不同分母混用。",
            "证据": "RESULT_CHECKPOINTS F01-F03；KOM-KB evidence/source sheets",
            "边界": "3260 是 RAG loaded subset，不一定等于数据库总证据单元数。",
        },
        {
            "编号": "S7",
            "标题": "KOM-RAG / GraphRAG 检索与生成质量审计",
            "方法正文": "KOM-RAG 使用 BAAI/bge-m3 生成 1024 维向量表示。检索基准包含 480 条查询，其中 development 320 条、holdout 160 条。检索管线加载 3260 条证据，candidate_k=180，RRF k=30，并比较 GraphRAG 与 naive RAG。生成端输出被归档在 generation_out，并通过多 LLM judge pipeline 评价 appropriateness、completeness、safety、personalization、actionability 和 overall 六个维度。该 judge pipeline 使用 6 个评委模型、9 个 arm、每评委 3 次重复评分。",
            "结果正文": "GraphRAG 在检索指标上优于 naive RAG：P@10 为 0.6763 vs 0.3025，Hit@10 为 1.0000 vs 0.6875，MRR 为 0.7483 vs 0.1588，nDCG@10 为 0.6902 vs 0.2367。生成质量 tiny/pilot 审计包含 2 个病例、9 个 arm、6 个 judge、3 次重复，共 330 条 raw judge calls 和 108 条 aggregated judge rows。overall 排名中，单智能体强基线 overall 5.750±0.878，完整系统_优化版 overall 5.694±0.979，无结构化 overall 5.639±0.784；无RAG最低，overall 3.694±0.771。评委一致性方面，overall ICC(2,k)=0.7653（95% CI 0.4643-0.9078），actionability ICC(2,k)=0.9147（95% CI 0.7817-0.9681）。引用标记代理审计显示 18 个 final outputs 中 explicit source-id/URL/DOI/PMID marker 为 0/18，broad guideline/citation marker 为 1/18。",
            "证据": "SUM_168_10_KOM_RAG_GRAPHRA；RAG生成多评委评分已补；KOMRAG生成引用审计；judge_out；generation_out",
            "边界": "多评委结果是 tiny/pilot generation-quality audit，不是 full-scale claim-level faithfulness/citation-support/unsupported-claim rate。",
        },
        {
            "编号": "S8",
            "标题": "KOM-MDT / KOM-Treat 多专科治疗建议系统",
            "方法正文": "KOM-Treat 由 R0-R8 多专科智能体组成。R0 完成病例标准化和缺失信息门控；R1-R4 生成药物、手术、运动、营养、心理/行为等专科草案；R5-R8 完成自审、中心审计、交叉复核和最终共识。消融实验在相同病例、相同 schema 和相同评分流程下比较 Full KOM、without RAG、without MDT 和 direct LLM。RAW_Treat_AblationScores 包含 2880 条 ok 评分记录，覆盖 4 个 arm 和 2 个 evaluator models。配对统计按 case_id、quadrant、sample 和 evaluator_model 匹配，计算 paired mean difference、95% CI、Cohen dz 和 Benjamini-Hochberg FDR。",
            "结果正文": "SUM_174 locked headline score 显示：Full KOM overall quality 84.6，safety 91.1，corrected rule score 84.3，safety-critical error 0；without RAG 为 65.6/79.9/71.6；without MDT 为 64.4/79.1/81.7；direct LLM 为 54.7/70.7/44.8。RAW evaluator overall_0_100 口径下，Full KOM vs without RAG 为 72.3 vs 67.8，benefit 4.5（95% CI 4.0-5.0），dz=0.65，q=4.8e-67，n=707；Full KOM vs without MDT 为 72.3 vs 70.0，benefit 2.3（95% CI 1.8-2.8），dz=0.34，q=1.1e-19，n=702；Full KOM vs direct LLM 为 72.3 vs 62.4，benefit 9.9（95% CI 9.3-10.5），dz=1.21，q=3.2e-224，n=706。",
            "证据": "SUM_174_16_KOM_TREAT_ABLATIO；RAW_Treat_AblationScores；KOMTreat配对统计已补；KOMTreat口径一致性检查",
            "边界": "84.6 是 locked headline score，72.3 是 raw evaluator paired-stat endpoint，二者不能混写。",
        },
        {
            "编号": "S9",
            "标题": "KOM-Safe 安全审计",
            "方法正文": "KOM-Safe 对治疗建议执行规则化安全审计和人工可读审计。核心门控包括 missing-information-first、口服 NSAID 前的 renal/eGFR、GI bleeding、anticoagulant/current meds 和 cardiovascular risk 检查、注射边界、手术转诊边界、运动 stop rules，以及 safety-critical 输出的人审和修复路径。",
            "结果正文": "Safety-rule library、PASS/WARN/FAIL safety records 和 safety-critical repair/human-review checkpoint 均已 source confirmed。SUM_169 中记录了 missing information、oral NSAID、injection、surgery escalation 和 exercise boundary 等明确安全门控。",
            "证据": "SUM_169_11_KOM_MDT_RX_SAFE；RESULT_CHECKPOINTS I01-I03",
            "边界": "只能写 recommendation safety/audit behavior，不能写真实患者安全结局。",
        },
        {
            "编号": "S10",
            "标题": "KOM-Score 评价体系",
            "方法正文": "KOM-Score 整合专家原始评分、规则评分、可靠性分析、错误分类、agreement/kappa 评价和训练/仲裁材料，为治疗建议质量、安全性、可读性、证据可追溯性和错误类型提供统一评价框架。",
            "结果正文": "Expert raw scores、expert reliability、rule score outputs、error taxonomy/agreement 和 arbitration materials 均通过 checkpoint review。KOM-Score 因此可以作为 KOM-Treat 消融和 KOM-Sim 人机交互评价的共同评分基础。",
            "证据": "SUM_170_12_KOM_SCORE_EVAL；RESULT_CHECKPOINTS J01-J05",
            "边界": "KOM-Score 衡量处方/建议质量，不是患者疗效终点。",
        },
        {
            "编号": "S11",
            "标题": "KOM-Sim 医生交互模拟",
            "方法正文": "KOM-Sim 评估医生端交互场景下系统对处方质量和工作流程的影响。实验包含 26 名 clinicians 和 30 个 tasks，形成 780 条 final locked records。比较条件包括 clinician alone、clinician + KOM、clinician + KOM-R 和 KOM standalone。任务分配、最终记录过滤、primary/secondary comparisons 和 experience/LOWESS analysis 均已归档。",
            "结果正文": "Clinician alone 的 overall quality 为 48.7，rule score 30.1，safety-critical errors 为 19.7 per 100，high-quality rate 为 4.6%。Clinician + KOM 的 overall quality 提高至 73.4，rule score 63.7，safety-critical errors 降至 8.8 per 100，high-quality rate 为 51.5%。Clinician + KOM-R 的 overall quality 为 70.1，rule score 61.0，safety-critical errors 为 14.5 per 100。KOM standalone 的 overall quality 为 84.6，rule score 70，safety-critical error 为 0。",
            "证据": "SUM_173_15_KOM_SIM_CLINICIAN；SIM_final_record_filter_log；RESULT_CHECKPOINTS K01-K06",
            "边界": "这是 simulated clinician interaction / prescription-quality evaluation，不是真实世界患者疗效研究。",
        },
        {
            "编号": "S12",
            "标题": "主图、补图和源数据审计",
            "方法正文": "所有论文关键数值和图源数据被索引到 master workbook。图件包记录 main/figure source-data manifest、figure denominators、supplementary method figures，以及 no bitmap-only figure rule。每个图件应优先由源表、预测 CSV、审计记录或 locked package 支撑。",
            "结果正文": "Main figure/source-data checkpoint 均通过。最新工作簿包含病例、画像、影像、风险、RAG、治疗、安全、评分、医生交互和最终性能汇总等模块源表，并通过 numeric traceability 记录 manuscript-critical values。",
            "证据": "RESULT_CHECKPOINTS L01-L04；SUM_175_17_FINAL_MODEL_PERFO；SUM_181_23_NUMERIC_TRACEABIL；figure master",
            "边界": "依赖未恢复 checkpoint 或未评分 generation faithfulness 的图件应标 pending 或从最终 claim 中移除。",
        },
    ]


def write_markdown(path, sections):
    lines = [
        "# KOM 膝骨关节炎医生端决策支持系统：补充方法学与对应结果完整正文",
        "",
        "## 补充方法学总览",
        "",
        "KOM 被组织为一个模块化的医生端膝骨关节炎决策支持系统。补充方法学按照病例构建、患者画像、影像分级、风险预测、特征解释、证据数据库、GraphRAG 检索、MDT 治疗生成、安全审计、评分体系、医生交互评价和图源数据审计的顺序展开。每一个方法模块后均给出对应结果，使方法中承诺的设计、训练、审计和评价均能回到总表、locked package 或本地审计文件。",
        "",
    ]
    for s in sections:
        lines.extend([
            f"## {s['编号']}. {s['标题']}",
            "",
            "### 方法",
            "",
            s["方法正文"],
            "",
            "### 对应结果",
            "",
            s["结果正文"],
            "",
            "### 证据与边界",
            "",
            f"证据：{s['证据']}。边界：{s['边界']}",
            "",
        ])
    lines.extend([
        "## 总体写作边界",
        "",
        "当前补充方法学可以直接作为 Supplementary Materials 初稿使用。可直接写入的模块包括标准化病例、KOM-Profile 锁定汇总、OAKNet 性能与图源、KOM-Risk locked training 和指标、SHAP/feature lineage、KOM-KB、GraphRAG 检索、KOM-Treat 消融、KOM-Safe、KOM-Score、KOM-Sim 和主图源数据。仍需保留的边界为：RAG generation faithfulness/citation-support/unsupported-claim rate 需要 claim-level annotation；OAKNet checkpoint-dependent Grad-CAM 和 new-case inference 需要 checkpoint 恢复或重训；KOM-Treat 的 locked headline score 与 raw evaluator paired-stat endpoint 必须分开命名。",
        "",
    ])
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def main():
    sections = build_sections()
    wb = load_workbook(WB_PATH)
    sheet_name = "补充材料完整正文中文稿"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 2)
    headers = ["编号", "标题", "方法正文", "结果正文", "证据", "边界"]
    write_table(ws, sections, headers)
    style(ws)
    wb.save(WB_PATH)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M")
    md_path = OUT_DIR / f"KOM_补充材料_Methods_Results_完整正文_中文正式稿_{ts}.md"
    write_markdown(md_path, sections)
    print(md_path)


if __name__ == "__main__":
    main()
